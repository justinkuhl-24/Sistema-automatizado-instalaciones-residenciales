# ============================================================
# INFORME ELÉCTRICO - genera el Excel con circuitos y materiales
# ============================================================
# Este script le pregunta al usuario los datos de una instalación
# eléctrica (empalme, tablero, circuitos, climatización, agua caliente...)
# y al final arma un Excel con 3 hojas:
#   - Informe: resumen de la instalación y de cada circuito
#   - Materiales: la lista completa de materiales a comprar
#   - Base Normativa: qué artículo del RIC justifica cada material
#
# Cómo está armado el archivo, de arriba hacia abajo:
#   1) Tabla con la normativa RIC (BLOQUES_NORMATIVA) y la función que
#      arma la hoja "Base Normativa" con los links.
#   2) Funciones de cálculo (conductores, conduit, alimentador, acometida,
#      ferrules, etc.) sacadas de las tablas del RIC.
#   3) Funciones para preguntar datos por consola (pedir_float_positivo...).
#   4) build_materiales_df(): la función más grande de todas. Recorre los
#      circuitos y arma, línea por línea, todo el listado de materiales.
#   5) Al final, el "script" en sí: acá ya no hay funciones, es el código
#      que corre de corrido preguntándole cosas al usuario por consola
#      y al final llama a build_materiales_df y escribe el Excel.
#
# Si es primera vez que lo lees: las funciones de cálculo del principio
# son fáciles de entender solas (reciben datos, devuelven un resultado).
# build_materiales_df es la más grande y compleja, por eso está separada
# en bloques con títulos tipo "# ===== NOMBRE SECCIÓN =====".
# El bloque final es el que realmente conversa con el usuario.
# ============================================================

import pandas as pd
from datetime import datetime
import os, sys
import math
import numpy as np

# =========================================================
# BASE NORMATIVA + HIPERVÍNCULOS PARA HOJA "MATERIALES"
# =========================================================

BASE_NORMATIVA_ROWS = []  # generado dinámicamente desde BLOQUES_NORMATIVA

BLOQUES_NORMATIVA = [
    # ── RIC 1 - EMPALME ──────────────────────────────────────────────────────
    ("Disyuntor termomagnético empalme",             "RIC 1 (5.3, 8.1, 8.2, Anexo 1.3) / RIC 5 (8.5, 8.7.1, 8.7.2, 8.7.3, 8.7.6.3, 8.7.7.4, 8.7.7.5) / RIC 10 (5.2.1)"),
    ("Portafusible de loza empalme",                 "RIC 1 (8.1, 8.2, Anexo 1.3) / RIC 5 (8.7.6.3)"),
    ("Sellador de roscas con teflón",                "RIC 1 (6.1) / RIC 4 (5.18)"),
    ("Abrazaderas tipo caddy",                       "RIC 1 (6.2) / RIC 4 (5.12.4, 5.44, 7.1.3, 7.1.10, 7.1.15, 7.16.2.1, 7.16.2.2, 7.16.2.3, 7.16.2.4, 7.16.2.5, 7.16.4.3, 7.16.4.4)"),
    ("Conector HUB acero galvanizado",               "RIC 1 (6.2) / RIC 2 (6.1.21.6) / RIC 4 (5.6, 5.12.4, 5.14, 5.24, 5.44, 7.1.3, 7.1.10, 7.1.15, 7.16.1.11, 7.16.2.1, 7.16.2.2, 7.16.2.3, 7.16.2.4, 7.16.2.5) / RIC 5 (8.6.4) / RIC 6 (7.2) / RIC 11 (10.3.5)"),
    ("Terminal ferrul acometida",                    "RIC 1 (6.2) / RIC 4 (5.9, 5.11.4)"),
    ("Terminal ferrul alimentador",                  "RIC 1 (6.2) / RIC 4 (5.9, 5.11.4)"),
    ("Terminal ferrul doble alimentador",            "RIC 1 (6.2) / RIC 4 (5.9, 5.11.4)"),
    ("Terminal ferrul tablero interior",             "RIC 4 (5.9, 5.11.4)"),
    ("Terminal de compresión tipo ojo",              "RIC 1 (6.2) / RIC 4 (5.13) / RIC 5 (8.6.4)"),
    ("Cable acometida concéntrico aéreo",            "RIC 1 (6.2) / RIC 3 (5.1.3) / RIC 4 (5.2, 5.5, 6.1.1, 6.1.2, 6.2.1, 6.2.2, 6.2.5, 6.2.6, 6.2.7, Tabla 4.4)"),
    ("Cable acometida concéntrico subterráneo",      "RIC 1 (6.2, 7.20) / RIC 3 (5.1.3) / RIC 4 (5.2, 5.5, 6.1.1, 6.1.2, 6.2.1, 6.2.2, 6.2.5, 6.2.6, 6.2.7, Tabla 4.4)"),
    ("Mordaza para alimentador aéreo",               "RIC 1 (6.2)"),
    ("Cabeza de servicio",                           "RIC 1 (6.2)"),
    ("Caja de empalme metálica",                     "RIC 1 (6.3, 6.5, Anexo 1.1) / RIC 2 (6.1.21.3) / RIC 4 (5.38, 7.16.1.11) / RIC 5 (8.6.4) / RIC 6 (7.2) / RIC 10 (5.2.4) / RIC 11 (10.3.5)"),
    ("Abrazadera conduit PVC acometida",             "RIC 1 (6.3) / RIC 4 (5.2, 5.5, 5.6, 5.12.4, 5.44, 7.1.3, 7.1.8, 7.1.9, 7.1.10, 7.1.15, 7.16.3.1, 7.16.3.2, 7.16.3.3, 7.16.3.5, 7.16.3.6, 7.16.4.2, 7.16.4.3, 7.16.4.5)"),
    ("Terminal PVC conduit con 2 tuercas",           "RIC 1 (6.3) / RIC 2 (6.1.21.6) / RIC 4 (5.14, 5.24, 7.16.1.11)"),
    ("Unidad de medida monofásica",                  "RIC 1 (6.6, 7.1, 7.2, 7.3, 7.5) / RIC 10 (5.2.5)"),
    ("Conduit PVC alimentador",                      "RIC 1 (7.15, 7.19) / RIC 4 (5.2, 5.5, 5.6, 5.12.4, 5.34, 5.42, 5.43, 5.44, 5.46, 7.1.3, 7.1.8, 7.1.9, 7.1.10, 7.1.15, 7.15.1.1, 7.15.1.2, 7.16.1.2, 7.16.1.3, 7.16.1.4, 7.16.1.5, 7.16.1.6, 7.16.1.7, 7.16.1.9, 7.16.1.10, 7.16.3.1, 7.16.3.2, 7.16.3.3, 7.16.3.5, 7.16.3.6, 7.16.4.2, Tabla 4.17, Tabla 4.18, Tabla 4.19, Tabla 4.20, Tabla 4.23)"),
    ("Abrazadera conduit PVC alimentador",           "RIC 1 (7.15, 7.19) / RIC 4 (5.2, 5.5, 5.6, 5.12.4, 5.44, 7.1.3, 7.1.8, 7.1.9, 7.1.10, 7.1.15, 7.16.3.1, 7.16.3.2, 7.16.3.3, 7.16.3.5, 7.16.3.6, 7.16.4.2, 7.16.4.3, 7.16.4.5)"),
    ("Conduit PVC acometida subterránea",            "RIC 1 (7.20) / RIC 4 (5.2, 5.5, 5.6, 5.12.4, 5.34, 5.42, 5.43, 5.44, 5.46, 7.1.3, 7.1.8, 7.1.9, 7.1.10, 7.1.15, 7.9.1, 7.9.2, 7.9.3, 7.9.4, 7.9.5, 7.9.6, 7.9.7.1, 7.9.7.2, 7.9.7.3, 7.9.7.4, 7.9.7.5, 7.16.7, 7.16.7.1, 7.16.7.2, 7.16.7.3, Tabla 4.29)"),
    # ── RIC 2 - TABLERO ──────────────────────────────────────────────────────
    ("Tablero embutido de PVC",                      "RIC 2 (5.2, 5.3, 5.3.6, 6.1.1, 6.1.3, 6.1.4, 6.1.8, 6.1.9, 6.1.10, 6.1.11, 6.1.12, 6.1.16.3, 6.1.19, 6.1.21.2, 6.1.24) / RIC 4 (5.34, 7.16.1.11) / RIC 10 (5.1.3.1, 5.1.3.2, 5.1.3.3)"),
    ("Tablero sobrepuesto de PVC",                   "RIC 2 (5.2, 5.3, 5.3.6, 6.1.1, 6.1.3, 6.1.4, 6.1.8, 6.1.9, 6.1.10, 6.1.11, 6.1.12, 6.1.16.3, 6.1.19, 6.1.21.2, 6.1.24) / RIC 4 (5.34, 7.16.1.11) / RIC 10 (5.1.3.1, 5.1.3.2, 5.1.3.3)"),
    ("Interruptor diferencial",                      "RIC 2 (5.3.7, 6.2.6) / RIC 5 (5.6, 7.8.2, 8.5, 8.7.1, 8.7.2, 8.7.3, 8.7.6.3, 8.7.7.4) / RIC 6 (7.9) / RIC 7 (7.3.7, 7.4.5) / RIC 10 (5.1.3.5, 5.1.3.6, 5.1.3.7, 5.1.3.8, 5.2.1)"),
    ("Interruptor diferencial 10mA agua caliente",   "RIC N°11 (6, 6.4.3, Tabla Vol.1) / RIC N°07 (7.4.5, 7.6.5.4) — Volumen 1 baño: sensibilidad ≤10mA"),
    ("Tablero desconexión externo agua caliente",    "RIC N°07 (7.2.8, 7.3.3, 7.4.1, 7.4.2) / RIC N°02 (5.2, 6.1) / RIC N°11 (6, Tabla Vol.3) — Fuera Vol.0,1,2 a la vista del equipo"),
    ("TM bipolar desconexión externo agua caliente", "RIC N°07 (7.2.8, 7.3.4, 7.4.2, 7.4.5) — Interruptor de desconexión a la vista del equipo; conductor ≥ I×1,25"),
    ("Bornera PE tablero externo agua caliente",     "RIC N°02 (6.2.7) / RIC N°06 (5.11, 5.14) — Continuidad conductor de protección"),
    ("Prensaestopa tablero externo agua caliente",   "RIC N°04 (5.15, 5.24) / RIC N°07 (5.2.8) — Fijación canalización en tablero"),
    # ── RIC 7 - CLIMATIZACIÓN ────────────────────────────────────────────────
    ("Enchufe climatización",                        "RIC 7 (7.1.1, 7.1.3, 7.3.1, 7.3.2, 7.4.4) — Circuito exclusivo climatización; enchufe coordinado con TM; capacidad nominal 16/20/25/32 A"),
    ("Disyuntor termomagnético climatización",         "RIC 7 (7.1.2, 7.2.8, 7.3.3, 7.3.4, 7.3.6, 7.4.5) — Circuito exclusivo; conductor ≥ I×1,25; mín 2,5 mm²; TM coordinado con conductor; diferencial exclusivo"),
    ("Interruptor diferencial climatización",        "RIC 7 (7.1.2, 7.3.7, 7.4.5, 7.6.5.4) — Diferencial exclusivo por circuito de climatización; ≤ 30 mA"),
    ("Canalización climatización",                   "RIC 7 (7.5.1, 7.5.2) / RIC 4 (5.2, 5.5, 5.6, 5.34, 5.44, 7.1.3, 7.1.8, 7.1.9, 7.1.10, 7.1.15) — Canalización según RIC 4 sec.7; corrección temperatura si T>30°C"),
    ("Conductor climatización",                      "RIC 7 (7.3.4, 7.5.1, 7.5.2) / RIC 3 (5.1.3) / RIC 4 (5.2, 5.4, 5.5, 5.34, 6.2.5) — Conductor ≥ I×1,25; mín 2,5 mm²; factor corrección si T>30°C"),
    ("Interruptor general omnipolar",                "RIC 2 (5.3.7, 6.5.3, 6.6.1, 6.6.2) / RIC 3 (5.2.5) / RIC 5 (8.5, 8.7.1, 8.7.2) / RIC 6 (6.5) / RIC 10 (5.1.3.3, 5.2.1)"),
    ("Supresor de transiente (SPD)",                 "RIC 5 (8.7.7) / RIC 6 (6.2.2)"),
    ("Protector sobrevoltaje y corriente",            "RIC 6 (6.6.2)"),
    ("Barra repartidora bipolar",                     "RIC 2 (6.2.1, 6.2.4, 6.2.7)"),
    ("Disyuntor termomagnético",                       "RIC 2 (5.3.7) / RIC 5 (8.5, 8.7.1, 8.7.2, 8.7.3, 8.7.6.3, 8.7.7.4, 8.7.7.5) / RIC 7 (7.2.8, 7.3.1, 7.3.6, 7.4.2, 7.4.4, 7.4.5) / RIC 10 (5.1.2.9, 5.1.3.4, 5.1.3.6, 5.1.3.7, 5.1.3.8, 5.1.4.1, 5.1.4.2, 5.2.1)"),
    ("Riel DIN",                                     "RIC 2 (6.1.15, 6.1.23)"),
    ("Salida de caja conduit de PVC",                "RIC 2 (6.1.21.6) / RIC 4 (5.2, 5.5, 5.6, 5.12.4, 5.14, 5.24, 5.44, 7.1.3, 7.1.8, 7.1.9, 7.1.10, 7.1.15, 7.16.1.11)"),
    ("Conductor alimentador llegada tablero",         "RIC 2 (6.2.1) / RIC 10 (5.2.1)"),
    ("Barra repartidora tetrapolar",                 "RIC 2 (6.2.1, 6.2.4, 6.2.7)"),
    ("Conductor interior tablero ≤6mm²",             "RIC 2 (6.2.2, 6.2.11) / RIC 4 (5.2, 5.4, 5.5, 5.32, 5.34, 5.37.4, 6.1.1, 6.1.2, 6.2.1, 6.2.2, 6.2.6, 6.2.7, Tabla 4.4) / RIC 6 (6.6.4)"),
    ("Conductor interior tablero >6mm²",             "RIC 2 (6.2.2, 6.2.11) / RIC 4 (5.2, 5.4, 5.5, 5.8, 5.32, 5.34, 5.37.4, 6.1.1, 6.1.2, 6.2.1, 6.2.2, 6.2.6, 6.2.7, Tabla 4.4) / RIC 6 (6.6.4)"),
    ("Barra unipolar verde PE",                      "RIC 2 (6.2.7) / RIC 6 (5.11, 5.14)"),
    ("Bornera de conexión",                          "RIC 2 (6.2.12)"),
    ("Luz piloto",                                   "RIC 2 (6.2.14, 6.2.15)"),
    ("Portafusible tablero",                         "RIC 2 (6.2.15, 6.3.6) / RIC 5 (8.7.6.3)"),
    ("Fusible cilíndrico tablero",                   "RIC 2 (6.2.15, 6.3.6) / RIC 5 (8.7.6.3, 8.7.7.4, 8.7.7.5)"),
    ("Barra copperweld + conector bronce",           "RIC 2 (6.4.1) / RIC 5 (5.1, 5.6, 6.7.1.2, 8.7.3, 8.7.7.1) / RIC 6 (5.4, 5.5, 5.6, 6.1, 6.4, 7.1, 7.3, 7.8, 7.9, 8.1, 8.3.1, 8.3.2, 8.5, 8.11, 8.13, 8.14, 11.1, 11.2, 11.3, 11.4, 12.1, 12.2, 12.3, Tabla 6.2, Tabla 6.3, Tabla 6.4) / RIC 10 (5.2.1) / RIC 11 (10.3.5)"),
    ("Camarilla PVC naranjo",                        "RIC 2 (6.4.1) / RIC 6 (5.4, 5.6, 5.15, 6.4, 7.8, 7.9, 8.2) / RIC 10 (5.2.1) / RIC 11 (10.3.5)"),
    ("Conductor THWN-2 verde PT ≤6mm²",               "RIC 2 (6.4.1) / RIC 3 (5.2.3) / RIC 4 (5.2, 5.4, 5.5, 5.32, 5.34, 5.37.4, 6.1.1, 6.1.2, 6.2.1, 6.2.2, 6.2.6, 6.2.7, Tabla 4.4) / RIC 5 (5.1, 5.6, 6.7.1.2, 8.7.3, 8.7.7.1) / RIC 6 (5.4, 5.6, 6.1, 6.4, 6.6.1, 6.6.2, 6.6.3, 7.3, 7.5, 7.6, 7.7, 7.8, 7.9, 11.1, 11.2, 11.3, 11.4, 12.1, 12.2, 12.3) / RIC 10 (5.2.1) / RIC 11 (10.3.5)"),
    ("Conductor THWN-2 verde PT >6mm²",               "RIC 2 (6.4.1) / RIC 3 (5.2.3) / RIC 4 (5.2, 5.4, 5.5, 5.8, 5.32, 5.34, 5.37.4, 6.1.1, 6.1.2, 6.2.1, 6.2.2, 6.2.6, 6.2.7, Tabla 4.4) / RIC 5 (5.1, 5.6, 6.7.1.2, 8.7.3, 8.7.7.1) / RIC 6 (5.4, 5.6, 6.1, 6.4, 6.6.1, 6.6.2, 6.6.3, 7.3, 7.5, 7.6, 7.7, 7.8, 7.9, 11.1, 11.2, 11.3, 11.4, 12.1, 12.2, 12.3) / RIC 10 (5.2.1) / RIC 11 (10.3.5)"),
    ("Conductor THWN-2 blanco PT ≤6mm²",              "RIC 4 (5.2, 5.4, 5.5, 5.32, 5.34, 5.37.4, 6.1.1, 6.1.2, 6.2.1, 6.2.2, 6.2.6, 6.2.7, Tabla 4.4) / RIC 5 (6.4.1, 6.4.2, 6.4.3, 6.5.1, 6.5.2, 6.7.1.2) / RIC 6 (5.4, 5.6, 6.1, 6.3, 6.4, 6.5, 6.6.1, 6.6.2, 6.6.3, 6.6.4, 7.7, 7.8, 11.1, 11.2, 11.3, 11.4, 12.1, 12.2, 12.3) / RIC 10 (5.2.1)"),
    ("Conductor THWN-2 blanco PT >6mm²",              "RIC 4 (5.2, 5.4, 5.5, 5.8, 5.32, 5.34, 5.37.4, 6.1.1, 6.1.2, 6.2.1, 6.2.2, 6.2.6, 6.2.7, Tabla 4.4) / RIC 5 (6.4.1, 6.4.2, 6.4.3, 6.5.1, 6.5.2, 6.7.1.2) / RIC 6 (5.4, 5.6, 6.1, 6.3, 6.4, 6.5, 6.6.1, 6.6.2, 6.6.3, 6.6.4, 7.7, 7.8, 11.1, 11.2, 11.3, 11.4, 12.1, 12.2, 12.3) / RIC 10 (5.2.1)"),
    ("Conductor desnudo Cu 16mm²",                   "RIC 2 (6.4.1) / RIC 4 (Tabla 4.3) / RIC 6 (8.3.2, 8.5, 8.9)"),
    # ── RIC 3 - CONDUCTORES ──────────────────────────────────────────────────
    ("Alimentador RV-K Cu ≤6mm²",                   "RIC 3 (5.1.2, 5.1.3, 5.2.2, 5.2.6, 6.1, 6.2, 6.3, 6.4.1) / RIC 4 (5.2, 5.4, 5.5, 5.32, 5.34, 5.37.4, 6.1.1, 6.1.2, 6.2.1, 6.2.2, 6.2.5, 6.2.6, 6.2.7, Tabla 4.4) / RIC 6 (6.6.4)"),
    ("Alimentador RV-K Cu >6mm²",                   "RIC 3 (5.1.2, 5.1.3, 5.2.2, 5.2.6, 6.1, 6.2, 6.3, 6.4.1) / RIC 4 (5.2, 5.4, 5.5, 5.8, 5.32, 5.34, 5.37.4, 6.1.1, 6.1.2, 6.2.1, 6.2.2, 6.2.5, 6.2.6, 6.2.7, Tabla 4.4) / RIC 6 (6.6.4)"),
    ("Conductores circuitos interiores ≤6mm²",      "RIC 3 (5.1.3) / RIC 4 (5.2, 5.4, 5.5, 5.32, 5.34, 5.37.4, 6.1.1, 6.1.2, 6.2.1, 6.2.2, 6.2.5, 6.2.6, 6.2.7, Tabla 4.4) / RIC 6 (6.6.4) / RIC 7 (7.1.2, 7.1.3, 7.3.4, 7.5.1, 7.5.2) / RIC 10 (5.1.1.1, 5.1.1.3, 5.1.3.4, 5.1.3.5, 5.1.4.1, 5.1.5.1, 5.1.5.2, 5.1.5.3) / RIC 11 (10.3.4)"),
    ("Conductores circuitos interiores >6mm²",      "RIC 3 (5.1.3) / RIC 4 (5.2, 5.4, 5.5, 5.8, 5.32, 5.34, 5.37.4, 6.1.1, 6.1.2, 6.2.1, 6.2.2, 6.2.5, 6.2.6, 6.2.7, Tabla 4.4) / RIC 6 (6.6.4) / RIC 7 (7.1.2, 7.1.3, 7.3.4, 7.5.1, 7.5.2) / RIC 10 (5.1.1.1, 5.1.1.3, 5.1.3.4, 5.1.3.5, 5.1.4.1, 5.1.5.1, 5.1.5.2, 5.1.5.3) / RIC 11 (10.3.4)"),
    # ── RIC 4 - CANALIZACIONES Y ACCESORIOS ──────────────────────────────────
    ("Canalización embutida conduit PVC",            "RIC 4 (5.2, 5.5, 5.6, 5.12.4, 5.34, 5.42, 5.43, 5.44, 5.46, 7.1.3, 7.1.8, 7.1.9, 7.1.10, 7.1.15, 7.15.1.1, 7.15.1.2, 7.16.1.2, 7.16.1.3, 7.16.1.4, 7.16.1.5, 7.16.1.6, 7.16.1.7, 7.16.1.9, 7.16.1.10, 7.16.1.16, 7.16.3.1, 7.16.3.2, 7.16.3.3, 7.16.3.5, 7.16.3.6, 7.16.4.2, Tabla 4.17, Tabla 4.18, Tabla 4.19, Tabla 4.20, Tabla 4.23) / RIC 7 (7.2.2, 7.5.1) / RIC 10 (5.1.2.1, 5.1.2.6) / RIC 11 (10.3.3, 10.3.4)"),
    ("Canalización sobrepuesta canaleta PVC",        "RIC 4 (5.2, 5.5, 5.6, 5.12.4, 5.34, 5.42, 5.43, 5.44, 5.46, 7.1.3, 7.1.8, 7.1.9, 7.1.10, 7.1.15, 7.7.1, 7.7.2, 7.7.3, 7.7.4, 7.7.5, 7.7.6, 7.7.7, 7.7.9, 7.16.3.1, 7.16.3.2, 7.16.4.1, 7.16.4.6, 7.16.4.7, 7.16.4.8, 7.16.4.9, Tabla 4.23) / RIC 7 (7.2.2, 7.5.1) / RIC 10 (5.1.2.1, 5.1.2.6) / RIC 11 (10.3.3, 10.3.4)"),
    ("Abrazadera conduit PVC circuitos",             "RIC 4 (5.2, 5.5, 5.6, 5.12.4, 5.44, 7.1.3, 7.1.8, 7.1.9, 7.1.10, 7.1.15, 7.16.3.1, 7.16.3.2, 7.16.3.3, 7.16.3.5, 7.16.3.6, 7.16.4.2, 7.16.4.3, 7.16.4.5) / RIC 10 (5.1.2.1, 5.1.2.6) / RIC 11 (10.3.3, 10.3.4)"),
    ("Boquilla conduit PVC",                         "RIC 4 (5.2, 5.5, 5.6, 5.12.4, 5.44, 7.1.3, 7.1.8, 7.1.9, 7.1.10, 7.1.15, 7.16.3.1, 7.16.3.2, 7.16.3.3, 7.16.3.5, 7.16.3.6, 7.16.4.2) / RIC 10 (5.1.2.1, 5.1.2.6) / RIC 11 (10.3.3, 10.3.4)"),
    ("Unión copla canaleta PVC",                     "RIC 4 (5.2, 5.5, 5.6, 5.12.4, 5.44, 7.1.3, 7.1.8, 7.1.9, 7.1.10, 7.1.15, 7.7.1, 7.7.2, 7.7.3, 7.7.4, 7.7.5, 7.7.6, 7.7.7, 7.7.9, 7.16.3.1, 7.16.3.2, 7.16.4.1, 7.16.4.6, 7.16.4.7, 7.16.4.8, 7.16.4.9, Tabla 4.23) / RIC 10 (5.1.2.1, 5.1.2.6) / RIC 11 (10.3.3, 10.3.4)"),
    ("Curvas internas 90° canaleta PVC",             "RIC 4 (5.2, 5.5, 5.6, 5.12.4, 5.44, 7.1.3, 7.1.8, 7.1.9, 7.1.10, 7.1.15, 7.7.1, 7.7.2, 7.7.3, 7.7.4, 7.7.5, 7.7.6, 7.7.7, 7.7.9, 7.16.3.1, 7.16.3.2, 7.16.4.1, 7.16.4.6, 7.16.4.7, 7.16.4.8, 7.16.4.9, Tabla 4.23) / RIC 10 (5.1.2.1, 5.1.2.6) / RIC 11 (10.3.3, 10.3.4)"),
    ("Curvas planas 90° canaleta PVC",               "RIC 4 (5.2, 5.5, 5.6, 5.12.4, 5.44, 7.1.3, 7.1.8, 7.1.9, 7.1.10, 7.1.15, 7.7.1, 7.7.2, 7.7.3, 7.7.4, 7.7.5, 7.7.6, 7.7.7, 7.7.9, 7.16.3.1, 7.16.3.2, 7.16.4.1, 7.16.4.6, 7.16.4.7, 7.16.4.8, 7.16.4.9, Tabla 4.23) / RIC 10 (5.1.2.1, 5.1.2.6) / RIC 11 (10.3.3, 10.3.4)"),
    ("Curva T canaleta PVC",                         "RIC 4 (5.2, 5.5, 5.6, 5.12.4, 5.44, 7.1.3, 7.1.8, 7.1.9, 7.1.10, 7.1.15, 7.7.1, 7.7.2, 7.7.3, 7.7.4, 7.7.5, 7.7.6, 7.7.7, 7.7.9, 7.16.3.1, 7.16.3.2, 7.16.4.1, 7.16.4.6, 7.16.4.7, 7.16.4.8, 7.16.4.9, Tabla 4.23) / RIC 10 (5.1.2.1, 5.1.2.6) / RIC 11 (10.3.3, 10.3.4)"),
    ("Caja derivación embutida PVC",                 "RIC 4 (5.2, 5.5, 5.6, 5.12.1, 5.12.3, 5.12.4, 5.14, 5.16, 5.21, 5.23, 5.27, 5.34, 5.45, 5.48.1, 7.16.1.11, 7.16.1.13) / RIC 10 (5.1.2.2, 5.1.2.3, 5.1.2.11, 5.1.2.12, 5.1.2.14, 5.2.2) / RIC 11 (10.3.4)"),
    ("Caja de paso estanca",                         "RIC 4 (5.2, 5.5, 5.6, 5.12.1, 5.12.3, 5.12.4, 5.14, 5.16, 5.21, 5.23, 5.27, 5.34, 5.45, 5.48.1, 7.16.1.11, 7.16.1.13) / RIC 10 (5.1.2.2, 5.1.2.3, 5.1.2.11, 5.1.2.12, 5.1.2.14, 5.2.2) / RIC 11 (10.3.4)"),
    ("Caja derivación sobrepuesta PVC",              "RIC 4 (5.2, 5.5, 5.6, 5.12.1, 5.12.3, 5.12.4, 5.14, 5.16, 5.21, 5.23, 5.27, 5.34, 5.45, 5.48.1, 7.16.1.11) / RIC 10 (5.1.2.2, 5.1.2.3, 5.1.2.14, 5.2.2) / RIC 11 (10.3.4)"),
    ("Caja derivación sobrepuesta chuqui PVC",       "RIC 4 (5.2, 5.5, 5.6, 5.12.1, 5.12.3, 5.12.4, 5.14, 5.16, 5.21, 5.23, 5.27, 5.34, 5.45, 5.48.1, 7.16.1.11) / RIC 10 (5.1.2.2, 5.1.2.3, 5.1.2.14, 5.2.2) / RIC 11 (10.3.4)"),
    ("Caja derivación octogonal PVC",                "RIC 4 (5.2, 5.5, 5.6, 5.12.1, 5.12.3, 5.12.4, 5.14, 5.16, 5.21, 5.23, 5.27, 5.34, 5.45, 5.48.1, 7.16.1.11) / RIC 10 (5.1.2.2, 5.1.2.3, 5.1.2.11, 5.1.2.12, 5.1.2.13, 5.1.2.14, 5.2.2) / RIC 11 (10.3.4)"),
    ("Tapa ciega octogonal PVC",                     "RIC 4 (5.2, 5.5, 5.6, 5.12.4, 5.17, 5.22)"),
    ("Tapa ciega PVC",                               "RIC 4 (5.2, 5.5, 5.6, 5.12.4, 5.17, 5.22)"),
    ("Tapa ciega chuqui PVC",                        "RIC 4 (5.2, 5.5, 5.6, 5.12.4, 5.17, 5.22)"),
    ("Tubo conduit galvanizado",                     "RIC 4 (5.6, 5.12.4, 5.38, 5.42, 5.44, 5.46, 7.1.3, 7.1.10, 7.1.15, 7.16.2.1, 7.16.2.2, 7.16.2.3, 7.16.2.4, 7.16.2.5, Tabla 4.17, Tabla 4.18, Tabla 4.19, Tabla 4.20) / RIC 5 (8.6.4) / RIC 6 (7.2) / RIC 11 (10.3.5)"),
    ("Caja derivación metálica",                     "RIC 4 (5.6, 5.12.2, 5.12.3, 5.12.4, 5.13, 5.14, 5.16, 5.21, 5.23, 5.27, 5.34, 5.38, 5.39, 5.45, 7.16.1.11) / RIC 5 (8.6.4) / RIC 6 (7.2) / RIC 11 (10.3.5)"),
    ("Enchufe",                                       "RIC 4 (5.34, 5.45) / RIC 7 (7.2.1, 7.3.2, 7.4.4) / RIC 10 (5.1.2.7, 5.1.2.8, 5.1.2.9, 5.1.4.6, 5.2.2) / RIC 11 (10.3.4)"),
    ("Interruptor de circuito",                      "RIC 4 (5.34, 5.45) / RIC 10 (5.1.2.4, 5.1.2.9, 5.2.2) / RIC 11 (10.3.4)"),
    ("Portalámpara / luminaria circuito interior",   "RIC 10 (5.1.2.11, 5.1.2.12, 5.1.2.13, 5.1.4.4, 5.1.4.5, 5.2.2) — Centro de iluminación en caja; soporte independiente del conductor; con PT"),
    ("Conector cónico",                              "RIC 4 (5.11.3)"),
    ("Cámara tipo C subterránea",                    "RIC 4 (7.9.5, 7.9.7.8, 7.9.7.9, 7.9.7.10, 7.9.7.12, 7.9.8.1, 7.9.8.2, 7.9.8.4.3, 7.9.8.5, 7.9.8.7)"),
    ("Marco metálico cámara tipo C",                 "RIC 4 (7.9.5, 7.9.7.8, 7.9.7.9, 7.9.7.10, 7.9.7.12, 7.9.8.1, 7.9.8.2, 7.9.8.4.3, 7.9.8.5, 7.9.8.7)"),
    ("Boquilla PVC cámara tipo C",                   "RIC 4 (7.9.8.9)"),
    ("Tornillo conexión tierra caja metálica",       "RIC 4 (5.13)"),
    ("Prensaestopa",                                 "RIC 4 (5.15, 5.24)"),
    ("Tubo de estaño",                               "RIC 4 (4.4.1, 5.11.1)"),
    ("Pasta para soldar",                            "RIC 4 (4.4.1, 5.11.1)"),
]


# Generar BASE_NORMATIVA_ROWS desde BLOQUES_NORMATIVA
# Cada entrada de BLOQUES_NORMATIVA trae 2 datos juntos: el nombre del
# material y su norma correspondiente
for _mat, _norma in BLOQUES_NORMATIVA:
    BASE_NORMATIVA_ROWS.append({
        "Material / Elemento": _mat,
        "Norma": _norma,
    })


def normalizar_ric_materiales(norma, descripcion=""):
    """
    Deja limpia la columna 'Norma / RIC' en la hoja Materiales.
    Ejemplo: 'RIC 4.7.2' se deja como 'RIC 4'.
    Si no detecta RIC, conserva el texto original.
    """
    s = str(norma or "").strip()  # texto original de la norma
    d = str(descripcion or "").lower()  # descripción del material, para adivinar el RIC si hace falta

    if s == "" or s.lower() in ["nan", "none"]:
        return ""  # no hay norma, no hay nada que limpiar

    # Un "-" explícito significa "no aplica norma específica" — se respeta tal
    # cual, sin intentar adivinar un RIC por palabras clave de la descripción.
    if s == "-":
        return "-"

    import re
    m = re.search(r"ric\s*(n°|nº|n|#)?\s*0?(\d+)", s, flags=re.IGNORECASE)  # busca "RIC" seguido de un número
    if m:
        return f"RIC {int(m.group(2))}"  # deja solo "RIC N" (sin puntos ni artículo)

    # Casos escritos como SEC u observaciones generales: si no venía un RIC
    # explícito, se adivina cuál corresponde según palabras clave del texto
    if "empalme" in d or "acometida" in d or "medidor" in d:
        return "RIC 1"
    if "tablero" in d or "riel din" in d or "barra" in d or "luz piloto" in d:
        return "RIC 2"
    if "alimentador" in d:
        return "RIC 3"
    if "conductor" in d or "canaleta" in d or "conduit" in d or "caja" in d or "prensaestopa" in d or "boquilla" in d:
        return "RIC 4"
    if "diferencial" in d or "automático" in d or "termomagnético" in d or "protección" in d or "supresor" in d:
        return "RIC 5"
    if "tierra" in d or "jabalina" in d or "copperweld" in d or "camarilla" in d:
        return "RIC 6"
    if "enchufe" in d or "interruptor" in d or "luminaria" in d or "ampolleta" in d:
        return "RIC 10"

    return s  # no se pudo adivinar nada, se deja el texto tal cual llegó

def aplicar_base_normativa_e_hipervinculos(writer, materiales_df, sheet_materiales="Materiales", sheet_base="Base Normativa"):
    """
    Crea hoja 'Base Normativa' con 2 columnas: Material / Elemento | Norma
    Agrega hipervínculos en columna 'Norma / RIC' de la hoja Materiales.
    """
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb     = writer.book
    ws_mat = writer.sheets[sheet_materiales]

    # ── Crear hoja ────────────────────────────────────────────────────────────
    if sheet_base in wb.sheetnames:
        del wb[sheet_base]  # si ya existía (por una corrida anterior), la borra primero
    ws_base = wb.create_sheet(sheet_base)

    thin        = Side(border_style="thin", color="BFBFBF")
    fill_titulo = PatternFill(start_color="9DC3E6", end_color="9DC3E6", fill_type="solid")
    fill_header = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    fill_blanco = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    borde_gris  = Border(top=thin, left=thin, right=thin, bottom=thin)

    # ── Fila 1: título ────────────────────────────────────────────────────────
    ws_base.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
    tc = ws_base["A1"]
    tc.value     = "BASE NORMATIVA RIC PARA TRAZABILIDAD DE MATERIALES"
    tc.font      = Font(bold=True, size=13, color="000000")
    tc.alignment = Alignment(horizontal="center", vertical="center")
    tc.fill      = fill_titulo
    for c in range(1, 3):
        ws_base.cell(row=1, column=c).fill   = fill_titulo
        ws_base.cell(row=1, column=c).border = borde_gris
    ws_base.row_dimensions[1].height = 22

    # ── Fila 2: encabezados ───────────────────────────────────────────────────
    for j, h in enumerate(["Material / Elemento", "Norma"]):
        cell = ws_base.cell(row=2, column=j+1)
        cell.value     = h
        cell.font      = Font(bold=True, color="000000")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill      = fill_header
        cell.border    = borde_gris
    ws_base.row_dimensions[2].height = 16

    # ── Escribir filas desde BLOQUES_NORMATIVA ────────────────────────────────
    mat_anchor = {}   # relaciona cada material con la fila del Excel donde quedó (para armar los hipervínculos)
    current_row = 3   # las filas 1 y 2 son título y encabezado, los datos parten en la 3

    # recorre toda la tabla normativa y la va escribiendo fila por fila
    for mat, norma in BLOQUES_NORMATIVA:
        mat_anchor[mat.lower()] = current_row  # guarda en qué fila quedó este material (para el link después)
        for j, v in enumerate([mat, norma]):
            cell = ws_base.cell(row=current_row, column=j+1)
            cell.value     = v
            cell.font      = Font(bold=(j==0), color="000000")  # la columna del material va en negrita
            cell.alignment = Alignment(vertical="center", wrap_text=True,
                                       horizontal="left")
            cell.fill      = fill_blanco
            cell.border    = borde_gris
        ws_base.row_dimensions[current_row].height = 50
        current_row += 1  # avanza a la siguiente fila

    # ── Anchos de columna ─────────────────────────────────────────────────────
    ws_base.column_dimensions["A"].width = 45
    ws_base.column_dimensions["B"].width = 115
    ws_base.freeze_panes = "A3"  # deja fijas las filas 1 y 2 cuando se mueve hacia abajo en la pantalla

    # Botón volver
    ws_base["D1"].value     = "← Volver a Materiales"
    ws_base["D1"].hyperlink = f"#'{sheet_materiales}'!A1"  # link interno a la otra hoja
    ws_base["D1"].font      = Font(color="0000FF", underline="single", bold=True)
    ws_base["D1"].alignment = Alignment(horizontal="center")
    ws_base.column_dimensions["D"].width = 22

    # ── Hipervínculos en hoja Materiales ──────────────────────────────────────
    # A partir de la descripción y el circuito, busca a qué fila de la hoja
    # normativa corresponde ese material.
    # Solo devuelve una fila si el material tiene entrada en BLOQUES_NORMATIVA.
    def get_row_from_material(desc, circ=""):
        # busca a qué fila de la hoja "Base Normativa" corresponde este material,
        # para armar el link. Compara palabras clave del texto (tipo de
        # material + tipo de circuito). Si no encuentra nada, devuelve None.
        #
        # De acá para abajo es puro "si el texto dice tal cosa, busca esta norma".
        # Está ordenado por secciones (RIC 1, RIC 2, RIC 4, etc.) y cada línea
        # ya se explica solita: el texto que busca y el nombre de la norma que
        # devuelve están en español. No hay vuelta que darle línea por línea,
        # es una lista larga de casos.
        import re
        d = str(desc).lower()
        c = str(circ).lower()

        def extraer_sec(texto):
            """Extrae sección en mm² desde el descriptor. Retorna float o None."""
            m = re.search(r"(\d+(?:[.,]\d+)?)\s*mm", texto.lower())  # busca un número seguido de "mm"
            if m:
                return float(m.group(1).replace(",", "."))  # lo convierte a número (cambia la coma por punto)
            return None  # no encontró ningún número de sección en el texto

        # ── RIC 7 - Climatización ─────────────────────────────────────────────
        if ("enchufe" in d or "2p+t" in d) and "sin enchufe" not in d and any(k in c for k in ("clima", "split", "aire")):
            return mat_anchor.get("enchufe climatización")
        if ("interruptor automático" in d or "interruptor automatico" in d) and any(k in c for k in ("clima", "split", "aire")):
            return mat_anchor.get("interruptor automático climatización")
        if "interruptor diferencial" in d and any(k in c for k in ("clima", "split", "aire")):
            return mat_anchor.get("interruptor diferencial climatización")
        if ("canalización" in d or "canalizacion" in d or "conduit" in d or "canaleta" in d) and any(k in c for k in ("clima", "split", "aire")):
            return mat_anchor.get("canalización climatización")
        if "conductor" in d and ("rojo =" in d or "blanco =" in d) and any(k in c for k in ("clima", "split", "aire")):
            return mat_anchor.get("conductor climatización")

        # ── RIC 7 - Agua caliente ─────────────────────────────────────────────
        _kw_agua_c = ("ducha", "termo", "calefon", "calefón", "calentador", "agua caliente")
        _es_agua_c = any(k in c for k in _kw_agua_c)
        if "diferencial" in d and "10" in d and "ma" in d.lower() and _es_agua_c:
            return mat_anchor.get("interruptor diferencial 10ma agua caliente")
        if "tablero sobrepuesto" in d and "desconexión" in d.replace("o","o") and _es_agua_c:
            return mat_anchor.get("tablero desconexión externo agua caliente")
        if ("interruptor automático" in d or "1p+n" in d) and "desconexión" in d and _es_agua_c:
            return mat_anchor.get("tm bipolar desconexión externo agua caliente")
        if "bornera" in d and "pe" in d and _es_agua_c:
            return mat_anchor.get("bornera pe tablero externo agua caliente")
        if "prensaestopa" in d and _es_agua_c:
            return mat_anchor.get("prensaestopa tablero externo agua caliente")
        if ("canalización" in d or "canalizacion" in d or "conduit" in d or "canaleta" in d) and _es_agua_c:
            return mat_anchor.get("canalización climatización")
        if "conductor" in d and ("rojo =" in d or "blanco =" in d) and _es_agua_c:
            return mat_anchor.get("conductor climatización")

        if "portalámpara" in d or "portalampara" in d:
            return mat_anchor.get("portalámpara / luminaria circuito interior")
        if any(x in d for x in ["foco", "luminaria", "ampolleta", "panel led", "tubo led",
                                  "tubo fluorescente", "aplique led", "aplique"]):
            return mat_anchor.get("portalámpara / luminaria circuito interior")
        if "supresor" in d or "spd" in d:
            return mat_anchor.get("supresor de transiente (spd)")
        if "protector sobrevoltaje" in d or "protector de sobrevoltaje" in d:
            return mat_anchor.get("protector sobrevoltaje y corriente")
        if "barra repartidora bipolar" in d or "barra bipolar" in d:
            return mat_anchor.get("barra repartidora bipolar")

        # ── RIC 1 ────────────────────────────────────────────────────────────
        if "disyuntor" in d:
            return mat_anchor.get("disyuntor termomagnético empalme")
        if "portafusible de loza" in d:
            return mat_anchor.get("portafusible de loza empalme")
        if "sellador" in d and ("teflón" in d or "teflon" in d):
            return mat_anchor.get("sellador de roscas con teflón")
        if "abrazadera" in d and "caddy" in d:
            return mat_anchor.get("abrazaderas tipo caddy")
        if "conector hub" in d:
            return mat_anchor.get("conector hub acero galvanizado")
        if "terminal ferrul doble" in d:
            return mat_anchor.get("terminal ferrul doble alimentador")
        if "terminal ferrul" in d and "acometida" in d:
            return mat_anchor.get("terminal ferrul acometida")
        if "terminal ferrul" in d and "alimentador" in d:
            return mat_anchor.get("terminal ferrul alimentador")
        if "terminal ferrul" in d:
            return mat_anchor.get("terminal ferrul tablero interior")
        if "terminal" in d and ("compresión" in d or "compresion" in d or " ojo" in d):
            return mat_anchor.get("terminal de compresión tipo ojo")
        if "cable" in d and ("concéntrico" in d or "concentrico" in d):
            if "sub" in d or "subterráneo" in d or "subterraneo" in d:
                return mat_anchor.get("cable acometida concéntrico subterráneo")
            return mat_anchor.get("cable acometida concéntrico aéreo")
        if "mordaza" in d:
            return mat_anchor.get("mordaza para alimentador aéreo")
        if "cabeza de servicio" in d:
            return mat_anchor.get("cabeza de servicio")
        if "caja de empalme" in d:
            return mat_anchor.get("caja de empalme metálica")
        if "abrazadera" in d and "pvc" in d:
            if "acometida" in c:
                return mat_anchor.get("abrazadera conduit pvc acometida")
            if "alimentador" in c:
                return mat_anchor.get("abrazadera conduit pvc alimentador")
            return mat_anchor.get("abrazadera conduit pvc circuitos")
        if "terminal pvc conduit" in d or ("terminal" in d and "pvc" in d and "conduit" in d):
            return mat_anchor.get("terminal pvc conduit con 2 tuercas")
        if "salida de caja" in d:
            return mat_anchor.get("salida de caja conduit de pvc")
        if "unidad de medida" in d:
            return mat_anchor.get("unidad de medida monofásica")
        if ("conduit de pvc" in d or "conduit pvc" in d) and "abrazadera" not in d:
            if "alimentador" in c:
                return mat_anchor.get("conduit pvc alimentador")
            if "acometida" in c:
                return mat_anchor.get("conduit pvc acometida subterránea")
            return None
        if "tubo conduit galvanizado" in d:
            return mat_anchor.get("tubo conduit galvanizado")
        if ("marco metálico" in d or "marco metalico" in d) and ("cámara" in d or "camara" in d):
            return mat_anchor.get("marco metálico cámara tipo c")
        if "boquilla" in d and ("cámara" in d or "camara" in d):
            return mat_anchor.get("boquilla pvc cámara tipo c")
        if "cámara tipo c" in d or "camara tipo c" in d:
            return mat_anchor.get("cámara tipo c subterránea")

        # ── RIC 3 + RIC 4 conductores (con lógica de sección) ────────────────
        if "rv-k" in d and "alimentador" in c:
            sec = extraer_sec(d)
            if sec and sec > 6:
                return mat_anchor.get("alimentador rv-k cu >6mm²")
            return mat_anchor.get("alimentador rv-k cu ≤6mm²")
        if ("flexible libre de halógenos" in d or "flexible libre de halogenos" in d) and "tablero" in c:
            sec = extraer_sec(d)
            if sec and sec > 6:
                return mat_anchor.get("conductor interior tablero >6mm²")
            return mat_anchor.get("conductor interior tablero ≤6mm²")
        if ("flexible libre de halógenos" in d or "flexible libre de halogenos" in d):
            return None
        if ("rojo =" in d or "blanco =" in d) and "conductor" in d:
            sec = extraer_sec(d)
            if sec and sec > 6:
                return mat_anchor.get("conductores circuitos interiores >6mm²")
            return mat_anchor.get("conductores circuitos interiores ≤6mm²")
        if "thwn-2" in d and "verde" in d:
            sec = extraer_sec(d)
            if sec and sec > 6:
                return mat_anchor.get("conductor thwn-2 verde pt >6mm²")
            return mat_anchor.get("conductor thwn-2 verde pt ≤6mm²")
        if "thwn-2" in d and "blanco" in d:
            sec = extraer_sec(d)
            if sec and sec > 6:
                return mat_anchor.get("conductor thwn-2 blanco pt >6mm²")
            return mat_anchor.get("conductor thwn-2 blanco pt ≤6mm²")

        # ── RIC 2 ────────────────────────────────────────────────────────────
        if "tablero embutido" in d:
            return mat_anchor.get("tablero embutido de pvc")
        if "tablero sobrepuesto" in d:
            return mat_anchor.get("tablero sobrepuesto de pvc")
        if "interruptor general omnipolar" in d or "general omnipolar" in d:
            return mat_anchor.get("interruptor general omnipolar")
        if "interruptor diferencial" in d:
            return mat_anchor.get("interruptor diferencial")
        if "interruptor automático" in d or "interruptor automatico" in d:
            return mat_anchor.get("interruptor automático")
        if "luz piloto" in d:
            return mat_anchor.get("luz piloto")
        if "riel din" in d:
            return mat_anchor.get("riel din")
        if "conductor" in d and "alimentador" in d and ("llegada" in d or "tablero" in d):
            return mat_anchor.get("conductor alimentador llegada tablero")
        if ("flexible libre de halógenos" in d or "flexible libre de halogenos" in d) and "tablero" in c:
            sec = extraer_sec(d)
            if sec and sec > 6:
                return mat_anchor.get("conductor interior tablero >6mm²")
            return mat_anchor.get("conductor interior tablero ≤6mm²")
        if "barra repartidora tetrapolar" in d or "barra tetrapolar" in d:
            return mat_anchor.get("barra repartidora tetrapolar")
        if "barra unipolar verde" in d:
            return mat_anchor.get("barra unipolar verde pe")
        if "bornera de conexión" in d or "bornera de conexion" in d:
            return mat_anchor.get("bornera de conexión")
        if "portafusible" in d and ("1p" in d or "32a" in d or "10x38" in d):
            return mat_anchor.get("portafusible tablero")
        if "fusible cilíndrico" in d or ("fusible" in d and "10x38" in d):
            return mat_anchor.get("fusible cilíndrico tablero")
        if "copperweld" in d:
            return mat_anchor.get("barra copperweld + conector bronce")
        if "camarilla" in d and "naranjo" in d:
            return mat_anchor.get("camarilla pvc naranjo")
        if "conductor desnudo cu" in d:
            return mat_anchor.get("conductor desnudo cu 16mm²")

        # ── RIC 4 canalizaciones y accesorios ────────────────────────────────
        if "canalización embutida" in d or "canalizacion embutida" in d:
            return mat_anchor.get("canalización embutida conduit pvc")
        if "canalización sobrepuesta" in d or "canalizacion sobrepuesta" in d:
            return mat_anchor.get("canalización sobrepuesta canaleta pvc")
        if "boquilla" in d and "conduit" in d:
            return mat_anchor.get("boquilla conduit pvc")
        if "unión copla" in d or "union copla" in d:
            return mat_anchor.get("unión copla canaleta pvc")
        if "curvas internas" in d:
            return mat_anchor.get("curvas internas 90° canaleta pvc")
        if "curvas planas" in d:
            return mat_anchor.get("curvas planas 90° canaleta pvc")
        if "curva t" in d:
            return mat_anchor.get("curva t canaleta pvc")
        if "caja de derivación embutida" in d or "caja de derivacion embutida" in d:
            if "octogon" in d:
                return mat_anchor.get("caja derivación octogonal pvc")
            return mat_anchor.get("caja derivación embutida pvc")
        if "caja de paso" in d:
            # Cajas de paso (interior o alimentador/estanca) — entrada propia
            # en el catálogo, con la misma cita RIC 7.16.1.13 que la caja
            # embutida general.
            return mat_anchor.get("caja de paso estanca")
        if "caja de derivación sobrepuesta" in d or "caja de derivacion sobrepuesta" in d:
            if "chuqui" in d:
                return mat_anchor.get("caja derivación sobrepuesta chuqui pvc")
            return mat_anchor.get("caja derivación sobrepuesta pvc")
        if "caja de derivación metálica" in d or "caja de derivacion metalica" in d:
            return mat_anchor.get("caja derivación metálica")
        if "tapa ciega" in d:
            if "octogon" in d:
                return mat_anchor.get("tapa ciega octogonal pvc")
            if "chuqui" in d:
                return mat_anchor.get("tapa ciega chuqui pvc")
            return mat_anchor.get("tapa ciega pvc")
        if ("enchufe" in d or "2p+t" in d) and "sin enchufe" not in d:
            return mat_anchor.get("enchufe")
        if "interruptor 9/" in d:
            return mat_anchor.get("interruptor de circuito")
        if "conector cónico" in d or "conector conico" in d:
            return mat_anchor.get("conector cónico")
        if "tierra" in d and ("caja metálica" in d or "caja metalica" in d):
            return mat_anchor.get("tornillo conexión tierra caja metálica")
        if "prensaestopa" in d:
            return mat_anchor.get("prensaestopa")
        if "tubo de estaño" in d or "tubo de estano" in d:
            return mat_anchor.get("tubo de estaño")
        if "pasta para soldar" in d:
            return mat_anchor.get("pasta para soldar")

        # Si no hay entrada en BLOQUES_NORMATIVA, no lleva hipervínculo
        return None

    # busca en qué columna de la hoja Materiales está cada dato que necesitamos
    try:
        col_desc = list(materiales_df.columns).index("Descripción técnica") + 1
    except ValueError:
        col_desc = None

    try:
        col_norma = list(materiales_df.columns).index("Norma / RIC") + 1
    except ValueError:
        col_norma = None

    try:
        col_sello = list(materiales_df.columns).index("Sello SEC") + 1
    except ValueError:
        col_sello = None

    try:
        col_circ = list(materiales_df.columns).index("Circuito") + 1
    except ValueError:
        col_circ = None

    if col_norma and col_desc:
        header_row_mat = 3  # en la hoja Materiales, los datos empiezan en la fila 3
        last_row_mat   = header_row_mat + len(materiales_df)
        # recorre cada fila de materiales y le pone el link si corresponde
        for r in range(header_row_mat + 1, last_row_mat + 1):
            val_desc   = str(ws_mat.cell(row=r, column=col_desc).value or "")
            val_circ   = str(ws_mat.cell(row=r, column=col_circ).value or "") if col_circ else ""
            cell_norma = ws_mat.cell(row=r, column=col_norma)
            destino    = get_row_from_material(val_desc, val_circ)  # busca a qué fila de la norma corresponde
            if destino:
                cell_norma.value     = "Ver normativa"
                cell_norma.hyperlink = f"#'{sheet_base}'!A{destino}"  # link a la hoja Base Normativa
                cell_norma.font      = Font(color="0000FF", underline="single")
                cell_norma.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
            else:
                cell_norma.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
            # Centrar columna Sello SEC
            if col_sello:
                cell_sello = ws_mat.cell(row=r, column=col_sello)
                cell_sello.alignment = Alignment(horizontal="center", vertical="top")

# =========================
# HELPERS (NO BORRAN NADA)
# =========================

def conduit_por_tabla(seccion_mm2, n_cond):
    """
    Tabla N°4.17 (H07V-U/R/K, etc.) -> Ø nominal ducto mm según sección y Nº conductores (1..5)
    Si n_cond > 5, lo limita a 5 (puedes ampliarlo después si quieres).
    """
    # Tabla: sección : {n_cond: ducto_mm}
    tabla = {
        1.5: {1:16, 2:16, 3:16, 4:20, 5:20},
        2.5: {1:16, 2:16, 3:20, 4:20, 5:20},
        4.0: {1:16, 2:16, 3:20, 4:20, 5:25},   # según Tabla N°4.17 RIC (2 cond. = 16mm)
        6.0: {1:16, 2:20, 3:20, 4:25, 5:25},
        10.0:{1:16, 2:20, 3:25, 4:32, 5:32},
        16.0:{1:20, 2:25, 3:32, 4:40, 5:40},
        # si después necesitas más secciones, las agregamos
    }

    # elegir la sección "igual o superior" disponible en tabla
    secciones = sorted(tabla.keys())
    sec = None
    for s in secciones:
        if float(seccion_mm2) <= float(s):
            sec = s  # ya encontró la sección de tabla que alcanza a cubrir la sección real
            break
    if sec is None:
        sec = secciones[-1]  # la sección real es más grande que todo lo que hay en la tabla, usa la máxima

    n = int(n_cond)
    if n < 1: n = 1   # mínimo 1 conductor
    if n > 5: n = 5   # la tabla solo llega hasta 5 conductores

    return tabla[sec][n]


def canalizacion_recomendada_por_conductores(tipo_canalizacion, seccion_mm2, n_cond):
    """
    - Si es embutida: usa la tabla de conduit según sección y N° de conductores.
    - Si es sobrepuesta: usa el cálculo de canaleta por área real del conductor.
    """
    tipo = (tipo_canalizacion or "").strip().lower()

    if "embut" in tipo:
        mm = conduit_por_tabla(seccion_mm2, n_cond)  # busca el diámetro del conduit en la tabla
        return f"Embutida (PVC conduit {mm}mm)"

    # sobrepuesta: usa cálculo dinámico por área real del conductor
    return canalizacion_recomendada(tipo_canalizacion, seccion_mm2, n_cond)

def n_conductores_iluminacion_para_circuito(circuito_items, ambientes_df):
    """
    Decide el N° de conductores para dimensionar canalización en iluminación,
    usando tu regla:
      - conmutado => 3
      - no conmutado: simple=3, doble=4, triple=5 (según luminarias del ambiente en ese circuito)
    Retorna el MÁXIMO requerido dentro del circuito.
    """
    if not isinstance(circuito_items, list) or len(circuito_items) == 0:
        return 3  # sin items, se asume el caso más simple (3 conductores)

    # mapa conmutadas por ambiente desde ambientes_df
    ncon_map = {}
    if ambientes_df is not None and "Ambiente" in ambientes_df.columns:
        amb_col = ambientes_df["Ambiente"].astype(str).str.strip().str.lower()
        if "N_conmutadas_924 (u)" in ambientes_df.columns:
            vals = pd.to_numeric(ambientes_df["N_conmutadas_924 (u)"], errors="coerce").fillna(0).astype(int)
            for a, v in zip(amb_col, vals):
                ncon_map[a] = int(max(0, v))  # cuántas luminarias conmutadas tiene cada ambiente

    # contar luminarias por ambiente dentro del circuito (desde _items)
    lum_by_amb = {}
    for it in circuito_items:
        if not isinstance(it, dict):
            continue
        # excluir enchufes y especiales
        if ("id_ench" in it) or ("modulos" in it) or ("n_ench" in it) or ("nombre" in it):
            continue
        amb = str(it.get("amb", "")).strip().lower() or "sin_amb"
        lum_by_amb[amb] = lum_by_amb.get(amb, 0) + 1  # suma una luminaria más a este ambiente

    n_max = 3  # mínimo posible (3 conductores)

    # recorre cada ambiente y ve cuántos conductores necesita, se queda con el peor caso
    for amb, n_lum in lum_by_amb.items():
        n_conmutadas = int(ncon_map.get(amb, 0))
        n_conmutadas = max(0, min(n_conmutadas, int(n_lum)))  # no puede haber más conmutadas que luminarias
        n_restantes = int(n_lum) - n_conmutadas  # luminarias que NO son conmutadas
        # Base: 9/12 => 3 conductores
        n_req = 3
        # Si hay conmutado (9/24)
        if n_conmutadas > 0:
            n_req = max(n_req, 3)
        # Para el resto de luminarias
        if n_restantes > 0:
            c12, c15, c32 = descomponer_interruptores(int(n_restantes))  # cómo se agrupan las luminarias restantes
            if c32 > 0:
                n_req = max(n_req, 5)  # existe 9/32
            elif c15 > 0:
                n_req = max(n_req, 4)  # existe 9/15
            else:
                n_req = max(n_req, 3)  # solo 9/12
        n_max = max(n_max, int(n_req))  # se queda con el máximo requerido entre todos los ambientes
    return int(n_max)


def binpack_items(items, Pmax):
    """
    Agrupa items (cargas) en bins/subcircuitos para que cada bin no supere Pmax [W].
    items: lista de dicts, mínimo {"amb": "...", "potencia": float}
    """
    # ordena las cargas de mayor a menor potencia (First-Fit Decreasing)
    items_sorted = sorted(items, key=lambda x: float(x.get("potencia", 0.0)), reverse=True)
    bins = []
    # recorre cada carga ya ordenada
    for it in items_sorted:
        p = float(it.get("potencia", 0.0))  # potencia de la carga
        placed = False  # ¿ya se asignó a algún circuito?
        for b in bins:  # recorre los circuitos ya creados
            if b["potencia_total"] + p <= Pmax:  # ¿cabe sin superar el máximo?
                b["items"].append(it)  # agrega la carga al circuito
                b["potencia_total"] += p  # actualiza la potencia total
                placed = True
                break  # sale al encontrar el primer circuito que sirve
        if not placed:  # si no cupo en ninguno, abre un circuito nuevo
            bins.append({"items": [it], "potencia_total": p})
    return bins

def pedir_longitud_sub(nombre, default_L):
    # pregunta la longitud real de un tramo (en metros). Si el usuario
    # deja vacío (solo ENTER), usa default_L. Insiste hasta que le den
    # un número válido.
    while True:
        s = input(
            f"       • Longitud REAL del '{nombre}' en metros (ENTER = usar {default_L} m)\n"
            f"         (En tramos horizontales: recorridos a 0,30m del cielo y 0,20m del piso - RIC N°4 7.16.1.16): "
        ).strip()  # lee lo que escribió el usuario
        if s == "":
            return float(default_L)  # dejó vacío, usa el valor por defecto
        try:
            v = float(s)  # intenta convertir a número
            if v <= 0:
                print("         ! Debe ser mayor que 0.")
                continue  # no sirve, vuelve a preguntar
            return v  # número válido
        except:
            print("         ! Ingrese un número válido.")  # no era un número

def pedir_float_positivo(prompt):
    # pregunta un número que tiene que ser mayor que 0. Si escriben
    # cualquier cosa que no sirva, vuelve a preguntar.
    while True:
        s = input(prompt).strip()  # lee lo que escribió el usuario
        try:
            v = float(s)  # intenta convertirlo a número
            if v <= 0:
                print("         ! Debe ser un número mayor que 0.")
                continue  # vuelve a preguntar
            return v  # número válido, listo
        except:
            print("         ! Ingrese un número válido (ej: 7.8).")  # no era un número

def pedir_float_opcional(prompt, valor_defecto=0.0, min_val=0.0):
    # pregunta un número, pero es opcional: si dejan vacío usa "valor_defecto".
    # Si escriben algo, tiene que ser >= min_val o vuelve a preguntar.
    while True:
        s = input(prompt).strip()  # lee lo que escribió el usuario
        if s == "":
            return float(valor_defecto)  # dejó vacío, usa el valor por defecto
        try:
            v = float(s)  # intenta convertirlo a número
            if v < min_val:
                print(f"         ! Debe ser >= {min_val}.")
                continue  # no cumple el mínimo, vuelve a preguntar
            return v  # número válido
        except:
            print("         ! Ingrese un número válido.")  # no era un número

def limpiar_nombre_circuito(txt):
    # limpia el nombre del circuito que escribió el usuario: saca espacios
    # de más y arregla el típico error de tipeo "eenc..." (lo deja como "enc...").
    t = (txt or "").strip()  # saca espacios al inicio/final
    low = t.lower()
    if low.startswith("eenc"):  # error típico de tipeo
        t = t[1:].strip()  # saca la "e" repetida
    t = " ".join(t.split())  # deja un solo espacio entre palabras
    return t

def sugerir_nombre_circuito(nombre):
    """Solo sugiere corrección si el nombre parece ser iluminacion,
    enchufes o climatizacion pero está mal escrito."""
    import difflib, unicodedata

    def normalizar(s):
        # pasa a minúsculas y saca las tildes, así "Iluminación" e "iluminacion"
        # se pueden comparar como si fueran lo mismo.
        s = s.lower().strip()
        s = unicodedata.normalize("NFD", s)
        return "".join(c for c in s if unicodedata.category(c) != "Mn")

    tipos_clave = ["ilumin", "enchufe", "clima", "aire", "split", "ac ", "a/c"]
    base_norm = normalizar(nombre)  # nombre del circuito sin tildes ni mayúsculas

    # Si ya contiene una clave reconocida, no hacer nada
    if any(k in base_norm for k in tipos_clave):
        return nombre  # está bien escrito, no hay que sugerir nada

    # Solo comparar contra los 3 tipos que importan
    tipos_canonicos = ["iluminacion", "enchufes", "climatizacion"]
    matches = difflib.get_close_matches(base_norm, tipos_canonicos, n=1, cutoff=0.55)  # busca el más parecido
    if matches:
        sugerido = matches[0]  # el tipo más parecido al nombre escrito
        resp = input(
            f"       ! No se reconoció '{nombre}' como tipo de circuito. "
            f"¿Quisiste decir '{sugerido}'? (si/no): "
        ).strip().lower()
        if resp == "si":
            return sugerido  # el usuario aceptó la corrección
    return nombre  # se deja el nombre tal cual lo escribió

def canalizacion_recomendada(tipo_canalizacion, seccion_mm2, n_cond=3):
    """
    Devuelve el texto final para la columna Canalización:
    - Embutida (PVC conduit XXmm)
    - Sobrepuesta (Canaleta de PVC AAxBBmm)
    """
    conduit_map = {
        1.5: 16, 2.08: 16,
        2.5: 20, 3.31: 20,
        4.0: 25, 5.26: 25, 6.0: 25,
        8.37: 32, 10.0: 32,
        13.3: 40, 16.0: 40
    }  # para cada sección en mm², su diámetro de conduit en mm
    # Área real del conductor con aislación H07Z1-K, valores aproximados de catálogo
    AREA_CON_AISLACION = {
        1.5:  9.08,
        2.08: 5.31,
        2.5:  13.20,
        3.31: 7.16,
        4.0:  18.10,
        5.26: 11.40,
        6.0:  22.06,
        8.37: 23.67,
        10.0: 36.32,
        13.3: 33.59,
        16.0: 51.53,
    }

    # Canaletas disponibles: para cada dimensión, su área total en mm²
    CANALETAS = [
        ('20x10',  200),
        ('32x12',  384),
        ('40x16',  640),
        ('60x25', 1500),
    ]
    FACTOR_OCUPACION = 0.40  # solo se puede ocupar el 40% del área de la canaleta

    def seleccionar_canaleta(seccion_mm2, n_conductores):
        """Selecciona la canaleta más pequeña que cumpla con el factor de ocupación."""
        # seccion_mm2 = sección del conductor (ej: 2.5)
        # n_conductores = cuántos conductores van juntos dentro de la misma canaleta

        # Buscar área del conductor más cercana
        # claves = las secciones que existen en la tabla AREA_CON_AISLACION (1.5, 2.5, 4.0, etc.)
        claves = sorted(AREA_CON_AISLACION.keys())
        key = claves[-1]  # por si seccion_mm2 es más grande que todo lo que hay en la tabla
        for k in claves:
            if float(seccion_mm2) <= float(k):
                key = k  # primera sección de la tabla que alcanza a cubrir la real
                break
        area_cond = AREA_CON_AISLACION[key]  # área de UN conductor (mm²)
        area_necesaria = area_cond * n_conductores  # área total ocupada por todos los conductores juntos
        # Seleccionar canaleta más pequeña que cumpla
        # dim = nombre de la canaleta (ej: "20x10"), area_total = su área en mm²
        for dim, area_total in CANALETAS:
            if area_total * FACTOR_OCUPACION >= area_necesaria:
                return dim  # esta canaleta ya alcanza con el 40% de ocupación
        return CANALETAS[-1][0]  # ninguna alcanzó: usa la más grande de la lista

    tipo = (tipo_canalizacion or "").strip().lower()
    claves = sorted(AREA_CON_AISLACION.keys())

    key = claves[-1]
    for k in claves:
        if float(seccion_mm2) <= float(k):
            key = k  # sección de la tabla que corresponde a la sección real
            break

    if "embut" in tipo:
        mm = conduit_map.get(key, 20)  # diámetro de conduit (20mm si no está en la tabla)
        return f"Embutida (PVC conduit {mm}mm)"
    else:
        dim = seleccionar_canaleta(seccion_mm2, n_cond if n_cond else 3)
        return f"Sobrepuesta (Canaleta de PVC {dim}mm)"

def resumen_items_por_ambiente(items, modo="enchufe"):
    """
    items: lista de dicts con al menos {"amb","potencia"}.
    Devuelve texto detallado por ambiente.
    """

    if modo == "especial":
        # circuitos especiales: cada item es una carga con nombre propio (ej: "horno")
        partes = []
        for it in items:
            amb = it.get("amb", "sin_amb")
            nombre = it.get("nombre", "carga")
            potencia = int(round(float(it.get("potencia", 0.0))))
            partes.append(f"{amb}: {nombre} ({potencia}W)")
        return "; ".join(partes) if partes else ""

    # para enchufes o iluminación: se suma potencia y cantidad por ambiente
    res = {}
    for it in items:
        a = it.get("amb", "sin_amb")  # nombre del ambiente
        p = float(it.get("potencia", 0.0))  # potencia de este item

        if modo == "enchufe":
            n = int(it.get("n_ench", 1))  # cantidad de enchufes de este item
        elif modo == "iluminacion":
            n = int(it.get("n_lum", 1))  # cantidad de luminarias de este item
        else:
            n = 0

        if a not in res:
            res[a] = {"p": 0.0, "n": 0}  # primera vez que aparece este ambiente

        res[a]["p"] += p  # va sumando la potencia del ambiente
        res[a]["n"] += n  # va sumando la cantidad del ambiente

    # arma el texto final, un trozo por ambiente
    partes = []
    for a, v in res.items():
        ptxt = int(round(v["p"]))
        if modo == "enchufe":
            partes.append(f"{a}: {v['n']} ench ({ptxt}W)")
        elif modo == "iluminacion":
            partes.append(f"{a}: {v['n']} lum ({ptxt}W)")
        else:
            partes.append(f"{a}: {ptxt}W")

    return ", ".join(partes) if partes else ""

def parse_in_tm(tm_text):
    # saca el número de corriente (In) de un texto tipo "1x16A curva C" (da 16).
    # Si no logra entenderlo, devuelve None.
    # ejemplo: "1x16A curva C" da 16
    if tm_text is None:
        return None
    s = str(tm_text).lower().replace(" ", "")  # saca espacios y pasa a minúsculas
    try:
        if "x" in s and "a" in s:  # tiene forma "1x16a..."
            after_x = s.split("x", 1)[1]  # se queda con lo que hay después de la "x"
            num = ""
            for ch in after_x:  # va juntando los dígitos del principio
                if ch.isdigit():
                    num += ch
                else:
                    break  # se corta apenas encuentra algo que no es número (la "a")
            return int(num) if num else None
    except:
        return None
    return None

def descomponer_interruptores(n_lum):
    """
    Regla (según cuántas luminarias hay):
      1 luminaria:  9/12
      2 luminarias: 9/15
      3 luminarias: 9/32
      4 luminarias: 9/32 + 9/12
      5 luminarias: 9/32 + 9/15
      6 luminarias: 9/32 + 9/32
      ... (primero se arman grupos de a 3, después de a 2, al final de a 1)
    Devuelve (c12, c15, c32)
    """
    n = int(max(0, n_lum))  # cantidad de luminarias, nunca negativo
    c32 = n // 3      # cuántos interruptores 9/32 entran (grupos de 3)
    rem = n % 3        # lo que sobra después de armar los grupos de 3
    c15 = 1 if rem == 2 else 0  # si sobran 2, va un 9/15
    c12 = 1 if rem == 1 else 0  # si sobra 1, va un 9/12
    return c12, c15, c32

def _get_amb_row(ambientes_df, amb_name):
    # busca la fila del ambiente "amb_name" dentro de ambientes_df
    # (sin importar mayúsculas ni espacios). Si no la encuentra, None.
    if ambientes_df is None or "Ambiente" not in ambientes_df.columns:
        return None  # no hay datos de ambientes, no se puede buscar
    key = str(amb_name).strip().lower()  # normaliza el nombre que estamos buscando
    m = ambientes_df[ambientes_df["Ambiente"].astype(str).str.strip().str.lower() == key]  # filtra la fila que calza
    if len(m) == 0:
        return None  # no existe ese ambiente
    return m.iloc[0]  # devuelve la primera (y única) coincidencia

# descripción automática SIN PREGUNTAR NADA MÁS
def _lum_necesita_estaño(tipo_txt, desc_txt=""):
    """Retorna True si la luminaria se conecta con estaño (LED con cable), False si va a bornes (portalámpara/ampolleta)."""
    t = str(tipo_txt or "").strip().lower()  # tipo de luminaria, en minúsculas
    d = str(desc_txt or "").strip().lower()  # descripción extra, en minúsculas
    if any(x in t or x in d for x in ["ampol", "incand", "portalamp", "portalámp", "e27", "e14"]):
        return False  # va a bornes (ampolleta/portalámpara), no lleva estaño
    if any(x in t or x in d for x in ["foco", "panel", "led", "aplique", "tubo", "fluores"]):
        return True  # va con cable, se conecta con estaño
    return False  # por defecto, no requiere estaño


def desc_luminaria_auto(tipo_txt, montaje_txt, potencia_w):
    """
    No pregunta nada extra. Solo usa lo que ya ingresas:
    - tipo (texto que el usuario escribió)
    - montaje (embutida/sobrepuesta)
    - potencia (W ingresada)
    Devuelve la descripción "bonita".
    """
    t = (tipo_txt or "").strip().lower()  # tipo que escribió el usuario
    m = (montaje_txt or "").strip().lower()  # embutida o sobrepuesta
    try:
        pw = int(round(float(potencia_w)))  # potencia redondeada, en watts
    except:
        pw = 0  # si viene mal, se deja en 0

    # detecta qué tipo de luminaria es, buscando palabras clave en el texto
    es_foco = ("foco" in t) or ("panel" in t)
    es_aplique = ("aplique" in t)
    es_tubo = ("tubo" in t) or ("fluores" in t)
    es_ampolleta = ("ampol" in t) or ("incand" in t)

    incand = ("incand" in t)  # ¿el texto menciona incandescente? (usada también más abajo por ampolleta)
    fluor = ("fluores" in t)  # ¿el texto menciona fluorescente?

    # tecnología para foco/panel y aplique: por defecto LED, salvo que el
    # texto que escribió el usuario diga explícitamente incandescente o
    # fluorescente (tubo y ampolleta calculan su propia tecnología más abajo)
    if incand:
        tecnologia = "incandescente"
    elif fluor:
        tecnologia = "fluorescente"
    else:
        tecnologia = "LED"

    if es_foco:
        if "embut" in m:
            return f"Foco panel {tecnologia} cuadrado/redondo embutido {pw} W luz cálida/fría/neutro"
        else:
            return f"Foco panel {tecnologia} cuadrado/redondo sobrepuesto {pw} W luz cálida/fría/neutro"

    if es_aplique:
        # arma el texto indicando si el aplique va embutido o sobrepuesto
        montaje_apl = "embutido" if "embut" in m else "sobrepuesto"
        return f"Aplique {tecnologia} {montaje_apl} {pw} W luz cálida/fría/neutro"

    # tubo y ampolleta calculan su propia tecnología (fluorescente/incandescente)
    # más abajo, por eso no usan la variable "tecnologia" definida arriba

    if es_tubo:
        # arma el texto indicando si el tubo va embutido o sobrepuesto
        montaje_tubo = "embutido" if "embut" in m else "sobrepuesto"
        if ("fluores" in t):
            return f"Tubo fluorescente {montaje_tubo} {pw} W luz cálida/fría/neutro"
        else:
            return f"Tubo LED {montaje_tubo} {pw} W luz cálida/fría/neutro"

    if es_ampolleta:
        montaje_amp = "embutida" if "embut" in m else "sobrepuesta"
        if incand:
            return f"Ampolleta incandescente {pw} W {montaje_amp} luz cálida/fría/neutro"
        else:
            return f"Ampolleta LED {pw} W {montaje_amp} luz cálida/fría/neutro"

    # no calzó con ningún tipo conocido: descripción genérica
    if "embut" in m:
        return f"Luminaria {pw} W (embutida) luz cálida/fría/neutro"
    return f"Luminaria {pw} W (sobrepuesta) luz cálida/fría/neutro"

# =========================================================
# ALIMENTADOR (selección por caída de tensión + tabla)
# =========================================================
SECCIONES_MM2_ALIM = [1.5, 2.08, 2.5, 3.31, 4, 5.26, 6, 8.37, 10, 13.3, 16, 21.1, 25, 26.7, 33.6, 35, 42.4, 50]
TABLA_AMPACIDAD_ALIM = {
    "B1": {  # ducto
        1.5: 18, 2.08: 24, 2.5: 24, 3.31: 31, 4: 37, 5.26: 39, 6: 48, 8.37: 59,
        10: 66, 13.3: 79, 16: 88, 21.1: 105, 25: 117, 26.7: 122, 33.6: 141, 35: 144,
        42.4: 163, 50: 175
    },
    "D1": {  # subterraneo
        1.5: 19, 2.08: 30, 2.5: 33, 3.31: 38, 4: 42, 5.26: 48, 6: 52, 8.37: 63,
        10: 68, 13.3: 80, 16: 89, 21.1: 103, 25: 113, 26.7: 117, 33.6: 132, 35: 136,
        42.4: 150, 50: 159
    },
    "E": {   # aereo
        1.5: 19, 2.08: 28, 2.5: 32, 3.31: 38, 4: 42, 5.26: 50, 6: 54, 8.37: 67,
        10: 75, 13.3: 89, 16: 100, 21.1: 114, 25: 127, 26.7: 133, 33.6: 154, 35: 158,
        42.4: 178, 50: 192
    }
}

TABLA_AMPACIDAD_ACOM = {
    "E": {   # Método E (aéreo) — 70°C, Tabla 4.4 RIC 4
        1.5: 19,
        2.08: 22,
        2.5: 24,
        3.31: 30,
        4: 31,
        5.26: 38,
        6: 43,
        8.37: 53,
        10: 60,
        13.3: 71,
        16: 80,
        21.1: 91,
        25: 101,
        26.7: 106,
        33.6: 122,
        35: 126,
        42.4: 142,
        50: 159
    },
    "D1": {   # Método D1 (ducto enterrado)
        1.5: 19,
        2.08: 30,
        2.5: 33,
        3.31: 38,
        4: 42,
        5.26: 48,
        6: 52,
        8.37: 63,
        10: 68,
        13.3: 80,
        16: 89,
        21.1: 103,
        25: 113,
        26.7: 117,
        33.6: 132,
        35: 136,
        42.4: 150,
        50: 159
    }
}

def factor_temperatura_ft(temp_c, metodo):
    """Factor de corrección de ampacidad por temperatura — RIC N°4 Tabla N°4.7 / art. 6.2.6.
    Uso correcto: Ic = Iz × ft  (ft MULTIPLICA sobre la ampacidad de tabla, NO divide I_diseno)
    ft = 1.00 a 30°C · ft < 1.0 para T > 30°C · ft > 1.0 para T < 30°C
    El método E (aéreo, sol directo) usa ft_e < ft_b1 para T > 30°C.
    """
    t = float(temp_c)
    # Tabla N°4.7 — tres columnas:
    # ft_b1: A1/B1 (embutido o en conducto)
    # ft_e:  E (aéreo expuesto al sol — más caliente que B1)
    # ft_d:  D1/D2 (subterráneo — temperatura del suelo)
    if t <= 10:
        ft_b1 = 1.22;  ft_e = 1.22;  ft_d = 1.07
    elif t <= 15:
        ft_b1 = 1.17;  ft_e = 1.17;  ft_d = 1.04
    elif t <= 20:
        ft_b1 = 1.12;  ft_e = 1.12;  ft_d = 1.00
    elif t <= 25:
        ft_b1 = 1.06;  ft_e = 1.06;  ft_d = 0.96
    elif t <= 30:
        ft_b1 = 1.00;  ft_e = 1.00;  ft_d = 0.93
    elif t <= 35:
        ft_b1 = 0.94;  ft_e = 0.90;  ft_d = 0.89
    elif t <= 40:
        ft_b1 = 0.87;  ft_e = 0.82;  ft_d = 0.85
    elif t <= 45:
        ft_b1 = 0.79;  ft_e = 0.74;  ft_d = 0.80
    elif t <= 50:
        ft_b1 = 0.71;  ft_e = 0.65;  ft_d = 0.76
    elif t <= 55:
        ft_b1 = 0.61;  ft_e = 0.55;  ft_d = 0.71
    else:
        ft_b1 = 0.50;  ft_e = 0.43;  ft_d = 0.65
    if metodo in ("D1", "D2"):
        return ft_d   # subterráneo: usa la columna de temperatura del suelo
    if metodo == "E":
        return ft_e   # aéreo: usa su propia columna, distinta de B1
    return ft_b1  # embutido o en conducto (A1/B1): columna por defecto

def metodo_por_tipo_alim(tipo_alimentador: str) -> str:
    # según si el alimentador es aéreo, subterráneo o por ducto, devuelve
    # el "método de instalación" del RIC (E, D1 o B1) para usar la tabla
    # de corriente correcta.
    t = str(tipo_alimentador).strip().lower()
    if "aer" in t:
        return "E"
    if "sub" in t:
        return "D1"
    if "duc" in t:
        return "B1"
    return "B1"  # default
def siguiente_seccion_alim(smin: float) -> float:
    # a partir de la sección mínima que dio el cálculo (smin), busca la
    # sección comercial más chica que alcance a cubrirla.
    for s in SECCIONES_MM2_ALIM:
        if s >= smin:
            return s  # esta ya alcanza a cubrir la sección mínima
    return SECCIONES_MM2_ALIM[-1]  # ninguna alcanzó, se usa la más grande de la lista
def dv_volts_alim(L_m: float, I_A: float, fp: float, rho: float, S_mm2: float) -> float:
    # fórmula típica de caída de tensión monofásica: ΔV = 2·L·I·fp·ρ / S
    return (2 * L_m * I_A * fp * rho) / S_mm2

def seleccionar_alimentador(L_m: float, I_demanda_A: float, I_empalme_A: float, fp: float, V_nom: float,
                           tipo_alimentador: str,
                           dv_max_volts: float = None,
                           dv_circuitos: list = None,
                           temp_override: float = None,
                           ):
    # elige la sección del alimentador (y su ducto) probando calibres desde
    # el mínimo hacia arriba, hasta que cumpla: soporte la corriente (Iz),
    # la caída de tensión no pase el 3% y sumado a los circuitos de abajo
    # no pase el 5% total (límites del RIC). Devuelve la sección elegida
    # junto con todos los datos del cálculo.
    # Paso 1: Smin
    rho = 0.0179  # resistividad del cobre
    if dv_max_volts is None:
        dv_max_volts = V_nom * 0.03  # tope de caída de tensión: 3% de la tensión nominal
    smin = (2 * L_m * I_demanda_A * fp * rho) / dv_max_volts  # sección mínima teórica
    # Paso 2: método por tipo
    metodo = metodo_por_tipo_alim(tipo_alimentador)  # E, D1 o B1 según sea aéreo/subterráneo/ducto
    tabla = TABLA_AMPACIDAD_ALIM[metodo]  # tabla de corriente admisible de ese método
    # sección comercial >= smin, mínimo 4mm² según RIC
    smin = max(smin, 4.0)
    S = siguiente_seccion_alim(smin)  # primera sección comercial candidata
    # Temperatura efectiva: usa temp_override si se proporcionó (ej. T° suelo para subterráneo)
    temp_calc = temp_override if temp_override is not None else temperatura
    # Paso 3 y 4: subir hasta cumplir Iz>=I, ΔV%<=3 y ΔV_total<=5% (RIC)
    _dv_circ_validos = [v for v in (dv_circuitos or []) if v is not None and not (v != v)]  # saca los None/NaN
    _dv_max_circ = max(_dv_circ_validos) if _dv_circ_validos else 0.0  # peor caída de tensión de los circuitos de abajo
    while True:
        Iz_base = tabla.get(S, 0)  # corriente admisible de esta sección (sin corregir)
        ft = factor_temperatura_ft(temp_calc, metodo)  # factor de corrección por temperatura
        Iz = Iz_base * ft  # corriente admisible real
        dv_v = dv_volts_alim(L_m, I_demanda_A, fp, rho, S)  # caída de tensión con esta sección
        dv_pct = (dv_v / V_nom) * 100.0  # caída de tensión en %
        cumple_ampacidad = (Iz >= I_demanda_A) and (Iz >= I_empalme_A)  # aguanta la corriente de demanda y de empalme
        cumple_caida = (dv_pct <= 3.0)  # no supera el 3% solo del alimentador
        cumple_total = (dv_pct + _dv_max_circ) <= 5.0  # sumado a los circuitos, no supera el 5% total
        if cumple_ampacidad and cumple_caida and cumple_total:
            # esta sección cumple todo, se queda con ella
            return {"metodo": metodo, "Smin": smin, "S": S, "Iz_base": Iz_base, "ft": ft, "Iz": Iz,
                    "I_demanda_A": I_demanda_A, "I_empalme_A": I_empalme_A, "dV_V": dv_v, "dV_pct": dv_pct,
                    "dV_max_circ": _dv_max_circ, "dV_total": round(dv_pct + _dv_max_circ, 2)}

        idx = SECCIONES_MM2_ALIM.index(S)
        if idx >= len(SECCIONES_MM2_ALIM) - 1:
            # ya no hay secciones más grandes, avisa que no se pudo cumplir todo
            return {"metodo": metodo, "Smin": smin, "S": S, "Iz_base": Iz_base, "ft": ft, "Iz": Iz,
                    "I_demanda_A": I_demanda_A, "I_empalme_A": I_empalme_A, "dV_V": dv_v, "dV_pct": dv_pct,
                    "dV_max_circ": _dv_max_circ, "dV_total": round(dv_pct + _dv_max_circ, 2),
                    "Advertencia": "No cumple Iz > I_demanda y/o Iz >= I_empalme y/o ΔV% <= 3 y/o ΔV_total <= 5% con secciones disponibles"}
        S = SECCIONES_MM2_ALIM[idx + 1]  # sube al siguiente calibre y prueba de nuevo

def seleccionar_acometida(L_m: float, I_empalme_A: float, fp: float, V_nom: float,
                          tipo_acometida: str,
                          dv_max_volts: float = None,
                          temp_override: float = None):
    # lo mismo que seleccionar_alimentador pero para la acometida (el tramo
    # entre el empalme y el medidor): calcula la sección que aguante la
    # corriente y no pase la caída de tensión máxima.
    rho = 0.0179  # resistividad del cobre
    if dv_max_volts is None:
        dv_max_volts = V_nom * 0.03  # por defecto, tope de caída de tensión = 3% de la tensión nominal
    smin = (2 * L_m * I_empalme_A * fp * rho) / dv_max_volts  # sección mínima teórica según caída de tensión
    # Elegir método según tipo de acometida
    if tipo_acometida.lower() == "aereo":
        metodo = "E"
    elif tipo_acometida.lower() == "subterraneo":
        metodo = "D1"
    else:
        raise ValueError(f"Tipo de acometida no válido: {tipo_acometida}")
    tabla = TABLA_AMPACIDAD_ACOM[metodo]  # tabla de corriente admisible según el método
    # mínimo 4mm² por consistencia con alimentador
    smin = max(smin, 4.0)
    S = siguiente_seccion_alim(smin)  # primera sección comercial candidata
    temp_calc = temp_override if temp_override is not None else temperatura
    while True:
        Iz_base = tabla.get(S, 0)  # corriente admisible de esta sección (sin corregir)
        ft = factor_temperatura_ft(temp_calc, metodo)  # factor de corrección por temperatura
        Iz = Iz_base * ft  # corriente admisible real
        dv_v = dv_volts_alim(L_m, I_empalme_A, fp, rho, S)  # caída de tensión con esta sección
        dv_pct = (dv_v / V_nom) * 100.0  # caída de tensión en %
        if (Iz >= I_empalme_A) and (dv_pct <= 3.0):
            # esta sección cumple ambos requisitos, se queda con ella
            return {"metodo": metodo,"Smin": smin,"S": S,"Iz_base": Iz_base,"ft": ft,"Iz": Iz,"dV_V": dv_v,
                    "dV_pct": dv_pct}
        idx = SECCIONES_MM2_ALIM.index(S)
        if idx >= len(SECCIONES_MM2_ALIM) - 1:
            # ya no hay secciones más grandes disponibles, avisa que no se pudo cumplir
            return {"metodo": metodo,"Smin": smin,"S": S,"Iz_base": Iz_base,"ft": ft,"Iz": Iz,"dV_V": dv_v,
                    "dV_pct": dv_pct,"warning": "No cumple Iz y/o ΔV% con secciones disponibles"}
        S = SECCIONES_MM2_ALIM[idx + 1]  # sube al siguiente calibre y prueba de nuevo

# =========================================================
# CANALIZACIÓN ALIMENTADOR - Tablas N°4.19 (mm²) y N°4.20 (AWG)
# =========================================================
# Tabla N°4.19 (mm²)
DUCTO_N419_MM2 = {
    1.5:  {1:16, 2:16, 3:16, 4:20, 5:25},
    2.5:  {1:16, 2:20, 3:20, 4:32, 5:32},
    4.0:  {1:16, 2:25, 3:25, 4:32, 5:40},
    6.0:  {1:16, 2:25, 3:32, 4:32, 5:40},
    10.0: {1:20, 2:32, 3:32, 4:40, 5:50},
    16.0: {1:25, 2:32, 3:40, 4:50, 5:50},
    25.0: {1:25, 2:40, 3:50, 4:50, 5:63},
    35.0: {1:32, 2:40, 3:50, 4:63, 5:63},
    50.0: {1:32, 2:50, 3:63, 4:63, 5:75},
    70.0: {1:40, 2:50, 3:63, 4:75, 5:75},
    95.0: {1:40, 2:63, 3:75, 4:100, 5:100},
    120.0:{1:50, 2:63, 3:75, 4:100, 5:100},
    150.0:{1:50, 2:75, 3:100,4:100, 5:125},
    185.0:{1:63, 2:75, 3:100,4:125, 5:125},
    240.0:{1:63, 2:100,3:125,4:125, 5:150},
}

# Tabla N°4.29 - Cables para uso en tuberías de canalizaciones subterráneas
# Solo hasta 25mm² y máximo 3 conductores (uso acometida y alimentador subterráneo)
DUCTO_N429_SUBTERRANEO = {
    1.5:  {1: 25, 2: 25, 3: 25},
    2.5:  {1: 25, 2: 25, 3: 32},
    4.0:  {1: 25, 2: 32, 3: 40},
    6.0:  {1: 32, 2: 32, 3: 50},
    10.0: {1: 40, 2: 50, 3: 63},
    16.0: {1: 50, 2: 50, 3: 63},
    25.0: {1: 63, 2: 63, 3: 75},
}
# Tabla N°4.20 (AWG/kcmil) - lo que aparece en tu imagen
DUCTO_N420_AWG = {
    "14":  {1:16, 2:16, 3:20, 4:25, 5:32},
    "12":  {1:16, 2:20, 3:25, 4:32, 5:32},
    "10":  {1:16, 2:25, 3:32, 4:32, 5:40},
    "8":   {1:20, 2:25, 3:32, 4:40, 5:40},
    "6":   {1:20, 2:32, 3:40, 4:40, 5:50},
    "4":   {1:25, 2:32, 3:40, 4:50, 5:50},
    "2":   {1:32, 2:40, 3:50, 4:63, 5:63},
    "1":   {1:32, 2:40, 3:50, 4:63, 5:63},
    "1/0": {1:32, 2:50, 3:63, 4:63, 5:75},
    "2/0": {1:40, 2:50, 3:63, 4:75, 5:75},
    "3/0": {1:40, 2:63, 3:75, 4:75, 5:100},
    "4/0": {1:50, 2:63, 3:75, 4:100, 5:100},
    "250": {1:50, 2:75, 3:100, 4:100, 5:125},
    "300": {1:50, 2:75, 3:100, 4:100, 5:125},
    "350": {1:63, 2:75, 3:100, 4:125, 5:125},
    "400": {1:63, 2:100, 3:100, 4:125, 5:125},
    "500": {1:63, 2:100, 3:125, 4:125, 5:150},
    "600": {1:75, 2:100, 3:125, 4:150, 5:150},
    "750": {1:75, 2:125, 3:150, 4:150, 5:175},
    "1000":{1:100, 2:125, 3:150, 4:175, 5:200},
}
# Conversión de mm² a AWG (tabla de sección comercial)
MM2_A_AWG = {
    2.08: "14",
    3.31: "12",
    5.26: "10",   # sección usada también en la tabla de conectores cónicos
    8.37: "8",
    13.3: "6",
    21.1: "4",
    33.6: "2",
    42.4: "1",
}
def ducto_nominal_tablas(S_mm2: float, n_cond: int, tipo_alimentador: str):
    """
    Devuelve Ø ducto nominal (mm) usando:
    - aereo          => None (sin canalización)
    - subterraneo    => Tabla N°4.29 (hasta 25mm², máx 3 conductores)
    - embutido/ducto => Tabla N°4.19 mm² o N°4.20 AWG si no existe en N°4.19
    """
    t = str(tipo_alimentador).strip().lower()
    if "aer" in t:
        return None  # aéreo no lleva ducto
    n = max(1, min(5, int(n_cond)))  # la tabla solo cubre entre 1 y 5 conductores
    S_key = round(float(S_mm2), 2)

    # Si es subterráneo, se usa la Tabla N°4.29
    if "sub" in t:
        n_sub = max(1, min(3, n))   # tabla solo tiene hasta 3 conductores
        S_val = float(S_mm2)
        # buscar clave más cercana por arriba en la tabla
        claves = sorted(DUCTO_N429_SUBTERRANEO.keys())
        clave = None
        for c in claves:
            if S_val <= c + 0.01:
                clave = c  # primer umbral que alcanza a cubrir la sección
                break
        if clave is None:
            clave = claves[-1]   # usar la mayor si supera 25mm²
        return DUCTO_N429_SUBTERRANEO[clave][n_sub]

    # Si es embutido o va por ducto, se usa la Tabla N°4.19
    if S_key in DUCTO_N419_MM2:
        d = DUCTO_N419_MM2[S_key][n]  # la sección está directo en la tabla
    else:
        # Intentar por AWG equivalente (Tabla N°4.20)
        awg = MM2_A_AWG.get(S_key)  # busca el calibre AWG equivalente a esta sección
        if awg is None:
            return None  # no se encontró equivalencia, no se puede calcular
        if awg not in DUCTO_N420_AWG:
            return None
        d = DUCTO_N420_AWG[awg][n]
    return max(32, d)  # nunca menor a 32mm (mínimo práctico)

# =========================
# HOJA 2: LISTADO DE MATERIALES (PRO + MARCAS + CONTEO REAL)
# =========================
def build_materiales_df(circuitos_df, texto_omni, ambientes_df, group_info, tipo_canalizacion,tipo_alimentador,
                        tipo_acometida, tipo_instalacion_empalme, requiere_mastil, acometida_txt, interruptor_texto,
                        longitud_transformador_empalme, canalizacion_txt, dist_empalme_pt1, dist_tda_pt2,
                        longitud_subterraneo_medidor, longitud_mastil, altura_acometida_aerea, longitud_subterraneo_medidor2,
                        longitud_llegada_aerea_tda, longitud_alimentador, longitud_abrazaderas_alimentador,
                        longitud_poste_alimentador_aereo, dist_vertical_acometida, alim_txt,
                        circuitos_climatizacion=None, items_por_nombre=None, cajas_adic_por_nombre=None,
                        tipo_dif_por_circ=None,
                        n_barras_pt1=1, n_barras_pt2=1,
                        long_cond_desnudo_pt1=0, long_cond_desnudo_pt2=0):
    """
    Esta es la función grande: arma toda la hoja "Materiales" del Excel
    a partir de los circuitos que ya se definieron antes.

    Recibe los circuitos (circuitos_df), los ambientes (ambientes_df),
    los datos del empalme/acometida/alimentador, la climatización, y
    algunos datos auxiliares más (los ítems de cada circuito, las cajas
    adicionales, cómo quedaron agrupados los diferenciales, etc.)

    Va calculando todo circuito por circuito, en este orden (cada bloque
    está marcado en el código con "# ===== NOMBRE SECCIÓN ====="):
      1. Canalización (conduit o canaleta), abrazaderas y salidas de caja.
      2. Conductores de cada circuito y sus chicotes.
      3. Cajas de derivación, uniones, cajas para enchufes/interruptores/
         luminarias.
      4. Protecciones: TM y diferencial de cada circuito y del tablero.
      5. Ferrules, borneras y cableado interior del tablero.
      6. Materiales de empalme y de acometida/alimentador.
      7. Puesta a tierra (barra copperweld, camarilla, conductor desnudo).
      8. Tornillos, tarugos, prensaestopas.
      9. Con todo eso calculado, arma cada fila final con add_row().

    Al final devuelve un DataFrame de pandas: una fila por material, listo
    para escribirse en la hoja "Materiales".
    """
    filas = []
    item = 1
    # =========================
    # ACUMULADORES TORNILLERÍA
    # =========================
    _abrazaderas_total = 0       # embutida: cantidad de abrazaderas
    _clima_items = []             # materiales climatización distribuidos en secciones
    _long_sobrepuesta_total = 0  # sobrepuesta: 1 tornillo por metro (ceil L)
    _cajas_total = 0             # total cajas derivación
    _tapas_ciegas_total = 0      # total tapas ciegas
    _cajas_paso_total = 0        # cajas de paso por tramos > 20m (RIC 7.16.1.13)
    _mm_conduit_ilumin = 20      # diámetro conduit circuito iluminación (default 20mm)
    _riel_m = 0                  # 1 o 2 (metros)
    _puestos_tablero = None      # tamaño tablero comercial


    # ====== MARCAS SUGERIDAS ======
    marcas = {
        "Protecciones": "Legrand, Schneider Electric, Bticino",
        "Conductores": "Madeco, Cosesa, Revi",
        "Conectores cónicos": "Globaltronic, Lexo, Mec",
        "Ferrule": "Lexo, Soliot, Genérico",
        "Canaleta PVC (sobrepuesta)": "Grantt, Schneider Electric, Hoffens",
        "Accesorios canaleta PVC": "Grantt, Schneider Electric, Hoffens",
        "Conduit PVC (embutida)": "Hoffens, Revi, Halux",
        "Boquilla bordes redondos": "Halux, Hoffens, Tigre",
        "Prensaestopas": "Lexo, Genérico, Soliot",
        "Abrazaderas": "Hoffens, Tigre, Halux",
        "Tablero embutido": "Lexo, Stanford, Tibox",
        "Tablero sobrepuesto": "Lexo, Stanford, Tibox",
        "Barra repartidora": "Lexo, DTK, Cabur",
        "Riel DIN": "Lexo, Mec, Stanford",
        "Enchufes": "Bticino, MEC, Schneider Electric",
        "Interruptores": "MEC, Schneider Electric, Bticino",
        "Iluminarias": "Dairu, Eglo, AVC",
        "Portalamparas": "Bticino, Schneider Electric, Genérico",
        "Cajas derivación embutidas": "Schneider Electric, Bticino, Hoffens",
        "Cajas derivación sobrepuestas": "Schneider Electric, Bticino, Legrand",
        "Cajas de paso estancas": "Gewiss, Scame, Lexo",
        "Tapa ciega": "Schneider Electric, Lexo, Bticino",
        "Tapa ciega octogonal": "Schneider Electric, Epem, Bticino",
        # (embutida accesorios)
        "Salida de caja conduit": "Halux, Hoffens, Tigre",
        "Abrazadera conduit": "Hoffens, Tigre, Halux",
        "Barra unipolar verde": "Lexo, DTK, Cabur",
        "Supresor de transiente": "Schneider Electric, Legrand, ABB",
        "Luz Piloto": "Schneider Electric, Lexo, Bticino",
        "Portafusible": "Lexo, Legrand, Bticino",
        "Fusible": "Bussmann, Mersen, Schneider Electric",
        "Tornillos": "Fixser; Mamut; Fischer",
        "Espuma expansiva PU": "Sika Boom; Fischer; Ceresita",
        "Medidor empalme": "Lexo, Metertek, DRL",
        "Caja empalme": "Metertek, Saime, Stanford",
        "Disyuntor empalme": "Saime, Legrand, Schneider Electric",
        "Cable acometida": "Madeco, Covisa, Elexor",
        "Tubo conduit galvanizado acometida": "Lexo, Ekoline, Rhona",
        "Cabeza de servicio": "Ekoline, Elexor, Rhona",
        "Cancamo abierto": "Mamut, Fixser, Rhona",
        "Granpa de retención": "Elexor, Tecnored, Standford",
        "Mordaza acometida": "Generico, Elexor, Rema",
        "Conector HUB": "Bekam, Enelux, Ekoline",
        "Terminal PVC conduit con 2 tuercas": "Hoffens, Halux, Revi",
        "Conductor THWN-2 blanco": "Madeco, Covisa, Nexans",
        "Conductor THWN-2 verde": "Madeco, Covisa, Nexans",
        "Barra copperweld": "Ekoline, Generico, Indutref",
        "Prensaestopa": "Legrand, Bticino, Generico",
        # ── AGUA CALIENTE ──────────────────────────────────────────────────────
        "Diferencial 10mA agua caliente": "Legrand, Schneider Electric, Bticino",
        "Tablero externo agua caliente":  "Lexo, Stanford, Tibox",
        "TM bipolar agua caliente":       "Legrand, Schneider Electric, Bticino",
        "Bornera PE agua caliente":       "Legrand, Phoenix Contact, Weidmuller",
        "Bornera de conexion":            "Lexo, DTK, Cabur",
        "Estaño": "Golden, Macrotel, Altronics",
        "Pasta para soldar": "Metalfer, Indepp, MCT",
        "Conductor desnudo Cu": "Madeco, Covisa, Nexans",
        "Camarilla PVC naranjo": "Generico, DTK, Tigre",
        "Abrazadera tipo caddy": "Lexo, Rhona, Ekoline",
        "Terminal compresion tipo ojo": "Generico, Soliot, Sofamel",
        "Sellador de roscas": "Dura, Permatex, Topex",
        "Terminal ferrul acometida": "Generico, Soliot, Lexo",
        "Terminal ferrul alimentador": "Generico, Soliot, Lexo",
        "Terminal ferrul doble alimentador": "Generico, Soliot, Lexo",
        "Tirafondo hexagonal madera": "Mamut, Fixser, Fischer",
        "Tornillo autoperforante hexagonal": "Mamut, Fixser, Fischer",
        "Alimentador RV-K": "Nexans, Cocesa, Madeco",
        "Conduit PVC": "Hoffens, Halux, Revi",
        "Abrazadera conduit PVC alimentador": "Hoffens, Tigre, Halux",
        "Tornillo punta fina madera": "Fixser, Mamut, Fischer",
        "Tornillo autoperforante broca": "Fixser, Mamut, Fischer",
        "Caja derivacion metalica": "Ectray, Generico, Standford",
        "Tarugo paloma": "Mamut, Fixser, Fischer",
        "Tornillo para tarugo paloma": "Mamut, Fixser, Fischer",
        "Portafusible de loza": "Lexo, Fujian, Kersting",
        "Poste madera": "Imperial, Lifewood, Sodimac",
        "Pilar metalico": "Genérico, ACMA, Kupfer",
        "Protector sobrevoltaje": "CNC, TOMZN, Sinotimer",
        "Cinta autofundente goma": "3M, Lexo, Truper",
        "Cinta aislante PVC": "3M, Lexo, Truper",
        "Camara tipo C": "Prefabricados del Sur, Hormipret, Genérico",
        "Marco metalico camara C": "Genérico, ACMA, Prefabricados del Sur",
        "Boquilla camara tipo C": "Hoffens, Halux, Tigre",
    }

    def formatear_canalizacion_material(canal_txt: str, es_sobrepuesta: bool) -> str:
        """
        Ejemplo: "Embutida (PVC conduit 16mm)" queda como "Canalización embutida PVC conduit 16mm"
        Ejemplo: "Sobrepuesta (Canaleta de PVC 20x10mm)" queda como "Canalización sobrepuesta Canaleta de PVC 20x10mm, 2mts"
        """
        t = (canal_txt or "").strip()
        if not t:
            return "Canalización no definida"  # no vino texto
        t = " ".join(t.split())  # saca espacios dobles

        if "(" in t and ")" in t:
            tipo = t.split("(")[0].strip().lower()  # "embutida" o "sobrepuesta"
            dentro = t.split("(", 1)[1].split(")")[0].strip()  # lo que va entre paréntesis (ej: "PVC conduit 16mm")
            base = f"Canalización {tipo} {dentro}"
        else:
            base = f"Canalización {t.lower()}"  # no tenía paréntesis, se deja tal cual

        if es_sobrepuesta:
            return f"{base}, 2mts"  # la canaleta sobrepuesta se vende en tramos de 2 metros
        return base

    def _extraer_medida_canaleta(canal_txt: str) -> str:
        """
        Saca la medida "AAxBBmm" del texto de canalización, para que los
        accesorios (unión copla, curvas, curva T) usen la MISMA medida que
        la canaleta real de ese circuito, en vez de un valor fijo.
        Ejemplo: "Sobrepuesta (Canaleta de PVC 40x16mm)" da "40x16mm"
        Si no encuentra el patrón, usa "20x10mm" como valor por defecto.
        """
        import re
        m = re.search(r'\d+x\d+mm', canal_txt or "")
        return m.group(0) if m else "20x10mm"

    def _extraer_mm_conduit(canal_txt: str):
        """
        Ejemplo: "Embutida (PVC conduit 20mm)" da 20
        """
        t = (canal_txt or "").lower().replace(" ", "")  # sin espacios, para buscar más fácil
        # buscar "...conduit20mm" o "conduit16mm"
        if "conduit" in t and "mm" in t:
            try:
                seg = t.split("conduit", 1)[1]  # lo que viene después de la palabra "conduit"
                num = ""
                for ch in seg:
                    if ch.isdigit():
                        num += ch  # va juntando los dígitos del número
                    else:
                        break  # se acabó el número
                return int(num) if num else None
            except:
                return None
        return None  # el texto no tenía "conduit...mm"

    def add_section(nombre):
        # agrega una fila "separadora" con solo el título de la sección
        # (ej: "Empalme"), para que se vea ordenado el Excel.
        filas.append({
            "Ítem": "",
            "Descripción técnica": nombre,
            "Marcas sugeridas": "",
            "Sello SEC": "",
            "Norma / RIC": "",
            "Circuito": "",
            "Unidad": "",
            "K": "",
            "Longitud (m) / Unidad": "",
            "Cantidad": ""
        })

    def add_row(desc, marcas_txt, norma, circuito, unidad, k, longitud_m, cantidad):
        # agrega una fila de material al listado final: le pone el número de
        # ítem, revisa si necesita Sello SEC y guarda todos los datos.
        nonlocal item
        sello = "SEC" if str(norma).strip() == "SEC" else "-"  # marca si el material requiere Sello SEC
        filas.append({
            "Ítem": item,
            "Descripción técnica": desc,
            "Marcas sugeridas": marcas_txt,
            "Sello SEC": sello,
            "Norma / RIC": norma,
            "Circuito": circuito,
            "Unidad": unidad,
            "K": k,
            "Longitud (m) / Unidad": longitud_m,
            "Cantidad": cantidad
        })
        item += 1  # sube el número de ítem para la próxima fila

    # -----------------------------
    # Conectores cónicos
    # -----------------------------
    def extraer_seccion_mm2(txt_conductor: str):
        """
        Ejemplo: "H07V 2.5 mm²" da 2.5
        Ejemplo: "THWN-2 1.5 mm2" da 1.5
        """
        s = (txt_conductor or "").lower().replace("mm²", "mm2")  # unifica el símbolo mm²
        toks = s.split()  # separa el texto en palabras
        for i, tok in enumerate(toks):
            if tok == "mm2" and i > 0:
                try:
                    return float(toks[i-1].replace(",", "."))  # el número justo antes de "mm2"
                except:
                    pass
        # si no lo encontró así, junta cualquier número suelto en el texto
        num = ""
        for ch in s:
            if ch.isdigit() or ch in [".", ","]:
                num += ch  # va juntando dígitos y separadores decimales
            elif num:
                break  # ya encontró un número, se detiene
        try:
            return float(num.replace(",", "."))
        except:
            return None  # no se pudo sacar ningún número

    def conico_por_seccion(seccion_mm2, n_cables=3):
        """
        Tabla de colores de conectores cónicos, según catálogo real, en
        función de la SECCIÓN del cable y la CANTIDAD DE CABLES que se
        unen en ese punto (2 o 3):

          Sección      | 2 cables            | 3 cables
          -------------|---------------------|---------------------
          1.5mm²       | N°33 (P73) Naranja  | N°33 (P73) Naranja
          2.5mm²       | N°44 (P74) Amarillo | N°44 (P74) Amarillo
          4.0mm²       | N°66 (P75) Rojo     | N°66 (P75) Rojo
          5.26mm²(10AWG)| N°66 (P75) Rojo    | N°88 (SP8/P78) Gris Grande
          6.0mm²       | N°66 (P75) Rojo     | N°88 (SP8/P78) Gris Grande

        No existen conectores cónicos para secciones mayores a 6.0mm².
        La sección real se redondea hacia arriba a la más cercana de la tabla.
        """
        try:
            s = float(seccion_mm2)  # s: la sección real del cable, convertida a número
        except:
            s = None

        if s is None:
            return ("(definir)", None, "")  # sección inválida, no se puede determinar

        if s > 6.0:
            return ("(definir)", None, f"{s} mm² (fuera de rango cónico)")

        # Redondea hacia arriba a la sección comercial de la tabla
        secciones_tabla = [1.5, 2.5, 4.0, 5.26, 6.0]
        sec_com = None  # sec_com: la sección "comercial" que le toca (la más
                         # chica de la tabla que alcanza a cubrir la real)
        for sec in secciones_tabla:
            if s <= sec:
                sec_com = sec
                break
        if sec_com is None:
            sec_com = 6.0

        tabla_2_cables = {
            1.5:  ("Naranja", 33, "P73"),
            2.5:  ("Amarillo", 44, "P74"),
            4.0:  ("Rojo", 66, "P75"),
            5.26: ("Rojo", 66, "P75"),
            6.0:  ("Rojo", 66, "P75"),
        }
        tabla_3_cables = {
            1.5:  ("Naranja", 33, "P73"),
            2.5:  ("Amarillo", 44, "P74"),
            4.0:  ("Rojo", 66, "P75"),
            5.26: ("Gris Grande", 88, "SP8/P78"),
            6.0:  ("Gris Grande", 88, "SP8/P78"),
        }

        tabla = tabla_2_cables if int(n_cables) == 2 else tabla_3_cables
        color, num, pcode = tabla.get(sec_com, ("(definir)", None, ""))  # pcode: código de catálogo (ej. "P73")
        _sec_txt = f"{sec_com}".rstrip("0").rstrip(".") if "." in f"{sec_com}" else f"{sec_com}"
        rango = f"{_sec_txt}mm², {int(n_cables)} cables ({pcode})"
        return (color, num, rango)

    # =========================
    # ACCESORIOS CANALETA SOBREPUESA (PVC, medida variable según sección)
    # Reglas usuario:
    #  ILUMINACIÓN (sobrepuesta):
    #    - Unión copla: cada 2 m, redondeando siempre hacia arriba (ceil(L/2))
    #    - Curvas internas 90°: 4 por ambiente + 1 por interruptor
    #    - Curvas planas 90°: 2 por interruptor del circuito (1 salida tablero
    #      + 2 por cada caja troncal, excepto la última que lleva 1 = 2×n_int)
    #  ENCHUFES (sobrepuesta):
    #    - Unión copla: cada 2 m, redondeando siempre hacia arriba (ceil(L/2))
    #    - Curvas internas 90°: 3 por ambiente
    #    - Curva T: (n° enchufes - 1)
    #    - Curvas planas 90°: 1 salida del circuito + 1 por ambiente
    # =========================
    def _contar_ambientes_en_items(items):
        # cuenta cuántos ambientes distintos hay en los items de un circuito
        ambs = set()  # set para no repetir ambientes
        if isinstance(items, list):
            for it in items:
                if isinstance(it, dict):
                    a = str(it.get("amb", "")).strip()  # nombre del ambiente de este item
                    if a:
                        ambs.add(a.strip().lower())  # lo agrega (sin mayúsculas, sin espacios)
        return len(ambs)  # cuántos ambientes distintos quedaron

    def _contar_enchufes_en_items(items):
        # suma los enchufes del circuito (n_ench). Si un item no trae n_ench,
        # se cuenta como 1
        n = 0
        if isinstance(items, list):
            for it in items:
                if not isinstance(it, dict):
                    continue  # item raro, se salta
                if ("id_ench" in it) or ("modulos" in it) or ("n_ench" in it):
                    n += int(it.get("n_ench", 1) or 1)  # suma la cantidad de este item
        return int(n)

    # ==========================================================
    # CONTEOS POR CIRCUITO PARA "Salida de caja conduit"
    # Reglas embutida:
    #  - abrazaderas: 1 por metro
    #  - salidas:
    #       1 por circuito
    #       2 por caja de enchufe
    #       3 por caja de derivación unión
    #       1 por caja de foco/ampolleta
    #       1 por caja de interruptor (9/12, 9/15, 9/24, 9/32) -> se cuenta #interruptores físicos
    #
    # IMPORTANTÍSIMO:
    #  - NO asumimos uniones en circuitos de enchufes (si no hay, es 0).
    #  - Uniones solo las modelamos para iluminación: 1 unión por luminaria del circuito (como tu modelo).
    # ==========================================================
    # =========================
    # 1) CANALIZACIONES (UNA FILA POR CIRCUITO, NO AGRUPADO)
    #    accesorios embutida (abrazaderas + salidas de caja)
    #    accesorios sobrepuesta (canaleta PVC, medida variable según sección)
    # =========================


    # Definir clima_rows y llenar _clima_items ANTES de todas las secciones
    clima_rows = pd.DataFrame()
    if circuitos_df is not None and "Circuito" in circuitos_df.columns:
        # detecta los circuitos de climatización por el nombre (contiene "climatiz", "aire acond" o "split")
        mask_nombre = circuitos_df["Circuito"].astype(str).str.contains(
            "climatiz|aire.acond|split", case=False, na=False
        )
        if "_es_climatizacion" in circuitos_df.columns:
            # o también por una marca explícita que se le puso al crear el circuito
            mask_flag = circuitos_df["_es_climatizacion"].apply(
                lambda x: bool(x) if x is not None and x is not np.nan else False
            )
        else:
            mask_flag = pd.Series(False, index=circuitos_df.index)
        clima_rows = circuitos_df[mask_nombre | mask_flag].copy()  # se queda con cualquiera de las dos formas

    # ── Llenar _agua_items (análogo a _clima_items) ──────────────────────────
    _agua_items = []
    _KEYWORDS_AGUA = ("ducha", "termo", "calefon", "calefón", "calentador", "agua caliente")
    if circuitos_df is not None and "Circuito" in circuitos_df.columns:
        # mismo criterio que climatización, pero buscando palabras de agua caliente
        mask_agua_nombre = circuitos_df["Circuito"].astype(str).str.contains(
            "|".join(_KEYWORDS_AGUA), case=False, na=False
        )
        if "_es_agua_caliente" in circuitos_df.columns:
            mask_agua_flag = circuitos_df["_es_agua_caliente"].apply(
                lambda x: bool(x) if x is not None and x is not np.nan else False
            )
        else:
            mask_agua_flag = pd.Series(False, index=circuitos_df.index)
        agua_rows = circuitos_df[mask_agua_nombre | mask_agua_flag].copy()

        # recorre cada circuito de agua caliente encontrado y arma su canalización
        for _, r_ac in agua_rows.iterrows():
            circ_ac_i  = str(r_ac.get("Circuito", "")).strip()
            L_ac_i     = float(r_ac.get("Longitud (m)", 0) or 0)
            canal_ac_i = str(r_ac.get("Canalización", "")).strip()
            tm_ac_i    = str(r_ac.get("Disyuntor termomagnético", "")).strip()
            in_tm_ac_i = parse_in_tm(tm_ac_i)  # None si no se puede leer — no se asume 20A
            if L_ac_i <= 0:
                continue  # sin longitud, no hay nada que calcular para este equipo
            canal_low_ac  = canal_ac_i.lower()

            # Forzar conduit embutido si el equipo está en Vol.1
            # RIC N°11 art. 6.4.3: en Volumen 1 se exige mínimo IPX4, y la canaleta NO cumple ese requisito
            # Art. 6.5.3: cable bajo tubo aislante con IPX5 garantizado
            _vol1_ac = False
            if "_vol1_bano_agua" in agua_rows.columns:
                _vol1_ac = bool(r_ac.get("_vol1_bano_agua", False))
            if not _vol1_ac:
                # si no viene ese dato, se asume por el nombre: la ducha siempre está en Vol.1
                _vol1_ac = "ducha" in circ_ac_i.lower()

            if _vol1_ac and ("canaleta" in canal_low_ac or "sobrepuesta" in canal_low_ac):
                # está en volumen 1 pero eligieron canaleta: se fuerza a conduit embutido por norma
                canal_low_ac = "conduit embutida"
                canal_ac_i   = "Embutida (conduit forzado — RIC N°11 6.4.3 Vol.1 IPX4)"

            es_emb_ac_i   = ("conduit" in canal_low_ac) or ("embutida" in canal_low_ac)
            mm_ac_i       = _extraer_mm_conduit(canal_ac_i) or 20  # diámetro del conduit (20mm si no se pudo leer)
            # Buscar si lleva tablero externo
            _lleva_te = False
            for eq_ac_i in circuitos_agua_caliente:
                if eq_ac_i.get("nombre_circ", "").lower() in circ_ac_i.lower():
                    _lleva_te = eq_ac_i.get("lleva_tablero_externo", False)
                    break
            _tiene_20m_ac = bool(r_ac.get("_tiene_tramo_20m", True))
            _cajas_paso_ac = int(L_ac_i // 20) if (es_emb_ac_i and _tiene_20m_ac) else 0
            _agua_items.append({
                "circ":               circ_ac_i,
                "L":                  L_ac_i,
                "canal":              canal_ac_i,
                "in_tm":              in_tm_ac_i,
                "es_emb":             es_emb_ac_i,
                "mm":                 mm_ac_i,
                "lleva_tab_ext":      _lleva_te,
                "cajas_paso":         _cajas_paso_ac,
            })  # guarda todos los datos de este equipo de agua caliente
            _cajas_paso_total += _cajas_paso_ac  # cajas de paso de agua caliente también cuentan para el total
            # Este ítem se suma más abajo vía _agua_cajas_emb/_agua_cajas_sob,
            # una vez que ya se sabe si la canalización es embutida o sobrepuesta.

    def _buscar_datos_clima(nombre_circ):
        # busca el equipo de climatización que corresponde a este circuito,
        # para sacar sus datos técnicos (LRA, tipo de compresor, etc.)
        if not circuitos_climatizacion:
            return {}  # no hay equipos de clima definidos
        for eq in circuitos_climatizacion:
            if eq.get("nombre_circ", "").lower() in nombre_circ.lower():  # coincide el nombre
                return eq
        return {}  # no se encontró ningún equipo para este circuito

    # recorre cada circuito de climatización encontrado y arma su canalización
    if len(clima_rows) > 0:
        for _, r_cl in clima_rows.iterrows():
            circ_cl  = str(r_cl.get("Circuito", "")).strip()
            L_cl     = float(r_cl.get("Longitud (m)", 0) or 0)
            canal_cl = str(r_cl.get("Canalización", "")).strip()
            tm_cl    = str(r_cl.get("Disyuntor termomagnético", "")).strip()
            in_tm_cl_raw = parse_in_tm(tm_cl)
            in_tm_cl = in_tm_cl_raw if in_tm_cl_raw is not None else 16  # si no se pudo leer, asume 16A
            if L_cl <= 0:
                continue  # sin longitud, no hay nada que calcular
            canal_low   = canal_cl.lower()
            es_emb_cl   = ("conduit" in canal_low) or ("embutida" in canal_low)
            mm_cl       = _extraer_mm_conduit(canal_cl) or 20
            # Si no se pudo leer el TM, no se adivina si tiene enchufe o no —
            # con_enchufe queda en None para que más abajo (en la parte que
            # recorre los cónicos y en el bloque C2 de ferrules) se genere un
            # aviso "(definir)" en vez de asumir un valor que puede estar mal.
            con_enchufe = (in_tm_cl <= 16) if in_tm_cl_raw is not None else None
            _tiene_20m_cl = bool(r_cl.get("_tiene_tramo_20m", True))
            _cajas_paso_cl = int(L_cl // 20) if (es_emb_cl and _tiene_20m_cl) else 0
            _clima_items.append({
                "circ":        circ_cl,
                "L":           L_cl,
                "canal":       canal_cl,
                "in_tm":       in_tm_cl,
                "es_emb":      es_emb_cl,
                "mm":          mm_cl,
                "con_enchufe": con_enchufe,
                "cajas_paso":  _cajas_paso_cl,
            })
            _cajas_paso_total += _cajas_paso_cl  # cajas de paso de climatización también cuentan para el total
            # Este ítem se suma más abajo vía _clima_cajas_emb/_clima_cajas_sob,
            # una vez que ya se sabe si la canalización es embutida o sobrepuesta.
            # _tapas_ciegas_total se calcula correctamente en bloque TAPAS CIEGAS más abajo

    # Cajas de paso cada 20m (RIC 7.16.1.13) para enchufes, iluminación y
    # especiales. Climatización y agua caliente no entran aquí porque ya
    # calculan sus propias cajas de paso más arriba (para no duplicarlas).
    if {"Canalización", "Longitud (m)", "Circuito"}.issubset(set(circuitos_df.columns)):
        _tmp_pre = circuitos_df.copy()
        _tmp_pre["Longitud (m)"] = pd.to_numeric(_tmp_pre["Longitud (m)"], errors="coerce").fillna(0)
        for _, _r_pre in _tmp_pre.iterrows():
            _canal_pre = str(_r_pre.get("Canalización", "")).strip().lower()
            _circ_pre  = str(_r_pre.get("Circuito", "")).strip().lower()
            _L_pre     = float(_r_pre.get("Longitud (m)", 0))
            _es_emb_pre  = ("conduit" in _canal_pre) or ("embutida" in _canal_pre)
            _es_cli_pre  = any(k in _circ_pre for k in ("climatiz", "aire acond", "split"))
            _es_agua_pre = any(k in _circ_pre for k in _KEYWORDS_AGUA)
            # Si no hay ningún tramo continuo >=20m (confirmado en el input),
            # no se cuenta ninguna caja de paso aunque el LARGO TOTAL del
            # circuito supere 20m (puede ser la suma de varios tramos cortos)
            _tiene_20m_pre = bool(_r_pre.get("_tiene_tramo_20m", True))
            if _es_emb_pre and not _es_cli_pre and not _es_agua_pre and _L_pre > 0 and _tiene_20m_pre:
                _cajas_paso_total += int(_L_pre // 20)  # 1 caja de paso cada 20m de canalización embutida

    add_section("Canalizaciones")
    # ncon_map_local solo depende de ambientes_df, disponible desde el inicio de la función.
    ncon_map_local = {}
    if ambientes_df is not None and "Ambiente" in ambientes_df.columns:
        amb_col_l = ambientes_df["Ambiente"].astype(str).str.strip()
        if "N_conmutadas_924 (u)" in ambientes_df.columns:
            ncon_col_l = pd.to_numeric(ambientes_df["N_conmutadas_924 (u)"], errors="coerce").fillna(0).astype(int)
            for a_l, nval_l in zip(amb_col_l, ncon_col_l):
                ncon_map_local[a_l.lower()] = int(max(0, nval_l))  # cuántas luminarias conmutadas tiene cada ambiente

    # recorre TODOS los circuitos (uno por fila del Excel) y arma su canalización
    if {"Canalización", "Longitud (m)", "Circuito"}.issubset(set(circuitos_df.columns)):
        circuitos_tmp = circuitos_df.copy()
        circuitos_tmp["Longitud (m)"] = pd.to_numeric(circuitos_tmp["Longitud (m)"], errors="coerce").fillna(0)
        for _row_idx, r in circuitos_tmp.iterrows():
            canal = str(r.get("Canalización", "")).strip()
            circ = str(r.get("Circuito", "")).strip()
            L = float(r.get("Longitud (m)", 0))

            if not canal or not circ or L <= 0:
                continue  # circuito sin datos válidos, se salta
            canal_low = canal.lower()
            es_sobrepuesta = ("canaleta" in canal_low) or ("sobrepuesta" in canal_low)
            es_embutida = ("conduit" in canal_low) or ("embutida" in canal_low)

            # Climatización: el conduit/canaleta se genera aquí normalmente,
            # pero los accesorios (abrazaderas, salidas, coplas, curvas) se generan
            # en la sección "Climatización" con la lógica correcta, así que acá se saltan los accesorios
            _es_clima_canal = False
            if "_es_climatizacion" in circuitos_tmp.columns:
                _es_clima_canal = bool(r.get("_es_climatizacion", False))
            if not _es_clima_canal:
                _es_clima_canal = any(k in circ.lower() for k in ("climatiz", "aire acond", "split"))

            # Agua caliente: igual que climatización — accesorios (boquillas, salidas de caja)
            # se generan en su propio bloque _agua_items, así que acá se saltan las salidas genéricas
            _es_agua_canal = False
            if "_es_agua_caliente" in circuitos_tmp.columns:
                _es_agua_canal = bool(r.get("_es_agua_caliente", False))
            if not _es_agua_canal:
                _es_agua_canal = any(k in circ.lower() for k in _KEYWORDS_AGUA)
            # fila de canalización
            _norma_canal = (
                "RIC 7 (7.5.1, 7.5.2) / RIC 4 (7.1.3, 7.1.8, 7.1.9, 7.1.10, 7.1.15)"
                if (_es_clima_canal or _es_agua_canal)
                else "RIC 4.7.2"
            )
            if es_sobrepuesta:
                # canaleta: se compra en tramos de 2 metros, se redondea hacia arriba
                tramo_m = 2.0
                cantidad_tramos = int(math.ceil(L / tramo_m))
                add_row(
                    desc=formatear_canalizacion_material(canal, es_sobrepuesta=True),
                    marcas_txt=marcas.get("Canaleta PVC (sobrepuesta)", ""),
                    norma=_norma_canal,
                    circuito=circ,
                    unidad="u",
                    k=1,
                    longitud_m=round(L, 4),
                    cantidad=cantidad_tramos
                )
                _long_sobrepuesta_total += cantidad_tramos * 2  # 1 tornillo por metro FÍSICO comprado (no el mínimo teórico)

                # Climatización: los accesorios (coplas, curvas) se generan en la sección Climatización, así que acá se saltan
                if _es_clima_canal:
                    continue
                # Agua caliente: los accesorios se generan en el bloque _agua_items, así que acá se saltan
                if _es_agua_canal:
                    continue

                # =========================
                # ACCESORIOS CANALETA PVC (SOBREPUESTA, medida variable según sección)
                # =========================
                # Obtener _items desde items_por_idx (más fiable que desde el DataFrame)
                _circ_key = str(r.get("Circuito",""))
                items = (items_por_nombre or {}).get(_circ_key, [])  # los puntos (luminarias/enchufes) de este circuito
                if not isinstance(items, list):
                    items = []
                # Unión copla: 1 copla por cada unión entre tramos de 2m, o sea n_tramos - 1
                _n_tramos = int(math.ceil(L / 2.0))
                union_copla = max(0, _n_tramos - 1)
                # contar ambientes / focos / enchufes del circuito
                n_amb = _contar_ambientes_en_items(items)
                n_ench = _contar_enchufes_en_items(items)
                # determinar si el circuito es iluminación o enchufe (por nombre)
                circ_low = (circ or "").lower()
                es_circ_ilum = ("ilumin" in circ_low)
                es_circ_ench = ("enchufe" in circ_low)
                # Unión copla (sirve para ambos: ilum/ench)
                if union_copla > 0:
                    add_row(
                        desc=f"Unión copla para canaleta de PVC {_extraer_medida_canaleta(canal)}",
                        marcas_txt=marcas.get("Accesorios canaleta PVC", ""),
                        norma="RIC 4.7.2",
                        circuito=circ,
                        unidad="u",
                        k=1,
                        longitud_m=f"{union_copla} unid",
                        cantidad=union_copla
                    )
                # Accesorios según tipo de circuito
                if es_circ_ilum:
                    # Curvas internas 90°: 4 por ambiente + 1 por interruptor
                    # (1 curva por bajada desde caja rect hacia primera oct del grupo)
                    _n_int_circ_real = 0  # cuenta cuántos interruptores físicos hay en total en el circuito
                    _ambs_vistos = set()  # para no procesar el mismo ambiente dos veces
                    for _it in items:
                        if not isinstance(_it, dict): continue
                        if "nombre" in _it or "id_ench" in _it or "modulos" in _it or "n_ench" in _it: continue  # salta enchufes/especiales
                        _amb_k = str(_it.get("amb", "")).strip().lower() or "sin_amb"
                        if _amb_k in _ambs_vistos: continue  # este ambiente ya se contó
                        _ambs_vistos.add(_amb_k)
                        # cuenta cuántas luminarias tiene este ambiente en este circuito
                        _n_lum_amb_c = sum(1 for x in items if isinstance(x, dict) and
                                          str(x.get("amb","")).strip().lower() == _amb_k and
                                          "nombre" not in x and "id_ench" not in x and
                                          "modulos" not in x and "n_ench" not in x)
                        _n_conm_amb_c = int(min(_n_lum_amb_c, ncon_map_local.get(_amb_k, 0)))  # cuántas son conmutadas
                        _n_rest_amb_c = _n_lum_amb_c - _n_conm_amb_c  # el resto, no conmutadas
                        _c12c, _c15c, _c32c = descomponer_interruptores(_n_rest_amb_c)  # cómo se agrupan
                        _n_int_circ_real += _c12c + _c15c + _c32c + (2 if _n_conm_amb_c > 0 else 0)  # 2 interruptores por conmutado

                    curvas_int_90 = int(4 * int(n_amb) + _n_int_circ_real)  # 4 curvas por ambiente + 1 por interruptor
                    # Curvas planas 90°:
                    # 1 salida tablero + 2 por cada caja troncal (excepto última del circuito que lleva 1)
                    # = 1 + 2*(n_int-1) + 1 = 2*n_int
                    curvas_plan_90 = 2 * _n_int_circ_real
                    if curvas_int_90 > 0:
                        add_row(
                            desc=f"Curvas internas en 90° para canaletas de PVC {_extraer_medida_canaleta(canal)}",
                            marcas_txt=marcas.get("Accesorios canaleta PVC", ""),
                            norma="RIC 4.7.2",
                            circuito=circ,
                            unidad="u",
                            k=1,
                            longitud_m=f"{curvas_int_90} unid",
                            cantidad=curvas_int_90
                        )
                    if curvas_plan_90 > 0:
                        add_row(
                            desc=f"Curvas planas en 90° para canaletas de PVC {_extraer_medida_canaleta(canal)}",
                            marcas_txt=marcas.get("Accesorios canaleta PVC", ""),
                            norma="RIC 4.7.2",
                            circuito=circ,
                            unidad="u",
                            k=1,
                            longitud_m=f"{curvas_plan_90} unid",
                            cantidad=curvas_plan_90
                        )
                elif es_circ_ench:
                    # Curvas internas 90°: 3 por ambiente
                    curvas_int_90 = int(3 * int(n_amb))
                    # Curva T: n° enchufes - 1 (mínimo 0)
                    curva_t = int(max(0, int(n_ench) - 1))
                    # Curvas planas 90°: 1 salida del circuito + 1 por cada
                    # ambiente (el último enchufe de cada ambiente)
                    curvas_plan_90 = 1 + int(n_amb)
                    if curvas_int_90 > 0:
                        add_row(
                            desc=f"Curvas internas en 90° para canaletas de PVC {_extraer_medida_canaleta(canal)}",
                            marcas_txt=marcas.get("Accesorios canaleta PVC", ""),
                            norma="RIC 4.7.2",
                            circuito=circ,
                            unidad="u",
                            k=1,
                            longitud_m=f"{curvas_int_90} unid",
                            cantidad=curvas_int_90
                        )
                    if curva_t > 0:
                        add_row(
                            desc=f"Curva T para canaletas de PVC {_extraer_medida_canaleta(canal)}",
                            marcas_txt=marcas.get("Accesorios canaleta PVC", ""),
                            norma="RIC 4.7.2",
                            circuito=circ,
                            unidad="u",
                            k=1,
                            longitud_m=f"{curva_t} unid",
                            cantidad=curva_t
                        )
                    if curvas_plan_90 > 0:
                        add_row(
                            desc=f"Curvas planas en 90° para canaletas de PVC {_extraer_medida_canaleta(canal)}",
                            marcas_txt=marcas.get("Accesorios canaleta PVC", ""),
                            norma="RIC 4.7.2",
                            circuito=circ,
                            unidad="u",
                            k=1,
                            longitud_m=f"{curvas_plan_90} unid",
                            cantidad=curvas_plan_90
                        )
                else:
                    # Especiales genéricos (lavadora, horno, etc. — no
                    # iluminación ni enchufes por nombre, y clima/agua ya se
                    # excluyeron arriba con continue): mismo criterio que
                    # enchufes — 3 curvas internas por ambiente, sin curva T
                    # (1 solo punto de conexión, no hay cadena entre varios
                    # enchufes), 1 salida del circuito + 1 por ambiente
                    curvas_int_90 = int(3 * int(n_amb))
                    curvas_plan_90 = 1 + int(n_amb)
                    if curvas_int_90 > 0:
                        add_row(
                            desc=f"Curvas internas en 90° para canaletas de PVC {_extraer_medida_canaleta(canal)}",
                            marcas_txt=marcas.get("Accesorios canaleta PVC", ""),
                            norma="RIC 4.7.2",
                            circuito=circ,
                            unidad="u",
                            k=1,
                            longitud_m=f"{curvas_int_90} unid",
                            cantidad=curvas_int_90
                        )
                    if curvas_plan_90 > 0:
                        add_row(
                            desc=f"Curvas planas en 90° para canaletas de PVC {_extraer_medida_canaleta(canal)}",
                            marcas_txt=marcas.get("Accesorios canaleta PVC", ""),
                            norma="RIC 4.7.2",
                            circuito=circ,
                            unidad="u",
                            k=1,
                            longitud_m=f"{curvas_plan_90} unid",
                            cantidad=curvas_plan_90
                        )
                continue
            elif es_embutida:
                # conduit: se compra en tramos de 3 metros, se redondea hacia arriba
                tramo_m = 3.0
                cantidad_tramos = int(math.ceil(L / tramo_m))
                desc_base = formatear_canalizacion_material(canal, es_sobrepuesta=False)
                desc_final = f"{desc_base}, 3mts"
                add_row(
                    desc=desc_final,
                    marcas_txt=marcas.get("Conduit PVC (embutida)", ""),
                    norma=_norma_canal,
                    circuito=circ,
                    unidad="u",
                    k=1,
                    longitud_m=f"{round(L, 4)}m",
                    cantidad=cantidad_tramos
                )
                # accesorios canalización embutida por circuito
                mm = _extraer_mm_conduit(canal) or 20
                # Capturar mm conduit de iluminación para cajas y tapas
                if "ilumin" in circ.lower():
                    _mm_conduit_ilumin = mm
                # el conduit se compra en tramos de 3m completos, así que la
                # cantidad real instalada es cantidad_tramos*3 (no el largo
                # medido L) — las abrazaderas van sobre esa cantidad
                L_efectiva = cantidad_tramos * tramo_m
                # 1) Abrazaderas: separación según Tabla N°4.24 (tuberías no metálicas)
                if mm <= 25:
                    sep_abraz = 1.20   # de 16 a 25mm, van cada 1,20m
                else:
                    sep_abraz = 1.50   # de 32 a 63mm, van cada 1,50m
                cant_abraz = int(math.ceil(L_efectiva / sep_abraz))  # cantidad de abrazaderas para todo el tramo instalado
                add_row(
                    desc=f"Abrazadera conduit de PVC de {mm}mm",
                    marcas_txt=marcas.get("Abrazadera conduit", ""),
                    norma="RIC 4 (Tabla N°4.24)",
                    circuito=circ,
                    unidad="u",
                    k=1,
                    longitud_m=f"{cant_abraz} unid",
                    cantidad=cant_abraz
                )
                _abrazaderas_total += cant_abraz

                # Cajas de paso intermedias (RIC 7.16.1.13): 1 cada 20m de tramo
                # Se agrupan con las demás cajas rectangulares en una sola fila

                # Climatización embutida: las salidas de caja se generan en la sección Climatización, así que acá se saltan
                if _es_clima_canal:
                    continue
                # Agua caliente embutida: las boquillas y salidas se generan en el bloque _agua_items, así que acá se saltan
                if _es_agua_canal:
                    continue

                # Salida de caja conduit: reglas por circuito (con 2 por enchufe)
                # Obtener _items desde items_por_idx
                _circ_key2 = str(r.get("Circuito",""))
                items = (items_por_nombre or {}).get(_circ_key2, [])
                if not isinstance(items, list):
                    items = []
                n_ench_circ = 0  # cantidad de enchufes del circuito
                n_lum_circ = 0   # cantidad de luminarias del circuito
                # contar enchufes/luminarias del circuito desde _items
                if isinstance(items, list):
                    for it in items:
                        if not isinstance(it, dict):
                            continue
                        # enchufes
                        if ("id_ench" in it) or ("modulos" in it) or ("n_ench" in it):
                            n_ench_circ += int(it.get("n_ench", 1) or 1)
                            continue
                        # especiales NO cuentan
                        if "nombre" in it:
                            continue
                        # luminaria
                        n_lum_circ += 1
                # Calcular salidas de caja iterando por tipo de caja
                # Cajas de paso intermedias (RIC 7.16.1.13): 1 cada 20m de
                # tramo, solo si el circuito tiene un tramo continuo >=20m
                # confirmado en el input (no basta con que el largo total
                # del circuito supere 20m)
                _tiene_20m_salidas = bool(r.get("_tiene_tramo_20m", True))
                cajas_paso_circ = int(L // 20) if _tiene_20m_salidas else 0
                salidas = 1  # 1 salida tablero origen

                if "ilumin" in circ.lower():
                    # Iluminación: iterar por ambiente y por interruptor
                    lum_by_amb_s = {}
                    for it in items:
                        if not isinstance(it, dict): continue
                        if "nombre" in it or "id_ench" in it or "modulos" in it or "n_ench" in it: continue
                        amb = str(it.get("amb", "")).strip().lower() or "sin_amb"
                        if amb not in lum_by_amb_s: lum_by_amb_s[amb] = []
                        lum_by_amb_s[amb].append(it)  # agrupa las luminarias por ambiente

                    n_amb_s = len(lum_by_amb_s)
                    for i_amb_s, (amb_key_s, lums_s) in enumerate(lum_by_amb_s.items()):
                        es_ultimo_amb_s = (i_amb_s == n_amb_s - 1)  # ¿es el último ambiente del circuito?
                        n_lum_s = len(lums_s)
                        n_conm_s = int(min(n_lum_s, ncon_map_local.get(amb_key_s, 0)))  # cuántas conmutadas
                        n_rest_s = n_lum_s - n_conm_s  # el resto, no conmutadas
                        ints_s = []  # lista de interruptores que hay que instalar en este ambiente
                        if n_conm_s > 0: ints_s.append(('9/24', n_conm_s))
                        if n_rest_s > 0:
                            c12s, c15s, c32s = descomponer_interruptores(n_rest_s)
                            for _ in range(c32s): ints_s.append(('9/32', 3))
                            for _ in range(c15s): ints_s.append(('9/15', 2))
                            for _ in range(c12s): ints_s.append(('9/12', 1))
                        n_ints_s = len(ints_s)

                        for i_int_s, (tipo_s, n_lum_g_s) in enumerate(ints_s):
                            es_ultimo_int_s = (i_int_s == n_ints_s - 1)  # ¿es el último interruptor de este ambiente?

                            # Caja troncal: la cantidad de salidas depende de si es el último
                            # interruptor y/o el último ambiente del circuito
                            if n_ints_s == 1 and es_ultimo_amb_s:
                                salidas += 3  # entrada + salida hacia el interruptor + salida hacia la caja octogonal
                            elif n_ints_s == 1 and not es_ultimo_amb_s:
                                salidas += 4  # entrada + salida hacia el interruptor + salida hacia la caja octogonal + salida hacia el siguiente ambiente
                            elif es_ultimo_int_s and es_ultimo_amb_s:
                                salidas += 3  # entrada + salida hacia el interruptor + salida hacia la caja octogonal
                            elif es_ultimo_int_s and not es_ultimo_amb_s:
                                salidas += 4  # entrada + salida hacia el interruptor + salida hacia la caja octogonal + salida hacia el siguiente ambiente
                            else:
                                salidas += 4  # primera o intermedia: entrada + salida hacia el interruptor + salida hacia la caja octogonal + salida hacia el siguiente tramo troncal

                            # Caja interruptor: 9/24 son 2 cajas físicas (viajeros
                            # entre ambas), cada una con entrada+salida = 4 en
                            # total. Los demás (9/12, 9/15, 9/32) son 1 sola caja,
                            # el retorno comparte el mismo conduit = 1.
                            salidas += 4 if tipo_s == '9/24' else 1

                            # Cajas octogonales: 1 por luminaria de este grupo.
                            # 9/24: TODAS llevan 2 salidas (incluida la última,
                            # no aplica la regla de "última=1" de los demás
                            # tipos). Resto (9/12, 9/15, 9/32): +1 extra si no
                            # es la última (para seguir la cadena), la última
                            # solo lleva 1 (solo entrada).
                            for i_lum_s in range(n_lum_g_s):
                                es_ultima_lum_s = (i_lum_s == n_lum_g_s - 1)
                                if tipo_s == '9/24':
                                    salidas += 2  # entrada + salida, siempre, incluida la última
                                elif es_ultima_lum_s:
                                    salidas += 1  # solo entrada
                                else:
                                    salidas += 2  # entrada + salida hacia la siguiente caja octogonal

                else:
                    # Enchufes: intermedio=2, último=1. Si el circuito se
                    # ramifica (caja adicional entre ambientes), hay 1 último
                    # por cada ambiente, no solo 1 para todo el circuito.
                    # Horno, lavadora y otras cargas especiales también
                    # cuentan como 1 último, porque igual necesitan su caja
                    # de conexión al artefacto.
                    n_ench_total = sum(
                        int(it.get("n_ench", 1) or 1)
                        for it in items
                        if isinstance(it, dict) and ("id_ench" in it or "modulos" in it or "n_ench" in it)
                    ) + sum(
                        1
                        for it in items
                        if isinstance(it, dict) and "nombre" in it
                    )
                    _cajas_adic_s = int((cajas_adic_por_nombre or {}).get(circ, 0))
                    if _cajas_adic_s > 0:
                        # el circuito se divide en un ramal por ambiente, y
                        # cada ramal termina en su propio último
                        _ambs_ench_s = set(
                            str(it.get("amb", "")).strip().lower()
                            for it in items
                            if isinstance(it, dict) and ("id_ench" in it or "modulos" in it or "n_ench" in it or "nombre" in it)
                        )
                        n_ultimos_s = max(1, len(_ambs_ench_s))
                    else:
                        n_ultimos_s = 1 if n_ench_total > 0 else 0
                    n_intermedios_s = max(0, n_ench_total - n_ultimos_s)
                    salidas += 2 * n_intermedios_s + n_ultimos_s + 3 * _cajas_adic_s

                # Cajas de paso: 2 por caja (entrada + salida)
                salidas += 2 * cajas_paso_circ
                add_row(
                    desc=f"Salida de caja conduit de PVC de {mm}mm",
                    marcas_txt=marcas.get("Salida de caja conduit", ""),
                    norma="RIC 4.7.2",
                    circuito=circ,
                    unidad="u",
                    k=1,
                    longitud_m=f"{salidas} unid",
                    cantidad=salidas
                )
                continue
            # si no calza como sobrepuesta ni como embutida, se cobra por metro lineal
            add_row(
                desc=formatear_canalizacion_material(canal, es_sobrepuesta=False),
                marcas_txt=marcas.get("Conduit PVC (embutida)", ""),
                norma="RIC 4.7.2",
                circuito=circ,
                unidad="m",
                k=1,
                longitud_m=math.ceil(L),
                cantidad=math.ceil(L)
            )

    # Salidas de caja conduit de climatización
    for _cl in _clima_items:
        if _cl["es_emb"]:
            _n_salidas_cl = 2 + 2 * _cl.get("cajas_paso", 0)  # tablero + equipo, más 2 por cada caja de paso intermedia (>=20m)
            add_row(
                desc=f"Salida de caja conduit de PVC de {_cl['mm']}mm",
                marcas_txt=marcas.get("Salida de caja conduit", ""),
                norma="RIC 4.7.2",
                circuito=_cl["circ"],
                unidad="u",
                k=1,
                longitud_m=f"{_n_salidas_cl} unid",
                cantidad=_n_salidas_cl
            )
        else:
            # Canaleta sobrepuesta: coplas + curvas
            cant_copla_cl = max(0, int(math.ceil(_cl["L"] / 2.0)) - 1)
            if cant_copla_cl > 0:
                add_row(
                    desc=f"Unión copla para canaleta de PVC {_extraer_medida_canaleta(_cl['canal'])}",
                    marcas_txt=marcas.get("Accesorios canaleta PVC", ""),
                    norma="RIC 4.7.2",
                    circuito=_cl["circ"],
                    unidad="u",
                    k=1,
                    longitud_m=f"{cant_copla_cl} unid",
                    cantidad=cant_copla_cl
                )
            add_row(
                desc=f"Curvas planas en 90° para canaletas de PVC {_extraer_medida_canaleta(_cl['canal'])}",
                marcas_txt=marcas.get("Accesorios canaleta PVC", ""),
                norma="RIC 4.7.2",
                circuito=_cl["circ"],
                unidad="u",
                k=1,
                longitud_m="2 unid",
                cantidad=2
            )
            add_row(
                desc=f"Curvas internas en 90° para canaletas de PVC {_extraer_medida_canaleta(_cl['canal'])}",
                marcas_txt=marcas.get("Accesorios canaleta PVC", ""),
                norma="RIC 4.7.2",
                circuito=_cl["circ"],
                unidad="u",
                k=1,
                longitud_m="2 unid",
                cantidad=2
            )

    # ── Salidas de caja conduit de AGUA CALIENTE ─────────────────────────────
    # Embutida sin tab.ext: 2 salidas (tablero principal + equipo)
    # Embutida con tab.ext: 4 salidas (2 tramos)
    # + 2 por cada caja de paso intermedia (tramo continuo >=20m)
    # Sobrepuesta: coplas + curvas (igual que climatización)
    for _ac in _agua_items:
        if _ac["es_emb"]:
            _n_bouq_ac = (4 if _ac["lleva_tab_ext"] else 2) + 2 * _ac.get("cajas_paso", 0)
            add_row(
                desc=f"Salida de caja conduit de PVC de {_ac['mm']}mm",
                marcas_txt=marcas.get("Salida de caja conduit", ""),
                norma="RIC 4.7.2",
                circuito=_ac["circ"],
                unidad="u",
                k=1,
                longitud_m=f"{_n_bouq_ac} unid",
                cantidad=_n_bouq_ac
            )
        else:
            # Canaleta sobrepuesta: coplas + curvas (igual que climatización)
            cant_copla_ac = max(0, int(math.ceil(_ac["L"] / 2.0)) - 1)
            if cant_copla_ac > 0:
                add_row(
                    desc=f"Unión copla para canaleta de PVC {_extraer_medida_canaleta(_ac['canal'])}",
                    marcas_txt=marcas.get("Accesorios canaleta PVC", ""),
                    norma="RIC 4.7.2",
                    circuito=_ac["circ"],
                    unidad="u",
                    k=1,
                    longitud_m=f"{cant_copla_ac} unid",
                    cantidad=cant_copla_ac
                )
            # Si hay tablero externo, el circuito se parte en 2 tramos (igual que
            # en el caso embutido, donde las salidas de caja se duplican de 2 a 4)
            _mult_ac = 2 if _ac["lleva_tab_ext"] else 1
            _curvas_plan_ac = 2 * _mult_ac
            _curvas_int_ac  = 2 * _mult_ac
            add_row(
                desc=f"Curvas planas en 90° para canaletas de PVC {_extraer_medida_canaleta(_ac['canal'])}",
                marcas_txt=marcas.get("Accesorios canaleta PVC", ""),
                norma="RIC 4.7.2",
                circuito=_ac["circ"],
                unidad="u",
                k=1,
                longitud_m=f"{_curvas_plan_ac} unid",
                cantidad=_curvas_plan_ac
            )
            add_row(
                desc=f"Curvas internas en 90° para canaletas de PVC {_extraer_medida_canaleta(_ac['canal'])}",
                marcas_txt=marcas.get("Accesorios canaleta PVC", ""),
                norma="RIC 4.7.2",
                circuito=_ac["circ"],
                unidad="u",
                k=1,
                longitud_m=f"{_curvas_int_ac} unid",
                cantidad=_curvas_int_ac
            )

    # =========================
    # 2) CONDUCTORES (POR CIRCUITO)
    # =========================
    add_section("Conductores")
    if {"Conductor", "Longitud (m)", "Circuito"}.issubset(set(circuitos_df.columns)):
        circuitos_tmp = circuitos_df.copy()
        circuitos_tmp["Longitud (m)"] = pd.to_numeric(circuitos_tmp["Longitud (m)"], errors="coerce").fillna(0)
        for _, r in circuitos_tmp.iterrows():
            cond = str(r.get("Conductor", "")).strip()
            circ = str(r.get("Circuito", "")).strip()
            L_total = float(r.get("Longitud (m)", 0))
            if (not cond) or (not circ) or L_total <= 0:
                continue  # circuito sin datos válidos, se salta
            es_ilum    = ("ilumin" in circ.lower())
            es_especial_cond = any(k in circ.lower() for k in ("especial", "horno", "encimera", "lavadora",
                                   "lavavajilla", "climatiz", "aire", "split", "ac ", "a/c",
                                   "ducha", "termo", "calefon", "calefón", "calentador", "agua caliente"))

            # --- Calcular chicotes según puntos de conexión (RIC) ---
            items_circ = (items_por_nombre or {}).get(circ, [])
            if not isinstance(items_circ, list):
                items_circ = []
            # Contar enchufes e interruptores del circuito
            n_ench_cond  = 0
            for it in items_circ:
                if not isinstance(it, dict):
                    continue
                if "n_ench" in it or "id_ench" in it or "modulos" in it:
                    n_ench_cond += int(it.get("n_ench", 1) or 1)

            if es_ilum:
                # ── Cálculo exacto de chicotes por punto de conexión ──────────────
                # Lógica validada ejemplo a ejemplo:
                #
                # CAJA RECTANGULAR por ambiente:
                #   - Intermedio: 3 entrada + 3 salida + 1 fase hacia el interruptor + 1N hacia la luminaria + 1T hacia la luminaria = 9
                #   - Último:     3 entrada + 1 fase hacia el interruptor + 1N hacia la luminaria + 1T hacia la luminaria             = 6
                #
                # CAJA INTERRUPTOR (por tipo):
                #   - 9/12: 1 fase + 1 retorno                = 2
                #   - 9/15: 1 fase + 2 retornos               = 3
                #   - 9/32: 1 fase + 3 retornos               = 4
                #   - 9/24 primer int: 1 fase + 2 viajeros    = 3
                #   - 9/24 segundo int: 2 viajeros + 1 retorno = 3
                #
                # CAJA OCTOGONAL (no conmutada):
                #   - Intermedia: 1 retorno + 3N + 3T = 7
                #   - Última:     1 retorno + 1N + 1T = 3
                #
                # CAJA OCTOGONAL (conmutada 9/24):
                #   - Intermedia: 3F + 3N + 3T = 9
                #   - Última:     1F + 1N + 1T = 3

                # Obtener luminarias por ambiente y conmutadas
                # ncon_map_local ya se construyó antes del bucle principal (ver arriba)
                lum_by_amb_cond = {}
                for it in items_circ:
                    if not isinstance(it, dict): continue
                    if "n_ench" in it or "id_ench" in it or "modulos" in it or "nombre" in it:
                        continue
                    amb = str(it.get("amb", "")).strip() or "sin_amb"
                    key = amb.lower()
                    lum_by_amb_cond[key] = lum_by_amb_cond.get(key, 0) + 1  # cuenta las luminarias de este ambiente

                n_amb_cond = len(lum_by_amb_cond)
                n_puntos = 0  # total de chicotes (puntos de conexión) del circuito

                for i_amb, (amb_key, n_lum_amb) in enumerate(lum_by_amb_cond.items()):
                    es_ultimo_amb = (i_amb == n_amb_cond - 1)  # ¿es el último ambiente del circuito?
                    n_conm_amb = int(min(n_lum_amb, ncon_map_local.get(amb_key, 0)))  # cuántas conmutadas tiene
                    n_rest_amb = int(n_lum_amb) - n_conm_amb  # el resto, no conmutadas

                    # Caja rectangular del ambiente
                    if es_ultimo_amb:
                        n_puntos += 6   # 3 entrada + 1F hacia el interruptor + 1N hacia la luminaria + 1T hacia la luminaria
                    else:
                        n_puntos += 9   # 6 + 3 salida hacia siguiente ambiente

                    # Cajas de interruptores del ambiente
                    if n_conm_amb > 0:
                        # 9/24: siempre 1 PAR de interruptores por ambiente (no uno por luminaria)
                        # primer interruptor: 1 fase + 2 viajeros = 3 chicotes
                        # segundo interruptor: 2 viajeros + 1 retorno = 3 chicotes
                        n_puntos += 6  # 1 par × (3+3)
                    if n_rest_amb > 0:
                        c12, c15, c32 = descomponer_interruptores(n_rest_amb)
                        n_puntos += c12 * 2   # 9/12: 1 fase + 1 retorno
                        n_puntos += c15 * 3   # 9/15: 1 fase + 2 retornos
                        n_puntos += c32 * 4   # 9/32: 1 fase + 3 retornos

                    # Cajas octogonales del ambiente
                    for i_lum in range(int(n_lum_amb)):
                        es_ultima_lum = (i_lum == int(n_lum_amb) - 1)
                        es_conm = (i_lum < n_conm_amb)
                        if es_ultima_lum:
                            n_puntos += 3   # 1F + 1N + 1T (última siempre 3)
                        else:
                            if es_conm:
                                n_puntos += 9   # 3F + 3N + 3T (conmutada intermedia)
                            else:
                                n_puntos += 7   # 1F + 3N + 3T (no conmutada intermedia)

            elif es_especial_cond:
                # Recordatorio: n_puntos cuenta los chicotes del circuito.
                # Un chicote es el pedazo de cable (15cm) que se pierde cada
                # vez que el conductor se corta para conectarlo en una caja;
                # más adelante n_puntos*0.15 se suma a los metros de cable
                # que hay que comprar.
                # Especiales, clima, agua caliente: 1 chicote base (caja
                # junto al equipo) + 2 por cada caja de paso (entrada y
                # salida), solo si hay un tramo continuo >=20m
                _canal_circ_esp = str(r.get("Canalización", "")).strip().lower()
                _es_emb_circ_esp = ("conduit" in _canal_circ_esp) or ("embutida" in _canal_circ_esp)
                _tiene_20m_circ_esp = bool(r.get("_tiene_tramo_20m", True))
                n_cajas_paso_esp = int(L_total // 20) if (_es_emb_circ_esp and _tiene_20m_circ_esp) else 0
                n_puntos = 1 + n_cajas_paso_esp * 2
            else:
                # Enchufes generales, por ambiente: intermedio=3 (entrada,
                # salida, al artefacto), último de CADA ambiente=1 (no solo
                # el último de todo el circuito). Caja adicional entre
                # ambientes=3 (entrada, salida 1, salida 2). Caja de paso=2
                # (entrada, salida), solo si hay un tramo continuo >=20m.
                n_cajas_adic_circ = int((cajas_adic_por_nombre or {}).get(circ, 0))
                _canal_circ = str(r.get("Canalización", "")).strip().lower()
                _es_emb_circ = ("conduit" in _canal_circ) or ("embutida" in _canal_circ)
                _tiene_20m_circ = bool(r.get("_tiene_tramo_20m", True))
                n_cajas_paso_circ = int(L_total // 20) if (_es_emb_circ and _tiene_20m_circ) else 0
                _n_ench_por_amb = {}
                for it in items_circ:
                    if not isinstance(it, dict): continue
                    if "n_ench" in it or "id_ench" in it or "modulos" in it:
                        _a = str(it.get("amb", "")).strip().lower() or "sin_amb"
                        _n_ench_por_amb[_a] = _n_ench_por_amb.get(_a, 0) + int(it.get("n_ench", 1) or 1)
                if _n_ench_por_amb:
                    n_puntos = sum((n_amb_ench - 1) * 3 + 1 for n_amb_ench in _n_ench_por_amb.values() if n_amb_ench > 0)
                else:
                    n = int(max(1, n_ench_cond))
                    n_puntos = (n - 1) * 3 + 1
                n_puntos += n_cajas_adic_circ * 3 + n_cajas_paso_circ * 2
            chicotes = n_puntos * 0.15  # cada chicote son 15cm de conductor extra
            L_con_chicotes = L_total + chicotes  # longitud real del circuito + lo que se pierde en los chicotes

            # --- NO iluminación: 3 conductores iguales ---
            if not es_ilum:
                # Los conductores se venden por metro entero, así que el largo se
                # redondea siempre hacia arriba (nunca hacia abajo)
                # +10% extra de holgura en los tres conductores (rojo, blanco, verde)
                Lr = math.ceil(L_con_chicotes * 1.10)
                Lb = math.ceil(L_con_chicotes * 1.10)
                Lv = math.ceil(L_con_chicotes * 1.10)
                total = Lr + Lb + Lv
                desc = f"Conductor {cond.replace('.', ',')} (Rojo = {Lr} m, Blanco = {Lb} m, Verde = {Lv} m)"
                _es_clima_cond = any(k in circ.lower() for k in ("climatiz", "aire", "split", "ac ", "a/c"))
                _es_agua_cond  = any(k in circ.lower() for k in ("ducha", "termo", "calefon", "calefón", "calentador", "agua caliente"))
                _norma_cond = (
                    "RIC 7 (7.3.4, 7.5.1, 7.5.2) / RIC 3 (5.1.3) / RIC 4 (5.2, 5.4, 5.5, 5.34, 6.2.5)"
                    if (_es_clima_cond or _es_agua_cond)
                    else "RIC 4.1.1"
                )
                # Desglose "valor crudo (x1,10) = resultado", en una sola
                # línea porque en no-iluminación Rojo=Blanco=Verde siempre
                # son iguales (no hace falta repetir 3 veces lo mismo)
                _crudo_txt = f"{L_con_chicotes:.4f}".replace(".", ",")
                _long_txt = f"Rojo=Blanco=Verde {_crudo_txt} (x1,10) = {Lr}m"
                add_row(
                    desc=desc,
                    marcas_txt=marcas.get("Conductores", ""),
                    norma=_norma_cond,
                    circuito=circ,
                    unidad="m",
                    k=1.10,
                    longitud_m=_long_txt,
                    cantidad=total
                )
                continue

            #  Iluminación: calcular Lr, Lb, Lv con chicotes por color
            items = (items_por_nombre or {}).get(circ, [])
            if not isinstance(items, list):
                items = []

            # Obtener ambientes del circuito
            ambs_circ = {}  # {amb_key: n_lum}
            for it in items:
                if not isinstance(it, dict): continue
                if "n_ench" in it or "id_ench" in it or "modulos" in it or "nombre" in it: continue
                a = str(it.get("amb", "")).strip().lower()
                if a: ambs_circ[a] = ambs_circ.get(a, 0) + 1  # cuenta las luminarias de este ambiente

            # ── Calcular chicotes por color ──────────────────────────────────
            # Lógica: recorre uno por uno los interruptores (= caja troncal) dentro de cada ambiente
            #
            # CAJA TRONCAL (1 por interruptor):
            #   n_ints==1, amb interm:  chic_R+=3  chic_N+=3
            #   n_ints==1, último amb:  chic_R+=2  chic_N+=2
            #   primera o intermedia:   chic_R+=3  chic_N+=3  (va hacia el siguiente tramo troncal)
            #   última, amb interm:     chic_R+=3  chic_N+=3  (va hacia el siguiente ambiente)
            #   última, último amb:     chic_R+=2  chic_N+=2
            #
            # CAJA INTERRUPTOR (solo rojo):
            #   9/24 suma 6 · 9/32 suma 4 · 9/15 suma 3 · 9/12 suma 2
            #
            # CAJA OCTOGONAL:
            #   Última: chic_R+=1 chic_N+=1
            #   Interm conm 9/24: chic_R+=3 chic_N+=3
            #   Interm no conm: chic_R+=1 chic_N+=3
            chic_R = 0
            chic_N = 0

            n_amb_circ = len(ambs_circ)
            for i_amb, (amb_key, n_lum_amb) in enumerate(ambs_circ.items()):
                n_conm_amb = int(min(n_lum_amb, ncon_map_local.get(amb_key, 0)))
                n_rest_amb = n_lum_amb - n_conm_amb
                es_ultimo_amb = (i_amb == n_amb_circ - 1)

                # Lista de interruptores del ambiente
                ints_amb = []
                if n_conm_amb > 0:
                    ints_amb.append(('9/24', n_conm_amb))
                if n_rest_amb > 0:
                    c12, c15, c32 = descomponer_interruptores(n_rest_amb)
                    for _ in range(c32): ints_amb.append(('9/32', 3))
                    for _ in range(c15): ints_amb.append(('9/15', 2))
                    for _ in range(c12): ints_amb.append(('9/12', 1))
                n_ints = len(ints_amb)

                # Cajas troncales (1 por interruptor)
                for i_int, (tipo, n_lum_grupo) in enumerate(ints_amb):
                    es_primer = (i_int == 0)
                    es_ultimo = (i_int == n_ints - 1)
                    if n_ints == 1:
                        # único interruptor del ambiente
                        if not es_ultimo_amb:
                            # entrada + salida hacia el interruptor + salida hacia el siguiente ambiente
                            chic_R += 3; chic_N += 3
                        else:
                            # último ambiente: entrada + salida hacia el interruptor (no hay a dónde más salir)
                            chic_R += 2; chic_N += 2
                    elif es_primer:
                        # Primera troncal: entrada + salida hacia el interruptor + salida hacia el siguiente tramo troncal
                        chic_R += 3; chic_N += 3
                    elif not es_ultimo:
                        # Troncal intermedia: entrada + salida hacia el interruptor + salida hacia el siguiente tramo troncal
                        chic_R += 3; chic_N += 3
                    else:
                        # Última troncal del ambiente
                        if not es_ultimo_amb:
                            # Amb intermedio: entrada + salida hacia el interruptor + salida hacia el siguiente ambiente
                            chic_R += 3; chic_N += 3
                        else:
                            # Último ambiente: entrada + salida hacia el interruptor
                            chic_R += 2; chic_N += 2

                # Cajas interruptores (solo rojo, porque acá solo llega la fase)
                for tipo, n_lum_grupo in ints_amb:
                    # 9/24 (conmutado): son 6 chicotes, 3 van al primer interruptor,
                    # 3 al viajero del segundo interruptor
                    if tipo == '9/24':   chic_R += 6
                    # 9/32: son 4 chicotes, 1 de entrada + 3 retornos (1 por cada luminaria del grupo)
                    elif tipo == '9/32': chic_R += 4
                    # 9/15: son 3 chicotes, 1 de entrada + 2 retornos
                    elif tipo == '9/15': chic_R += 3
                    # 9/12: son 2 chicotes, 1 de entrada + 1 retorno
                    elif tipo == '9/12': chic_R += 2

                # Cajas octogonales (donde va cada luminaria)
                for tipo, n_lum_grupo in ints_amb:
                    for i_lum in range(n_lum_grupo):
                        es_ultima_lum = (i_lum == n_lum_grupo - 1)
                        es_conm = (tipo == '9/24')
                        if es_ultima_lum:
                            # última luminaria del grupo: solo llega 1 chicote (el retorno desde el interruptor)
                            chic_R += 1; chic_N += 1
                        else:
                            # luminaria intermedia:
                            # si es conmutada (9/24): son 3 chicotes, 1 que llega, 1 que va a la
                            #   luminaria y 1 que sigue a la siguiente caja octogonal
                            # si NO es conmutada: es 1 chicote, el que llega del retorno
                            chic_R += 3 if es_conm else 1
                            chic_N += 3

            # Caja de paso (1 cada 20m del circuito, solo canalización
            # embutida, y solo si el circuito tiene un tramo continuo >=20m
            # confirmado en el input): 2 chicotes rojo + 2 neutro/tierra por
            # caja. Se cuenta por circuito completo (L_total), no por ambiente.
            _canal_circ_ilum = str(r.get("Canalización", "")).strip().lower()
            _es_emb_circ_ilum = ("conduit" in _canal_circ_ilum) or ("embutida" in _canal_circ_ilum)
            _tiene_20m_circ_ilum = bool(r.get("_tiene_tramo_20m", True))
            if _es_emb_circ_ilum and _tiene_20m_circ_ilum:
                n_cajas_paso_ilum = int(L_total // 20)
                chic_R += n_cajas_paso_ilum * 2
                chic_N += n_cajas_paso_ilum * 2

            # ── Calcular extra longitud rojo vs neutro/tierra ────────────────
            # Rojo recorre tramos adicionales al interruptor y viajeros
            # Neutro/tierra no van al interruptor, así que hay que restar esos tramos
            # Si hay múltiples interruptores en el mismo ambiente, se multiplica
            # por n_interruptores (todos usan el mismo L_ida)
            # El rojo (fase) pasa por cada interruptor y vuelve (ida y
            # vuelta), por eso recorre más metros que la longitud medida
            # del circuito. El neutro y la tierra van directo a la
            # luminaria sin pasar por el interruptor, por eso recorren
            # menos. extra_R y extra_NT corrigen esa diferencia.
            extra_R = 0.0   # metros extra que recorre el rojo (a SUMAR)
            extra_NT = 0.0  # metros a RESTAR del neutro/tierra

            for amb_key in ambs_circ:
                amb_row = _get_amb_row(ambientes_df, amb_key)
                if amb_row is None: continue
                ncon = int(pd.to_numeric(amb_row.get("N_conmutadas_924 (u)", 0), errors="coerce") or 0)
                n_lum_amb = ambs_circ[amb_key]
                n_rest_amb = n_lum_amb - ncon

                if ncon > 0:
                    # 9/24: rojo suma L_via, resta L_troncal_primera_oct (ese
                    # tramo va directo de la caja a la luminaria, el rojo no
                    # lo recorre porque se desvía por los interruptores) · neutro/tierra
                    # resta L_fas + L_via + L_ret
                    # L_via: entre los 2 interruptores
                    L_via = float(pd.to_numeric(amb_row.get("L_viajeros_924 (m)", 0.0), errors="coerce") or 0.0)
                    # L_ret: del interruptor a la lámpara
                    L_ret = float(pd.to_numeric(amb_row.get("L_retorno_lampara (m)", 0.0), errors="coerce") or 0.0)
                    # L_fas: de la caja al primer interruptor
                    L_fas = float(pd.to_numeric(amb_row.get("L_fase_caja_primer_int (m)", 0.0), errors="coerce") or 0.0)
                    # L_tr_oct: de la caja troncal a la primera caja octogonal (1 vez por ambiente)
                    L_tr_oct = float(pd.to_numeric(amb_row.get("L_troncal_primera_oct_924 (m)", 0.0), errors="coerce") or 0.0)
                    extra_R  += L_via - L_tr_oct
                    extra_NT += L_fas + L_via + L_ret

                if n_rest_amb > 0:
                    # Un ambiente puede tener más de 1 interruptor (por
                    # ejemplo un 9/32 para las luminarias del centro y un
                    # 9/12 para un foco aparte), y cada interruptor puede
                    # estar a una distancia distinta de la caja troncal.
                    # "Grupos_interruptor_ilum" guarda esa lista: 1 grupo
                    # por cada interruptor del ambiente, con sus propias
                    # distancias (L_ida, L_tr_oct1, L_o1_o2).
                    grupos_amb = amb_row.get("Grupos_interruptor_ilum", None)
                    if isinstance(grupos_amb, list) and len(grupos_amb) > 0:
                        # Se calcula interruptor por interruptor, cada uno
                        # con su propia distancia. Si en cambio se usara un
                        # solo promedio para todo el ambiente, el resultado
                        # queda inflado cuando las distancias son distintas
                        # entre sí — por eso se recorre grupo por grupo.
                        for _g in grupos_amb:
                            _tipo_g  = _g.get("tipo", "")       # 9/12, 9/15 o 9/32
                            _L_ida_g = float(_g.get("L_ida", 0.0) or 0.0)       # distancia de la caja troncal a este interruptor
                            _L_tr1_g = float(_g.get("L_tr_oct1", 0.0) or 0.0)   # distancia de la caja troncal a la primera luminaria de este grupo
                            _L_o12_g = float(_g.get("L_o1_o2", 0.0) or 0.0)     # entre la 1ra y 2da luminaria de este grupo
                            if _tipo_g == '9/32':
                                # 3 luminarias, entonces 3 retornos (una vuelta por cada una)
                                # + el tramo hasta la 1ra luminaria y hasta la 2da
                                extra_R += _L_ida_g * 3 + (_L_tr1_g + (_L_tr1_g + _L_o12_g))
                            elif _tipo_g == '9/15':
                                # 2 luminarias, entonces 2 retornos + el tramo hasta la 1ra
                                # (L_o1_o2 no aplica en este grupo, son solo 2 luminarias)
                                extra_R += _L_ida_g * 2 + _L_tr1_g
                            elif _tipo_g == '9/12':
                                # 1 sola luminaria, entonces 1 retorno, sin tramo extra
                                extra_R += _L_ida_g * 1
                            # el neutro y la tierra no pasan por el interruptor,
                            # así que solo se restan el tramo de ida hasta él
                            # (no van y vuelven como el rojo)
                            extra_NT += _L_ida_g
                    else:
                        # si no viene el detalle por grupo, se usan los totales
                        # agregados del ambiente. Exacto solo si el ambiente
                        # tiene 1 único interruptor.
                        L_ida     = float(pd.to_numeric(amb_row.get("L_caja_int_fase_ida (m)", 0.0), errors="coerce") or 0.0)
                        L_tr_oct1 = float(pd.to_numeric(amb_row.get("L_troncal_oct1 (m)", 0.0), errors="coerce") or 0.0)
                        L_o1_o2   = float(pd.to_numeric(amb_row.get("L_oct1_oct2 (m)", 0.0), errors="coerce") or 0.0)
                        c12, c15, c32 = descomponer_interruptores(n_rest_amb)
                        n_retornos = c32 * 3 + c15 * 2 + c12 * 1
                        extra_R  += L_ida * n_retornos
                        extra_R  += c32 * (L_tr_oct1 + (L_tr_oct1 + L_o1_o2))
                        extra_R  += c15 * L_tr_oct1
                        extra_NT += L_ida

            # ── Longitudes finales por color ─────────────────────────────────
            # Lr = ceil((L_total + chicotes_R×0.15 + extra_R) × 1.10)
            # Lb = Lv = ceil((L_total + chicotes_N×0.15 - extra_NT) × 1.10)
            _crudo_r  = L_total + chic_R * 0.15 + extra_R
            _crudo_bv = max(0.0, L_total + chic_N * 0.15 - extra_NT)
            Lr = math.ceil(_crudo_r  * 1.10)
            Lb = math.ceil(_crudo_bv * 1.10)
            Lv = Lb
            total = Lr + Lb + Lv
            desc = f"Conductor {cond.replace('.', ',')} (Rojo = {Lr} m, Blanco = {Lb} m, Verde = {Lv} m)"
            _crudo_r_txt  = f"{_crudo_r:.4f}".replace(".", ",")
            _crudo_bv_txt = f"{_crudo_bv:.4f}".replace(".", ",")
            _long_txt = (
                f"Rojo {_crudo_r_txt} (x1,10) = {Lr}m / "
                f"Blanco {_crudo_bv_txt} (x1,10) = {Lb}m / "
                f"Verde {_crudo_bv_txt} (x1,10) = {Lv}m"
            )
            add_row(
                desc=desc,
                marcas_txt=marcas.get("Conductores", ""),
                norma="RIC 4.1.1",
                circuito=circ,
                unidad="m",
                k=1.10,
                longitud_m=_long_txt,
                cantidad=total
            )

    # =========================
    # 3) PROTECCIONES
    # =========================
    add_section("Protecciones")
    add_row(
        desc=f"Interruptor general omnipolar {texto_omni}",
        marcas_txt=marcas.get("Protecciones", ""),
        norma="RIC 6.4.1",
        circuito="General",
        unidad="u",
        k=1,
        longitud_m="1 unid",
        cantidad=1
    )
    # Supresor de transiente (SPD)
    add_row(
        desc="Supresor de transiente Tipo 2 (Clase II), 2P 230 V 20kA, Up ≤ 2.5kV",
        marcas_txt=marcas.get("Supresor de transiente", marcas.get("Protecciones", "")),
        norma="RIC 5 (8.7.7) / RIC 6 (6.2.2)",
        circuito="General",
        unidad="u",
        k=1,
        longitud_m="1 unid",
        cantidad=1
    )
    # Protector sobrevoltaje y corriente ajustable
    add_row(
        desc="Protector sobrevoltaje y corriente ajustable 63A 220V",
        marcas_txt=marcas.get("Protector sobrevoltaje", ""),
        norma="RIC 6 (6.6.2)",
        circuito="General",
        unidad="u",
        k=1,
        longitud_m="1 unid",
        cantidad=1
    )
    # DIFERENCIALES: 1 por grupo
    if group_info and "Circuito" in circuitos_df.columns:
        for gid, meta in group_info.items():
            sens     = meta.get("sensibilidad_dif", "30mA")
            dif_txt  = f"2X{meta['dif']} {sens} / Tipo A"
            circuitos_grupo = []
            for idx in meta["indices"]:
                if 0 <= idx < len(circuitos_df):
                    circuitos_grupo.append(str(circuitos_df.loc[idx, "Circuito"]))  # nombres de los circuitos de este grupo
            lista_circuitos = "\n".join(circuitos_grupo)
            # Marca diferente para 10mA
            marca_dif = (marcas.get("Diferencial 10mA agua caliente", marcas.get("Protecciones", ""))
                         if sens == "10mA" else marcas.get("Protecciones", ""))
            add_row(
                desc=f"Interruptor diferencial {dif_txt.replace(' / Tipo A', '')} Tipo A",
                marcas_txt=marca_dif,
                norma="RIC N°11 (6.4.3)" if sens == "10mA" else "RIC 6.5.1",
                circuito=lista_circuitos if lista_circuitos else "Varios",
                unidad="u",
                k=1,
                longitud_m="1 unid",
                cantidad=1
            )

    # Termomagnéticos: 1 por circuito
    if {"Disyuntor termomagnético", "Circuito"}.issubset(set(circuitos_df.columns)):
        tms_df = circuitos_df[["Disyuntor termomagnético", "Circuito"]].copy()
        tms_df["Disyuntor termomagnético"] = tms_df["Disyuntor termomagnético"].astype(str).str.strip()
        tms_df["Circuito"] = tms_df["Circuito"].astype(str).str.strip()
        tms_df = tms_df[tms_df["Disyuntor termomagnético"] != ""]  # se saltan los circuitos sin TM definido
        for _, rr in tms_df.iterrows():
            tm_txt = rr["Disyuntor termomagnético"]
            circ2 = rr["Circuito"]
            add_row(
                desc=f"Disyuntor termomagnético {tm_txt}",
                marcas_txt=marcas.get("Protecciones", ""),
                norma="RIC 6.4.1",
                circuito=circ2,
                unidad="u",
                k=1,
                longitud_m="1 unid",
                cantidad=1
            )

    # =========================
    # 4) ACCESORIOS (cajas + tapas + portalamparas + interruptores + enchufes + especiales)
    # =========================
    add_section("Accesorios")
    tipo_can = (tipo_canalizacion or "").strip().lower()
    es_embutida = "embut" in tipo_can
    es_sobrepuesta = "sobre" in tipo_can
    total_luminarias = 0
    total_luminarias_sobrepuestas = 0  # solo las que su MONTAJE individual es sobrepuesto (llevan tornillo), las embutidas no
    total_enchufes = 0
    total_interruptores = 0
    total_ampolletas = 0

    # Salidas de caja octogonal encadenadas, por ambiente: en canalización
    # sobrepuesta TODA luminaria (sea el foco embutido o sobrepuesto) lleva su
    # propia caja octogonal embutida puntual para la conexión/derivación (no
    # "centro a centro"). Cada luminaria del ambiente lleva 2 salidas
    # (entrada + salida a la siguiente), excepto la última del ambiente que
    # lleva solo 1 (entrada). Misma regla que ya se usa para embutida
    # (líneas ~2090-2096).
    _salidas_oct_luminarias = 0
    if ambientes_df is not None:
        # ---- Luminarias (contar líneas de Detalle iluminación) ----
        if "Detalle iluminación" in ambientes_df.columns:
            for v in ambientes_df["Detalle iluminación"].fillna("").astype(str):
                lineas = [x.strip() for x in v.split("\n") if x.strip() and x.strip().lower() != "ninguna"]
                n_lum_amb = len(lineas)  # una línea de detalle = una luminaria
                total_luminarias += n_lum_amb
                if n_lum_amb > 0:
                    _salidas_oct_luminarias += 2 * (n_lum_amb - 1) + 1  # todas x2, última x1
                for ln in lineas:
                    lnl = ln.strip().lower()
                    if "sobrepuest" in lnl:
                        total_luminarias_sobrepuestas += 1
                    if ("ampolleta led" in lnl) or ("ampolleta incand" in lnl) or ("incandescente" in lnl and "ampolleta" in lnl):
                        total_ampolletas += 1  # cuenta cuántas son ampolletas (necesitan portalámpara)

        # Enchufes (cantidad enchufes por ambiente)
        if "Cantidad enchufes (u)" in ambientes_df.columns:
            total_enchufes = int(
                pd.to_numeric(ambientes_df["Cantidad enchufes (u)"], errors="coerce").fillna(0).sum()
            )
        # Interruptores reales por grupo (normal y 9/24)
        total_interruptores_pared = 0  # 1 por grupo normal, 2 por grupo 9/24 (sus 2 mecanismos físicos)
        if ("Cantidad luminarias (u)" in ambientes_df.columns) and ("N_conmutadas_924 (u)" in ambientes_df.columns):
            for _, ar in ambientes_df.iterrows():
                n_lum = int(pd.to_numeric(ar.get("Cantidad luminarias (u)", 0), errors="coerce") or 0)
                n_con = int(pd.to_numeric(ar.get("N_conmutadas_924 (u)", 0), errors="coerce") or 0)
                n_con = max(0, min(n_con, n_lum))
                n_rest = max(0, n_lum - n_con)
                c12, c15, c32 = descomponer_interruptores(n_rest)
                grupo_924 = 1 if n_con > 0 else 0  # el par 9/24 es 1 solo grupo, sin importar cuántas luminarias controle
                total_interruptores += grupo_924 + c12 + c15 + c32  # grupos totales (para cónicos/curvas/caja troncal)
                total_interruptores_pared += (2 * grupo_924) + c12 + c15 + c32  # 9/24 pesa 2 (2 mecanismos)

    # Sumar cajas adicionales de enchufes (misma fuente que usan los cónicos)
    _cajas_adic_enchufes = sum(int(v) for v in (cajas_adic_por_nombre or {}).values())

    # Sumar cajas y tapas de climatización a los contadores
    # para que queden en la misma fila que las cajas normales
    _clima_cajas_emb  = sum(1 for _cl in _clima_items if _cl["es_emb"])
    _clima_cajas_sob  = sum(1 for _cl in _clima_items if not _cl["es_emb"])
    _clima_tapas      = sum(1 for _cl in _clima_items if _cl["es_emb"] and not _cl["con_enchufe"])
    _clima_circ_str   = ", ".join(_cl["circ"].split("(")[0].strip() for _cl in _clima_items) if _clima_items else ""

    # Sumar cajas y tapas de agua caliente (si la conexión es fija, siempre lleva tapa)
    _agua_cajas_emb  = sum(1 for _ac in _agua_items if _ac["es_emb"])
    _agua_cajas_sob  = sum(1 for _ac in _agua_items if not _ac["es_emb"])
    _agua_tapas      = sum(1 for _ac in _agua_items if _ac["es_emb"] and _ac["in_tm"] is not None and _ac["in_tm"] > 16)  # tapa solo si va sin enchufe (conexión fija) — mismo criterio que climatización (línea 2681)
    _agua_circ_str   = ", ".join(_ac["circ"].split("(")[0].strip() for _ac in _agua_items) if _agua_items else ""

    # Circuitos especiales genéricos (no clima/agua): cualquiera sea su
    # amperaje, todos necesitan su propia caja (con enchufe dedicado si es
    # ≤16A, o con prensaestopa si es >16A)
    _nombres_esp_genericos = []
    if circuitos_df is not None and "Circuito" in circuitos_df.columns:
        for _, _rr in circuitos_df.iterrows():
            _nombre_c = str(_rr.get("Circuito", "")).strip()
            if not _nombre_c:
                continue
            _es_clima_o_agua = any(k in _nombre_c.lower() for k in
                                    ("clima", "split", "aire", "ducha", "termo", "calefon", "agua caliente"))
            if _es_clima_o_agua:
                continue
            _items_c = (items_por_nombre or {}).get(_nombre_c, [])
            _es_generico = isinstance(_items_c, list) and any(
                isinstance(_it, dict) and ("nombre" in _it) for _it in _items_c
            )
            if _es_generico:
                _nombres_esp_genericos.append(_nombre_c)
    _esp_conexion_fija = len(_nombres_esp_genericos)  # usado para CAJAS (todos)
    # nombres de esos circuitos, para mostrar en la columna "Circuito" del Excel
    _esp_circ_str = ", ".join(n.split("(")[0].strip() for n in _nombres_esp_genericos)
    # Solo los >16A van SIN enchufe (conexión directa con prensaestopa) — esos
    # sí necesitan tapa ciega, porque su caja no queda cubierta por ningún
    # enchufe. Los ≤16A ya quedan tapados por el enchufe dedicado, sin tapa.
    _esp_conexion_fija_tapas = 0
    if circuitos_df is not None and "Circuito" in circuitos_df.columns:
        for _, _rr in circuitos_df.iterrows():
            _nombre_c = str(_rr.get("Circuito", "")).strip()
            if _nombre_c in _nombres_esp_genericos:
                _tm_c = parse_in_tm(_rr.get("Disyuntor termomagnético", "")) or 0
                if _tm_c > 16:
                    _esp_conexion_fija_tapas += 1

    # cajas SOLO para UNIONES (para tapas ciegas)
    if es_embutida:
        cajas_octogonales = total_luminarias
        # 2 cajas rectangulares por interruptor:
        # 1 caja troncal (derivación del troncal) + 1 caja interruptor (mecanismo)
        cajas_rect_troncal      = total_interruptores  # 1 troncal por GRUPO de interruptor (9/24 cuenta 1 grupo, no 2)
        cajas_rect_interruptor  = total_interruptores_pared  # 1 caja mecanismo por grupo (9/24 pesa 2, sus 2 interruptores físicos)
        cajas_rect_union        = cajas_rect_troncal + cajas_rect_interruptor
        cajas_rect_enchufe      = total_enchufes + _cajas_adic_enchufes
        if cajas_octogonales > 0:
            # Tipo de caja octogonal según diámetro conduit iluminación
            if _mm_conduit_ilumin <= 20:
                desc_oct  = 'Caja de derivación embutida octogonal de PVC 100x41 mm 4" (12 salidas)'
            else:
                desc_oct  = "Caja de derivación embutida octogonal grande de PVC 109x70x45 mm (12 salidas)"
            add_row(
                desc=desc_oct,
                marcas_txt=marcas.get("Cajas derivación embutidas", ""),
                norma="RIC 4.3.1",
                circuito="Iluminación",
                unidad="u",
                k=1,
                longitud_m=f"{cajas_octogonales} unid",
                cantidad=cajas_octogonales
            )
        # Un circuito especial con conexión fija (sin enchufe, ej. horno)
        # también necesita su caja de derivación en canalización embutida
        # (en sobrepuesta ya se suma, ver cajas_chuqui más abajo)
        total_rectangulares = cajas_rect_union + cajas_rect_enchufe + _clima_cajas_emb + _agua_cajas_emb + _esp_conexion_fija
        total_rectangulares += _cajas_paso_total
        # Las cajas de paso se suman a _cajas_total para que espuma/tornillos/tarugos
        # de Panel SIP (que usan _cajas_total) las contemplen correctamente.
        _cajas_total += int(cajas_octogonales) + int(total_rectangulares)
        if total_rectangulares > 0:
            circ_cajas = "Enchufes / Interruptores / Uniones"
            if _cajas_paso_total > 0:
                circ_cajas += " / Cajas de paso"
            if _clima_cajas_emb > 0:
                circ_cajas += f" / {_clima_circ_str}"
            if _agua_cajas_emb > 0:
                circ_cajas += f" / {_agua_circ_str}"
            if _esp_conexion_fija > 0:
                circ_cajas += f" / Especiales ({_esp_circ_str})"
            add_row(
                desc="Caja de derivación embutida de PVC para tabiques 110x110x67 mm (12 salidas)",
                marcas_txt=marcas.get("Cajas derivación embutidas", ""),
                norma="RIC 4.3.1",
                circuito=circ_cajas,
                unidad="u",
                k=1,
                longitud_m=f"{total_rectangulares} unid",
                cantidad=total_rectangulares
            )
        # NOTA: las salidas de caja de las cajas de paso interior (RIC 7.16.1.13)
        # ya se generan por circuito, más arriba, dentro de la fila "Salida de
        # caja conduit" de cada circuito embutido (línea ~2129, "salidas += 2 *
        # cajas_paso_circ") — no hace falta (ni corresponde) generarlas de
        # nuevo acá agregadas, se duplicaría.
    elif es_sobrepuesta:
        # Cajas chuqui por interruptor: 1 troncal por grupo + 1 mecanismo por
        # interruptor físico (el 9/24 pesa 2 mecanismos, igual que en embutida)
        cajas_chuqui = total_enchufes + _cajas_adic_enchufes + total_interruptores + total_interruptores_pared + _cajas_paso_total + _esp_conexion_fija + _clima_cajas_sob + _agua_cajas_sob
        _cajas_total += int(total_luminarias) + int(cajas_chuqui)
        if cajas_chuqui > 0:
            circ_chuqui = "Enchufes / Interruptores / Uniones"
            if _clima_cajas_sob > 0:
                circ_chuqui += f" / {_clima_circ_str}"
            if _agua_cajas_sob > 0:
                circ_chuqui += f" / {_agua_circ_str}"
            if _esp_conexion_fija > 0:
                circ_chuqui += f" / Especiales ({_esp_circ_str})"
            add_row(
                desc="Caja de derivación sobrepuesta chuqui de PVC 12x8,5 cm",
                marcas_txt=marcas.get("Cajas derivación sobrepuestas", ""),
                norma="RIC 4.3.1",
                circuito=circ_chuqui,
                unidad="u",
                k=1,
                longitud_m=f"{cajas_chuqui} unid",
                cantidad=cajas_chuqui
            )
        # Focos: en canalización sobrepuesta, TODA luminaria (sea el foco
        # embutido o sobrepuesto) lleva su propia caja octogonal EMBUTIDA
        # puntual (no chuqui, porque el foco sobrepuesto delgado no calza
        # sobre una caja que sobresale) para poder hacer la conexión/
        # derivación dentro de la caja y no "centro a centro". Misma regla de
        # cadena que en embutida: 2 salidas por luminaria (entrada + salida a
        # la siguiente), 1 sola (solo entrada) para la última del ambiente.
        if total_luminarias > 0:
            if _mm_conduit_ilumin <= 20:
                desc_oct_sob = 'Caja de derivación embutida octogonal de PVC 100x41 mm 4" (12 salidas)'
            else:
                desc_oct_sob = "Caja de derivación embutida octogonal grande de PVC 109x70x45 mm (12 salidas)"
            add_row(
                desc=desc_oct_sob,
                marcas_txt=marcas.get("Cajas derivación embutidas", ""),
                norma="RIC 4.3.1",
                circuito="Iluminación",
                unidad="u",
                k=1,
                longitud_m=f"{total_luminarias} unid",
                cantidad=total_luminarias
            )
            add_row(
                desc=f"Salida de caja conduit de PVC de {_mm_conduit_ilumin}mm",
                marcas_txt=marcas.get("Salida de caja conduit", ""),
                norma="RIC 4.7.2",
                circuito="Iluminación",
                unidad="u",
                k=1,
                longitud_m=f"{_salidas_oct_luminarias} unid",
                cantidad=_salidas_oct_luminarias
            )
    else:
        add_row(
            desc="Caja de derivación (definir embutida o sobrepuesta)",
            marcas_txt=f"{marcas.get('Cajas derivación embutidas','')} / {marcas.get('Cajas derivación sobrepuestas','')}",
            norma="RIC 4.3.1",
            circuito="Varios",
            unidad="u",
            k=1,
            longitud_m="Por definir",
            cantidad=""
        )

    # TAPAS CIEGAS
    tipo_can_tap = (tipo_canalizacion or "").strip().lower()
    es_emb_tap = "embut" in tipo_can_tap
    es_sob_tap = "sobre" in tipo_can_tap
    # Cajas de paso solo aplican para embutida.
    # NOTA: _cajas_paso_total ya quedó sumado a _cajas_total dentro del
    # bloque ACCESORIOS (vía total_rectangulares). Por eso acá NO se vuelve
    # a sumar: si se sumara de nuevo, tornillos/tarugos/espuma quedarían
    # sobreestimados.

    if es_emb_tap:
        # Embutida: octogonal por luminaria (incluye foco) + rectangular por
        # caja TRONCAL (no por el mecanismo del interruptor, que ya lo cubre
        # la placa) + cajas de paso + cajas adicionales entre ambientes + clima + agua
        tapas_octogonales   = int(max(0, total_luminarias))
        # La caja de conexión fija de un circuito especial también necesita su
        # tapa ciega — solo los >16A (sin enchufe), los ≤16A ya quedan
        # tapados por su propio enchufe dedicado
        tapas_rectangulares = int(max(0, cajas_rect_troncal)) + int(_cajas_adic_enchufes) + _clima_tapas + _agua_tapas + _cajas_paso_total + _esp_conexion_fija_tapas
        total_tapas_ciegas  = tapas_octogonales + tapas_rectangulares
        _tapas_ciegas_total = int(total_tapas_ciegas)
        # Tapa ciega octogonal
        if tapas_octogonales > 0:
            if _mm_conduit_ilumin <= 20:
                desc_tapa_oct = 'Tapa ciega octogonal de PVC 4" (10,2 cm diámetro)'
            else:
                desc_tapa_oct = "Tapa ciega octogonal grande de PVC (11,0 x 11,0 cm)"
            add_row(
                desc=desc_tapa_oct,
                marcas_txt=marcas.get("Tapa ciega octogonal", ""),
                norma="RIC 4.3.1",
                circuito="Iluminación",
                unidad="u",
                k=1,
                longitud_m=f"{tapas_octogonales} unid",
                cantidad=tapas_octogonales
            )
        # Tapa ciega rectangular (troncal + cajas adicionales + cajas de paso + clima + agua)
        if tapas_rectangulares > 0:
            circ_tapas = "Uniones"
            if _cajas_adic_enchufes > 0:
                circ_tapas += " / Cajas adicionales entre ambientes"
            if _cajas_paso_total > 0:
                circ_tapas += " / Cajas de paso"
            if _clima_tapas > 0:
                circ_tapas += f" / {_clima_circ_str}"
            if _agua_tapas > 0:
                circ_tapas += f" / {_agua_circ_str}"
            if _esp_conexion_fija_tapas > 0:
                circ_tapas += f" / Especiales ({_esp_circ_str})"
            add_row(
                desc="Tapa ciega de PVC de 110 x 67 mm",
                marcas_txt=marcas.get("Tapa ciega", ""),
                norma="RIC 4.3.1",
                circuito=circ_tapas,
                unidad="u",
                k=1,
                longitud_m=f"{tapas_rectangulares} unid",
                cantidad=tapas_rectangulares
            )
    elif es_sob_tap:
        # Sobrepuesta: tapa chuqui en caja troncal + cajas_adic + cajas_paso
        # NO en caja mecanismo (interruptor) ni enchufe (tienen mecanismo
        # instalado) — los especiales solo suman tapa si son >16A (sin
        # enchufe), los ≤16A ya quedan tapados por su propio enchufe dedicado
        tapas_chuqui = int(total_interruptores) + int(_cajas_adic_enchufes) + int(_cajas_paso_total) + _clima_tapas + _agua_tapas + _esp_conexion_fija_tapas
        # Luminarias: en sobrepuesta, cada foco ahora sí lleva su propia caja
        # octogonal embutida puntual (ver bloque de ACCESORIOS más arriba), así
        # que también necesita su tapa ciega octogonal, igual que en embutida.
        tapas_octogonales_sob = int(max(0, total_luminarias))
        total_tapas_ciegas  = tapas_chuqui + tapas_octogonales_sob
        _tapas_ciegas_total = int(total_tapas_ciegas)
        if tapas_octogonales_sob > 0:
            if _mm_conduit_ilumin <= 20:
                desc_tapa_oct_sob = 'Tapa ciega octogonal de PVC 4" (10,2 cm diámetro)'
            else:
                desc_tapa_oct_sob = "Tapa ciega octogonal grande de PVC (11,0 x 11,0 cm)"
            add_row(
                desc=desc_tapa_oct_sob,
                marcas_txt=marcas.get("Tapa ciega octogonal", ""),
                norma="RIC 4.3.1",
                circuito="Iluminación",
                unidad="u",
                k=1,
                longitud_m=f"{tapas_octogonales_sob} unid",
                cantidad=tapas_octogonales_sob
            )
        if tapas_chuqui > 0:
            circ_tapas = "Uniones"
            if _clima_tapas > 0:
                circ_tapas += f" / {_clima_circ_str}"
            if _agua_tapas > 0:
                circ_tapas += f" / {_agua_circ_str}"
            if _esp_conexion_fija_tapas > 0:
                circ_tapas += f" / Especiales ({_esp_circ_str})"
            add_row(
                desc="Tapa ciega chuqui de PVC 12x8,5 cm",
                marcas_txt=marcas.get("Tapa ciega", ""),
                norma="RIC 4.3.1",
                circuito=circ_tapas,
                unidad="u",
                k=1,
                longitud_m=f"{tapas_chuqui} unid",
                cantidad=tapas_chuqui
            )

    # PORTALÁMPARAS SOLO SI HAY AMPOLLETAS
    if total_ampolletas > 0:
        add_row(
            desc="Portalámpara plafón E27 redondo",
            marcas_txt=marcas.get("Portalamparas", ""),
            norma="RIC 10 (5.1.4.4, 5.1.4.5)",
            circuito="Iluminación",
            unidad="u",
            k=1,
            longitud_m=f"{total_ampolletas} unid",
            cantidad=total_ampolletas
        )

    # INTERRUPTORES (9/12 - 9/15 - 9/24 - 9/32)
    ncon_map = {}
    if ambientes_df is not None and "Ambiente" in ambientes_df.columns:
        amb_col = ambientes_df["Ambiente"].astype(str).str.strip()
        if "N_conmutadas_924 (u)" in ambientes_df.columns:
            ncon_col = pd.to_numeric(ambientes_df["N_conmutadas_924 (u)"], errors="coerce").fillna(0).astype(int)
            for a, nval in zip(amb_col, ncon_col):
                ncon_map[a.lower()] = int(max(0, nval))
        else:
            for a in amb_col:
                ncon_map[a.lower()] = 0
    if {"Circuito", "Disyuntor termomagnético"}.issubset(set(circuitos_df.columns)):
        _interruptores_acum = {}  # {desc: {"total": N, "detalle": {"circuito y ambiente": cant}}}

        def _acum_interruptor(desc, detalle_txt, cant):
            if desc not in _interruptores_acum:
                _interruptores_acum[desc] = {"total": 0, "detalle": {}}
            _interruptores_acum[desc]["total"] += cant
            _interruptores_acum[desc]["detalle"][detalle_txt] = _interruptores_acum[desc]["detalle"].get(detalle_txt, 0) + cant

        for _, rr in circuitos_df.iterrows():
            nombre_circ = str(rr.get("Circuito", "")).strip()
            if "ilumin" not in nombre_circ.lower():
                continue  # solo interesa iluminación
            tm_txt = rr.get("Disyuntor termomagnético", "")
            in_tm = parse_in_tm(tm_txt)
            amp_sw = "16A" if (in_tm == 16) else "10A"  # el amperaje del interruptor de pared depende del TM
            items = (items_por_nombre or {}).get(nombre_circ, [])
            if not isinstance(items, list) or not items:
                continue
            # agrupa las luminarias de este circuito por ambiente
            lum_by_amb = {}
            for it in items:
                if not isinstance(it, dict):
                    continue
                if ("modulos" in it) or ("id_ench" in it) or ("n_ench" in it) or ("nombre" in it):
                    continue
                amb = str(it.get("amb", "")).strip()
                if not amb:
                    amb = "sin_amb"
                key = amb.lower()
                lum_by_amb[key] = lum_by_amb.get(key, 0) + 1
            # recorre cada ambiente y arma la fila de interruptores correspondiente
            for amb_key, n_lum in lum_by_amb.items():
                amb_show = amb_key  # nombre "bonito" del ambiente (con mayúsculas originales)
                try:
                    match = ambientes_df[ambientes_df["Ambiente"].astype(str).str.strip().str.lower() == amb_key]
                    if len(match) > 0:
                        amb_show = str(match.iloc[0]["Ambiente"]).strip()
                except:
                    pass
                n_conmutadas = int(max(0, ncon_map.get(amb_key, 0)))
                n_conmutadas = min(n_conmutadas, int(n_lum))
                n_restantes = int(n_lum) - n_conmutadas
                if n_conmutadas > 0:
                    _acum_interruptor(
                        f"Interruptor 9/24 (conmutado) {amp_sw} 250V",
                        f"{nombre_circ} -> {amb_show}",
                        2  # siempre 2 (1 par), sin importar cuántas luminarias controle
                    )
                if n_restantes > 0:
                    c12, c15, c32 = descomponer_interruptores(int(n_restantes))  # cómo se agrupan las no conmutadas
                    if c12 > 0:
                        _acum_interruptor(f"Interruptor 9/12 {amp_sw} 250V", f"{nombre_circ} -> {amb_show}", c12)
                    if c15 > 0:
                        _acum_interruptor(f"Interruptor 9/15 {amp_sw} 250V", f"{nombre_circ} -> {amb_show}", c15)
                    if c32 > 0:
                        _acum_interruptor(f"Interruptor 9/32 {amp_sw} 250V", f"{nombre_circ} -> {amb_show}", c32)

        # emite 1 sola fila por descripción, sumando todos los circuitos/ambientes
        for desc, info in sorted(_interruptores_acum.items(), key=lambda x: x[0].lower()):
            detalle_txt = ", ".join(f"{k} ({v})" for k, v in info["detalle"].items())
            add_row(
                desc=desc,
                marcas_txt=marcas.get("Interruptores", ""),
                norma="RIC 7.3.2",
                circuito=detalle_txt,
                unidad="u",
                k=1,
                longitud_m=f"{info['total']} unid",
                cantidad=info["total"]
            )

    # ENCHUFES COMUNES: por circuito + ambiente + tipo
    ench_rows = []
    _conexion_directa_acum = {"total": 0, "detalle": {}}  # circuitos >16A sin enchufe común
    if {"Circuito", "Disyuntor termomagnético"}.issubset(set(circuitos_df.columns)):
        for _, rr in circuitos_df.iterrows():
            nombre_circ = str(rr.get("Circuito", "")).strip()
            tm_txt = rr.get("Disyuntor termomagnético", "")
            in_tm = parse_in_tm(tm_txt)
            # Usar items_por_nombre en vez de _items del DataFrame
            items = (items_por_nombre or {}).get(nombre_circ, [])
            if not isinstance(items, list):
                items = []
            if not nombre_circ or len(items) == 0:
                continue
            items_ench = []
            for it in items:
                if isinstance(it, dict) and (("modulos" in it) or ("id_ench" in it)):
                    items_ench.append(it)  # solo se queda con los items que son enchufes
            if not items_ench:
                continue
            if in_tm is not None and in_tm > 16:
                # circuito de más de 16A: no lleva enchufe común, va a conexión directa
                _conexion_directa_acum["total"] += len(items_ench)
                _conexion_directa_acum["detalle"][nombre_circ] = _conexion_directa_acum["detalle"].get(nombre_circ, 0) + len(items_ench)
                continue

            # arma una fila por cada enchufe, para después agruparlas
            for it in items_ench:
                amb = str(it.get("amb", "")).strip()
                mod = int(it.get("modulos", 1)) if str(it.get("modulos", "")).strip() != "" else 1
                if in_tm == 16:
                    nominal = "10/16A"
                else:
                    nominal = "10A"
                ench_rows.append({
                    "Circuito": nombre_circ,
                    "Ambiente": amb if amb else "Sin ambiente",
                    "Nominal": nominal,
                    "Modulos": mod
                })
    if _conexion_directa_acum["total"] > 0:
        detalle_txt = ", ".join(f"{k} ({v})" for k, v in _conexion_directa_acum["detalle"].items())
        add_row(
            desc="Punto para conexión directa (>16A) - sin enchufe",
            marcas_txt="",
            norma="RIC 13.5.1",
            circuito=detalle_txt,
            unidad="u",
            k=1,
            longitud_m=f"{_conexion_directa_acum['total']} unid",
            cantidad=_conexion_directa_acum["total"]
        )
    if ench_rows:
        # agrupa todos los enchufes por circuito+ambiente+amperaje+módulos, para
        # sacar la cantidad de cada combinación
        ench_df = pd.DataFrame(ench_rows)
        grp = ench_df.groupby(["Circuito", "Ambiente", "Nominal", "Modulos"]).size().reset_index(name="Cantidad")
        mod_order = {1: 0, 2: 1, 3: 2}
        grp["__ord"] = grp["Modulos"].map(lambda x: mod_order.get(int(x), 9))
        grp = grp.sort_values(["Nominal", "__ord", "Circuito", "Ambiente"]).drop(columns="__ord")
        # consolida en 1 fila por descripción (nominal+módulos), sumando todos
        # los circuitos/ambientes, con el detalle de cada uno en "Circuito"
        _enchufes_acum = {}  # {desc: {"total": N, "detalle": {"circuito y ambiente": cant}}}
        # recorre cada combinación de circuito+ambiente+amperaje+módulos y la
        # va sumando dentro de _enchufes_acum, según el nombre final que le
        # corresponda (simple/doble/triple)
        for _, g in grp.iterrows():
            circ = g["Circuito"]
            amb = g["Ambiente"]
            nominal = g["Nominal"]
            mod = int(g["Modulos"])
            cant = int(g["Cantidad"])
            if mod == 1:
                tipo = "Enchufe simple"
            elif mod == 2:
                tipo = "Enchufe doble"
            else:
                tipo = "Enchufe triple"
            desc = f"{tipo} 2P+T {nominal} 250V"
            detalle_txt = f"{circ} -> {amb}"
            if desc not in _enchufes_acum:
                _enchufes_acum[desc] = {"total": 0, "detalle": {}}
            _enchufes_acum[desc]["total"] += cant
            _enchufes_acum[desc]["detalle"][detalle_txt] = _enchufes_acum[desc]["detalle"].get(detalle_txt, 0) + cant
        for desc, info in _enchufes_acum.items():
            detalle_txt = ", ".join(f"{k} ({v})" for k, v in info["detalle"].items())
            add_row(
                desc=desc,
                marcas_txt=marcas.get("Enchufes", ""),
                norma="RIC 13.5.1",
                circuito=detalle_txt,
                unidad="u",
                k=1,
                longitud_m=f"{info['total']} unid",
                cantidad=info["total"] if info["total"] > 0 else ""
            )

    # Enchufe dedicado para climatización, siempre 2P+T 10/16A
    for _cl in _clima_items:
        if _cl["con_enchufe"]:
            add_row(
                desc="Enchufe dedicado para aire acondicionado simple 2P+T 10/16A 250V",
                marcas_txt=marcas.get("Enchufes", ""),
                norma="RIC 7 (7.3.1, 7.3.2, 7.4.4)",
                circuito=_cl["circ"],
                unidad="u",
                k=1,
                longitud_m="1 unid",
                cantidad=1
            )

    # Tabla prensaestopa según sección conductor (PG/Métrico)
    # Incluye calibres AWG equivalentes
    PRENSAESTOPA_POR_SECCION = {
        1.5:  ("PG11",   "M16"),   # 1,5mm²
        2.08: ("PG11",   "M16"),   # AWG 14
        2.5:  ("PG11",   "M20"),   # 2,5mm²
        3.31: ("PG11",   "M20"),   # AWG 12
        4.0:  ("PG13,5", "M20"),   # 4mm²
        5.26: ("PG13,5", "M20"),   # AWG 10
        6.0:  ("PG16",   "M25"),   # 6mm²
        8.37: ("PG16",   "M25"),   # AWG 8
        10.0: ("PG21",   "M25"),   # 10mm²
    }

    def prensaestopa_para_seccion(sec):
        # sec = sección del cable en mm² (ej: 4.0), este es el dato de entrada
        # pg  = tamaño del prensaestopa (ej: "PG11")
        # mt  = la rosca métrica equivalente a ese pg (ej: "M20")
        # c   = cada una de las secciones que hay en la tabla PRENSAESTOPA_POR_SECCION
        # claves = todas las secciones "c" de la tabla, ordenadas de menor a mayor
        claves = sorted(PRENSAESTOPA_POR_SECCION.keys())
        pg, mt = PRENSAESTOPA_POR_SECCION[claves[-1]]  # parte con el prensaestopa más grande, por si acaso
        for c in claves:
            # +0.01 es solo para evitar problemas de redondeo con decimales
            if sec <= c + 0.01:  # ¿el cable (sec) ya cabe en esta sección de tabla (c)?
                pg, mt = PRENSAESTOPA_POR_SECCION[c]  # sí cabe: se queda con el pg/mt de esta sección
                break
        sec_txt = str(round(sec, 2)).replace(".", ",")  # sección con coma decimal (2,5 en vez de 2.5)
        return f"Prensaestopa {pg} ({mt}) para cordón (cable 3x{sec_txt}mm²)"

    # ─────────────────────────────────────────────────────────────────────────
    # ACCESORIOS AGUA CALIENTE
    # Con enchufe (TM≤16A) o conexión fija sin enchufe (TM>16A) — RIC N°07 7.2.8 / 7.3
    # Genera prensaestopa en equipo (solo sin enchufe) + tablero externo si corresponde
    # NOTA: este bloque va DESPUÉS de definir PRENSAESTOPA_POR_SECCION
    # ─────────────────────────────────────────────────────────────────────────
    if {"Circuito", "Conductor", "Disyuntor termomagnético"}.issubset(set(circuitos_df.columns)):
        for _, rr_ac in circuitos_df.iterrows():
            nombre_ac = str(rr_ac.get("Circuito", "")).strip()
            es_agua_mat = any(k in nombre_ac.lower() for k in
                              ("ducha", "termo", "calefon", "calefón",
                               "calentador", "agua caliente"))
            if not es_agua_mat:
                continue  # no es un circuito de agua caliente, se salta

            sec_ac = extraer_seccion_mm2(str(rr_ac.get("Conductor", ""))) or 4.0
            in_tm_ac = parse_in_tm(str(rr_ac.get("Disyuntor termomagnético", "")))
            if in_tm_ac is None:
                # no se pudo leer el TM del circuito: se marca para que lo revisen a mano
                add_row(
                    desc=f"(definir) {nombre_ac} - TM no legible, revisar con/sin enchufe manualmente",
                    marcas_txt="",
                    norma="",
                    circuito=nombre_ac,
                    unidad="u",
                    k=1,
                    longitud_m="1 unid",
                    cantidad=1
                )
                continue

            # Buscar datos del equipo en circuitos_agua_caliente
            _datos_ac = {}
            for eq_ac in circuitos_agua_caliente:
                if eq_ac.get("nombre_circ", "").lower() in nombre_ac.lower():
                    _datos_ac = eq_ac
                    break
            lleva_tab_ext = _datos_ac.get("lleva_tablero_externo", False)

            # Calcular prensaestopa según sección real del conductor
            _pg_ac, _mt_ac = ("PG11", "M16")  # valor por defecto
            for _sec_k, (_pg_k, _mt_k) in sorted(PRENSAESTOPA_POR_SECCION.items()):
                if sec_ac <= _sec_k + 0.01:
                    _pg_ac, _mt_ac = _pg_k, _mt_k  # primer tamaño de la tabla que alcanza
                    break
            sec_txt_ac = str(round(sec_ac, 1)).replace(".", ",")

            # Sección bornera PE: igual a sección conductor (si es ≤16mm², usa esa misma sección)
            # Normalizar a sección comercial de bornera disponible
            if sec_ac <= 1.5:   sec_born_pe = 1.5
            elif sec_ac <= 2.5: sec_born_pe = 2.5
            elif sec_ac <= 4.0: sec_born_pe = 4.0
            elif sec_ac <= 6.0: sec_born_pe = 6.0
            else:               sec_born_pe = 10.0
            sec_born_pe_txt = str(sec_born_pe).replace(".", ",")

            # 1) Punto conexión fija (sin enchufe) o enchufe (TM≤16A), igual criterio que climatización
            con_enchufe_ac = (in_tm_ac <= 16)
            if con_enchufe_ac:
                # Los enchufes solo existen comercialmente en 10A o 16A — se
                # redondea el TM real al calibre comercial correcto
                _amp_ench_ac = 10 if in_tm_ac <= 10 else 16
                add_row(
                    desc=f"Enchufe 2P+T {_amp_ench_ac}A / 250V para {_datos_ac.get('tipo_equipo', 'equipo agua caliente')}",
                    marcas_txt=marcas.get("Enchufes", ""),
                    norma="RIC N°07 (7.3.1, 7.3.2, 7.4.4)",
                    circuito=nombre_ac,
                    unidad="u",
                    k=1,
                    longitud_m="1 unid",
                    cantidad=1
                )
            else:
                add_row(
                    desc=f"Punto para conexión fija — {_datos_ac.get('tipo_equipo', 'equipo agua caliente')} (sin enchufe)",
                    marcas_txt="",
                    norma="RIC N°07 (7.2.8) / RIC N°13 (5.1)",
                    circuito=nombre_ac,
                    unidad="u",
                    k=1,
                    longitud_m="1 unid",
                    cantidad=1
                )

            # 2) Prensaestopa en equipo — solo para conexión fija (sin enchufe);
            # con enchufe no hay cordón entrando a una caja que sellar
            if not con_enchufe_ac:
                add_row(
                    desc=f"Prensaestopa {_pg_ac} ({_mt_ac}) para cordón — entrada equipo (cable 3×{sec_txt_ac}mm²)",
                    marcas_txt=marcas.get("Prensaestopa", ""),
                    norma="RIC N°04 (5.15, 5.24) / RIC N°07 (5.2.8)",
                    circuito=nombre_ac,
                    unidad="u",
                    k=1,
                    longitud_m="1 unid",
                    cantidad=1
                )

            if not lleva_tab_ext:
                continue  # si no tiene tablero externo, este circuito termina aquí

            # ── TABLERO EXTERNO DE DESCONEXIÓN ──────────────────────────────
            # RIC N°07 art. 7.2.8: tablero de comando a la vista del equipo
            # RIC N°11 sección 6: fuera de Volúmenes 0, 1 y 2

            # 3) Tablero sobrepuesto 6 puestos IP41
            # TM bipolar 2p (2 puestos) + borneras F+N+PE en riel (2-3 puestos): 6 puestos alcanza holgado
            add_row(
                desc="Tablero sobrepuesto de PVC 6 puestos IP41 (desconexión agua caliente)",
                marcas_txt=marcas.get("Tablero externo agua caliente", marcas.get("Tablero sobrepuesto", "")),
                norma="RIC N°02 (5.2, 6.1) / RIC N°07 (7.2.8, 7.3.3, 7.4.1, 7.4.2) / RIC N°11 (6, Vol.3)",
                circuito=nombre_ac,
                unidad="u",
                k=1,
                longitud_m="1 unid",
                cantidad=1
            )

            # 4) Riel DIN 35×7,5mm tira de 10cm
            add_row(
                desc="Riel DIN 35×7,5mm tira de 10cm (tablero externo agua caliente)",
                marcas_txt=marcas.get("Riel DIN", ""),
                norma="RIC N°02 (6.1.15, 6.1.23)",
                circuito=nombre_ac,
                unidad="u",
                k=1,
                longitud_m="1 unid",
                cantidad=1
            )

            # 5) TM bipolar 1P+N — mismo calibre que TM del tablero principal
            add_row(
                desc=f"Disyuntor termomagnético 1P+N {in_tm_ac}A / 6kA / Curva C (desconexión local agua caliente)",
                marcas_txt=marcas.get("TM bipolar agua caliente", marcas.get("Protecciones", "")),
                norma="RIC N°07 (7.2.8, 7.3.4, 7.4.2, 7.4.5)",
                circuito=nombre_ac,
                unidad="u",
                k=1,
                longitud_m="1 unid",
                cantidad=1
            )

            # 6) Bornera PE tablero externo — sección según conductor real (RIC 6 tabla 6.4)
            add_row(
                desc=f"Bornera de conexión PE {sec_born_pe_txt}mm² (tierra tablero externo agua caliente)",
                marcas_txt=marcas.get("Bornera PE agua caliente", marcas.get("Barra unipolar verde", "")),
                norma="RIC N°02 (6.2.7) / RIC N°06 (5.11, 5.14)",
                circuito=nombre_ac,
                unidad="u",
                k=1,
                longitud_m="1 unid",
                cantidad=1
            )

            # 9) Fijación tablero externo — 6 puntos para el tablero a la pared
            #    + 2 puntos para el riel DIN = 8 puntos en total, SIEMPRE del
            #    mismo tipo de tornillo (no tiene sentido mezclar 2 tipos para
            #    fijar la misma pieza). En volcanita/fibrocemento, además,
            #    los 8 puntos necesitan tarugo (no solo un tornillo suelto),
            #    porque ahí van montados los equipos de protección y pesan —
            #    sin tarugo el tornillo se puede salir. En madera u otro
            #    forrado sólido, el riel DIN no necesita tarugo (el tornillo
            #    muerde bien directo), por lo que van 8 tornillos simples.
            _amb_row_ac = _get_amb_row(ambientes_df, _datos_ac.get("ambientes_str", nombre_ac))
            _mat_amb_ac = str(_amb_row_ac.get("Material forrado interior", "")).strip().lower() if _amb_row_ac is not None else ""
            if ("volcan" in _mat_amb_ac) or ("vulcan" in _mat_amb_ac) or ("fibro" in _mat_amb_ac):
                # 6 (tablero a la pared) + 2 (riel DIN) = 8 puntos, todos con tarugo
                add_row(
                    desc="Tarugo paloma 6mm",
                    marcas_txt=marcas.get("Tarugo paloma", ""),
                    norma="-",
                    circuito=nombre_ac,
                    unidad="u",
                    k=1,
                    longitud_m="8 unid",
                    cantidad=8
                )
                add_row(
                    desc='Tornillo volcanita punta fina 6x1 1/4"',
                    marcas_txt=marcas.get("Tornillo para tarugo paloma", ""),
                    norma="-",
                    circuito=nombre_ac,
                    unidad="u",
                    k=1,
                    longitud_m="8 unid",
                    cantidad=8
                )
            else:
                # Madera u otro forrado sólido: NO tiene sentido usar 2 tipos
                # de tornillo distintos (6 para el tablero + 2 "para tarugo"
                # para el riel DIN) si no hay tarugo de por medio en ningún
                # punto. Se unifica todo en 1 solo tipo de tornillo, 8 unidades
                # (6 tablero + 2 riel DIN), igual que en volcanita/fibrocemento
                # donde los 8 puntos también son del mismo tipo (ahí con tarugo).
                if "madera" in _mat_amb_ac:
                    _desc_torn_ext = 'Tornillo punta fina para madera cabeza lenteja 6x1/2"'
                else:
                    _desc_torn_ext = 'Tornillo (definir según forrado)'
                add_row(
                    desc=_desc_torn_ext,
                    marcas_txt=marcas.get("Tornillos", ""),
                    norma="Instalación",
                    circuito=nombre_ac,
                    unidad="u",
                    k=1,
                    longitud_m="8 unid",
                    cantidad=8
                )

            # 10) Nota normativa obligatoria en informe
            add_row(
                desc=(
                    f"NOTA NORMATIVA — Tablero externo {nombre_ac}: "
                    f"Instalar FUERA de Volúmenes 0, 1 y 2 (RIC N°11 sección 6). "
                    f"Debe quedar a la vista directa del equipo (RIC N°07 art. 7.4.2). "
                    f"IP mínimo 41 en interior seco, IP44 en ambiente húmedo."
                ),
                marcas_txt="",
                norma="RIC N°07 (7.2.8, 7.3.3, 7.4.1, 7.4.2) / RIC N°11 (6, Tabla Volúmenes)",
                circuito=nombre_ac,
                unidad="—",
                k=1,
                longitud_m="—",
                cantidad=""
            )
    # ─────────────────────────────────────────────────────────────────────────
    # PRENSAESTOPA CLIMATIZACIÓN — mismo criterio que agua caliente y especiales:
    # con enchufe (TM≤16A) no necesita, porque el cordón termina en un enchufe,
    # no entra a una caja. Sin enchufe (TM>16A) sí necesita, porque el cordón
    # del equipo entra a la caja de derivación y hay que sujetarlo/sellarlo ahí
    # (la conexión eléctrica en sí, dentro de la caja, ya se hace con cónico —
    # eso no cambia, esto es solo la entrada física del cordón).
    # ─────────────────────────────────────────────────────────────────────────
    for _cl_pe in _clima_items:
        if _cl_pe["es_emb"] and _cl_pe["con_enchufe"] is False:
            _fila_cl_pe = circuitos_df[circuitos_df["Circuito"].astype(str).str.contains(
                _cl_pe["circ"].split("(")[0].strip(), regex=False, na=False
            )]
            sec_cl_pe = extraer_seccion_mm2(str(_fila_cl_pe["Conductor"].iloc[0])) if len(_fila_cl_pe) > 0 else None
            sec_cl_pe = sec_cl_pe or 2.5
            _pg_cl, _mt_cl = ("PG11", "M16")  # valor por defecto
            for _sec_k, (_pg_k, _mt_k) in sorted(PRENSAESTOPA_POR_SECCION.items()):
                if sec_cl_pe <= _sec_k + 0.01:
                    _pg_cl, _mt_cl = _pg_k, _mt_k
                    break
            sec_txt_cl = str(round(sec_cl_pe, 1)).replace(".", ",")
            add_row(
                desc=f"Prensaestopa {_pg_cl} ({_mt_cl}) para cordón — entrada equipo (cable 3×{sec_txt_cl}mm²)",
                marcas_txt=marcas.get("Prensaestopa", ""),
                norma="RIC N°04 (5.15, 5.24) / RIC N°07 (5.2.8)",
                circuito=_cl_pe["circ"],
                unidad="u",
                k=1,
                longitud_m="1 unid",
                cantidad=1
            )

    # CIRCUITOS ESPECIALES GENÉRICOS (horno, lavadora, encimera, etc.):
    # punto de conexión (enchufe o directo si supera 16A) para cada equipo especial
    if {"Circuito", "Disyuntor termomagnético"}.issubset(set(circuitos_df.columns)):
        _esp_acum = {}  # {desc: {"total": N, "detalle": {"circuito y ambiente": cant}}}

        def _acum_especial(desc, detalle_txt, cant):
            if desc not in _esp_acum:
                _esp_acum[desc] = {"total": 0, "detalle": {}}
            _esp_acum[desc]["total"] += cant
            _esp_acum[desc]["detalle"][detalle_txt] = _esp_acum[desc]["detalle"].get(detalle_txt, 0) + cant

        for _, rr in circuitos_df.iterrows():
            nombre_circ = str(rr.get("Circuito", "")).strip()

            # Excluir circuitos de climatización
            es_clima_row = any(k in nombre_circ.lower() for k in ("climatiz", "aire acond", "split"))
            if es_clima_row:
                continue
            # Excluir circuitos de agua caliente (tienen su propia lógica de materiales)
            es_agua_row = any(k in nombre_circ.lower() for k in
                              ("ducha", "termo", "calefon", "calefón", "calentador", "agua caliente"))
            if es_agua_row:
                continue

            tm_txt = rr.get("Disyuntor termomagnético", "")
            in_tm = parse_in_tm(tm_txt)
            items = (items_por_nombre or {}).get(nombre_circ, [])
            if not isinstance(items, list):
                items = []
            if not nombre_circ or len(items) == 0:
                continue
            items_esp = [it for it in items if isinstance(it, dict) and ("nombre" in it)]  # solo los items "especiales"
            if not items_esp:
                continue
            for it in items_esp:
                amb = str(it.get("amb", "")).strip() or "Sin ambiente"
                equipo = str(it.get("nombre", "")).strip() or "equipo especial"
                tm_val = in_tm if in_tm is not None else 16
                detalle_txt = f"{nombre_circ} -> {amb}"
                if tm_val > 16:
                    # más de 16A: no lleva enchufe, va a conexión directa + prensaestopa
                    desc = f"Punto para conexión directa de {equipo} (>16A) - sin enchufe"
                    _acum_especial(desc, detalle_txt, 1)
                    # Prensaestopa según sección del conductor
                    sec_esp = extraer_seccion_mm2(str(rr.get("Conductor", "")))
                    desc_pg = prensaestopa_para_seccion(sec_esp) if sec_esp else "Prensaestopa PG13,5 (M20) para cordón (cable 3x4,0mm²)"
                    _acum_especial(desc_pg, detalle_txt, 1)
                else:
                    # 16A o menos: sí lleva enchufe dedicado — siempre simple,
                    # porque un enchufe "dedicado" es para 1 solo equipo
                    nominal = "10A" if tm_val <= 10 else "10/16A"
                    desc = f"Enchufe dedicado para {equipo} simple 2P+T {nominal} 250V"
                    _acum_especial(desc, detalle_txt, 1)

        # emite 1 sola fila por descripción, sumando todos los circuitos/ambientes
        for desc, info in sorted(_esp_acum.items(), key=lambda x: x[0].lower()):
            detalle_txt = ", ".join(f"{k} ({v})" for k, v in info["detalle"].items())
            if "prensaestopa" in desc.lower():
                marca_txt, norma_txt = marcas.get("Prensaestopa", ""), "RIC 4.5.1"
            else:
                marca_txt, norma_txt = marcas.get("Enchufes", ""), "RIC 13.5.1"
            add_row(
                desc=desc,
                marcas_txt=marca_txt,
                norma=norma_txt,
                circuito=detalle_txt,
                unidad="u",
                k=1,
                longitud_m=f"{info['total']} unid",
                cantidad=info["total"]
            )

    def texto_seccion(sec):
        # convierte 2.5 en "2,5" para que se vea con coma como acá en Chile
        return str(sec).replace(".", ",")

    def borneras_por_puesto(sec):
        """Cuántas borneras de una sección caben en 1 puesto del riel DIN."""
        if sec <= 1.5:  return 4
        if sec <= 2.5:  return 3
        if sec <= 4.0:  return 3
        if sec <= 6.0:  return 2
        return 1  # 10mm² y mayor

    # Nota: el cálculo de borneras de neutro/organización vive más abajo,
    # en _borneras_neutro_acum / _borneras_organizacion_acum (RIC 2, 6.2.12).

    # =========================
    # CONECTORES CÓNICOS (POR SECCIÓN)
    # Reglas:
    #   - 3 por enchufe
    #   - 3 por caja de derivación de UNIONES
    #   - 2 por foco (luminaria)
    # Se elige color según sección del circuito donde pertenece.
    # =========================
    add_section("Conectores cónicos")
    conicos_por_color = {}  # key=(color, num, rango) -> {"cantidad":int, "circuitos":set()}
    def _accum_conico(color, num, rango, cantidad, circuito_nombre):
        # va sumando cuántos conectores cónicos de cada color/tamaño se
        # necesitan, y anota de qué circuito son
        key = (color, num, rango)  # identifica el tipo de conector cónico
        if key not in conicos_por_color:
            conicos_por_color[key] = {"cantidad": 0, "circuitos": set()}  # primera vez que aparece
        conicos_por_color[key]["cantidad"] += int(max(0, cantidad))  # suma la cantidad
        if circuito_nombre:
            conicos_por_color[key]["circuitos"].add(str(circuito_nombre))  # anota de qué circuito viene
    n_circ_estaño = 0  # circuitos interiores con conductor >6mm² se conectan con estaño
    circ_estaño_txt = []
    _n_focos_led_estaño = 0  # focos LED que se estañan
    _circ_focos_led = []
    if {"Circuito", "Conductor"}.issubset(set(circuitos_df.columns)):
        for _, rr in circuitos_df.iterrows():
            circ_name = str(rr.get("Circuito", "")).strip()
            cond_txt = str(rr.get("Conductor", "")).strip()
            items = (items_por_nombre or {}).get(circ_name, [])
            if not isinstance(items, list):
                items = []
            secc = extraer_seccion_mm2(cond_txt)

            # Climatización, agua caliente y especiales (horno, lavadora, etc.) tienen
            # su propia lógica de conexión — se resuelven acá, antes de que el código
            # más abajo salte este circuito por no tener sus datos en "items"
            # (esos casos guardan sus datos aparte, en _clima_items, no en items_por_nombre)
            es_clima_row    = any(k in circ_name.lower() for k in ("climatiz","aire","split","ac ","a/c"))
            es_agua_row     = any(k in circ_name.lower() for k in ("ducha","termo","calefon","calefón","calentador","agua caliente"))
            es_especial_row = any(k in circ_name.lower() for k in ("especial","horno","encimera","lavadora","lavaplatos","secadora","jacuzzi","piscina"))
            _especial_generico = es_especial_row and not es_agua_row and not es_clima_row
            if es_agua_row:
                _in_tm_ac_row = parse_in_tm(str(rr.get("Disyuntor termomagnético", "")))
                if _in_tm_ac_row is None:
                    _accum_conico("(definir)", None, "", 1, f"{circ_name} - TM no legible, revisar con/sin enchufe manualmente")
                    continue
                if _in_tm_ac_row <= 16:
                    continue  # si tiene enchufe, los ferrules del enchufe se agregan en su propio bloque
                es_especial_row = True  # sin enchufe (TM>16A) = conexión fija, mismo caso que horno/lavadora
            if es_clima_row:
                _circ_base_cl = circ_name.split("(")[0].strip()
                _cl_match_cl = next((_cl for _cl in _clima_items if _cl["circ"].split("(")[0].strip() == _circ_base_cl), None)
                if _cl_match_cl is not None and _cl_match_cl["con_enchufe"] is None:
                    _accum_conico("(definir)", None, "", 1, f"{circ_name} - TM no legible, revisar con/sin enchufe manualmente")
                    continue
                if _cl_match_cl is not None and _cl_match_cl["con_enchufe"]:
                    continue  # ya se cuenta en la sección de ferrules del enchufe (C2)
                es_especial_row = True  # climatización sin enchufe = conexión fija, mismo caso que horno/lavadora
            if _especial_generico:
                # Mismo criterio que agua caliente y climatización (arriba): TM
                # decide con/sin enchufe, no la potencia. Si tiene enchufe (TM≤16A),
                # se salta acá, los ferrules del enchufe van en su propio bloque
                # (D3, más abajo en el archivo). Si no tiene enchufe (TM>16A), sigue
                # como es_especial_row=True y cae al tratamiento de conexión fija.
                # Si el TM no se puede leer, no se adivina — se deja un aviso
                # "(definir)" para que se revise a mano, igual que ya se hace
                # cuando falla la columna Conductor.
                _in_tm_esp_row = parse_in_tm(str(rr.get("Disyuntor termomagnético", "")))
                if _in_tm_esp_row is None:
                    _accum_conico("(definir)", None, "", 1, f"{circ_name} - TM no legible, revisar con/sin enchufe manualmente")
                    continue
                if _in_tm_esp_row <= 16:
                    continue
            if es_especial_row:
                # Conexión fija = cola de rata + soldadura/cónico (NO bornera con
                # ferrule — el equipo no tiene bornes, se empalma directo al cable).
                # Se suma acá mismo (antes de que se cierren los totales de cónico/
                # estaño/cintas más abajo) para no generar un rollo aparte por poco.
                # El corte "con/sin enchufe" ya se resolvió arriba por TM, igual
                # criterio que agua caliente y climatización — este bloque solo
                # se alcanza para circuitos sin enchufe (TM>16A), por lo que el
                # cónico/estaño se genera siempre, sin límite de potencia.
                if secc is not None and secc > 6.0:
                    n_circ_estaño += 1
                    circ_estaño_txt.append(f"{circ_name} - conexión fija equipo")
                else:
                    color_esp, num_esp, rango_esp = conico_por_seccion(secc, n_cables=3)
                    _accum_conico(color_esp, num_esp, rango_esp, 3, f"{circ_name} - conexión fija equipo")
                continue

            # Circuitos con conductor >6mm² se conectan con estaño, no con conector cónico
            if secc is not None and secc > 6.0:
                n_circ_estaño += 1
                circ_estaño_txt.append(circ_name)
                continue

            if not isinstance(items, list) or len(items) == 0:
                continue

            # Contar enchufes y luminarias por ambiente
            n_ench_circ = 0
            n_lum_estaño_circ = 0
            lum_by_amb_con = {}  # {amb_key: [items_lum]}

            for it in items:
                if not isinstance(it, dict):
                    continue
                if ("id_ench" in it) or ("modulos" in it) or ("n_ench" in it):
                    n_ench_circ += int(it.get("n_ench", 1) or 1)
                    continue
                if "nombre" in it:
                    continue
                # Es luminaria
                amb = str(it.get("amb", "")).strip().lower() or "sin_amb"
                if amb not in lum_by_amb_con:
                    lum_by_amb_con[amb] = []
                lum_by_amb_con[amb].append(it)
                if _lum_necesita_estaño(it.get("tipo_lum", ""), it.get("desc_lum", "")):
                    n_lum_estaño_circ += 1  # esta luminaria se conecta con estaño, no cónico

            # ── ILUMINACIÓN: cónicos ──────────────────────────────────────────
            # Caja troncal:            +3 cónicos POR CADA interruptor del ambiente
            # Caja octogonal intermedia (no última del grupo): +2 cónicos
            # Caja octogonal última del grupo: 0 cónicos
            # Caja interruptor (9/12/15/32/24): 0 cónicos
            #
            # De TODOS los cónicos de iluminación del circuito, los últimos 3
            # se unen con 2 cables (van con la tabla de 2 cables), el resto
            # (total - 3) se une con 3 cables (tabla de 3 cables).
            if lum_by_amb_con:
                con_ilum = 0
                for amb_key, lums_amb in lum_by_amb_con.items():
                    n_lum_amb = len(lums_amb)
                    n_conm_amb = int(min(n_lum_amb, ncon_map_local.get(amb_key, 0)))
                    n_rest_amb = n_lum_amb - n_conm_amb
                    _c32 = n_rest_amb // 3
                    _rem = n_rest_amb % 3
                    _c15 = 1 if _rem == 2 else 0
                    _c12 = 1 if _rem == 1 else 0

                    # Caja troncal: 3 por CADA interruptor NO conmutado del
                    # ambiente (grupos 9/12/15/32), más 3 si el ambiente tiene
                    # conmutadas — siempre 1 sola caja troncal para el par
                    # conmutado, sin importar cuántas luminarias conmutadas
                    # haya (mismo criterio de "1 par por ambiente" que ya se
                    # usa en el cálculo de chicotes para cinta aislante).
                    n_interruptores_amb = (1 if n_conm_amb > 0 else 0) + _c12 + _c15 + _c32
                    con_ilum += 3 * n_interruptores_amb

                    # Cajas octogonales intermedias: última del grupo: 0
                    # Grupo conmutadas 9/24: 3 cónicos c/u (lleva 1 cable extra
                    # por el conductor rojo que sigue hacia la siguiente
                    # luminaria conmutada), entonces son n_conm_amb-1 intermedias
                    # Grupos no conmutadas: 2 cónicos c/u, descomponer en c12/c15/c32
                    #   9/32 (3 lum): 2 intermedias · 9/15 (2 lum): 1 intermedia · 9/12 (1 lum): 0 intermedias
                    if n_conm_amb > 1:
                        con_ilum += 3 * (n_conm_amb - 1)
                    if n_rest_amb > 0:
                        con_ilum += 2 * (_c32 * 2)  # 9/32: 2 intermedias por grupo
                        con_ilum += 2 * (_c15 * 1)  # 9/15: 1 intermedia por grupo
                        # 9/12: 0 intermedias

                if con_ilum > 0:
                    con_ilum_2cables = min(3, con_ilum)
                    con_ilum_3cables = max(0, con_ilum - 3)
                    if con_ilum_3cables > 0:
                        color_i3, num_i3, rango_i3 = conico_por_seccion(secc, n_cables=3)
                        _accum_conico(color_i3, num_i3, rango_i3, con_ilum_3cables, f"{circ_name} (iluminación, 3 cables)")
                    if con_ilum_2cables > 0:
                        color_i2, num_i2, rango_i2 = conico_por_seccion(secc, n_cables=2)
                        _accum_conico(color_i2, num_i2, rango_i2, con_ilum_2cables, f"{circ_name} (iluminación, 2 cables)")

            # ── ENCHUFES: cónicos ─────────────────────────────────────────────
            # Enchufe intermedio: +3 cónicos
            # Último enchufe de CADA ambiente (no solo el último de todo el
            # circuito): 0 cónicos
            # Caja adicional entre ambientes: +3 cónicos
            # Enchufes siempre se unen con 3 cables (F/N/T)
            if n_ench_circ > 0:
                _n_ench_por_amb_conico = {}
                for it in items:
                    if not isinstance(it, dict): continue
                    if ("id_ench" in it) or ("modulos" in it) or ("n_ench" in it):
                        _a = str(it.get("amb", "")).strip().lower() or "sin_amb"
                        _n_ench_por_amb_conico[_a] = _n_ench_por_amb_conico.get(_a, 0) + int(it.get("n_ench", 1) or 1)
                _cajas_adic_circ = int((cajas_adic_por_nombre or {}).get(circ_name, 0))
                con_ench = sum(3 * max(0, n_amb_ench - 1) for n_amb_ench in _n_ench_por_amb_conico.values())
                con_ench += 3 * _cajas_adic_circ
                if con_ench > 0:
                    color, num, rango = conico_por_seccion(secc, n_cables=3)
                    _accum_conico(color, num, rango, con_ench, f"{circ_name} (enchufes)")

            # Acumular focos LED para estaño
            if n_lum_estaño_circ > 0:
                _n_focos_led_estaño += n_lum_estaño_circ
                _circ_focos_led.append(circ_name)
    # si por algún motivo no se generó ningún cónico (columnas faltantes, etc.),
    # se hace una estimación aproximada en base a los totales generales
    # (sin dato de sección real disponible acá, se asume la típica: 2.5mm²
    # para enchufes y 1.5mm² para iluminación)
    if not conicos_por_color:
        if total_enchufes > 0:
            _total_cajas_adic = sum(int(v) for v in (cajas_adic_por_nombre or {}).values())
            # Enchufes: 3*(n_ench-1) + 3*cajas_adic, siempre 3 cables
            _con_ench_fb = 3 * max(0, total_enchufes - 1) + 3 * _total_cajas_adic
            if _con_ench_fb > 0:
                _color_efb, _num_efb, _rango_efb = conico_por_seccion(2.5, n_cables=3)
                _accum_conico(_color_efb, _num_efb, _rango_efb, _con_ench_fb, "Enchufes (estimado)")
        if total_luminarias > 0:
            # Sin info de ambientes, se aproxima con 1 ambiente por circuito:
            # troncal (3×n_interruptores) + octogonales intermedias
            # (no se conoce cuántas son conmutadas, se asume 0)
            _c32_fb = total_luminarias // 3
            _rem_fb = total_luminarias % 3
            _c15_fb = 1 if _rem_fb == 2 else 0
            _c12_fb = 1 if _rem_fb == 1 else 0
            _n_int_fb = _c12_fb + _c15_fb + _c32_fb
            _con_ilum_fb = 3 * _n_int_fb + 2 * (_c32_fb * 2 + _c15_fb * 1)  # troncal + oct intermedias
            if _con_ilum_fb > 0:
                _con_ilum_fb_2 = min(3, _con_ilum_fb)
                _con_ilum_fb_3 = max(0, _con_ilum_fb - 3)
                if _con_ilum_fb_3 > 0:
                    _color_i3fb, _num_i3fb, _rango_i3fb = conico_por_seccion(1.5, n_cables=3)
                    _accum_conico(_color_i3fb, _num_i3fb, _rango_i3fb, _con_ilum_fb_3, "Iluminación (estimado, 3 cables)")
                if _con_ilum_fb_2 > 0:
                    _color_i2fb, _num_i2fb, _rango_i2fb = conico_por_seccion(1.5, n_cables=2)
                    _accum_conico(_color_i2fb, _num_i2fb, _rango_i2fb, _con_ilum_fb_2, "Iluminación (estimado, 2 cables)")
    # arma una fila por cada color/tamaño de conector cónico acumulado
    for (color, num, rango), info in sorted(conicos_por_color.items(), key=lambda x: (str(x[0][0]), str(x[0][1]))):
        cant = int(info["cantidad"])
        if cant <= 0:
            continue
        if color == "Amarillo" and num == 44:
            desc = f"Conector cónico Amarillo (N° {num}) para {rango}"
        elif num is not None:
            desc = f"Conector cónico {color} (N° {num}) para {rango}"
        else:
            desc = f"Conector cónico {color} para {rango}"
        circuitos_txt = "\n".join(sorted(info["circuitos"])) if info["circuitos"] else "Varios"
        add_row(
            desc=desc,
            marcas_txt=marcas.get("Conectores cónicos", ""),
            norma="RIC 4.4.1",
            circuito=circuitos_txt,
            unidad="u",
            k=1,
            longitud_m=f"{cant} unid",
            cantidad=cant
        )

    # ================================================================
    # ESTAÑO, PASTA Y CINTAS — TODO UNIFICADO
    # Fuentes:
    #   A) Circuitos >6mm²  dan conexiones_circ (conductor grueso)
    #   B) Puesta a tierra  da conexiones_pt1 + conexiones_pt2 + conexiones_desnudo (conductor grueso, 16mm², va con A al grupo grueso ÷4)
    #   C) Focos LED        dan _n_focos_led_estaño × 3 (conductor fino)
    # ================================================================
    conexiones_circ    = n_circ_estaño * 3
    conexiones_pt1     = n_barras_pt1 * 2
    conexiones_pt2     = n_barras_pt2 * 1
    conexiones_desnudo = ((n_barras_pt1 - 1) * 2) + ((n_barras_pt2 - 1) * 2)
    _conexiones_grueso = conexiones_circ + conexiones_pt1 + conexiones_pt2 + conexiones_desnudo
    _conexiones_focos  = _n_focos_led_estaño * 3  # conductor fino

    _hay_estaño = (_conexiones_grueso > 0) or (_conexiones_focos > 0)

    if _hay_estaño:
        # Tubos: conductor grueso (>6mm², incluye circuitos y PT) usa 1 cada 4 conexiones
        #        conductor fino (solo focos LED) usa 1 cada 15 conexiones
        _conexiones_fino = _conexiones_focos
        _conexiones_grueso_solo = _conexiones_grueso  # circuitos >6mm² + PT (pt1+pt2+desnudo)
        _tubos_total = math.ceil(_conexiones_grueso_solo / 4 + _conexiones_fino / 15) if (_conexiones_grueso_solo + _conexiones_fino) > 0 else 0
        _tubos_total = max(1, _tubos_total)
        # Pasta: 1 frasco cada 4 tubos
        _pasta_total = max(1, math.ceil(_tubos_total / 4))

        # Armar texto de circuitos
        _circ_estaño_all = []
        if circ_estaño_txt:
            _circ_estaño_all += circ_estaño_txt
        if conexiones_pt1 > 0 or conexiones_pt2 > 0:
            _circ_estaño_all.append("Puesta a tierra N°1 y N°2")
        if _circ_focos_led:
            _circ_estaño_all += _circ_focos_led
        _circ_estaño_str_unif = " / ".join(sorted(set(_circ_estaño_all))) if _circ_estaño_all else "General"

        add_row(
            desc="Tubo de estaño 1m / 17gr",
            marcas_txt=marcas.get("Estaño", ""),
            norma="RIC 4 (5.11.1)",
            circuito=_circ_estaño_str_unif,
            unidad="u",
            k=1,
            longitud_m=f"{_tubos_total} unid",
            cantidad=_tubos_total
        )
        add_row(
            desc="Pasta para soldar 50gr",
            marcas_txt=marcas.get("Pasta para soldar", ""),
            norma="-",
            circuito=_circ_estaño_str_unif,
            unidad="u",
            k=1,
            longitud_m=f"{_pasta_total} unid",
            cantidad=_pasta_total
        )

        # Cinta autofundente de goma — sobre estaño (focos LED y >6mm²)
        # 20cm por conexión, rollo 3m = 300cm, redondeando siempre hacia arriba (conexiones × 20 / 300)
        _todas_conexiones_goma = _conexiones_focos + (n_circ_estaño * 3)
        if _todas_conexiones_goma > 0:
            _rollos_goma = max(1, math.ceil(_todas_conexiones_goma * 20 / 300))
            _circ_goma_all = []
            if circ_estaño_txt:
                _circ_goma_all += circ_estaño_txt
            if _circ_focos_led:
                _circ_goma_all += _circ_focos_led
            _circ_goma_str = " / ".join(sorted(set(_circ_goma_all))) if _circ_goma_all else "General"
            add_row(
                desc="Cinta autofundente de goma 3m",
                marcas_txt=marcas.get("Cinta autofundente goma", ""),
                norma="-",
                circuito=_circ_goma_str,
                unidad="u",
                k=1,
                longitud_m=f"{_rollos_goma} unid",
                cantidad=_rollos_goma
            )

    # Cinta aislante PVC — unificada: cónicos (30cm c/u) + focos LED (20cm c/u) + circuitos >6mm² (35cm c/u)
    # No incluye puesta a tierra ni conductor desnudo (mismo criterio que la cinta de goma)
    # Rollo 20m = 2000cm
    total_conicos = sum(info["cantidad"] for info in conicos_por_color.values())
    _cm_conicos = total_conicos * 30
    _cm_focos   = _conexiones_focos * 20
    _cm_circ_grueso = conexiones_circ * 35
    _cm_total_pvc = _cm_conicos + _cm_focos + _cm_circ_grueso
    if _cm_total_pvc > 0:
        _rollos_pvc = max(1, math.ceil(_cm_total_pvc / 2000))
        _circs_pvc = set()
        for info in conicos_por_color.values():
            _circs_pvc.update(info["circuitos"])
        if _circ_focos_led:
            _circs_pvc.update(_circ_focos_led)
        if circ_estaño_txt:
            _circs_pvc.update(circ_estaño_txt)
        _circs_pvc_txt = " / ".join(sorted(_circs_pvc)) if _circs_pvc else "Varios circuitos"
        add_row(
            desc="Cinta aislante PVC 20m",
            marcas_txt=marcas.get("Cinta aislante PVC", ""),
            norma="-",
            circuito=_circs_pvc_txt,
            unidad="u",
            k=1,
            longitud_m=f"{_rollos_pvc} unid",
            cantidad=_rollos_pvc
        )

    # =========================
    # LUMINARIAS: 1 fila por tipo de luminaria (todos los ambientes juntos)
    # =========================
    if ambientes_df is not None and "Detalle iluminación" in ambientes_df.columns and "Ambiente" in ambientes_df.columns:
        add_section("Iluminarias")
        conteo = {}  # {desc_bonita: {"total": N, "por_ambiente": {amb: cant}}}
        # recorre cada ambiente y cuenta cuántas veces se repite cada descripción de luminaria
        for _, ar in ambientes_df.iterrows():
            amb_show = str(ar.get("Ambiente", "")).strip() or "Sin ambiente"
            det = str(ar.get("Detalle iluminación", "") or "").strip()
            if not det or det.lower() == "ninguna":
                continue
            lineas = [x.strip() for x in det.split("\n") if x.strip()]
            for ln in lineas:
                if ln not in conteo:
                    conteo[ln] = {"total": 0, "por_ambiente": {}}
                conteo[ln]["total"] += 1
                conteo[ln]["por_ambiente"][amb_show] = conteo[ln]["por_ambiente"].get(amb_show, 0) + 1
        # arma 1 fila por cada tipo de luminaria distinta, sumando todos los ambientes
        for desc_bonita, info in sorted(conteo.items(), key=lambda x: x[0].lower()):
            detalle_amb = ", ".join(
                f"{amb} ({cant})" for amb, cant in sorted(info["por_ambiente"].items(), key=lambda x: x[0].lower())
            )
            add_row(
                desc=desc_bonita,
                marcas_txt=marcas.get("Iluminarias", ""),
                norma="RIC 10 (5.1.4.4, 5.1.4.5)",
                circuito=f"Iluminación -> {detalle_amb}",
                unidad="u",
                k=1,
                longitud_m=f"{info['total']} unid",
                cantidad=info["total"]
            )

    # =========================================================
    # Barra unipolar verde según cantidad de circuitos
    # Regla usuario:
    #  1    circuito:   4 polos
    #  2-3  circuitos:  6 polos
    #  4-5  circuitos:  8 polos
    #  6-7  circuitos:  10 polos
    #  8-9  circuitos:  12 polos
    #  10+  circuitos:  15 polos
    # =========================================================
    n_circ = 0
    try:
        # cuenta circuitos reales (filas con "Circuito" no vacío)
        if circuitos_df is not None and "Circuito" in circuitos_df.columns:
            n_circ = int(
                circuitos_df["Circuito"].astype(str).str.strip().replace("", np.nan).dropna().shape[0]
            )
        else:
            n_circ = int(len(circuitos_df)) if circuitos_df is not None else 0
    except:
        n_circ = int(len(circuitos_df)) if circuitos_df is not None else 0
    polos_barra_verde = None
    # busca cuántos polos corresponde según el rango de circuitos (ver tabla arriba)
    if n_circ == 1:
        polos_barra_verde = 4
    elif 2 <= n_circ <= 3:
        polos_barra_verde = 6
    elif 4 <= n_circ <= 5:
        polos_barra_verde = 8
    elif 6 <= n_circ <= 7:
        polos_barra_verde = 10
    elif 8 <= n_circ <= 9:
        polos_barra_verde = 12
    elif n_circ >= 10:
        polos_barra_verde = 15
    if polos_barra_verde is not None:
        add_row(
            desc=f"Barra unipolar verde de {polos_barra_verde} polos 63A",
            marcas_txt=marcas.get("Barra unipolar verde", ""),
            norma="RIC 9.2.1",
            circuito="General",
            unidad="u",
            k=1,
            longitud_m=f"{polos_barra_verde} polos",
            cantidad=1
        )

    # =========================================================
    # BARRA REPARTIDORA TETRAPOLAR PRINCIPAL — siempre 4 polos fija
    # =========================================================
    n_dif = 0
    try:
        n_dif = int(len(group_info)) if group_info else 0  # cantidad de diferenciales (1 por grupo)
    except:
        n_dif = 0
    add_row(
        desc="Barra repartidora tetrapolar de 4 polos de 125[A]",
        marcas_txt=marcas.get("Barra repartidora", ""),
        norma="RIC 9.2.1",
        circuito="General",
        unidad="u",
        k=1,
        longitud_m="4 polos",
        cantidad=1
    )

    # =========================================================
    # BARRA REPARTIDORA BIPOLAR según nº de diferenciales
    #  1-3 diferenciales:  4 polos
    #  4-6 diferenciales:  7 polos
    #  7-10 diferenciales: 11 polos
    # =========================================================
    polos_bipolar = None
    if 1 <= n_dif <= 3:
        polos_bipolar = 4
    elif 4 <= n_dif <= 6:
        polos_bipolar = 7
    elif 7 <= n_dif <= 10:
        polos_bipolar = 11
    if polos_bipolar is not None:
        add_row(
            desc=f"Barra repartidora bipolar de {polos_bipolar} polos de 125[A]",
            marcas_txt=marcas.get("Barra repartidora", ""),
            norma="RIC 2 (6.2.1, 6.2.4, 6.2.7)",
            circuito="General",
            unidad="u",
            k=1,
            longitud_m=f"{polos_bipolar} polos",
            cantidad=1
        )
    # =========================================================
    # CONDICIÓN 2:
    # Barras repartidoras tetrapolares EXTRA de 4 polos 125A
    # Regla (2 grupos por barra, no 2 TM por barra):
    #  - Tomar SOLO diferenciales que tengan 2 o 3 TM
    #  - Contar cuántos GRUPOS válidos hay: eso da total_grupos_validos
    #  - Cada barra 4P alimenta 2 grupos
    #  - barras_extra = ceil(total_grupos_validos / 2)
    # Ejemplos:
    #   1 grupo (2 o 3 TM):  1 barra
    #   2 grupos (2 o 3 TM): 1 barra
    #   3 grupos (2 o 3 TM): 2 barras
    #   4 grupos (2 o 3 TM): 2 barras
    # =========================================================
    total_grupos_validos = 0
    try:
        if group_info:
            for meta in group_info.values():
                n_tm = int(len(meta.get("indices", [])))
                if n_tm in (2, 3):
                    total_grupos_validos += 1
    except:
        total_grupos_validos = 0
    barras_extra_4p = int(math.ceil(total_grupos_validos / 2.0)) if total_grupos_validos > 0 else 0
    if barras_extra_4p > 0:
        add_row(
            desc="Barra repartidora tetrapolar de 4 polos de 125[A]",
            marcas_txt=marcas.get("Barra repartidora", ""),
            norma="RIC 9.2.1",
            circuito="General",
            unidad="u",
            k=1,
            longitud_m=f"{barras_extra_4p} barra(s) de 4 polos",
            cantidad=barras_extra_4p
        )
    # =========================================================
    # TABLERO (PVC IP41) – CONDICIONES 1 + 2 + 3 + REDONDEO
    # (DEBE IR DESPUÉS DE calcular polos_barra_verde, polos_principal y barras_extra_4p)
    # =========================================================
    # contar TMs reales (1 puesto c/u)
    n_tm = 0
    try:
        if {"Disyuntor termomagnético", "Circuito"}.issubset(set(circuitos_df.columns)):
            tms_df = circuitos_df[["Disyuntor termomagnético", "Circuito"]].copy()
            tms_df["Disyuntor termomagnético"] = tms_df["Disyuntor termomagnético"].astype(str).str.strip()
            tms_df = tms_df[tms_df["Disyuntor termomagnético"] != ""]
            n_tm = int(len(tms_df))
    except:
        n_tm = 0
    # contar diferenciales (2 puestos c/u)
    n_dif = int(len(group_info)) if group_info else 0

    # CONDICIÓN 1: puestos por protecciones — se van sumando los puestos que
    # ocupa cada elemento del tablero (cada uno tiene un ancho fijo en puestos)
    puestos_omni = 2
    puestos_tm = n_tm * 1
    puestos_dif = n_dif * 2
    puestos_luz_piloto = 1 if n_circ >= 3 else 0
    puestos_portafusible = 1 if n_circ >= 3 else 0
    puestos_spd = 2
    puestos_protector_sobrevoltaje = 2
    puestos_barra_verde = 1 if polos_barra_verde is not None else 0
    puestos_barra_principal = 2  # siempre 4 polos = 2 puestos
    puestos_barra_bipolar = 0
    if polos_bipolar == 4:
        puestos_barra_bipolar = 2
    elif polos_bipolar == 7:
        puestos_barra_bipolar = 4
    elif polos_bipolar == 11:
        puestos_barra_bipolar = 6
    puestos_barras_extra = int(barras_extra_4p) * 2 if barras_extra_4p > 0 else 0
    # Puestos de bornera — se agrupan por sección, ya que varias borneras
    # chicas comparten el mismo puesto del riel (no es 1 puesto fijo por
    # circuito). Hay 2 tipos, con condiciones distintas, y NO se reemplazan
    # entre sí porque son para conductores distintos:
    #  - Bornera de neutro: existe siempre que haya un diferencial exclusivo
    #    (1:1), sin importar la cantidad total de circuitos del proyecto.
    #    Es solo para el conductor neutro.
    #  - Bornera de organización: solo si el proyecto tiene más de 8
    #    circuitos. Es para el conductor fase, y aplica a TODOS los
    #    circuitos en ese caso, tengan o no bornera de neutro.
    puestos_borneras = 0

    def _seccion_A1_local(corriente_A):
        # busca la sección de conductor (mm²) más chica que aguante esta
        # corriente, usando la tabla de ampacidad método A1 (embutido/ducto)
        secciones_A1 = [
            (1.5, 14), (2.5, 18), (4.0, 24), (6.0, 31),
            (10.0, 42), (16.0, 56), (25.0, 73), (35.0, 89),
        ]
        for sec, iz in secciones_A1:
            if iz >= corriente_A:
                return sec
        return 35.0

    _borneras_por_seccion = {}  # {sección: cantidad de borneras}

    # Bornera de neutro (siempre, sin condición de n_circ)
    if group_info:
        for _gid, _meta in group_info.items():
            _indices_grupo = _meta.get("indices", [])
            if len(_indices_grupo) == 1:
                # si el diferencial es exclusivo, lleva bornera de neutro, con la sección del diferencial
                _dif_a = _meta.get("dif", None)
                _sec_dif_local = _seccion_A1_local(_dif_a) if _dif_a else None
                if _sec_dif_local:
                    _borneras_por_seccion[_sec_dif_local] = _borneras_por_seccion.get(_sec_dif_local, 0) + 1

    # Bornera de organización (solo si n_circ>8) — aplica a TODOS los
    # circuitos, incluyendo los que ya tienen bornera de neutro, porque
    # es para el conductor fase, no el neutro.
    if n_circ > 8 and {"Conductor", "Circuito"}.issubset(set(circuitos_df.columns)):
        for _idx, _r in circuitos_df.iterrows():
            _cond_b = str(_r.get("Conductor", "")).strip()
            _sec_b = extraer_seccion_mm2(_cond_b)
            if _sec_b:
                _borneras_por_seccion[_sec_b] = _borneras_por_seccion.get(_sec_b, 0) + 1

    for _sec_b, _cantidad_b in _borneras_por_seccion.items():
        puestos_borneras += int(math.ceil(_cantidad_b / borneras_por_puesto(_sec_b)))
    puestos_base = int(puestos_omni + puestos_tm + puestos_dif + puestos_luz_piloto + puestos_portafusible + puestos_spd
        + puestos_protector_sobrevoltaje + puestos_barra_principal + puestos_barra_bipolar
        + puestos_barras_extra + puestos_barra_verde + puestos_borneras)

    # CONDICIÓN 2: reserva por circuitos — deja espacio libre para futuras ampliaciones
    puestos_reserva = int(math.ceil(n_circ * 0.25)) * 3 if n_circ > 0 else 0
    # CONDICIÓN 3: suma total
    puestos_total = int(puestos_base + puestos_reserva)
    # REDONDEO A TABLEROS COMERCIALES: busca el tamaño de tablero comercial
    # más chico que alcance a cubrir los puestos necesarios
    tamanos_tablero = [2, 4, 6, 8, 12, 16, 18, 24, 36, 42, 48, 54, 56, 72]
    puestos_tablero = None
    for t in tamanos_tablero:
        if t >= puestos_total:
            puestos_tablero = t
            break
    if puestos_tablero is None:
        puestos_tablero = tamanos_tablero[-1]  # se pasó de todos los tamaños, usa el más grande
    _puestos_tablero = int(puestos_tablero)
    # descripción según canalización (embutida/sobrepuesta)
    tipo_can = (tipo_canalizacion or "").strip().lower()
    es_embutida_tab = "embut" in tipo_can
    if es_embutida_tab:
        desc_tab = f"Tablero embutido de PVC de {puestos_tablero} puestos IP41"
        marca_tab = marcas.get("Tablero embutido", "")
    else:
        desc_tab = f"Tablero sobrepuesto de PVC de {puestos_tablero} puestos IP41"
        marca_tab = marcas.get("Tablero sobrepuesto", "")
    add_row(
        desc=desc_tab,
        marcas_txt=marca_tab,
        norma="RIC 9.1.1",
        circuito=f"General ({puestos_total} puestos requeridos)",
        unidad="u",
        k=1,
        longitud_m=f"{puestos_tablero} puestos",
        cantidad=1
    )
    # =========================================================
    # RIEL DIN según tamaño de tablero
    # - 1m para: 4,8,12,16,18,24,36,42 puestos
    # - 2m para: 56 o 72 puestos
    # =========================================================
    if puestos_tablero in [56, 72]:
        desc_riel = "Riel DIN de 35x7,5 mm tira de 2m"
        largo_riel = 2
    else:
        desc_riel = "Riel DIN de 35x7,5 mm tira de 1m"
        largo_riel = 1
    _riel_m = largo_riel
    add_row(
        desc=desc_riel,
        marcas_txt=marcas.get("Riel DIN", ""),
        norma="RIC 9.1.1",
        circuito="General",
        unidad="u",
        k=1,
        longitud_m=f"{largo_riel} m",
        cantidad=1
    )
    # =========================================================
    # LUZ PILOTO + PORTAFUSIBLE + FUSIBLE (solo si hay >=3 circuitos)
    # =========================================================
    if n_circ >= 3:
        # Luz piloto tablero
        add_row(
            desc="Luz piloto LED Riel Din 220 VAC IP44",
            marcas_txt=marcas.get("Luz Piloto", ""),
            norma="RIC 9.1.1",
            circuito="Tablero de alumbrado",
            unidad="u",
            k=1,
            longitud_m="1 Unid",
            cantidad=1
        )
        # Portafusible tablero
        add_row(
            desc="Portafusible 1P 32A 10x38mm 500V",
            marcas_txt=marcas.get("Portafusible", ""),
            norma="RIC 9.1.1",
            circuito="Tablero de alumbrado",
            unidad="u",
            k=1,
            longitud_m="1 Unid",
            cantidad=1
        )
        # Fusible tablero
        add_row(
            desc="Fusible cilíndrico 2A 10x38mm 500V",
            marcas_txt=marcas.get("Fusible", ""),
            norma="RIC 9.1.1",
            circuito="Tablero de alumbrado",
            unidad="u",
            k=1,
            longitud_m="1 Unid",
            cantidad=1
        )

        # =========================
    # TERMINALES FERRUL INTERIORES
    # =========================
    add_section("Terminales ferrul interiores")

    def normalizar_seccion_ferrul(seccion_mm2):
        # redondea la sección hacia arriba a la sección comercial de ferrul
        # más cercana (1.5, 2.5, 4.0, 6.0, 10.0, 16.0, 25.0 o 35.0mm²)
        try:
            s = float(seccion_mm2)  # intenta convertir a número
        except:
            return None  # no era un número válido

        secciones = [1.5, 2.5, 4.0, 6.0, 10.0, 16.0, 25.0, 35.0]  # secciones comerciales de ferrul

        for sec in secciones:
            if s <= sec:
                return sec  # primera sección comercial que alcanza a cubrirla

        return secciones[-1]  # era más grande que todas, se usa la máxima

    def color_ferrul_por_seccion(sec):
        # color del ferrul según la sección del conductor (código de colores
        # típico: 1.5mm²=rojo, 2.5mm²=azul, etc.)
        colores = {
            1.5: "rojo",
            2.5: "azul",
            4.0: "naranjo",
            6.0: "amarillo",
            10.0: "rojo",
            16.0: "azul",
            25.0: "amarillo",
            35.0: "gris"
        }  # tabla de colores por sección

        try:
            return colores.get(float(sec), "por definir")  # busca el color, si no está lo marca "por definir"
        except:
            return "por definir"  # la sección no era un número válido

    ferrules = {}

    def agregar_ferrules(seccion_mm2, cantidad, detalle):
        # suma ferrules al conteo total, según la sección que corresponda.
        # Si la sección no es válida o la cantidad es 0, no hace nada
        sec = normalizar_seccion_ferrul(seccion_mm2)  # sección comercial de ferrul

        if sec is None:
            return  # sección inválida, no hace nada

        cantidad = int(cantidad)

        if cantidad <= 0:
            return  # no hay nada que sumar

        if sec not in ferrules:
            ferrules[sec] = {}  # primera vez que se usa esta sección

        ferrules[sec][detalle] = ferrules[sec].get(detalle, 0) + cantidad  # suma al total de esa sección/detalle

    def seccion_circuito(cond_txt):
        # saca la sección en mm² del texto del conductor (ej: "3x2.5mm²" da 2.5)
        return extraer_seccion_mm2(str(cond_txt))

    def seccion_A1(corriente_A):
        """Selecciona la sección mínima del conductor interior del tablero
        usando método A1 (RIC 2 punto 6.2.2), secciones comerciales H07Z1-K."""
        secciones_A1 = [
            (1.5, 14), (2.5, 18), (4.0, 24), (6.0, 31),
            (10.0, 42), (16.0, 56), (25.0, 73), (35.0, 89),
        ]  # (sección mm², capacidad de corriente A)
        for sec, iz in secciones_A1:
            if iz >= corriente_A:
                return sec  # la primera que alcanza a soportar la corriente
        return 35.0  # máximo disponible

    def contar_enchufes_items(items):
        # suma los enchufes del circuito (misma idea que _contar_enchufes_en_items,
        # pero definida acá porque se usa en otra parte de la función)
        total = 0

        if isinstance(items, list):
            for it in items:
                if not isinstance(it, dict):
                    continue  # item raro, se salta

                if ("id_ench" in it) or ("modulos" in it) or ("n_ench" in it):
                    total += int(it.get("n_ench", 1) or 1)  # suma la cantidad de este item

        return int(total)

    def buscar_grupo_por_indice(idx_circuito):
        # busca a qué grupo de diferencial pertenece este circuito
        if not group_info:
            return None, None  # no hay agrupación definida

        for gid, meta in group_info.items():
            indices = meta.get("indices", [])

            if idx_circuito in indices:
                return gid, meta  # encontró el grupo al que pertenece

        return None, None  # no pertenece a ningún grupo

    try:
        sec_alimentador = float(res_alim["S"])  # sección del alimentador ya calculada
        # Sección conductor PE según RIC 6.7 (igual a fase para ≤16mm²)
        if sec_alimentador <= 16:
            sec_pt_tablero = sec_alimentador  # PE igual a la fase
        elif sec_alimentador <= 35:
            sec_pt_tablero = 16.0  # PE fijo en 16mm²
        else:
            sec_pt_tablero = sec_alimentador / 2.0  # PE a la mitad de la fase
    except:
        sec_alimentador = None
        sec_pt_tablero = 4.0  # valor de respaldo si algo falla

    sec_control = 1.5

    try:
        n_circ_ferrul = int(
            circuitos_df["Circuito"]
            .astype(str)
            .str.strip()
            .replace("", np.nan)
            .dropna()
            .shape[0]
        )
    except:
        n_circ_ferrul = int(len(circuitos_df)) if circuitos_df is not None else 0

    usa_luz_piloto = n_circ_ferrul >= 3

    # -------------------------
    # A) TABLERO FIJO
    # -------------------------
    # Todo lo que cuelga de la barra principal, protegido por el mismo
    # interruptor general omnipolar (entre la barra y el omnipolar, el SPD,
    # el protector de sobrevoltaje) se dimensiona según Método A1 (RIC 2,
    # 6.2.2) con la corriente nominal del omnipolar — NO con la sección del
    # alimentador.
    # Ej: alimentador 4mm² (RV-K) pero omnipolar 2x25A: el método A1 exige 6mm²
    # (4mm² solo aguanta hasta 24A, no alcanza para un interruptor de 25A).
    # Solo el tramo entre el alimentador y la barra principal sigue usando la
    # sección real del alimentador (es el mismo cable que ya viene puesto).
    sec_omni_A1 = seccion_A1(interruptor_empalme)
    agregar_ferrules(sec_alimentador, 2, "Alimentador F/N a barra principal")
    agregar_ferrules(sec_alimentador, 1, "Alimentador PE a barra tierra")
    agregar_ferrules(sec_pt_tablero, 1, "Barra verde PE tablero")
    agregar_ferrules(sec_omni_A1, 4, "Barra principal a omnipolar")
    agregar_ferrules(sec_omni_A1, 4, "Omnipolar salida a barra principal")
    agregar_ferrules(sec_omni_A1, 4, "Barra principal a SPD F/N")
    agregar_ferrules(sec_omni_A1, 2, "SPD a barra PE")
    agregar_ferrules(sec_omni_A1, 4, "Barra principal a protector sobrevoltaje")
    agregar_ferrules(sec_omni_A1, 4, "Protector sobrevoltaje a barra bipolar")

    if usa_luz_piloto:
        agregar_ferrules(sec_control, 6, "Luz piloto + portafusible")

    # -------------------------
    # B) DIFERENCIALES + TM + SALIDAS
    # -------------------------
    if {"Circuito", "Conductor", "Disyuntor termomagnético"}.issubset(set(circuitos_df.columns)):
        grupos_ya_contados = set()  # para no contar 2 veces los ferrules de entrada al mismo diferencial
        grupos_barra_extra_contados = set()
        _borneras_neutro_acum = {}  # {sec_dif: {"cantidad": N, "circuitos": [...]}} — se fusionan en 1 fila por sección
        _borneras_organizacion_acum = {}  # {sec_circ: {"cantidad": N, "circuitos": [...]}} — 1 por circuito, cuando n_circ>8

        for idx, r in circuitos_df.iterrows():
            circ = str(r.get("Circuito", "")).strip()
            cond = str(r.get("Conductor", "")).strip()
            tm = str(r.get("Disyuntor termomagnético", "")).strip()

            if not circ or not cond or not tm:
                continue

            sec_circ = seccion_circuito(cond)
            gid, meta = buscar_grupo_por_indice(idx)

            if meta is None:
                # este circuito no pertenece a ningún grupo de diferencial (caso raro/individual)
                agregar_ferrules(sec_circ, 2, f"{circ} - barra a TM")
                if n_circ > 8:
                    agregar_ferrules(sec_circ, 2, f"{circ} - salida TM a entrada bornera")
                    agregar_ferrules(sec_circ, 3, f"{circ} - salida bornera al circuito")
                    if sec_circ not in _borneras_organizacion_acum:
                        _borneras_organizacion_acum[sec_circ] = {"cantidad": 0, "circuitos": []}
                    _borneras_organizacion_acum[sec_circ]["cantidad"] += 1
                    _borneras_organizacion_acum[sec_circ]["circuitos"].append(circ)
                else:
                    agregar_ferrules(sec_circ, 3, f"{circ} - salida F/N/PE")
                continue

            indices_grupo = meta.get("indices", [])
            n_circuitos_grupo = int(len(indices_grupo))  # cuántos circuitos comparten este diferencial
            dif_a = meta.get("dif", None)
            sec_dif = seccion_A1(dif_a)  # tabla métrica limpia (1,5-2,5-4-6-10-16-25-35), igual que el omnipolar

            if gid not in grupos_ya_contados:
                # ferrules de entrada al diferencial, solo se cuentan 1 vez por grupo
                agregar_ferrules(sec_dif, 4, f"Diferencial G{gid} entrada desde barra bipolar")
                grupos_ya_contados.add(gid)

            if n_circuitos_grupo == 1:
                # el diferencial tiene 2 salidas: fase y neutro. La FASE va
                # del diferencial a la entrada del TM — este tramo usa la
                # sección del DIFERENCIAL (sec_dif), no la del circuito,
                # porque todavía es cable "grueso" antes de llegar al TM.
                agregar_ferrules(sec_dif, 2, f"{circ} - diferencial a TM")

                # El NEUTRO del diferencial no pasa por el TM (el TM solo
                # censura la fase) — va directo a una bornera propia, con
                # el mismo criterio: sección del diferencial, 1 ferrule en
                # la salida del diferencial + 1 en la entrada de la bornera.
                agregar_ferrules(sec_dif, 2, f"{circ} - diferencial neutro a bornera")
                if sec_dif not in _borneras_neutro_acum:
                    _borneras_neutro_acum[sec_dif] = {"cantidad": 0, "circuitos": []}
                _borneras_neutro_acum[sec_dif]["cantidad"] += 1
                _borneras_neutro_acum[sec_dif]["circuitos"].append(circ)

                # Salida de la FASE: si el proyecto tiene más de 8 circuitos,
                # también necesita su propia bornera de organización — es un
                # conductor distinto al neutro, así que ambas borneras
                # coexisten para este circuito, no se reemplazan entre sí.
                # Mismo criterio que el caso de circuitos compartidos: entrada
                # a la bornera (calibre del TM) + salida al circuito (calibre
                # del circuito).
                if n_circ > 8:
                    agregar_ferrules(sec_circ, 2, f"{circ} - salida TM a entrada bornera organización (fase)")
                    agregar_ferrules(sec_circ, 3, f"{circ} - salida al circuito (fase+neutro+tierra, cada una desde su propia bornera)")
                    if sec_circ not in _borneras_organizacion_acum:
                        _borneras_organizacion_acum[sec_circ] = {"cantidad": 0, "circuitos": []}
                    _borneras_organizacion_acum[sec_circ]["cantidad"] += 1
                    _borneras_organizacion_acum[sec_circ]["circuitos"].append(circ)
                else:
                    agregar_ferrules(sec_circ, 3, f"{circ} - salida F/N/PE")

            elif n_circuitos_grupo in (2, 3):
                # el diferencial comparte 2 o 3 circuitos: necesita una barra extra para repartir
                if gid not in grupos_barra_extra_contados:
                    agregar_ferrules(sec_dif, 4, f"Diferencial G{gid} a barra extra")
                    grupos_barra_extra_contados.add(gid)

                # "barra extra a TM" usa sec_dif (cable grueso), no sec_circ
                # — todavía es cable del calibre del diferencial, igual que
                # "diferencial a TM" en el caso de 1 circuito por diferencial
                agregar_ferrules(sec_dif, 2, f"{circ} - barra extra a TM")
                if n_circ > 8:
                    agregar_ferrules(sec_circ, 2, f"{circ} - salida TM a entrada bornera")
                    agregar_ferrules(sec_circ, 3, f"{circ} - salida bornera al circuito")
                    if sec_circ not in _borneras_organizacion_acum:
                        _borneras_organizacion_acum[sec_circ] = {"cantidad": 0, "circuitos": []}
                    _borneras_organizacion_acum[sec_circ]["cantidad"] += 1
                    _borneras_organizacion_acum[sec_circ]["circuitos"].append(circ)
                else:
                    agregar_ferrules(sec_circ, 3, f"{circ} - salida F/N/PE")

            else:
                # cualquier otro caso (no debería pasar, pero por seguridad se cubre igual)
                agregar_ferrules(sec_dif, 2, f"{circ} - diferencial a TM")
                if n_circ > 8:
                    agregar_ferrules(sec_circ, 2, f"{circ} - salida TM a entrada bornera")
                    agregar_ferrules(sec_circ, 3, f"{circ} - salida bornera al circuito")
                    if sec_circ not in _borneras_organizacion_acum:
                        _borneras_organizacion_acum[sec_circ] = {"cantidad": 0, "circuitos": []}
                    _borneras_organizacion_acum[sec_circ]["cantidad"] += 1
                    _borneras_organizacion_acum[sec_circ]["circuitos"].append(circ)
                else:
                    agregar_ferrules(sec_circ, 3, f"{circ} - salida F/N/PE")

        # Borneras de neutro (modo "1 diferencial por circuito"): 1 fila por
        # sección, con la cantidad total y el detalle de cada circuito en
        # la columna Circuito (en vez de 1 fila repetida por cada circuito).
        for _sec_b, _info_b in sorted(_borneras_neutro_acum.items()):
            _sec_b_txt = texto_seccion(_sec_b)
            _circ_txt = ", ".join(_info_b["circuitos"])
            add_row(
                desc=f"Bornera de conexión {_sec_b_txt}mm²",
                marcas_txt=marcas.get("Bornera de conexion", ""),
                norma="RIC 2 (6.2.12)",
                circuito=f"Neutro diferencial (1 por circuito): {_circ_txt}",
                unidad="u",
                k=1,
                longitud_m=f"{_info_b['cantidad']} unid",
                cantidad=_info_b["cantidad"]
            )

        # Borneras de organización (n_circ>8): 1 fila por sección, mismo
        # criterio que la bornera de neutro — 1 bornera por cada circuito
        # que use esta rama.
        for _sec_o, _info_o in sorted(_borneras_organizacion_acum.items()):
            _sec_o_txt = texto_seccion(_sec_o)
            _circ_o_txt = ", ".join(_info_o["circuitos"])
            add_row(
                desc=f"Bornera de conexión {_sec_o_txt}mm²",
                marcas_txt=marcas.get("Bornera de conexion", ""),
                norma="RIC 2 (6.2.12)",
                circuito=f"Organización salida circuitos (>8 circuitos): {_circ_o_txt}",
                unidad="u",
                k=1,
                longitud_m=f"{_info_o['cantidad']} unid",
                cantidad=_info_o["cantidad"]
            )

    # -------------------------
    # C) ENCHUFES COMUNES
    # -------------------------
    if {"Circuito", "Conductor"}.issubset(set(circuitos_df.columns)):
        fallback_enchufes = False

        for _, r in circuitos_df.iterrows():
            circ = str(r.get("Circuito", "")).strip()

            if "enchufe" not in circ.lower():
                continue  # solo interesan los circuitos de enchufes

            sec_ench = seccion_circuito(r.get("Conductor", ""))
            n_ench_circ = contar_enchufes_items((items_por_nombre or {}).get(circ, []))

            if n_ench_circ > 0:
                agregar_ferrules(sec_ench, 3 * n_ench_circ, f"{circ} - enchufes ({n_ench_circ})")
            else:
                fallback_enchufes = True  # no se pudo contar por items, hay que estimar más abajo

        # si algún circuito de enchufes no tenía items detallados, se estima
        # con el total general de enchufes por ambiente
        if fallback_enchufes and ambientes_df is not None and "Cantidad enchufes (u)" in ambientes_df.columns:
            total_enchufes = int(
                pd.to_numeric(
                    ambientes_df["Cantidad enchufes (u)"],
                    errors="coerce"
                ).fillna(0).sum()
            )

            secciones_ench = []

            for _, r in circuitos_df.iterrows():
                if "enchufe" in str(r.get("Circuito", "")).lower():
                    sec = seccion_circuito(r.get("Conductor", ""))

                    if sec is not None:
                        secciones_ench.append(float(sec))

            sec_enchufe = max(secciones_ench) if secciones_ench else None  # usa la sección más grande, para no quedar corto

            if total_enchufes > 0:
                agregar_ferrules(sec_enchufe, 3 * total_enchufes, f"Enchufes comunes ({total_enchufes})")

    # -------------------------
    # C2) ENCHUFE CLIMATIZACIÓN (solo cuando TM <= 16A)
    # -------------------------
    for _cl in _clima_items:
        if _cl["con_enchufe"]:
            # busca la sección del conductor de este circuito de climatización
            sec_cl = seccion_circuito(
                circuitos_df[circuitos_df["Circuito"].astype(str).str.contains(
                    _cl["circ"].split("(")[0].strip(), regex=False, na=False
                )]["Conductor"].iloc[0]
                if len(circuitos_df[circuitos_df["Circuito"].astype(str).str.contains(
                    _cl["circ"].split("(")[0].strip(), regex=False, na=False
                )]) > 0 else ""
            ) or 2.5
            agregar_ferrules(sec_cl, 3, f"{_cl['circ']} - enchufe climatización F/N/T")

    # NOTA: los circuitos especiales sin enchufe (horno, lavadora,
    # climatización sin enchufe van con conector cónico o estaño + cintas)
    # se calculan en la parte que recorre los cónicos (línea ~3398), para que
    # se sumen a los totales generales de cónico/estaño/cinta del proyecto,
    # en vez de generar un rollo aparte por cada circuito especial.

    # -------------------------
    # D2) AGUA CALIENTE — ferrules tablero externo + equipo con enchufe
    # Los conductores van directo a bornes del TM/bornera y, si tiene enchufe,
    # a los bornes propios del enchufe (TM≤16A, igual criterio que climatización):
    #   - 3 ferrules entrada equipo (F+N+PE) — SOLO si tiene enchufe (TM≤16A)
    #     Si no tiene enchufe (TM>16A), la conexión es cola de rata: cónico/
    #     estaño, ya contado en la parte que recorre los cónicos (línea ~3395)
    #   - Si además lleva tablero externo:
    #       2 ferrules entrada TM (F+N)
    #       2 ferrules salida TM (F+N)
    #       2 ferrules bornera PE (entrada+salida tierra)
    # -------------------------
    if {"Circuito", "Conductor"}.issubset(set(circuitos_df.columns)):
        for _, r_ac in circuitos_df.iterrows():
            circ_ac = str(r_ac.get("Circuito", "")).strip()
            _es_agua_ferrul = any(k in circ_ac.lower() for k in
                                  ("ducha", "termo", "calefon", "calefón",
                                   "calentador", "agua caliente"))
            if not _es_agua_ferrul:
                continue

            sec_ac_f = seccion_circuito(r_ac.get("Conductor", "")) or 4.0
            in_tm_ac_row2 = parse_in_tm(str(r_ac.get("Disyuntor termomagnético", "")))

            # Ferrules en los bornes del enchufe — solo si tiene enchufe (TM≤16A).
            # Si el TM no se pudo leer, no se genera nada acá (ya se avisó una
            # vez para este circuito en el bloque de la línea ~3090).
            if in_tm_ac_row2 is not None and in_tm_ac_row2 <= 16:
                agregar_ferrules(sec_ac_f, 3, f"{circ_ac} - enchufe equipo F/N/PE")

            # Buscar si lleva tablero externo
            _datos_ac_f = {}
            for eq_ac_f in circuitos_agua_caliente:
                if eq_ac_f.get("nombre_circ", "").lower() in circ_ac.lower():
                    _datos_ac_f = eq_ac_f
                    break

            if _datos_ac_f.get("lleva_tablero_externo", False) and in_tm_ac_row2 is not None:
                sec_int_tab = seccion_A1(in_tm_ac_row2)
                # 6 ferrules en el tablero externo (el TM ya trae sus propios
                # bornes, no se usa bornera F+N):
                # 2 entrada TM (F+N), 2 salida TM (F+N), 2 bornera PE (tierra)
                agregar_ferrules(sec_int_tab, 2, f"{circ_ac} - tablero externo entrada TM (F+N)")
                agregar_ferrules(sec_int_tab, 2, f"{circ_ac} - tablero externo salida TM (F+N)")
                agregar_ferrules(sec_ac_f,    2, f"{circ_ac} - tablero externo bornera PE (tierra)")

    # -------------------------
    # D3) ESPECIALES GENÉRICOS (horno, lavadora, encimera, etc.) — ferrules
    # equipo con enchufe. Mismo criterio y mismo formato que climatización (C2)
    # y agua caliente (D2): si TM≤16A, el equipo tiene enchufe, entonces van 3 ferrules
    # F/N/PE en los bornes de ese enchufe.
    # -------------------------
    if {"Circuito", "Conductor"}.issubset(set(circuitos_df.columns)):
        for _, r_esp in circuitos_df.iterrows():
            circ_esp = str(r_esp.get("Circuito", "")).strip()
            _es_esp_ferrul = any(k in circ_esp.lower() for k in
                                  ("especial", "horno", "encimera", "lavadora",
                                   "lavaplatos", "secadora", "jacuzzi", "piscina"))
            _es_agua_o_clima = any(k in circ_esp.lower() for k in
                                    ("ducha", "termo", "calefon", "calefón", "calentador",
                                     "agua caliente", "climatiz", "aire", "split", "ac ", "a/c"))
            if not _es_esp_ferrul or _es_agua_o_clima:
                continue

            sec_esp = seccion_circuito(r_esp.get("Conductor", "")) or 2.5
            in_tm_esp = parse_in_tm(str(r_esp.get("Disyuntor termomagnético", "")))

            # Si el TM no se pudo leer, no se asume "con enchufe" — el aviso
            # "(definir)" ya se generó una vez en la parte que recorre cónico/
            # estaño para este mismo circuito, no hace falta duplicarlo ni adivinar acá.
            if in_tm_esp is not None and in_tm_esp <= 16:
                agregar_ferrules(sec_esp, 3, f"{circ_esp} - enchufe equipo F/N/PE")

    # -------------------------
    # E) INTERRUPTORES
    # -------------------------
    if ambientes_df is not None:
        secciones_ilum = []

        # busca la sección más grande usada en los circuitos de iluminación
        if {"Circuito", "Conductor"}.issubset(set(circuitos_df.columns)):
            for _, r in circuitos_df.iterrows():
                if "ilumin" in str(r.get("Circuito", "")).lower():
                    sec = seccion_circuito(r.get("Conductor", ""))

                    if sec is not None:
                        secciones_ilum.append(float(sec))

        sec_iluminacion = max(secciones_ilum) if secciones_ilum else None

        # recorre cada ambiente y calcula cuántos ferrules necesitan sus interruptores
        for _, ar in ambientes_df.iterrows():
            amb = str(ar.get("Ambiente", "")).strip()
            n_lum = int(pd.to_numeric(ar.get("Cantidad luminarias (u)", 0), errors="coerce") or 0)
            n_con = int(pd.to_numeric(ar.get("N_conmutadas_924 (u)", 0), errors="coerce") or 0)

            n_con = max(0, min(n_con, n_lum))
            n_rest = max(0, n_lum - n_con)

            c12, c15, c32 = descomponer_interruptores(n_rest)

            # Ferrules por interruptor: 9/12 lleva 2 · 9/15 lleva 5 · 9/32 lleva 8 · conmutado 9/24 lleva 6 (fijo por grupo, sin importar cuántas luminarias controle)
            grupo_924_ferrul = 1 if n_con > 0 else 0
            total_int = (c12 * 2) + (c15 * 5) + (c32 * 8) + (grupo_924_ferrul * 6)

            if total_int > 0:
                agregar_ferrules(sec_iluminacion, total_int, f"Interruptores {amb}")

    # -------------------------
    # F) AGREGAR FILAS
    # -------------------------
    # arma una fila del Excel por cada sección de ferrul acumulada, con el
    # detalle de a qué corresponde cada cantidad (para que se pueda revisar)
    for sec, detalles in sorted(ferrules.items()):
        cantidad_total = int(sum(detalles.values()))
        color = color_ferrul_por_seccion(sec)

        detalle_txt = "\n".join(
            [f"{nombre} ({cant}u)" for nombre, cant in detalles.items()]
        )

        add_row(
            desc=f"Terminal ferrul color {color} {texto_seccion(sec)}mm²",
            marcas_txt=marcas.get("Ferrule", ""),
            norma="RIC 4.1.1",
            circuito=detalle_txt,
            unidad="u",
            k=1,
            longitud_m=f"{cantidad_total} unid",
            cantidad=cantidad_total
        )

        # =========================
    # CABLEADO INTERIOR DEL TABLERO
    # =========================
    # Cálculo basado en dimensiones reales del tablero y posición en riel DIN
    # Dimensiones tablero (alto x ancho en cm):
    TAB_DIMS = {
        2:  (17, 11),  4:  (20, 12),  6:  (20, 16),  8:  (20, 20),
        12: (25, 28), 16: (26, 36), 18: (26, 39), 24: (36, 30),
        36: (48, 30), 42: (54, 36), 48: (65, 35), 54: (54, 44),
        56: (75, 36), 72: (75, 46),
    }
    tab_h_cm, tab_w_cm = TAB_DIMS.get(_puestos_tablero, (36, 30))

    # ============================================================
    # REGLA de estimación de largo de cableado interior del tablero:
    #
    #  - Por cada circuito: 1m Rojo + 1m Blanco + 1m Verde, fijo, en la
    #    sección de ESE circuito (sec_int) — sin importar diferencial,
    #    barra extra, ni posición en el riel.
    #  - Luz piloto + portafusible: 1 sola vez para todo el tablero,
    #    +1m Rojo +1m Blanco (sin verde, no necesita tierra).
    #  - Tramo grueso del tablero (entre la barra principal y el omnipolar / SPD /
    #    protector sobrevoltaje / diferenciales — todos con la misma
    #    sección sec_omni_A1, ya que el calibre del diferencial siempre
    #    sale del mismo interruptor de empalme):
    #      Verde: 1m fijo (salida del SPD a tierra)
    #      Rojo = Blanco = (ancho + alto + ancho) del tablero, en metros,
    #             × cantidad total de circuitos del proyecto
    #  - El alimentador (F/N/PE hasta la barra principal) NO se cuenta acá
    #    — ya está incluido en el metraje propio del alimentador.
    # ============================================================
    # el alto se cuenta 2 veces y el ancho 1 vez, en metros
    _factor_dim_m = (tab_h_cm + tab_w_cm + tab_h_cm) / 100.0
    HOLGURA_TABLERO = 1.10  # +10% de holgura sobre el total final

    add_section("Cableado interior tablero")

    cableado_tablero = {}

    def agregar_cable_tablero(seccion_mm2, colores, longitud_base, detalle):
        # va sumando metros de cable interior del tablero por sección y color
        sec = normalizar_seccion_ferrul(seccion_mm2)  # redondea a sección comercial
        if sec is None:
            return  # sección inválida, no se puede agregar
        if isinstance(colores, str):
            colores = [colores]  # si mandaron un solo color como texto, lo mete en una lista
        if sec not in cableado_tablero:
            cableado_tablero[sec] = {"base_por_color": {}, "detalles": []}  # primera vez que aparece esta sección
        if detalle not in cableado_tablero[sec]["detalles"]:
            cableado_tablero[sec]["detalles"].append(detalle)  # anota de dónde viene este cable
        for color in colores:
            color = str(color).strip().capitalize()  # ej: "rojo" queda como "Rojo"
            cableado_tablero[sec]["base_por_color"][color] = (
                cableado_tablero[sec]["base_por_color"].get(color, 0.0)
                + float(longitud_base)  # suma los metros de este color
            )

    # n_circ ya viene calculado más arriba (cuenta solo filas con "Circuito"
    # no vacío) — se reutiliza acá para no contar de nuevo con un método
    # menos preciso.

    # -------------------------
    # A) TABLERO FIJO — tramo grueso (todo comparte sec_omni_A1)
    # -------------------------
    agregar_cable_tablero(sec_omni_A1, ["Rojo", "Blanco"],
                           _factor_dim_m * n_circ,
                           "Barra principal a omnipolar, SPD, protector sobrevoltaje y diferenciales (tramo grueso)")
    agregar_cable_tablero(sec_omni_A1, "Verde", 1.0, "SPD a barra PE (fijo)")

    # Luz piloto + portafusible: 1 sola vez para todo el tablero
    if usa_luz_piloto:
        agregar_cable_tablero(sec_control, "Rojo",   1.0, "Luz piloto + portafusible")
        agregar_cable_tablero(sec_control, "Blanco", 1.0, "Luz piloto + portafusible")

    # -------------------------
    # B) POR CADA CIRCUITO — 1m Rojo + 1m Blanco + 1m Verde, fijo
    # Usa el conductor que ya quedó definido para ese circuito (sec_circ),
    # no uno recalculado de nuevo con otro método.
    # -------------------------
    if {"Circuito", "Conductor", "Disyuntor termomagnético"}.issubset(set(circuitos_df.columns)):
        for idx, r in circuitos_df.iterrows():
            circ = str(r.get("Circuito", "")).strip()
            cond = str(r.get("Conductor", "")).strip()
            tm   = str(r.get("Disyuntor termomagnético", "")).strip()
            if not circ or not cond or not tm:
                continue

            sec_circ = seccion_circuito(cond)

            agregar_cable_tablero(sec_circ, "Rojo",   1.0, f"{circ} - cableado interior")
            agregar_cable_tablero(sec_circ, "Blanco", 1.0, f"{circ} - cableado interior")
            agregar_cable_tablero(sec_circ, "Verde",  1.0, f"{circ} - cableado interior")

        # -------------------------
    # C) AGREGAR FILAS
    # -------------------------
    orden_colores = ["Rojo", "Blanco", "Verde"]

    # arma una fila del Excel por cada sección de cable acumulada
    for sec, data in sorted(cableado_tablero.items()):
        base_por_color = data["base_por_color"]

        metros_base_total = round(sum(base_por_color.values()), 2)
        metros_final_total = round(metros_base_total * HOLGURA_TABLERO, 2)

        resumen_colores = []

        for color in orden_colores:
            if color in base_por_color:
                base_color = round(base_por_color[color], 2)
                _txt = f"{base_color}"
                if "." in _txt:
                    _txt = _txt.rstrip("0").rstrip(".")
                _base_color_txt = _txt.replace(".", ",")
                resumen_colores.append(
                    f"{color}: {_base_color_txt} = {math.ceil(base_color)}m"
                )

        detalle_longitud = "\n".join(resumen_colores)

        detalles_txt = "\n".join(data["detalles"]).lower()

        componentes = []

        if "barra principal" in detalles_txt:
            componentes.append("Barras repartidoras")

        if "omnipolar" in detalles_txt:
            componentes.append("Interruptor general")

        if "spd" in detalles_txt:
            componentes.append("SPD")

        if "portafusible" in detalles_txt:
            componentes.append("Portafusible")

        if "luz piloto" in detalles_txt:
            componentes.append("Luz piloto")

        if "diferencial" in detalles_txt:
            componentes.append("Interruptores diferenciales")

        if ("tm" in detalles_txt) or ("circuito" in detalles_txt):
            componentes.append("Interruptores termomagnéticos y salidas de circuitos")

        if not componentes:
            componentes.append("Cableado interior de tablero")

        detalle_circuito = "Tablero: " + ", ".join(componentes)

        add_row(
            desc=f"Conductor flexible libre de halógenos {tipo_cable_default} {texto_seccion(sec)}mm²",
            marcas_txt=marcas.get("Conductores", ""),
            norma="RIC 4.1.1",
            circuito=detalle_circuito,
            unidad="m",
            k=HOLGURA_TABLERO,
            longitud_m=detalle_longitud,
            cantidad=metros_final_total
        )

    # ── CABLEADO INTERIOR TABLERO EXTERNO AGUA CALIENTE ─────────────────────
    # Conductor método A1 dentro del tablero externo (mismo tipo de cable que
    # el resto de la instalación: THWN-2 en zona húmeda, H07Z1-K en zona seca —
    # ver tipo_cable_default, definido según la zona ingresada por el usuario):
    #   Tramo de entrada del prensaestopa al TM bipolar F (Rojo): 0.25m
    #   Tramo de entrada del prensaestopa al TM bipolar N (Blanco): 0.25m
    #   Tramo del TM bipolar a la salida del prensaestopa F (Rojo): 0.25m
    #   Tramo del TM bipolar a la salida del prensaestopa N (Blanco): 0.25m
    #   Tramo de la bornera PE de entrada a la bornera PE de salida (Verde): 0.25m
    # Total por color: Rojo=0.50m, Blanco=0.50m, Verde=0.25m
    # Sección F+N: seccion_A1(In_tm); Sección tierra: sección real conductor circuito
    # Se genera como fila SEPARADA del tablero principal (no se mezcla)
    if {"Circuito", "Conductor", "Disyuntor termomagnético"}.issubset(set(circuitos_df.columns)):
        for _, r_cab_ac in circuitos_df.iterrows():
            circ_cab = str(r_cab_ac.get("Circuito", "")).strip()
            _es_agua_cab = any(k in circ_cab.lower() for k in
                               ("ducha", "termo", "calefon", "calefón",
                                "calentador", "agua caliente"))
            if not _es_agua_cab:
                continue
            _datos_cab = {}
            for eq_cab in circuitos_agua_caliente:
                if eq_cab.get("nombre_circ", "").lower() in circ_cab.lower():
                    _datos_cab = eq_cab
                    break
            if not _datos_cab.get("lleva_tablero_externo", False):
                continue

            in_tm_cab = parse_in_tm(str(r_cab_ac.get("Disyuntor termomagnético", "")))
            if in_tm_cab is None:
                continue  # ya se avisó una vez para este circuito en el bloque de la línea ~3090
            sec_int_cab  = seccion_A1(in_tm_cab)   # F+N: método A1
            sec_ext_cab  = extraer_seccion_mm2(str(r_cab_ac.get("Conductor", ""))) or 4.0  # tierra: sección real
            L_fn  = round(0.25 * 2 * 1.15, 2)  # 2 tramos × 0.25m × holgura 1.15
            L_t   = round(0.25 * 1 * 1.15, 2)  # 1 tramo × 0.25m × holgura 1.15
            sec_fn_txt = str(sec_int_cab).replace(".", ",")
            sec_t_txt  = str(sec_ext_cab).replace(".", ",")
            add_row(
                desc=f"Conductor flexible libre de halógenos {tipo_cable_default} {sec_fn_txt}mm² (tablero externo agua caliente F+N)",
                marcas_txt=marcas.get("Conductores", ""),
                norma="RIC 4.1.1",
                circuito=f"{circ_cab} - tablero externo",
                unidad="m",
                k=1,
                longitud_m=f"Rojo {L_fn}m / Blanco {L_fn}m",
                cantidad=math.ceil(L_fn * 2)
            )
            add_row(
                desc=f"Conductor flexible libre de halógenos {tipo_cable_default} {sec_t_txt}mm² (tablero externo agua caliente tierra)",
                marcas_txt=marcas.get("Conductores", ""),
                norma="RIC 4.1.1",
                circuito=f"{circ_cab} - tablero externo",
                unidad="m",
                k=1,
                longitud_m=f"Verde {L_t}m",
                cantidad=L_t
            )
    add_section("Tornillería")
    tipo_can2 = (tipo_canalizacion or "").strip().lower()
    es_emb2 = "embut" in tipo_can2
    es_sob2 = "sobre" in tipo_can2
    # tornillos totales ----------
    tornillos_total = 0
    if es_emb2:
        tornillos_total += 2 * int(_abrazaderas_total)        # 2 por abrazadera
        tornillos_total += 4 * int(_cajas_total)              # 4 por caja
        tornillos_total += 2 * int(_tapas_ciegas_total)       # 2 por tapa ciega
        tornillos_total += 2 * int(total_luminarias_sobrepuestas)  # 2 por luminaria, SOLO las de montaje sobrepuesto (las embutidas no llevan)
        tornillos_total += (14 if int(_riel_m) == 2 else 7)   # riel DIN
        # Tornillos de fijación del tablero según su cantidad de puestos
        if _puestos_tablero in [2, 4, 6, 8, 12, 16]:
            tornillos_total += 6
        elif _puestos_tablero in [18, 24, 36, 42]:
            tornillos_total += 8
        elif _puestos_tablero in [48, 54, 56, 72]:
            tornillos_total += 10
    elif es_sob2:
        tornillos_total += int(_long_sobrepuesta_total)       # 1 por metro de canaleta
        tornillos_total += 4 * int(_cajas_total)              # 4 por caja
        tornillos_total += 2 * int(_tapas_ciegas_total)       # 2 por tapa ciega
        tornillos_total += 2 * int(total_luminarias_sobrepuestas)  # 2 por luminaria, SOLO las de montaje sobrepuesto (las embutidas no llevan)
        tornillos_total += (14 if int(_riel_m) == 2 else 7)   # riel DIN
        # Tornillos de fijación del tablero según su cantidad de puestos
        if _puestos_tablero in [2, 4, 6, 8, 12, 16]:
            tornillos_total += 6
        elif _puestos_tablero in [18, 24, 36, 42]:
            tornillos_total += 8
        elif _puestos_tablero in [48, 54, 56, 72]:
            tornillos_total += 10
    # Según tipo de material de construcción
    col_mat = "Material tabique" if es_emb2 else "Material forrado interior"
    # El reparto de tornillos se pondera por la cantidad real de cajas de
    # cada ambiente (luminarias + enchufes), así un ambiente con más cajas
    # aporta más al reparto de tornillos que uno con pocas.
    mats_peso = {}  # {material: peso acumulado (n° de cajas reales)}
    if ambientes_df is not None and col_mat in ambientes_df.columns:
        for _, _ar_mat in ambientes_df.iterrows():
            _m = str(_ar_mat.get(col_mat, "")).strip().lower()
            if not _m:
                continue
            _n_lum_amb = int(pd.to_numeric(_ar_mat.get("Cantidad luminarias (u)", 0), errors="coerce") or 0)
            _n_ench_amb = int(pd.to_numeric(_ar_mat.get("Cantidad enchufes (u)", 0), errors="coerce") or 0)
            _peso_amb = _n_lum_amb + _n_ench_amb  # cajas reales de este ambiente
            if _peso_amb <= 0:
                _peso_amb = 1  # ambiente sin cajas registradas: igual cuenta como mínimo 1, para no perderlo del reparto
            mats_peso[_m] = mats_peso.get(_m, 0) + _peso_amb
    # relaciona cada material con su tipo de tornillo
    def tornillo_por_material(m: str) -> str:
        # tornillo adecuado según el material de la superficie (metalcón,
        # volcanita, madera, etc.) y si es embutida o no
        mm = (m or "").lower()  # material en minúsculas
        if es_emb2:  # canalización embutida
            if "metalcon" in mm:
                return 'Tornillo autoperforante punta broca 6x1 1/2"'  # para estructura metálica
            return 'Tornillo punta fina para madera cabeza trompeta 6x1"'  # para madera, por defecto
        else:  # canalización sobrepuesta
            if ("volcan" in mm) or ("vulcan" in mm) or ("fibro" in mm):
                return 'Tornillo volcanita punta fina 1 1/4"'  # para volcanita/fibrocemento
            if "madera" in mm:
                return 'Tornillo punta fina para madera cabeza lenteja 6x1/2"'
            return 'Tornillo (definir según forrado)'  # material no identificado

    # salida al listado (UNA fila por tipo de tornillo)
    if tornillos_total > 0:
        if len(mats_peso) == 0:
            # no hay datos de material de construcción: se deja una fila genérica
            add_row(
                desc="Tornillo definir según construcción",
                marcas_txt=marcas.get("Tornillos", ""),
                norma="-",
                circuito="General",
                unidad="u",
                k=1,
                longitud_m=f"{tornillos_total} unid",
                cantidad=tornillos_total
            )
        else:
            # reparte el total de tornillos proporcionalmente según cuántas
            # cajas reales tiene cada material (no según cantidad de piezas)
            total_peso = sum(mats_peso.values())
            # Estas 2 porciones nunca necesitan tarugo:
            #  - Tapas ciegas (2 c/u): se atornillan a la caja, no a la pared.
            #  - Cajas octogonales en sobrepuesta (4 c/u): van fijadas
            #    directo al tabique, no a la superficie exterior con tarugo.
            # Se separan del resto antes de repartir, y se vuelven a sumar
            # al final como tornillo simple.
            _tornillos_tapas = 2 * int(_tapas_ciegas_total)
            _tornillos_oct_sob = 4 * int(total_luminarias) if es_sob2 else 0
            _tornillos_sin_tarugo_total = _tornillos_tapas + _tornillos_oct_sob
            tornillos_total_resto = max(0, tornillos_total - _tornillos_sin_tarugo_total)
            # agrupar por tipo de tornillo
            tornillos_por_tipo = {}
            for mat, peso in mats_peso.items():
                desc_tipo = tornillo_por_material(mat)
                # frac_material: qué fracción del total de cajas del proyecto
                # corresponde a este material puntual (ej. si "volcanita" tiene
                # 30 de 50 cajas totales, frac_material = 0.6 = 60%)
                frac_material = (peso / total_peso) if total_peso > 0 else 1.0
                # torn_estim: cuántos tornillos de ESTE material le tocan del
                # total elegible para tarugo, según su fracción del proyecto
                torn_estim = int(math.ceil(tornillos_total_resto * frac_material))
                # sin_tarugo_estim: mismo reparto proporcional, pero para la
                # porción que nunca lleva tarugo (tapas + octogonales sobrepuesta)
                sin_tarugo_estim = int(math.ceil(_tornillos_sin_tarugo_total * frac_material))
                if desc_tipo not in tornillos_por_tipo:
                    tornillos_por_tipo[desc_tipo] = {"tornillos": 0, "tornillos_sin_tarugo": 0, "mats": set()}
                tornillos_por_tipo[desc_tipo]["tornillos"] += torn_estim
                tornillos_por_tipo[desc_tipo]["tornillos_sin_tarugo"] += sin_tarugo_estim
                tornillos_por_tipo[desc_tipo]["mats"].add(mat)
            # escribir UNA fila por tipo (excepto volcanita/fibrocemento en
            # sobrepuesta, que usa tarugo+tornillo en vez de tornillo directo;
            # ese caso se calcula acá mismo, no se repite en otra sección)
            for desc_tipo, info in tornillos_por_tipo.items():
                cantidad_tornillos = int(info["tornillos"])
                cantidad_sin_tarugo = int(info["tornillos_sin_tarugo"])
                if cantidad_tornillos <= 0 and cantidad_sin_tarugo <= 0:
                    continue
                mats_txt = " + ".join(sorted(info["mats"]))
                if es_sob2 and desc_tipo == 'Tornillo volcanita punta fina 1 1/4"':
                    if cantidad_tornillos > 0:
                        add_row(
                            desc="Tarugo paloma 6mm",
                            marcas_txt=marcas.get("Tarugo paloma", ""),
                            norma="-",
                            circuito=f"General ({col_mat}: {mats_txt})",
                            unidad="u",
                            k=1,
                            longitud_m=f"{cantidad_tornillos} unid",
                            cantidad=cantidad_tornillos
                        )
                    # tornillo para tarugo (parte elegible) + tornillo simple
                    # de la porción sin tarugo (tapas ciegas + octogonales
                    # sobrepuesta), mismo tipo, se suman en la misma fila
                    cantidad_tornillo_total = cantidad_tornillos + cantidad_sin_tarugo
                    if cantidad_tornillo_total > 0:
                        add_row(
                            desc='Tornillo volcanita punta fina 6x1 1/4"',
                            marcas_txt=marcas.get("Tornillo para tarugo paloma", ""),
                            norma="-",
                            circuito=f"General ({col_mat}: {mats_txt})",
                            unidad="u",
                            k=1,
                            longitud_m=f"{cantidad_tornillo_total} unid",
                            cantidad=cantidad_tornillo_total
                        )
                    continue
                cantidad_total_tipo = cantidad_tornillos + cantidad_sin_tarugo
                add_row(
                    desc=desc_tipo,
                    marcas_txt=marcas.get("Tornillos", ""),
                    norma="-",
                    circuito=f"General ({col_mat}: {mats_txt})",
                    unidad="u",
                    k=1,
                    longitud_m=f"{cantidad_total_tipo} unid",
                    cantidad=cantidad_total_tipo
                )

    # NOTA: los tarugos paloma 6mm para volcanita/fibrocemento (canalización
    # sobrepuesta) ya se generan arriba, dentro de la sección "Tornillería".


    # SOLO para Panel SIP + canalización embutida
    # Criterio
    #  - SOLO por cajas de derivación
    # Regla:
    #  - 1 tubo cada 10 cajas de derivación
    #  - mínimo 1 tubo si aplica a material SIP
    # El tablero se fija con tornillos, no con espuma.
    # =========================
    tipo_can_esp = (tipo_canalizacion or "").strip().lower()
    es_emb_esp = "embut" in tipo_can_esp
    if es_emb_esp:
        # detectar Panel SIP en "Material tabique"
        mats_tab = []
        if ambientes_df is not None and "Material tabique" in ambientes_df.columns:
            mats_tab = (
                ambientes_df["Material tabique"].fillna("").astype(str).str.strip().str.lower().tolist())
        hay_sip = any("sip" in m for m in mats_tab)
        if hay_sip:
            # (opcional) sección nueva para que quede ordenado en el Excel
            add_section("Sellos / Aislación (Panel SIP)")
            # 1) SOLO cajas de derivación
            n_cajas = int(_cajas_total) if _cajas_total is not None else 0
            # 3) cálculo de tubos — la espuma es solo para sellar/rellenar el
            # hueco de las cajas embutidas en el núcleo del panel SIP, el
            # tablero no la necesita.
            tubos = int(math.ceil(n_cajas / 10.0)) if n_cajas > 0 else 0
            tubos = max(1, tubos)  # mínimo 1 si aplica SIP
            # 4) agregar al listado
            add_row(
                desc="Espuma expansiva de poliuretano 750 ml",
                marcas_txt=marcas.get("Espuma expansiva PU", ""),
                norma="Sellado de cajas de derivación embutidas en Panel SIP",
                circuito="Alumbrado (iluminación y enchufes)",
                unidad="u",
                k=1,
                longitud_m=f"{tubos} tubo(s) 750 ml",
                cantidad=tubos
            )

# =========================
# CLIMATIZACIÓN (RIC N°07 SEC. 7)
# Materiales exclusivos por equipo:
#  - Enchufe 2P+T 16A (si el equipo llega con enchufe)
#  - Caja de derivación cerca del equipo
#  - Canalización exclusiva (conduit o canaleta)
#  - Abrazaderas / accesorios de canalización
#  - Salidas de caja conduit (si embutida)
#  - Conector cónico 2.5 mm² (unión en caja)
#  - Boquilla bordes redondeados
#  - Prensaestopa (si conexión directa sin enchufe)
# Materiales que se actualizan automáticamente
# por ser circuitos más en circuitos_df:
#  - Conductor (3×L ya generado en sección Conductores)
#  - TM y diferencial (ya generados en sección Protecciones)
#  - Ferrules (ya generados en sección Ferrules)
#  - Tablero, riel, barra verde, barra repartidora
#    (n_circ ya incluye circuitos de climatización)
# =========================
# EMPALME
# =========================
    add_section("Empalme")
    # Empalme en fachada con o sin mástil
    if (str(tipo_instalacion_empalme).strip().lower() == "fachada"
        and "duct" in str(tipo_alimentador).strip().lower()
        and ("aer" in str(tipo_acometida).strip().lower() or "sub" in str(tipo_acometida).strip().lower())
        ):
        # Unidad de medida monofásica (medidor)
        add_row(
            desc="Unidad de medida monofásica 220V 50Hz 50A (medidor)",
            marcas_txt=marcas.get("Medidor empalme", ""),
            norma="SEC",
            circuito="Empalme",
            unidad="u",
            k=1,
            longitud_m="1 unid",
            cantidad=1
        )
        #Caja metálica del empalme
        add_row(
            desc="Caja de empalme metálica para medidor monofásico 405x200x137 mm IP54",
            marcas_txt=marcas.get("Caja empalme", ""),
            norma="SEC",
            circuito="Empalme",
            unidad="u",
            k=1,
            longitud_m="1 unid",
            cantidad=1
        )
        #Disyuntor termomagnético del empalme
        add_row(
            desc=f"Disyuntor termomagnético {interruptor_texto}",
            marcas_txt=marcas.get("Disyuntor empalme", ""),
            norma="SEC",
            circuito="Empalme",
            unidad="u",
            k=1,
            longitud_m="1 unid",
            cantidad=1
        )
        # Cable de la acometida
        add_row(
            desc=f"Cable {acometida_txt}",
            marcas_txt=marcas.get("Cable acometida", ""),
            norma="SEC",
            circuito="Acometida",
            unidad="m",
            k=1,
            longitud_m=f"{math.ceil(float(longitud_transformador_empalme))} m",
            cantidad=math.ceil(float(longitud_transformador_empalme))
        )
        # Tubo conduit galvanizado según sección de acometida
        sec_acom = float(acometida_txt.split("2x")[1].replace("mm²", "").strip())  # extrae el número de mm² del texto
        if sec_acom <= 6:
            desc_tubo_acom = "Tubo conduit galvanizado 25mm, 3mts"
        elif sec_acom <= 16:
            desc_tubo_acom = "Tubo conduit galvanizado 32mm, 3mts"
        elif sec_acom <= 35:
            desc_tubo_acom = "Tubo conduit galvanizado 40mm, 3mts"
        else:
            desc_tubo_acom = ""  # sección fuera de rango, no hay tubo definido
        if desc_tubo_acom:
            # SIN MÁSTIL
            if str(requiere_mastil).strip().lower() == "no":
                # Si la acometida es subterránea,
                # calcular cantidad según longitud hasta el medidor
                if "sub" in str(tipo_acometida).strip().lower():
                    largo_tubo_acom = float(longitud_subterraneo_medidor)
                    cantidad_tubo_acom = int(math.ceil(largo_tubo_acom / 3))
                else:
                    largo_tubo_acom = float(dist_vertical_acometida)
                    cantidad_tubo_acom = int(math.ceil(largo_tubo_acom / 3))
            # CON MÁSTIL
            elif str(requiere_mastil).strip().lower() == "si":
                largo_tubo_acom = float(longitud_mastil)
                cantidad_tubo_acom = int(math.ceil(largo_tubo_acom / 3))
            # Puestas a tierra
            largo_total_pt = float(dist_empalme_pt1) + float(dist_tda_pt2)
            cantidad_tubo_pt = int(math.ceil(largo_total_pt / 3))
            # Totales
            largo_total_tubos = largo_tubo_acom + largo_total_pt  # metros totales entre acometida y puesta a tierra
            cantidad_total_tubos = cantidad_tubo_acom + cantidad_tubo_pt
            add_row(
                desc=desc_tubo_acom,
                marcas_txt=marcas.get("Tubo conduit galvanizado acometida", ""),
                norma="SEC",
                circuito="Acometida / Puesta a tierra",
                unidad="u",
                k=1,
                longitud_m=f"{round(largo_total_tubos, 2)} m",
                cantidad=cantidad_total_tubos
            )

        # Cabeza de servicio según tubo galvanizado (solo acometida aérea)
        if "aer" in str(tipo_acometida).strip().lower():
            if "25mm" in desc_tubo_acom:
                desc_cabeza = 'Cabeza de servicio 3/4"'
            elif "32mm" in desc_tubo_acom:
                desc_cabeza = 'Cabeza de servicio 1"'
            elif "40mm" in desc_tubo_acom:
                desc_cabeza = 'Cabeza de servicio 1 1/4"'
            elif "50mm" in desc_tubo_acom:
                desc_cabeza = 'Cabeza de servicio 2"'
            else:
                desc_cabeza = ""
            if desc_cabeza:
                add_row(
                    desc=desc_cabeza,
                    marcas_txt=marcas.get("Cabeza de servicio", ""),
                    norma="SEC",
                    circuito="Acometida",
                    unidad="u",
                    k=1,
                    longitud_m="1 unid",
                    cantidad=1
                )
        # Cáncamo de acero galvanizado
        if "aer" in str(tipo_acometida).strip().lower():
            add_row(
                desc="Cáncamo abierto de acero galvanizado 7,8 x 110 mm",
                marcas_txt=marcas.get("Cancamo abierto", ""),
                norma="-",
                circuito="Acometida",
                unidad="u",
                k=1,
                longitud_m="1 unid",
                cantidad=1
            )
        #Granpa de retención — según sección de la acometida
        if "aer" in str(tipo_acometida).strip().lower():
            if sec_acom <= 10:
                desc_granpa = "Granpa de retención tipo cuña 2,5-25mm2 (6-10mm)"
            else:
                desc_granpa = "Granpa de retención tipo cuña 16-25mm2 (10-15mm)"
            add_row(
                desc=desc_granpa,
                marcas_txt=marcas.get("Granpa de retención", ""),
                norma="-",
                circuito="Acometida",
                unidad="u",
                k=1,
                longitud_m="1 unid",
                cantidad=1
            )
        # Conector HUB según tubo galvanizado
        if "25mm" in desc_tubo_acom:
            desc_hub = "Conector HUB 25mm de acero galvanizado"
        elif "32mm" in desc_tubo_acom:
            desc_hub = "Conector HUB 32mm de acero galvanizado"
        elif "40mm" in desc_tubo_acom:
            desc_hub = "Conector HUB 40mm de acero galvanizado"
        elif "50mm" in desc_tubo_acom:
            desc_hub = "Conector HUB 50mm de acero galvanizado"
        else:
            desc_hub = ""
        # cantidad cambia si acometida es subterránea
        if "sub" in str(tipo_acometida).strip().lower():
            cantidad_hub = 6
        else:
            cantidad_hub = 5
        if desc_hub:
            add_row(
                desc=desc_hub,
                marcas_txt=marcas.get("Conector HUB", ""),
                norma="SEC",
                circuito="Acometida, puesta a tierra 1, puesta a tierra 2",
                unidad="u",
                k=1,
                longitud_m=f"{cantidad_hub} unid",
                cantidad=cantidad_hub
            )

        # Terminal PVC conduit con 2 tuercas — igual criterio que en poste:
        # se separa acometida (solo si subterránea, Tabla N°4.29 con 2
        # conductores) y alimentador (siempre en ducto en fachada, Tabla
        # N°4.29 con 3 conductores). Si coinciden en diámetro, 1 fila;
        # si no, 2 filas separadas.
        _es_acom_sub_term_fach = "sub" in str(tipo_acometida).strip().lower()
        qty_term_acom_fach = 1 if _es_acom_sub_term_fach else 0
        qty_term_alim_fach = 4  # el alimentador en fachada siempre va en ducto

        diam_term_acom_fach = ducto_nominal_tablas(sec_acom, 2, "subterraneo") if qty_term_acom_fach else None

        import re
        def _extraer_diam_num_fach(txt):
            m = re.search(r'(\d+)\s*mm', str(txt))
            return int(m.group(1)) if m else None
        diam_term_alim_fach = _extraer_diam_num_fach(canalizacion_txt)

        if diam_term_acom_fach and diam_term_alim_fach and diam_term_acom_fach == diam_term_alim_fach:
            add_row(
                desc=f"Terminal PVC conduit con 2 tuercas {diam_term_acom_fach} mm",
                marcas_txt=marcas.get("Terminal PVC conduit con 2 tuercas", ""),
                norma="RIC 4.7.2",
                circuito="Acometida y alimentador en caja empalme, entrada y salida de caja metálica, entrada al TDA",
                unidad="u",
                k=1,
                longitud_m=f"{qty_term_acom_fach + qty_term_alim_fach} unid",
                cantidad=qty_term_acom_fach + qty_term_alim_fach
            )
        else:
            if diam_term_acom_fach and qty_term_acom_fach:
                add_row(
                    desc=f"Terminal PVC conduit con 2 tuercas {diam_term_acom_fach} mm",
                    marcas_txt=marcas.get("Terminal PVC conduit con 2 tuercas", ""),
                    norma="RIC 4.7.2",
                    circuito="Acometida (tramo PVC subterráneo)",
                    unidad="u",
                    k=1,
                    longitud_m=f"{qty_term_acom_fach} unid",
                    cantidad=qty_term_acom_fach
                )
            if diam_term_alim_fach and qty_term_alim_fach:
                add_row(
                    desc=f"Terminal PVC conduit con 2 tuercas {diam_term_alim_fach} mm",
                    marcas_txt=marcas.get("Terminal PVC conduit con 2 tuercas", ""),
                    norma="RIC 4.7.2",
                    circuito="Alimentador en caja empalme, entrada y salida de caja metálica, entrada al TDA",
                    unidad="u",
                    k=1,
                    longitud_m=f"{qty_term_alim_fach} unid",
                    cantidad=qty_term_alim_fach
                )

        # =========================
        # PUESTA A TIERRA
        # =========================
        #Conductor puesta a tierra
        sec_pt = "4" # mínima sección para que sea igual a la sección del alimentador
        try:
            sec_base = float(alim_txt.split("3x")[1].replace("mm²", "").strip())  # sección del alimentador
            # la sección de PT sigue a la del alimentador, redondeada a la comercial más cercana
            if sec_base <= 4:
                sec_pt = "4"
            elif sec_base <= 6:
                sec_pt = "6"
            elif sec_base <= 10:
                sec_pt = "10"
            elif sec_base <= 16:
                sec_pt = "16"
            elif sec_base <= 25:
                sec_pt = "25"
            else:
                sec_pt = "25"
        except:
            sec_pt = "4"  # no se pudo leer el alimentador, usa el mínimo
        # +2m: 1 chicote en la camarilla N°1 + 1 chicote en la caja metálica del empalme
        metros_pt = math.ceil(float(dist_empalme_pt1)) + 2
        add_row(
            desc=f"Conductor THWN-2 {sec_pt}mm2 blanco",
            marcas_txt=marcas.get("Conductor THWN-2 blanco", ""),
            norma="SEC",
            circuito="Puesta a tierra N°1 (empalme - camarilla N°1)",
            unidad="m",
            k=1,
            longitud_m=f"{metros_pt} m",
            cantidad=metros_pt
        )
        add_row(
            desc=f"Conductor THWN-2 {sec_pt}mm2 verde",
            marcas_txt=marcas.get("Conductor THWN-2 verde", ""),
            norma="SEC",
            circuito="Puesta a tierra N°1 (empalme - camarilla N°1)",
            unidad="m",
            k=1,
            longitud_m=f"{metros_pt} m",
            cantidad=metros_pt
        )
        # Barra copperweld PT1 (empalme - camarilla N°1)
        add_row(
            desc=desc_barra_pt,
            marcas_txt=marcas.get("Barra copperweld", ""),
            norma="RIC 6 (8.3.2, 8.5, 8.6, Tabla 6.1)",
            circuito="Puesta a tierra N°1 (empalme - camarilla N°1)",
            unidad="u",
            k=1,
            longitud_m=f"{n_barras_pt1} unid",
            cantidad=n_barras_pt1
        )
        # Camarilla PT1 según tubo conduit galvanizado
        diam_cam = ""
        if "25mm" in desc_tubo_acom:
            diam_cam = "25mm"
        elif "32mm" in desc_tubo_acom:
            diam_cam = "32mm"
        elif "40mm" in desc_tubo_acom:
            diam_cam = "40mm"
        elif "50mm" in desc_tubo_acom:
            diam_cam = "50mm"
        # RIC N°6 art. 8.3.2: si hay más de 1 barra, hay que unirlas con
        # un conductor desnudo de cobre, mínimo 16mm² de sección (fijo,
        # ese mínimo no depende de la sección del alimentador)
        if n_barras_pt1 > 1:
            add_row(
                desc=f"Conductor desnudo Cu 16mm² (unión entre barras PT1)",
                marcas_txt=marcas.get("Conductor desnudo Cu", ""),
                norma="RIC 6 (8.3.2, 8.7, 8.9)",
                circuito="Puesta a tierra N°1 (empalme - camarilla N°1)",
                unidad="m",
                k=1,
                longitud_m=f"{long_cond_desnudo_pt1:.1f} m",
                cantidad=long_cond_desnudo_pt1
            )
        # Barra copperweld PT2 (TDA - camarilla N°2)
        add_row(
            desc=desc_barra_pt,
            marcas_txt=marcas.get("Barra copperweld", ""),
            norma="RIC 6 (8.3.2, 8.5, 8.6, Tabla 6.1)",
            circuito="Puesta a tierra N°2 (TDA - camarilla N°2)",
            unidad="u",
            k=1,
            longitud_m=f"{n_barras_pt2} unid",
            cantidad=n_barras_pt2
        )
        # Camarilla PT1 + PT2: mismo diámetro siempre (diam_cam es único,
        # compartido entre ambas), así que se fusionan en 1 sola fila en
        # vez de 2 filas idénticas repetidas.
        if diam_cam:
            add_row(
                desc=f"Camarilla de registro con tapa PVC naranjo 160 x {diam_cam}",
                marcas_txt=marcas.get("Camarilla PVC naranjo", ""),
                norma="RIC 6 (5.15)",
                circuito="Puesta a tierra N°1 y N°2 (empalme - camarilla N°1 / TDA - camarilla N°2)",
                unidad="u",
                k=1,
                longitud_m=f"{n_barras_pt1 + n_barras_pt2} unid",
                cantidad=n_barras_pt1 + n_barras_pt2
            )
        # RIC N°6 art. 8.3.2: mismo criterio que PT1, mínimo 16mm² fijo
        # si hay más de 1 barra (esta es la puesta a tierra de PROTECCIÓN,
        # la del tablero, distinta de la de SERVICIO que es PT1)
        if n_barras_pt2 > 1:
            add_row(
                desc=f"Conductor desnudo Cu 16mm² (unión entre barras PT2)",
                marcas_txt=marcas.get("Conductor desnudo Cu", ""),
                norma="RIC 6 (8.3.2, 8.7, 8.9)",
                circuito="Puesta a tierra N°2 (TDA - camarilla N°2)",
                unidad="m",
                k=1,
                longitud_m=f"{long_cond_desnudo_pt2:.1f} m",
                cantidad=long_cond_desnudo_pt2
            )
        #Abrazaderas tipo caddy — 3 por cada tubo, con el diámetro de SU
        # propio tubo galvanizado de acometida
        def _desc_caddy_de_fach(diam_txt):
            if "25mm" in diam_txt:
                return "Abrazadera tipo caddy 25mm"
            elif "32mm" in diam_txt:
                return "Abrazadera tipo caddy 32mm"
            elif "40mm" in diam_txt:
                return "Abrazadera tipo caddy 40mm"
            elif "50mm" in diam_txt:
                return "Abrazadera tipo caddy 50mm"
            return "Abrazadera tipo caddy"
        cantidad_caddy = int(cantidad_total_tubos * 3)  # 3 abrazaderas por cada tubo galvanizado
        add_row(
            desc=_desc_caddy_de_fach(desc_tubo_acom),
            marcas_txt=marcas.get("Abrazadera tipo caddy", ""),
            norma="SEC",
            circuito="Puesta a tierra 1 y 2, acometida",
            unidad="u",
            k=1,
            longitud_m=f"{cantidad_caddy} unid",
            cantidad=cantidad_caddy
        )
        #Sellador de roscas: 1 frasco de 50ml alcanza para 6 conectores HUB
        cantidad_sellador = int(math.ceil(cantidad_hub / 6))
        add_row(
            desc="Sellador de roscas con teflón 50ml",
            marcas_txt=marcas.get("Sellador de roscas", ""),
            norma="SEC",
            circuito="Conectores Hub(cajas metálicas empalme y derivación)",
            unidad="u",
            k=1,
            longitud_m=f"{cantidad_sellador} unid",
            cantidad=cantidad_sellador
        )
        # Terminal ferrul acometida: mismo color/sección que el tablero (color_ferrul_por_seccion)
        try:
            sec_acom = float(acometida_txt.split("2x")[1].replace("mm²", "").strip())
        except:
            sec_acom = 4.0
        color_ferrul = color_ferrul_por_seccion(sec_acom)
        sec_ferrul = str(int(sec_acom)) if sec_acom == int(sec_acom) else texto_seccion(sec_acom)
        add_row(
            desc=f"Terminal ferrul color {color_ferrul} {sec_ferrul}mm (acometida)",
            marcas_txt=marcas.get("Terminal ferrul acometida", ""),
            norma="SEC",
            circuito="Acometida (Fase y Neutro)",
            unidad="u",
            k=1,
            longitud_m="2 unid",
            cantidad=2
        )
        # Terminal ferrul alimentador: mismo color/sección que el tablero
        try:
            sec_alim = float(alim_txt.split("3x")[1].replace("mm²", "").strip())
        except:
            sec_alim = 4.0
        color_ferrul_alim = color_ferrul_por_seccion(sec_alim)
        sec_ferrul_alim = str(int(sec_alim)) if sec_alim == int(sec_alim) else texto_seccion(sec_alim)
        add_row(
            desc=f"Terminal ferrul color {color_ferrul_alim} {sec_ferrul_alim}mm (alimentador)",
            marcas_txt=marcas.get("Terminal ferrul alimentador", ""),
            norma="SEC",
            circuito="Alimentador (Salida de medidor fase, entrada y salida termomagnético empalme)",
            unidad="u",
            k=1,
            longitud_m="3 unid",
            cantidad=3
        )
        #terminal ferrul doble alimentador
        add_row(
            desc=f"Terminal ferrul doble color {color_ferrul_alim} {sec_ferrul_alim}mm (alimentador)",
            marcas_txt=marcas.get("Terminal ferrul doble alimentador", ""),
            norma="SEC",
            circuito="Salida de medidor (Neutro del alimentador y neutro aterrizado)",
            unidad="u",
            k=1,
            longitud_m="1 unid",
            cantidad=1
        )
        # Cantidad de cajas de derivación metálica (se usa acá y en la fijación del caddy)
        if "sub" in str(tipo_acometida).strip().lower():
            cantidad_caja_derivacion = 2  # subterránea necesita una caja extra
        else:
            cantidad_caja_derivacion = 1
        #Terminal de compresión tipo ojo: 1 por caja de empalme + 1 por cada caja de derivación
        cantidad_terminal_ojo = 1 + cantidad_caja_derivacion
        add_row(
            desc=f"Terminal de compresión tipo ojo {sec_pt}mm",
            marcas_txt=marcas.get("Terminal compresion tipo ojo", ""),
            norma="SEC",
            circuito="Tramos metálicos (caja metálica empalme, caja de derivación, acometida, puesta a tierra 1 y 2)",
            unidad="u",
            k=1,
            longitud_m=f"{cantidad_terminal_ojo} unid",
            cantidad=cantidad_terminal_ojo
        )
        # Nota: el sistema de tornillos/tarugos de fijación (caja del empalme,
        # abrazaderas caddy, cajas de derivación y abrazaderas PVC) se calcula
        # más abajo, una vez que ya se conoce abrazaderas_fachada. Estas
        # variables se calculan acá porque el bloque de "caja de paso del
        # alimentador" (más abajo, pero antes del sistema de tornillos) también las necesita.
        hay_madera = "madera" in material_forrado_exterior
        hay_pvc = "pvc" in material_forrado_exterior  # ej. "siding pvc" — usa el mismo tornillo de madera, sin tarugo
        hay_siding_metalico = ("metál" in material_forrado_exterior) or ("metal" in material_forrado_exterior)  # con o sin tilde
        hay_fibro = "fibro" in material_forrado_exterior
        hay_forrado_valido = any(x in material_forrado_exterior for x in ("madera", "fibro", "siding"))
        #Alimentador (ya viene calculado en res_alim, de seleccionar_alimentador)
        add_row(
            desc=f'Alimentador RV-K Cu 3x{round(float(res_alim["S"]),2)}mm2',
            marcas_txt=marcas.get("Alimentador RV-K", ""),
            norma="SEC",
            circuito="Alimentador",
            unidad="m",
            k=1,
            longitud_m=f"{math.ceil(float(longitud_alimentador))} m",
            cantidad=math.ceil(float(longitud_alimentador))
        )
        # Conduit PVC para tramo subterráneo de acometida. Es un material
        # de PVC, no de acero — usa Tabla N°4.29 (2 conductores, F+N), NO
        # el diámetro de desc_tubo_acom (tubo de acero, tabla simple aparte).
        diam_pvc_sub = None  # inicializar siempre antes de usar
        if "sub" in str(tipo_acometida).strip().lower():
            diam_pvc_sub = ducto_nominal_tablas(sec_acom, 2, "subterraneo")
            if diam_pvc_sub:
                longitud_pvc_sub = (float(longitud_transformador_empalme) - float(longitud_subterraneo_medidor))
                cantidad_pvc_sub = int(math.ceil(longitud_pvc_sub / 3))
                add_row(desc=f"Conduit PVC {diam_pvc_sub}mm, 3mts",
                    marcas_txt=marcas.get("Conduit PVC", ""),
                    norma="SEC",
                    circuito="Acometida subterránea",
                    unidad="u",
                    k=1,
                    longitud_m=f"{round(longitud_pvc_sub, 2)} m",
                    cantidad=cantidad_pvc_sub)

        # Abrazadera adicional para PVC subterráneo
        if "sub" in str(tipo_acometida).strip().lower() and diam_pvc_sub:
            add_row(desc=f"Abrazadera conduit PVC {diam_pvc_sub}mm",
                marcas_txt=marcas.get("Abrazadera conduit PVC alimentador", ""),
                norma="SEC",
                circuito="Acometida subterránea",
                unidad="u",
                k=1,
                longitud_m="1 Unid",
                cantidad=1)

        # Conduit PVC según canalización alimentador (mínimo 32 mm)
        # Se define ANTES del bloque cámara para que diam_conduit esté disponible
        _m_diam_conduit = re.search(r'(\d+)\s*mm', str(canalizacion_txt))
        diam_conduit = max(32, int(_m_diam_conduit.group(1))) if _m_diam_conduit else 32
        metros_cond = math.ceil(float(longitud_alimentador))

        # =========================================================
        # CÁMARA TIPO C — CANALIZACIÓN SUBTERRÁNEA RESIDENCIAL
        # RIC N°04, sección 7.9, 7.9.5, 7.9.8.4.3 y Anexo 4.5
        # Regla RIC N°4 art. 7.9.7.8 / 7.9.7.9 / 7.9.7.10:
        #   - Si L es 20 m o menos: 0 cámaras (forma U, RIC N°4 art. 7.9.7.10)
        #   - Si L es mayor a 20 m: ceil(L / 90) cámaras
        #   Cada tramo (acometida / alimentador) se calcula de forma
        #   INDEPENDIENTE y se suma al total.
        # Dimensiones mínimas (Anexo 4.5 Lámina 2):
        #   tapa 440mm, marco 440x440mm, cámara 400x450mm, drenaje ø10mm
        # =========================================================
        if "sub" in str(tipo_acometida).strip().lower():
            # PVC enterrado = total acometida menos el tramo galvanizado (del suelo al medidor)
            long_pvc_acom = max(0.0,
                float(longitud_transformador_empalme) - float(longitud_subterraneo_medidor))

            # Cámaras según RIC N°4 art. 7.9.7.8 / 7.9.7.10
            # 20 m o menos: 0 cámaras (forma U)
            # más de 20 m: ceil(L / 90) cámaras
            if long_pvc_acom <= 20.0:
                camaras_tipo_c = 0
            else:
                camaras_tipo_c = int(math.ceil(long_pvc_acom / 90.0))

            # Diámetro boquilla = diámetro del PVC subterráneo de la acometida
            _diam_pvc_sub_safe_f = None
            try:
                _diam_pvc_sub_safe_f = diam_pvc_sub  # puede no existir si no es tramo subterráneo con tubo definido
            except NameError:
                _diam_pvc_sub_safe_f = None
            diam_cam_c = _diam_pvc_sub_safe_f if _diam_pvc_sub_safe_f else 25

            if camaras_tipo_c > 0:
                add_row(
                    desc="Cámara tipo C de hormigón prefabricado con tapa de acero diamantado 440x440mm",
                    marcas_txt=marcas.get("Camara tipo C", ""),
                    norma="RIC 4 (7.9, 7.9.5, 7.9.7.8, 7.9.7.10, 7.9.8.4.3, Anexo 4.5)",
                    circuito="Acometida subterránea",
                    unidad="u",
                    k=1,
                    longitud_m=f"{camaras_tipo_c} unid (L={round(long_pvc_acom,1)} m PVC sub.)",
                    cantidad=camaras_tipo_c
                )
                add_row(
                    desc="Marco metálico galvanizado para cámara tipo C 440x440mm",
                    marcas_txt=marcas.get("Marco metalico camara C", ""),
                    norma="RIC 4 (7.9.8, Anexo 4.5)",
                    circuito="Acometida subterránea",
                    unidad="u",
                    k=1,
                    longitud_m=f"{camaras_tipo_c} unid",
                    cantidad=camaras_tipo_c
                )
                boquillas_camara = camaras_tipo_c * 2
                add_row(
                    desc=(
                        f"Boquilla de PVC ø{diam_cam_c}mm con borde redondeado "
                        f"para entrada/salida conduit en cámara tipo C"
                    ),
                    marcas_txt=marcas.get("Boquilla camara tipo C", ""),
                    norma="RIC 4 (7.9.8.9, 5.14)",
                    circuito="Acometida subterránea",
                    unidad="u",
                    k=1,
                    longitud_m=f"{boquillas_camara} unid",
                    cantidad=boquillas_camara
                )
        cantidad_conduit_alim = int(math.ceil(metros_cond / 3))
        add_row(
            desc=f'Conduit de PVC de {diam_conduit}mm, 3mts',
            marcas_txt=marcas.get("Conduit PVC", ""),
            norma="RIC 4.7.2",
            circuito="Alimentador",
            unidad="u",
            k=1,
            longitud_m=f"{metros_cond} m",
            cantidad=cantidad_conduit_alim
        )
        # Abrazaderas según Tabla N°4.24
        sep_abraz_alim = 1.20 if diam_conduit <= 25 else 1.50
        cantidad_abrazaderas_alim = int(math.ceil(metros_cond / sep_abraz_alim))
        add_row(
            desc=f"Abrazadera conduit de PVC {diam_conduit}mm",
            marcas_txt=marcas.get("Abrazadera conduit PVC alimentador", ""),
            norma="RIC 4 (Tabla N°4.24)",
            circuito="Alimentador",
            unidad="u",
            k=1,
            longitud_m=f"{cantidad_abrazaderas_alim} unid",
            cantidad=cantidad_abrazaderas_alim
        )
        # Cajas de paso alimentador en ducto (RIC 7.16.1.13)
        # Caja de paso solo aplica si el alimentador va en ducto EMBUTIDO
        cajas_paso_alim = int(metros_cond // 20) if "duct" in str(tipo_alimentador).strip().lower() else 0
        if cajas_paso_alim > 0:
            # Caja de paso ESTANCA (el alimentador va por fachada/exterior, es
            # distinta a la caja de PVC de interior que se usa en los circuitos).
            # No lleva tapa aparte: la caja estanca viene con su tapa integrada
            # de fábrica (se atornilla directo a la caja, no es una pieza separada
            # como la tapa ciega de las cajas de interior).
            # El tamaño depende del diámetro real del conduit del alimentador:
            # con tubos de 32/40mm entra una caja compacta 150x110x70mm; con
            # 50mm hace falta una más grande 190x140x90mm.
            _dc_alim = 0
            try:
                _dc_alim = int(diam_conduit)
            except Exception:
                _dc_alim = 0
            if _dc_alim >= 50:
                _medida_caja_paso_alim = "190x140x90 mm"
            else:
                _medida_caja_paso_alim = "150x110x70 mm"
            add_row(
                desc=f"Caja de paso estanca IP65 de PVC/policarbonato para exterior {_medida_caja_paso_alim} (incluye tapa)",
                marcas_txt=marcas.get("Cajas de paso estancas", ""),
                norma="RIC 4 (7.16.1.13)",
                circuito="Alimentador - Caja de paso (tramo > 20m)",
                unidad="u",
                k=1,
                longitud_m=f"{cajas_paso_alim} unid",
                cantidad=cajas_paso_alim
            )
            # Salidas de caja PVC: 2 por caja de paso (entrada + salida conduit)
            salidas_paso_alim = 2 * cajas_paso_alim
            add_row(
                desc=f"Salida de caja conduit de PVC de {diam_conduit}mm",
                marcas_txt=marcas.get("Salida de caja conduit", ""),
                norma="RIC 4.7.2",
                circuito="Alimentador - Caja de paso (tramo > 20m)",
                unidad="u",
                k=1,
                longitud_m=f"{salidas_paso_alim} unid",
                cantidad=salidas_paso_alim
            )
            # Fijación de la caja de paso (fila propia, porque el acumulador
            # general de tornillos ya se calculó y escribió antes de llegar
            # acá — sumar solo a _cajas_total no alcanza a reflejarse en la
            # fila de "Tornillo" del Excel). Usa el mismo criterio por
            # material de forrado exterior que el resto de fijaciones del
            # alimentador (hay_madera/hay_pvc/hay_siding_metalico/hay_fibro,
            # calculadas más arriba) — con fibrocemento hace falta tarugo
            # antes del tornillo, no sirve un autoperforante directo.
            _tornillos_paso_alim = 4 * cajas_paso_alim
            if (hay_madera or hay_pvc) and hay_forrado_valido:
                add_row(
                    desc='Tirafondo hexagonal para madera de 1/4" x 1 1/2" (caja de paso alimentador)',
                    marcas_txt=marcas.get("Tirafondo hexagonal madera", ""),
                    norma="-",
                    circuito="Alimentador - Caja de paso (tramo > 20m)",
                    unidad="u",
                    k=1,
                    longitud_m=f"{_tornillos_paso_alim} unid",
                    cantidad=_tornillos_paso_alim
                )
            elif hay_siding_metalico and hay_forrado_valido:
                add_row(
                    desc='Tornillo autoperforante hexagonal 10 x 1-1/2" (caja de paso alimentador)',
                    marcas_txt=marcas.get("Tornillo autoperforante hexagonal", ""),
                    norma="-",
                    circuito="Alimentador - Caja de paso (tramo > 20m)",
                    unidad="u",
                    k=1,
                    longitud_m=f"{_tornillos_paso_alim} unid",
                    cantidad=_tornillos_paso_alim
                )
            elif hay_fibro:
                add_row(
                    desc="Tarugo paloma 8mm (caja de paso alimentador)",
                    marcas_txt=marcas.get("Tarugo paloma", ""),
                    norma="-",
                    circuito="Alimentador - Caja de paso (tramo > 20m)",
                    unidad="u",
                    k=1,
                    longitud_m=f"{_tornillos_paso_alim} unid",
                    cantidad=_tornillos_paso_alim
                )
                add_row(
                    desc="Tirafondo zincado punta fina 4,5 x 30 mm, rosca gruesa (caja de paso alimentador)",
                    marcas_txt=marcas.get("Tirafondo hexagonal madera", ""),
                    norma="-",
                    circuito="Alimentador - Caja de paso (tramo > 20m)",
                    unidad="u",
                    k=1,
                    longitud_m=f"{_tornillos_paso_alim} unid",
                    cantidad=_tornillos_paso_alim
                )
            else:
                add_row(
                    desc="Tornillo (definir según material de forrado exterior) — caja de paso alimentador",
                    marcas_txt=marcas.get("Tornillos", ""),
                    norma="-",
                    circuito="Alimentador - Caja de paso (tramo > 20m)",
                    unidad="u",
                    k=1,
                    longitud_m=f"{_tornillos_paso_alim} unid",
                    cantidad=_tornillos_paso_alim
                )
            # Sumar a acumuladores para otros usos (espuma SIP, etc.) aunque no
            # llegue a tiempo para la fila general de tornillos
            _cajas_total += cajas_paso_alim
        #caja de derivación metalica con tapa (cantidad_caja_derivacion ya calculada más arriba)
        add_row(
            desc="Caja de derivación metálica pregalvanizada 100x65x65mm con tapa",
            marcas_txt=marcas.get("Caja derivacion metalica", ""),
            norma="SEC",
            circuito="Empalme",
            unidad="u",
            k=1,
            longitud_m=f"{cantidad_caja_derivacion} unid",
            cantidad=cantidad_caja_derivacion
        )
        add_row(
            desc="Tornillo autoperforante punta broca 8 x 1/2\" cabeza lenteja + golilla (conexión tierra caja metálica)",
            marcas_txt=marcas.get("Tornillos", ""),
            norma="RIC 4 (5.13)",
            circuito="Empalme (caja de empalme + caja(s) de derivación)",
            unidad="u",
            k=1,
            longitud_m=f"{1 + cantidad_caja_derivacion} unid",
            cantidad=1 + cantidad_caja_derivacion
        )

        # Sistema completo de tornillos/tarugos de fijación (empalme en
        # fachada), según el material del forrado exterior de la casa.
        # Grupo A: caja metálica del empalme (fijo, siempre 6)
        # Grupo B: abrazaderas caddy (×1) + caja(s) de derivación (×4) +
        #          abrazaderas PVC del alimentador (×2)
        if "sub" in str(tipo_acometida).strip().lower():
            cajas_fachada = 2
            abrazaderas_fachada = cantidad_abrazaderas_alim + 1
        else:
            cajas_fachada = 1
            abrazaderas_fachada = cantidad_abrazaderas_alim

        cant_grupo_A = 6
        cant_grupo_B = (1 * cantidad_caddy) + (4 * cantidad_caja_derivacion) + (2 * abrazaderas_fachada)

        if (hay_madera or hay_pvc) and hay_forrado_valido:
            add_row(
                desc='Tirafondo hexagonal 1/4" x 1 1/2" + golilla 1/4"',
                marcas_txt=marcas.get("Tirafondo hexagonal madera", ""),
                norma="-",
                circuito="Caja metálica del empalme",
                unidad="u",
                k=1,
                longitud_m=f"{cant_grupo_A} unid",
                cantidad=cant_grupo_A
            )
            add_row(
                desc='Tornillo 8x1" cabeza lenteja punta fina',
                marcas_txt=marcas.get("Tornillo punta fina madera", ""),
                norma="-",
                circuito="Abrazaderas caddy, caja(s) de derivación metálica y abrazaderas PVC",
                unidad="u",
                k=1,
                longitud_m=f"{cant_grupo_B} unid",
                cantidad=cant_grupo_B
            )
        elif hay_fibro and hay_forrado_valido:
            # fibrocemento: necesita tarugo + tornillo juntos, en 1 solo grupo combinado
            cant_fibro_total = cant_grupo_A + cant_grupo_B
            add_row(
                desc='Tornillo 8x1 1/2" punta fina cabeza lenteja',
                marcas_txt=marcas.get("Tornillo punta fina madera", ""),
                norma="-",
                circuito="Caja metálica del empalme, abrazaderas caddy, caja(s) de derivación y abrazaderas PVC",
                unidad="u",
                k=1,
                longitud_m=f"{cant_fibro_total} unid",
                cantidad=cant_fibro_total
            )
            add_row(
                desc="Tarugo paloma N°8",
                marcas_txt=marcas.get("Tarugo paloma", ""),
                norma="-",
                circuito="Caja metálica del empalme, abrazaderas caddy, caja(s) de derivación y abrazaderas PVC",
                unidad="u",
                k=1,
                longitud_m=f"{cant_fibro_total} unid",
                cantidad=cant_fibro_total
            )
        elif hay_siding_metalico and hay_forrado_valido:
            add_row(
                desc='Tornillo autoperforante hexagonal 1/4" x 1 1/2" + golilla',
                marcas_txt=marcas.get("Tornillo autoperforante hexagonal", ""),
                norma="-",
                circuito="Caja metálica del empalme",
                unidad="u",
                k=1,
                longitud_m=f"{cant_grupo_A} unid",
                cantidad=cant_grupo_A
            )
            add_row(
                desc='Tornillo cabeza lenteja 8x1 1/4" punta broca',
                marcas_txt=marcas.get("Tornillo autoperforante broca", ""),
                norma="-",
                circuito="Abrazaderas caddy, caja(s) de derivación metálica y abrazaderas PVC",
                unidad="u",
                k=1,
                longitud_m=f"{cant_grupo_B} unid",
                cantidad=cant_grupo_B
            )
        elif not (hay_madera or hay_pvc or hay_siding_metalico or hay_fibro):
            cant_total_indef = cant_grupo_A + cant_grupo_B
            add_row(
                desc='Tornillo (definir según material de tabique)',
                marcas_txt=marcas.get("Tornillos", ""),
                norma="-",
                circuito="Caja metálica del empalme, abrazaderas caddy, caja(s) de derivación y abrazaderas PVC",
                unidad="u",
                k=1,
                longitud_m=f"{cant_total_indef} unid",
                cantidad=cant_total_indef
            )

        # Conductor puesta a tierra camarilla N°2 (TDA - puesta a tierra 2)
        # +2m: 1 chicote en la camarilla N°2 + 1 chicote en el TDA
        metros_pt2 = math.ceil(float(dist_tda_pt2)) + 2
        add_row(
            desc=f"Conductor THWN-2 {sec_pt}mm2 verde",
            marcas_txt=marcas.get("Conductor THWN-2 verde", ""),
            norma="SEC",
            circuito="Puesta a tierra N°2 (Tda - Camarilla N°2)",
            unidad="m",
            k=1,
            longitud_m=f"{metros_pt2} m",
            cantidad=metros_pt2
        )

        # Tarugos solo si el forrado es volcanita o fibrocemento (abrazaderas PVC del alimentador en fachada)
        # La caja de derivación metálica YA se cuenta en la fijación de abrazaderas caddy, no acá
        # hay_fibro ya se definió más arriba
        cantidad_tarugos = 0
        if hay_fibro:
            cantidad_tarugos = int(abrazaderas_fachada * 2)
            if "sub" in str(tipo_acometida).strip().lower() and diam_pvc_sub:
                cantidad_tarugos += 1
            add_row(
                desc="Tarugo paloma 6mm",
                marcas_txt=marcas.get("Tarugo paloma", ""),
                norma="-",
                circuito="Abrazaderas PVC en fachada",
                unidad="u",
                k=1,
                longitud_m=f"{cantidad_tarugos} unid",
                cantidad=cantidad_tarugos
            )
            add_row(
                desc='Tornillo punta plana 6x1 1/4" para tarugo',
                marcas_txt=marcas.get("Tornillo para tarugo paloma", ""),
                norma="-",
                circuito="Abrazaderas PVC en fachada",
                unidad="u",
                k=1,
                longitud_m=f"{cantidad_tarugos} unid",
                cantidad=cantidad_tarugos
            )

        # Portafusible aéreo loza según interruptor termomagnético del empalme
        tm_empalme_A = parse_in_tm(interruptor_texto)
        fusible_A = None
        if tm_empalme_A == 25:
            fusible_A = 30
        elif tm_empalme_A in [32, 40]:
            fusible_A = 60
        if fusible_A:
            add_row(
                desc=f"Portafusible de loza con fusibles cartucho {fusible_A}A",
                marcas_txt=marcas.get("Portafusible de loza", ""),
                norma="SEC",
                circuito="Empalme",
                unidad="u",
                k=1,
                longitud_m="1 unid",
                cantidad=1
            )

#============================================
#Empalme independiente poste madera/metálico
    elif (str(tipo_instalacion_empalme).strip().lower() == "independiente"
        and ("aer" in str(tipo_alimentador).strip().lower() or "sub" in str(tipo_alimentador).strip().lower())
        and ("aer" in str(tipo_acometida).strip().lower() or "sub" in str(tipo_acometida).strip().lower())):
        # Unidad de medida monofásico / medidor
        add_row(
            desc="Unidad de medida monofásica 220V 50Hz 50A (medidor)",
            marcas_txt=marcas.get("Medidor empalme", ""),
            norma="SEC",
            circuito="Empalme",
            unidad="u",
            k=1,
            longitud_m="1 unid",
            cantidad=1
        )
        # Caja metálica empalme
        add_row(
            desc="Caja de empalme metálica para medidor monofásico 405x200x137 mm IP54",
            marcas_txt=marcas.get("Caja empalme", ""),
            norma="SEC",
            circuito="Empalme",
            unidad="u",
            k=1,
            longitud_m="1 unid",
            cantidad=1
        )
        # Disyuntor termomagnético empalme
        add_row(
            desc=f"Disyuntor termomagnético {interruptor_texto}",
            marcas_txt=marcas.get("Disyuntor empalme", ""),
            norma="SEC",
            circuito="Empalme",
            unidad="u",
            k=1,
            longitud_m="1 unid",
            cantidad=1
        )
        # Conductor acometida
        add_row(
            desc=f"Cable {acometida_txt}",
            marcas_txt=marcas.get("Cable acometida", ""),
            norma="SEC",
            circuito="Acometida",
            unidad="m",
            k=1,
            longitud_m=f"{math.ceil(float(longitud_transformador_empalme))} m",
            cantidad=math.ceil(float(longitud_transformador_empalme))
        )
        # ============================================================
        # Tubo conduit galvanizado — dos diámetros posibles y distintos:
        #   - Tramo ACOMETIDA (+ puesta a tierra): según sección del cable
        #     de la acometida (sec_acom), sin piso mínimo (25/32/40mm).
        #   - Tramo ALIMENTADOR (aéreo O subterráneo, da igual): según
        #     sección del cable del alimentador (sec_alim), con PISO
        #     MÍNIMO de 32mm (nunca puede quedar en 25mm, aunque el cable
        #     sea delgado) — el alimentador SIEMPRE calcula su propio
        #     diámetro, no se fusiona en silencio con el de la acometida
        #     solo porque sea subterráneo.
        # Si ambos diámetros coinciden, se fusionan en 1 sola fila; si son
        # distintos, se generan 2 filas separadas.
        # ============================================================
        sec_acom = float(acometida_txt.split("2x")[1].replace("mm²", "").strip())
        _acom_es_aerea = "aer" in str(tipo_acometida).strip().lower()
        _alim_es_aereo = "aer" in str(tipo_alimentador).strip().lower()

        # Tubo de ACERO galvanizado: siempre tabla simple de umbrales, sea
        # aérea o subterránea — la Tabla N°4.29 no aplica acá, es solo para
        # el Terminal PVC / Conduit PVC (materiales de PVC, no de acero).
        if sec_acom <= 6:
            desc_tubo_acom = "Tubo conduit galvanizado 25mm, 3mts"
        elif sec_acom <= 16:
            desc_tubo_acom = "Tubo conduit galvanizado 32mm, 3mts"
        elif sec_acom <= 35:
            desc_tubo_acom = "Tubo conduit galvanizado 40mm, 3mts"
        else:
            desc_tubo_acom = ""

        # El alimentador SIEMPRE calcula su propio diámetro (con piso
        # mínimo 32mm), sea aéreo o subterráneo.
        try:
            sec_alim = float(alim_txt.split("3x")[1].replace("mm²", "").strip())
        except Exception:
            sec_alim = 0.0
        # Piso mínimo 32mm para el alimentador (no hay opción de 25mm)
        if sec_alim <= 16:
            desc_tubo_alim = "Tubo conduit galvanizado 32mm, 3mts"
        elif sec_alim <= 35:
            desc_tubo_alim = "Tubo conduit galvanizado 40mm, 3mts"
        else:
            desc_tubo_alim = ""

        cantidad_total_tubos = 0  # se usa después para abrazaderas caddy
        largo_total_acom_pt = 0.0
        cantidad_total_acom_pt = 0
        largo_tramo_alim = 0.0
        cantidad_tramo_alim = 0
        cantidad_caddy_pt2 = 0

        if desc_tubo_acom:
            # Largo del tramo de acometida (+PT). El alimentador (aéreo o
            # subterráneo) no se suma acá: tiene su propio diámetro y se
            # calcula aparte más abajo.
            if _acom_es_aerea:
                largo_tramo_acom = float(altura_acometida_aerea)
            else:
                largo_tramo_acom = float(longitud_subterraneo_medidor2)
            cantidad_tramo_acom = int(math.ceil(largo_tramo_acom / 3))
            # Puesta a tierra: mismo diámetro que la acometida
            largo_total_pt = float(dist_empalme_pt1) + float(dist_tda_pt2)
            cantidad_tubo_pt = int(math.ceil(largo_total_pt / 3))
            largo_total_acom_pt = largo_tramo_acom + largo_total_pt
            cantidad_total_acom_pt = cantidad_tramo_acom + cantidad_tubo_pt
            # PT2 (del TDA a la camarilla N°2) físicamente está en la fachada de la
            # casa, no en el poste — se guarda aparte para reasignar sus
            # tornillos de fijación de la abrazadera caddy hacia la fachada
            # más abajo (el material "Abrazadera tipo caddy" en sí no cambia,
            # solo QUIÉN paga el tornillo de fijación).
            cantidad_caddy_pt2 = int(math.ceil(float(dist_tda_pt2) / 3)) * 3

        if desc_tubo_alim:
            # Largo del tramo del alimentador: aéreo usa su propio poste,
            # subterráneo usa su propio "cuello".
            if _alim_es_aereo:
                largo_tramo_alim = float(longitud_poste_alimentador_aereo)
            else:
                largo_tramo_alim = float(longitud_subterraneo_medidor2)
            cantidad_tramo_alim = int(math.ceil(largo_tramo_alim / 3))

        # ¿Se fusionan en 1 fila (mismo diámetro) o van 2 filas separadas?
        if desc_tubo_acom and desc_tubo_alim and desc_tubo_acom == desc_tubo_alim:
            largo_total_tubos = largo_total_acom_pt + largo_tramo_alim
            cantidad_total_tubos = cantidad_total_acom_pt + cantidad_tramo_alim
            add_row(
                desc=desc_tubo_acom,
                marcas_txt=marcas.get("Tubo conduit galvanizado acometida", ""),
                norma="SEC",
                circuito="Acometida / Alimentador / Puesta a tierra",
                unidad="u",
                k=1,
                longitud_m=f"{round(largo_total_tubos, 2)} m",
                cantidad=cantidad_total_tubos
            )
        else:
            if desc_tubo_acom and cantidad_total_acom_pt:
                cantidad_total_tubos += cantidad_total_acom_pt
                add_row(
                    desc=desc_tubo_acom,
                    marcas_txt=marcas.get("Tubo conduit galvanizado acometida", ""),
                    norma="SEC",
                    circuito="Acometida / Puesta a tierra",
                    unidad="u",
                    k=1,
                    longitud_m=f"{round(largo_total_acom_pt, 2)} m",
                    cantidad=cantidad_total_acom_pt
                )
            if desc_tubo_alim and cantidad_tramo_alim:
                cantidad_total_tubos += cantidad_tramo_alim
                add_row(
                    desc=desc_tubo_alim,
                    marcas_txt=marcas.get("Tubo conduit galvanizado acometida", ""),
                    norma="SEC",
                    circuito="Alimentador",
                    unidad="u",
                    k=1,
                    longitud_m=f"{round(largo_tramo_alim, 2)} m",
                    cantidad=cantidad_tramo_alim
                )

        # Cabeza de servicio — 1 por cada tramo aéreo presente, con el
        # TAMAÑO DE SU PROPIO diámetro (no siempre el mismo tamaño para
        # los dos tramos, ya que acometida y alimentador pueden diferir).
        def _desc_cabeza_de(diam_txt):
            if "25mm" in diam_txt:
                return 'Cabeza de servicio 3/4"'
            elif "32mm" in diam_txt:
                return 'Cabeza de servicio 1"'
            elif "40mm" in diam_txt:
                return 'Cabeza de servicio 1 1/4"'
            elif "50mm" in diam_txt:
                return 'Cabeza de servicio 2"'
            return ""

        _cabezas = {}  # relaciona cada descripción con su cantidad (se fusionan si coinciden)
        if _acom_es_aerea and desc_tubo_acom:
            _d = _desc_cabeza_de(desc_tubo_acom)
            if _d:
                _cabezas[_d] = _cabezas.get(_d, 0) + 1
        if _alim_es_aereo and desc_tubo_alim:
            _d = _desc_cabeza_de(desc_tubo_alim)
            if _d:
                _cabezas[_d] = _cabezas.get(_d, 0) + 1
        for _desc_cab, _cant_cab in _cabezas.items():
            add_row(
                desc=_desc_cab,
                marcas_txt=marcas.get("Cabeza de servicio", ""),
                norma="SEC",
                circuito="Acometida y/o alimentador",
                unidad="u",
                k=1,
                longitud_m=f"{_cant_cab} unid",
                cantidad=_cant_cab
            )
        # Cáncamo de acero galvanizado
        # 1 por cada tramo aéreo presente (acometida y/o alimentador).
        # No tiene variantes de tamaño, por lo que no se ve afectado por
        # el split de diámetros de arriba.
        if _acom_es_aerea or _alim_es_aereo:
            cantidad_cancamo = (1 if _acom_es_aerea else 0) + (1 if _alim_es_aereo else 0)
            add_row(
                desc="Cáncamo abierto de acero galvanizado 7,8 x 110 mm",
                marcas_txt=marcas.get("Cancamo abierto", ""),
                norma="-",
                circuito="Acometida y/o alimentador",
                unidad="u",
                k=1,
                longitud_m=f"{cantidad_cancamo} unid",
                cantidad=cantidad_cancamo
                )
        #Granpa de retención — según sección de la acometida
        if "aer" in str(tipo_acometida).strip().lower():
            if sec_acom <= 10:
                desc_granpa = "Granpa de retención tipo cuña 2,5-25mm2 (6-10mm)"
            else:
                desc_granpa = "Granpa de retención tipo cuña 16-25mm2 (10-15mm)"
            add_row(
                desc=desc_granpa,
                marcas_txt=marcas.get("Granpa de retención", ""),
                norma="-",
                circuito="Acometida y/o alimentador",
                unidad="u",
                k=1,
                longitud_m="1 unid",
                cantidad=1
            )
        # Mordaza alimentador
        if "aer" in str(tipo_alimentador).strip().lower():
            add_row(
                desc="Mordaza para alimentador aéreo",
                marcas_txt=marcas.get("Mordaza acometida", ""),
                norma="SEC",
                circuito="Alimentador",
                unidad="u",
                k=1,
                longitud_m="1 unid",
                cantidad=1
            )
        # Conector HUB — total según combinación acometida/alimentador (tabla
        # fija), y de ese total, cuántos corresponden específicamente al tramo
        # del ALIMENTADOR (el resto es acometida + PT1 + PT2). Si el diámetro
        # del alimentador difiere del de la acometida (aéreo o subterráneo,
        # da igual), van en 2 filas separadas (una por diámetro); si no, se
        # fusiona en 1 sola fila.
        _acom_aer_hub = "aer" in str(tipo_acometida).strip().lower()
        _alim_aer_hub = "aer" in str(tipo_alimentador).strip().lower()
        if _acom_aer_hub and _alim_aer_hub:
            cantidad_hub_total, cantidad_hub_alim = 6, 1
        elif _acom_aer_hub and not _alim_aer_hub:
            cantidad_hub_total, cantidad_hub_alim = 7, 2
        elif (not _acom_aer_hub) and _alim_aer_hub:
            cantidad_hub_total, cantidad_hub_alim = 7, 1
        else:
            cantidad_hub_total, cantidad_hub_alim = 8, 2
        cantidad_hub_acom = cantidad_hub_total - cantidad_hub_alim

        def _desc_hub_de(diam_txt):
            if "25mm" in diam_txt:
                return "Conector HUB 25mm de acero galvanizado"
            elif "32mm" in diam_txt:
                return "Conector HUB 32mm de acero galvanizado"
            elif "40mm" in diam_txt:
                return "Conector HUB 40mm de acero galvanizado"
            elif "50mm" in diam_txt:
                return "Conector HUB 50mm de acero galvanizado"
            return ""

        if desc_tubo_alim and desc_tubo_alim != desc_tubo_acom:
            # Diámetros distintos (alimentador aéreo o subterráneo, da
            # igual): 2 filas separadas
            _hub_acom_desc = _desc_hub_de(desc_tubo_acom)
            if _hub_acom_desc and cantidad_hub_acom:
                add_row(
                    desc=_hub_acom_desc,
                    marcas_txt=marcas.get("Conector HUB", ""),
                    norma="SEC",
                    circuito="Acometida, puesta a tierra 1, puesta a tierra 2",
                    unidad="u",
                    k=1,
                    longitud_m=f"{cantidad_hub_acom} unid",
                    cantidad=cantidad_hub_acom
                )
            _hub_alim_desc = _desc_hub_de(desc_tubo_alim)
            if _hub_alim_desc and cantidad_hub_alim:
                add_row(
                    desc=_hub_alim_desc,
                    marcas_txt=marcas.get("Conector HUB", ""),
                    norma="SEC",
                    circuito="Alimentador",
                    unidad="u",
                    k=1,
                    longitud_m=f"{cantidad_hub_alim} unid",
                    cantidad=cantidad_hub_alim
                )
        else:
            # Mismo diámetro (acometida y alimentador coinciden): se
            # mantiene fusionado en 1 sola fila
            _hub_desc = _desc_hub_de(desc_tubo_acom)
            if _hub_desc and cantidad_hub_total:
                add_row(
                    desc=_hub_desc,
                    marcas_txt=marcas.get("Conector HUB", ""),
                    norma="SEC",
                    circuito="Acometida, puesta a tierra 1, puesta a tierra 2",
                    unidad="u",
                    k=1,
                    longitud_m=f"{cantidad_hub_total} unid",
                    cantidad=cantidad_hub_total
                )

        # Terminal PVC conduit con 2 tuercas — se separa en acometida y
        # alimentador, cada uno con su propia cantidad y diámetro (son 2
        # tramos de PVC enterrado distintos, no comparten medida):
        #  - Acometida subterránea: 1 terminal, con la Tabla N°4.29 (2
        #    conductores, F+N sin tierra) — este es el diámetro del ducto
        #    PVC real, DISTINTO al tubo de acero galvanizado (que usa la
        #    tabla simple) — si es aérea, no aplica (0).
        #  - Alimentador: 4 terminales si es subterráneo (diámetro real del
        #    ducto, Tabla N°4.29 con 3 conductores) o 3 si es aéreo (mismo
        #    piso mínimo de 32mm que ya usa el tubo galvanizado del
        #    alimentador, sin ser un valor fijo).
        # Si ambos diámetros coinciden, se fusionan en 1 fila; si no, 2.
        import re
        def _extraer_diam_num(txt):
            m = re.search(r'(\d+)\s*mm', str(txt))
            return int(m.group(1)) if m else None

        _es_acom_sub_term = "sub" in str(tipo_acometida).strip().lower()
        _es_alim_aer_term = "aer" in str(tipo_alimentador).strip().lower()

        qty_term_acom = 1 if _es_acom_sub_term else 0
        qty_term_alim = 3 if _es_alim_aer_term else 4

        diam_term_acom = ducto_nominal_tablas(sec_acom, 2, "subterraneo") if qty_term_acom else None
        if qty_term_alim:
            diam_term_alim = _extraer_diam_num(desc_tubo_alim) if _es_alim_aer_term else _extraer_diam_num(canalizacion_txt)
        else:
            diam_term_alim = None

        if diam_term_acom and diam_term_alim and diam_term_acom == diam_term_alim:
            add_row(
                desc=f"Terminal PVC conduit con 2 tuercas {diam_term_acom} mm",
                marcas_txt=marcas.get("Terminal PVC conduit con 2 tuercas", ""),
                norma="RIC 4.7.2",
                circuito="Acometida y alimentador en caja empalme, entrada y salida de caja metálica, entrada al TDA",
                unidad="u",
                k=1,
                longitud_m=f"{qty_term_acom + qty_term_alim} unid",
                cantidad=qty_term_acom + qty_term_alim
            )
        else:
            if diam_term_acom and qty_term_acom:
                add_row(
                    desc=f"Terminal PVC conduit con 2 tuercas {diam_term_acom} mm",
                    marcas_txt=marcas.get("Terminal PVC conduit con 2 tuercas", ""),
                    norma="RIC 4.7.2",
                    circuito="Acometida (tramo PVC subterráneo)",
                    unidad="u",
                    k=1,
                    longitud_m=f"{qty_term_acom} unid",
                    cantidad=qty_term_acom
                )
            if diam_term_alim and qty_term_alim:
                add_row(
                    desc=f"Terminal PVC conduit con 2 tuercas {diam_term_alim} mm",
                    marcas_txt=marcas.get("Terminal PVC conduit con 2 tuercas", ""),
                    norma="RIC 4.7.2",
                    circuito="Alimentador en caja empalme, entrada y salida de caja metálica, entrada al TDA",
                    unidad="u",
                    k=1,
                    longitud_m=f"{qty_term_alim} unid",
                    cantidad=qty_term_alim
                )

        # =========================
        # PUESTA A TIERRA
        # =========================
        #Conductor thwn-2 puesta a tierra empalme
        sec_pt = "4" # mínima sección para que sea igual a la sección del alimentador
        try:
            sec_base = float(alim_txt.split("3x")[1].replace("mm²", "").strip())
            if sec_base <= 4:
                sec_pt = "4"
            elif sec_base <= 6:
                sec_pt = "6"
            elif sec_base <= 10:
                sec_pt = "10"
            elif sec_base <= 16:
                sec_pt = "16"
            elif sec_base <= 25:
                sec_pt = "25"
            else:
                sec_pt = "25"
        except:
            sec_pt = "4"
        # +2m: 1 chicote en la camarilla N°1 + 1 chicote en la caja metálica del empalme
        metros_pt = math.ceil(float(dist_empalme_pt1)) + 2
        add_row(
            desc=f"Conductor THWN-2 {sec_pt}mm2 blanco",
            marcas_txt=marcas.get("Conductor THWN-2 blanco", ""),
            norma="SEC",
            circuito="Puesta a tierra N°1 (empalme - camarilla N°1)",
            unidad="m",
            k=1,
            longitud_m=f"{metros_pt} m",
            cantidad=metros_pt
        )
        #Conductor thwn-2 puesta a tierra empalme
        add_row(
            desc=f"Conductor THWN-2 {sec_pt}mm2 verde",
            marcas_txt=marcas.get("Conductor THWN-2 verde", ""),
            norma="SEC",
            circuito="Puesta a tierra N°1 (empalme - camarilla N°1)",
            unidad="m",
            k=1,
            longitud_m=f"{metros_pt} m",
            cantidad=metros_pt
        )

        # Barra copperweld PT1 (empalme - camarilla N°1)
        add_row(
            desc=desc_barra_pt,
            marcas_txt=marcas.get("Barra copperweld", ""),
            norma="RIC 6 (8.3.2, 8.5, 8.6, Tabla 6.1)",
            circuito="Puesta a tierra N°1 (empalme - camarilla N°1)",
            unidad="u",
            k=1,
            longitud_m=f"{n_barras_pt1} unid",
            cantidad=n_barras_pt1
        )
        # Camarilla PT1 según tubo conduit galvanizado
        diam_cam = ""
        if "25mm" in desc_tubo_acom:
            diam_cam = "25mm"
        elif "32mm" in desc_tubo_acom:
            diam_cam = "32mm"
        elif "40mm" in desc_tubo_acom:
            diam_cam = "40mm"
        elif "50mm" in desc_tubo_acom:
            diam_cam = "50mm"
        # RIC N°6 art. 8.3.2: si hay más de 1 barra, hay que unirlas con
        # un conductor desnudo de cobre, mínimo 16mm² de sección (fijo,
        # ese mínimo no depende de la sección del alimentador)
        if n_barras_pt1 > 1:
            add_row(
                desc=f"Conductor desnudo Cu 16mm² (unión entre barras PT1)",
                marcas_txt=marcas.get("Conductor desnudo Cu", ""),
                norma="RIC 6 (8.3.2, 8.7, 8.9)",
                circuito="Puesta a tierra N°1 (empalme - camarilla N°1)",
                unidad="m",
                k=1,
                longitud_m=f"{long_cond_desnudo_pt1:.1f} m",
                cantidad=long_cond_desnudo_pt1
            )
        # Barra copperweld PT2 (TDA - camarilla N°2)
        add_row(
            desc=desc_barra_pt,
            marcas_txt=marcas.get("Barra copperweld", ""),
            norma="RIC 6 (8.3.2, 8.5, 8.6, Tabla 6.1)",
            circuito="Puesta a tierra N°2 (TDA - camarilla N°2)",
            unidad="u",
            k=1,
            longitud_m=f"{n_barras_pt2} unid",
            cantidad=n_barras_pt2
        )
        # Camarilla PT1 + PT2: mismo diámetro siempre (diam_cam es único,
        # compartido entre ambas), así que se fusionan en 1 sola fila en
        # vez de 2 filas idénticas repetidas.
        if diam_cam:
            add_row(
                desc=f"Camarilla de registro con tapa PVC naranjo 160 x {diam_cam}",
                marcas_txt=marcas.get("Camarilla PVC naranjo", ""),
                norma="RIC 6 (5.15)",
                circuito="Puesta a tierra N°1 y N°2 (empalme - camarilla N°1 / TDA - camarilla N°2)",
                unidad="u",
                k=1,
                longitud_m=f"{n_barras_pt1 + n_barras_pt2} unid",
                cantidad=n_barras_pt1 + n_barras_pt2
            )
        # RIC N°6 art. 8.3.2: mismo criterio que PT1, mínimo 16mm² fijo
        # si hay más de 1 barra (esta es la puesta a tierra de PROTECCIÓN,
        # la del tablero, distinta de la de SERVICIO que es PT1)
        if n_barras_pt2 > 1:
            add_row(
                desc=f"Conductor desnudo Cu 16mm² (unión entre barras PT2)",
                marcas_txt=marcas.get("Conductor desnudo Cu", ""),
                norma="RIC 6 (8.3.2, 8.7, 8.9)",
                circuito="Puesta a tierra N°2 (TDA - camarilla N°2)",
                unidad="m",
                k=1,
                longitud_m=f"{long_cond_desnudo_pt2:.1f} m",
                cantidad=long_cond_desnudo_pt2
            )

        #Abrazaderas tipo caddy — 3 por cada tubo, con el tamaño de SU
        # PROPIO diámetro (acometida+PT vs. alimentador pueden diferir).
        # Se fusiona en 1 fila si ambos diámetros coinciden.
        def _desc_caddy_de(diam_txt):
            if "25mm" in diam_txt:
                return "Abrazadera tipo caddy 25mm"
            elif "32mm" in diam_txt:
                return "Abrazadera tipo caddy 32mm"
            elif "40mm" in diam_txt:
                return "Abrazadera tipo caddy 40mm"
            elif "50mm" in diam_txt:
                return "Abrazadera tipo caddy 50mm"
            return ""

        cantidad_caddy_acom = cantidad_total_acom_pt * 3
        cantidad_caddy_alim = cantidad_tramo_alim * 3
        cantidad_caddy = cantidad_caddy_acom + cantidad_caddy_alim  # total, usado más abajo en tornillos de fijación

        if desc_tubo_acom and desc_tubo_alim and desc_tubo_acom == desc_tubo_alim:
            # Mismo diámetro: 1 sola fila con el total
            _desc_caddy = _desc_caddy_de(desc_tubo_acom)
            if _desc_caddy and cantidad_caddy:
                add_row(
                    desc=_desc_caddy,
                    marcas_txt=marcas.get("Abrazadera tipo caddy", ""),
                    norma="SEC",
                    circuito="Puesta a tierra 1 y 2, acometida, alimentador",
                    unidad="u",
                    k=1,
                    longitud_m=f"{cantidad_caddy} unid",
                    cantidad=cantidad_caddy
                )
        else:
            # Diámetros distintos: 2 filas separadas, una por cada tramo
            # (acometida y alimentador siempre tienen su propio diámetro)
            _desc_caddy_acom = _desc_caddy_de(desc_tubo_acom)
            if _desc_caddy_acom and cantidad_caddy_acom:
                add_row(
                    desc=_desc_caddy_acom,
                    marcas_txt=marcas.get("Abrazadera tipo caddy", ""),
                    norma="SEC",
                    circuito="Puesta a tierra 1 y 2, acometida",
                    unidad="u",
                    k=1,
                    longitud_m=f"{cantidad_caddy_acom} unid",
                    cantidad=cantidad_caddy_acom
                )
            _desc_caddy_alim = _desc_caddy_de(desc_tubo_alim)
            if _desc_caddy_alim and cantidad_caddy_alim:
                add_row(
                    desc=_desc_caddy_alim,
                    marcas_txt=marcas.get("Abrazadera tipo caddy", ""),
                    norma="SEC",
                    circuito="Alimentador",
                    unidad="u",
                    k=1,
                    longitud_m=f"{cantidad_caddy_alim} unid",
                    cantidad=cantidad_caddy_alim
                )

        #Sellador de roscas: 1 frasco de 50ml alcanza para 6 conectores HUB
        cantidad_sellador = int(math.ceil(cantidad_hub_total / 6))
        add_row(
            desc="Sellador de roscas con teflón 50ml",
            marcas_txt=marcas.get("Sellador de roscas", ""),
            norma="SEC",
            circuito="Conectores Hub(cajas metálicas empalme y derivación)",
            unidad="u",
            k=1,
            longitud_m=f"{cantidad_sellador} unid",
            cantidad=cantidad_sellador
        )

        # Terminal ferrul acometida
        # Terminal ferrul acometida: mismo color/sección que el tablero (color_ferrul_por_seccion)
        try:
            sec_acom = float(acometida_txt.split("2x")[1].replace("mm²", "").strip())
        except:
            sec_acom = 4.0
        color_ferrul = color_ferrul_por_seccion(sec_acom)
        sec_ferrul = str(int(sec_acom)) if sec_acom == int(sec_acom) else texto_seccion(sec_acom)
        add_row(
            desc=f"Terminal ferrul color {color_ferrul} {sec_ferrul}mm (acometida)",
            marcas_txt=marcas.get("Terminal ferrul acometida", ""),
            norma="SEC",
            circuito="Acometida (Fase y Neutro)",
            unidad="u",
            k=1,
            longitud_m="2 unid",
            cantidad=2
        )

        # Terminal ferrul alimentador: mismo color/sección que el tablero
        try:
            sec_alim = float(alim_txt.split("3x")[1].replace("mm²", "").strip())
        except:
            sec_alim = 4.0
        color_ferrul_alim = color_ferrul_por_seccion(sec_alim)
        sec_ferrul_alim = str(int(sec_alim)) if sec_alim == int(sec_alim) else texto_seccion(sec_alim)
        add_row(
            desc=f"Terminal ferrul color {color_ferrul_alim} {sec_ferrul_alim}mm (alimentador)",
            marcas_txt=marcas.get("Terminal ferrul alimentador", ""),
            norma="SEC",
            circuito="Alimentador (Salida de medidor fase, entrada y salida termomagnético empalme)",
            unidad="u",
            k=1,
            longitud_m="3 unid",
            cantidad=3
        )

        #terminal ferrul doble alimentador
        add_row(
            desc=f"Terminal ferrul doble color {color_ferrul_alim} {sec_ferrul_alim}mm (alimentador)",
            marcas_txt=marcas.get("Terminal ferrul doble alimentador", ""),
            norma="SEC",
            circuito="Salida de medidor (Neutro del alimentador y neutro aterrizado)",
            unidad="u",
            k=1,
            longitud_m="1 unid",
            cantidad=1
        )

        # Cajas metálicas de derivación por ubicación (se usa acá y en la fijación del caddy del poste)
        # esto es específico del empalme "independiente": puede haber caja en
        # el poste (si algún tramo es subterráneo) además de la de la fachada
        if "aer" in str(tipo_acometida).strip().lower() and "aer" in str(tipo_alimentador).strip().lower():
            cajas_poste = 0
            cajas_fachada = 1
        elif "aer" in str(tipo_acometida).strip().lower() and "sub" in str(tipo_alimentador).strip().lower():
            cajas_poste = 1
            cajas_fachada = 1
        elif "sub" in str(tipo_acometida).strip().lower() and "aer" in str(tipo_alimentador).strip().lower():
            cajas_poste = 1
            cajas_fachada = 1
        elif "sub" in str(tipo_acometida).strip().lower() and "sub" in str(tipo_alimentador).strip().lower():
            cajas_poste = 2
            cajas_fachada = 1
        else:
            cajas_poste = 0
            cajas_fachada = 0

        # Tornillos según material y ubicación (cajas_poste/cajas_fachada ya calculadas más arriba)
        # - Fachada/casa: abrazaderas PVC del alimentador + cajas metálicas en fachada
        # - Poste: cajas metálicas en poste + abrazaderas PVC en poste
        # Diámetro real del ducto del alimentador (Tabla N°4.19/4.29, ya
        # calculado más arriba en canalizacion_txt). La separación entre
        # abrazaderas sigue la Tabla N°4.24 según ese diámetro: 1,20m si el
        # conduit es ≤25mm, 1,50m si es mayor.
        _m_diam_conduit_tornillos = re.search(r'(\d+)\s*mm', str(canalizacion_txt))
        _diam_conduit_tornillos = max(32, int(_m_diam_conduit_tornillos.group(1))) if _m_diam_conduit_tornillos else 32
        _sep_abraz_tornillos = 1.20 if _diam_conduit_tornillos <= 25 else 1.50
        # Abrazaderas PVC por ubicación
        if "aer" in str(tipo_acometida).strip().lower() and "aer" in str(tipo_alimentador).strip().lower():
            abrazaderas_poste = 0
            abrazaderas_fachada = int(math.ceil(float(longitud_llegada_aerea_tda) / _sep_abraz_tornillos))
        elif "aer" in str(tipo_acometida).strip().lower() and "sub" in str(tipo_alimentador).strip().lower():
            abrazaderas_poste = 1
            abrazaderas_fachada = int(math.ceil(float(longitud_abrazaderas_alimentador) / _sep_abraz_tornillos))
        elif "sub" in str(tipo_acometida).strip().lower() and "aer" in str(tipo_alimentador).strip().lower():
            abrazaderas_poste = 1
            abrazaderas_fachada = int(math.ceil(float(longitud_llegada_aerea_tda) / _sep_abraz_tornillos))
        elif "sub" in str(tipo_acometida).strip().lower() and "sub" in str(tipo_alimentador).strip().lower():
            abrazaderas_poste = 2
            abrazaderas_fachada = int(math.ceil(float(longitud_abrazaderas_alimentador) / _sep_abraz_tornillos))
        else:
            abrazaderas_poste = 0
            abrazaderas_fachada = 0
        # Tornillos fachada/casa
        # + cantidad_caddy_pt2: la abrazadera caddy del tramo PT2 (del TDA a la camarilla
        # N°2) físicamente está en la fachada, no en el poste — se fija con
        # 1 tornillo (y 1 tarugo si el forrado lo requiere) por cada abrazadera
        cant_tornillos_fachada = int((abrazaderas_fachada * 2) + (cajas_fachada * 4) + cantidad_caddy_pt2)

        # Grupo A: caja metálica del empalme (fijo, siempre 6), según el
        # material del POSTE (madera/metálico) — con golilla solo acá.
        cant_grupo_A_poste = 6
        if "madera" in str(tipo_poste).strip().lower():
            add_row(
                desc='Tirafondo hexagonal 1/4" x 1 1/2" + golilla 1/4"',
                marcas_txt=marcas.get("Tirafondo hexagonal madera", ""),
                norma="-",
                circuito="Caja metálica del empalme",
                unidad="u",
                k=1,
                longitud_m=f"{cant_grupo_A_poste} unid",
                cantidad=cant_grupo_A_poste
            )
        elif "metal" in str(tipo_poste).strip().lower():
            add_row(
                desc='Tornillo autoperforante 1/4" x 1 1/2" + golilla 1/4"',
                marcas_txt=marcas.get("Tornillo autoperforante broca", ""),
                norma="-",
                circuito="Caja metálica del empalme",
                unidad="u",
                k=1,
                longitud_m=f"{cant_grupo_A_poste} unid",
                cantidad=cant_grupo_A_poste
            )

        # Grupo B (poste): abrazaderas caddy del poste + caja(s) de derivación
        # del poste + abrazaderas PVC del poste — usa el material del POSTE
        # (tipo_poste), NO el de la fachada, ya que físicamente van montadas
        # sobre el poste (mismo criterio de textos que fachada, pero
        # seleccionado por tipo_poste en vez de material_forrado_exterior).
        cant_grupo_B_poste = (1 * (cantidad_caddy - cantidad_caddy_pt2)) + (4 * cajas_poste) + (2 * abrazaderas_poste)
        if cant_grupo_B_poste > 0:
            if "madera" in str(tipo_poste).strip().lower():
                add_row(
                    desc='Tornillo 8x1" cabeza lenteja punta fina',
                    marcas_txt=marcas.get("Tornillo punta fina madera", ""),
                    norma="-",
                    circuito="Poste: abrazaderas caddy, caja(s) de derivación metálica y abrazaderas PVC",
                    unidad="u",
                    k=1,
                    longitud_m=f"{cant_grupo_B_poste} unid",
                    cantidad=cant_grupo_B_poste
                )
            elif "metal" in str(tipo_poste).strip().lower():
                add_row(
                    desc='Tornillo cabeza lenteja 8x1 1/4" punta broca',
                    marcas_txt=marcas.get("Tornillo autoperforante broca", ""),
                    norma="-",
                    circuito="Poste: abrazaderas caddy, caja(s) de derivación metálica y abrazaderas PVC",
                    unidad="u",
                    k=1,
                    longitud_m=f"{cant_grupo_B_poste} unid",
                    cantidad=cant_grupo_B_poste
                )

        #Alimentador
        add_row(
            desc=f'Alimentador RV-K Cu 3x{round(float(res_alim["S"]),2)}mm2',
            marcas_txt=marcas.get("Alimentador RV-K", ""),
            norma="SEC",
            circuito="Alimentador",
            unidad="m",
            k=1,
            longitud_m=f"{math.ceil(float(longitud_alimentador))} m",
            cantidad=math.ceil(float(longitud_alimentador))
        )

        # Conduit PVC para tramo subterráneo solo de acometida en empalme.
        # Es un material de PVC, no de acero — usa Tabla N°4.29 (2
        # conductores, F+N), NO el diámetro de desc_tubo_acom (que es del
        # tubo de acero galvanizado y va con tabla simple aparte).
        diam_pvc_sub = None  # inicializar siempre antes de usar
        if "sub" in str(tipo_acometida).strip().lower():
            diam_pvc_sub = ducto_nominal_tablas(sec_acom, 2, "subterraneo")
            if diam_pvc_sub:
                longitud_pvc_sub = (float(longitud_transformador_empalme) - float(longitud_subterraneo_medidor2))
                cantidad_pvc_sub = int(math.ceil(longitud_pvc_sub / 3))
                add_row(desc=f"Conduit PVC {diam_pvc_sub}mm, 3mts",
                    marcas_txt=marcas.get("Conduit PVC", ""),
                    norma="SEC",
                    circuito="Acometida subterránea",
                    unidad="u",
                    k=1,
                    longitud_m=f"{round(longitud_pvc_sub, 2)} m",
                    cantidad=cantidad_pvc_sub)

        # Abrazadera adicional para PVC subterráneo acometida, distinta medida que las del alimentador
        if "sub" in str(tipo_acometida).strip().lower() and diam_pvc_sub:
            add_row(desc=f"Abrazadera conduit PVC {diam_pvc_sub}mm",
                marcas_txt=marcas.get("Abrazadera conduit PVC alimentador", ""),
                norma="SEC",
                circuito="Acometida subterránea",
                unidad="u",
                k=1,
                longitud_m="1 Unid",
                cantidad=1)

        # Conduit PVC según canalización alimentador (mínimo 32 mm)
        # Se define ANTES del bloque cámara para que diam_conduit esté disponible
        _m_diam_conduit2 = re.search(r'(\d+)\s*mm', str(canalizacion_txt))
        diam_conduit = max(32, int(_m_diam_conduit2.group(1))) if _m_diam_conduit2 else 32
        # Longitud conduit según tipo alimentador
        if "aer" in str(tipo_alimentador).strip().lower():
            metros_cond = math.ceil(float(longitud_llegada_aerea_tda))
        elif "sub" in str(tipo_alimentador).strip().lower():
            metros_cond = math.ceil(float(longitud_alimentador) - float(longitud_subterraneo_medidor2))  # descuenta el tramo que ya va con su propio tubo
        else:
            metros_cond = math.ceil(float(longitud_alimentador))

        if metros_cond > 0:
            cantidad_conduit_alim = int(math.ceil(metros_cond / 3))
            add_row(
                desc=f'Conduit de PVC de {diam_conduit}mm, 3mts',
                marcas_txt=marcas.get("Conduit PVC", ""),
                norma="RIC 4.7.2",
                circuito="Alimentador",
                unidad="u",
                k=1,
                longitud_m=f"{metros_cond} m",
                cantidad=cantidad_conduit_alim
            )

        #Abrazaderas para canalizacion del alimentador según Tabla N°4.24
        sep_abraz_alim = 1.20 if diam_conduit <= 25 else 1.50
        if "aer" in str(tipo_alimentador).strip().lower():
            longitud_abrazaderas = metros_cond
            cantidad_abrazaderas_alim = int(math.ceil(longitud_abrazaderas / sep_abraz_alim))
            circ_abraz_alim = "Alimentador (tramo llegada aérea al TDA, poste)"
        elif "sub" in str(tipo_alimentador).strip().lower():
            longitud_abrazaderas = float(longitud_abrazaderas_alimentador)
            cantidad_abrazaderas_alim = int(math.ceil(longitud_abrazaderas / sep_abraz_alim)) + 1  # +1 del poste
            circ_abraz_alim = "Alimentador (tramo salida subterránea al TDA, poste)"
        else:
            cantidad_abrazaderas_alim = int(math.ceil(metros_cond / sep_abraz_alim))
            circ_abraz_alim = "Alimentador (tramo en ducto)"
        if cantidad_abrazaderas_alim > 0:
            add_row(
                desc=f"Abrazadera conduit de PVC {diam_conduit}mm",
                marcas_txt=marcas.get("Abrazadera conduit PVC alimentador", ""),
                norma="RIC 4 (Tabla N°4.24)",
                circuito=circ_abraz_alim,
                unidad="u",
                k=1,
                longitud_m=f"{cantidad_abrazaderas_alim} unid",
                cantidad=cantidad_abrazaderas_alim
            )
        # =========================================================
        # CÁMARA TIPO C — CANALIZACIÓN SUBTERRÁNEA RESIDENCIAL
        # RIC N°04, sección 7.9, 7.9.5, 7.9.8.4.3 y Anexo 4.5
        # Regla RIC N°4 art. 7.9.7.8 / 7.9.7.9 / 7.9.7.10:
        #   Cada tramo (acometida / alimentador) se calcula INDEPENDIENTE:
        #   - Si L es 20 m o menos: 0 cámaras (forma U, RIC N°4 art. 7.9.7.10)
        #   - Si L es mayor a 20 m: ceil(L / 90) cámaras
        #   El total es la SUMA de cámaras de cada tramo.
        # =========================================================
        _hay_sub_acom = "sub" in str(tipo_acometida).strip().lower()
        _hay_sub_alim = "sub" in str(tipo_alimentador).strip().lower()

        if _hay_sub_acom or _hay_sub_alim:

            # --- Longitudes por tramo ---
            _long_acom = max(0.0,
                float(longitud_transformador_empalme) - float(longitud_subterraneo_medidor2)
            ) if _hay_sub_acom else 0.0

            _long_alim = max(0.0,
                float(longitud_alimentador) - float(longitud_subterraneo_medidor2)
            ) if _hay_sub_alim else 0.0

            # --- Cámaras por tramo (independientes) ---
            _cam_acom = int(math.ceil(_long_acom / 90.0)) if _long_acom > 20.0 else 0
            _cam_alim = int(math.ceil(_long_alim / 90.0)) if _long_alim > 20.0 else 0
            _camaras_c = _cam_acom + _cam_alim

            # --- Diámetros de boquilla por tramo ---
            # Si la acometida es subterránea, usa diam_pvc_sub; si no está definido, usa 25mm (mínimo RIC N°4 art. 7.9.7.1)
            _diam_boq_acom = None
            try:
                _diam_boq_acom = diam_pvc_sub if diam_pvc_sub else 25
            except NameError:
                _diam_boq_acom = 25
            # Si el alimentador es subterráneo, usa diam_conduit (ya definido antes del bloque)
            _diam_boq_alim = diam_conduit

            _circ_cam = []
            if _hay_sub_acom:
                _circ_cam.append("Acometida subterránea")
            if _hay_sub_alim:
                _circ_cam.append("Alimentador subterráneo")
            _circ_cam_txt = " / ".join(_circ_cam)

            if _camaras_c > 0:
                add_row(
                    desc="Cámara tipo C de hormigón prefabricado con tapa de acero diamantado 440x440mm",
                    marcas_txt=marcas.get("Camara tipo C", ""),
                    norma="RIC 4 (7.9, 7.9.5, 7.9.7.8, 7.9.7.10, 7.9.8.4.3, Anexo 4.5)",
                    circuito=_circ_cam_txt,
                    unidad="u",
                    k=1,
                    longitud_m=(
                        f"{_camaras_c} unid "
                        f"(Acom={round(_long_acom,1)}m→{_cam_acom}u / "
                        f"Alim={round(_long_alim,1)}m→{_cam_alim}u)"
                    ),
                    cantidad=_camaras_c
                )
                add_row(
                    desc="Marco metálico galvanizado para cámara tipo C 440x440mm",
                    marcas_txt=marcas.get("Marco metalico camara C", ""),
                    norma="RIC 4 (7.9.8, Anexo 4.5)",
                    circuito=_circ_cam_txt,
                    unidad="u",
                    k=1,
                    longitud_m=f"{_camaras_c} unid",
                    cantidad=_camaras_c
                )

                # --- Boquillas por tramo ---
                # CASO 1: ambos sub, ambos con cámara y distinto diámetro: van 2 filas separadas
                if _hay_sub_acom and _hay_sub_alim \
                        and _cam_acom > 0 and _cam_alim > 0 \
                        and _diam_boq_acom != _diam_boq_alim:
                    add_row(
                        desc=(
                            f"Boquilla de PVC ø{_diam_boq_acom}mm con borde redondeado "
                            f"para entrada/salida conduit en cámara tipo C (acometida)"
                        ),
                        marcas_txt=marcas.get("Boquilla camara tipo C", ""),
                        norma="RIC 4 (7.9.8.9, 5.14)",
                        circuito="Acometida subterránea",
                        unidad="u",
                        k=1,
                        longitud_m=f"{_cam_acom * 2} unid",
                        cantidad=_cam_acom * 2
                    )
                    add_row(
                        desc=(
                            f"Boquilla de PVC ø{_diam_boq_alim}mm con borde redondeado "
                            f"para entrada/salida conduit en cámara tipo C (alimentador)"
                        ),
                        marcas_txt=marcas.get("Boquilla camara tipo C", ""),
                        norma="RIC 4 (7.9.8.9, 5.14)",
                        circuito="Alimentador subterráneo",
                        unidad="u",
                        k=1,
                        longitud_m=f"{_cam_alim * 2} unid",
                        cantidad=_cam_alim * 2
                    )
                else:
                    # CASO 2: solo acometida sub: usa diam_boq_acom
                    # CASO 3: solo alimentador sub: usa diam_boq_alim
                    # CASO 4: ambos sub mismo diámetro: una sola fila con el total
                    if _hay_sub_acom and not _hay_sub_alim:
                        _diam_boq = _diam_boq_acom
                    elif _hay_sub_alim and not _hay_sub_acom:
                        _diam_boq = _diam_boq_alim
                    else:
                        # ambos sub mismo diámetro
                        _diam_boq = _diam_boq_acom
                    _boquillas = _camaras_c * 2
                    add_row(
                        desc=(
                            f"Boquilla de PVC ø{_diam_boq}mm con borde redondeado "
                            f"para entrada/salida conduit en cámara tipo C"
                        ),
                        marcas_txt=marcas.get("Boquilla camara tipo C", ""),
                        norma="RIC 4 (7.9.8.9, 5.14)",
                        circuito=_circ_cam_txt,
                        unidad="u",
                        k=1,
                        longitud_m=f"{_boquillas} unid",
                        cantidad=_boquillas
                    )

        # Cajas de paso alimentador (RIC 7.16.1.13)
        # Solo aplica cuando el empalme es en FACHADA (bloque de arriba,
        # línea ~5960) — en poste NUNCA se generan, aunque el alimentador
        # sea "en ducto", porque acá la caja de paso es exclusiva de fachada.
        cajas_paso_alim = 0
        if cajas_paso_alim > 0:
            # Caja de paso ESTANCA (el alimentador va exterior, tipo empalme
            # independiente, no corresponde la caja de PVC de interior).
            # No lleva tapa aparte: la caja estanca viene con su tapa
            # integrada de fábrica, no es una pieza separada.
            # Tamaño según diámetro real del conduit del alimentador: con
            # tubos de 32/40mm entra una caja compacta 150x110x70mm; con
            # 50mm hace falta una más grande 190x140x90mm.
            _dc_alim = 0
            try:
                _dc_alim = int(diam_conduit)
            except Exception:
                _dc_alim = 0
            if _dc_alim >= 50:
                _medida_caja_paso_alim = "190x140x90 mm"
            else:
                _medida_caja_paso_alim = "150x110x70 mm"
            add_row(
                desc=f"Caja de paso estanca IP65 de PVC/policarbonato para exterior {_medida_caja_paso_alim} (incluye tapa)",
                marcas_txt=marcas.get("Cajas de paso estancas", ""),
                norma="RIC 4 (7.16.1.13)",
                circuito="Alimentador - Caja de paso (tramo > 20m)",
                unidad="u",
                k=1,
                longitud_m=f"{cajas_paso_alim} unid",
                cantidad=cajas_paso_alim
            )
            # Salidas de caja PVC: 2 por caja de paso (entrada + salida conduit)
            salidas_paso_alim = 2 * cajas_paso_alim
            add_row(
                desc=f"Salida de caja conduit de PVC de {diam_conduit}mm",
                marcas_txt=marcas.get("Salida de caja conduit", ""),
                norma="RIC 4.7.2",
                circuito="Alimentador - Caja de paso (tramo > 20m)",
                unidad="u",
                k=1,
                longitud_m=f"{salidas_paso_alim} unid",
                cantidad=salidas_paso_alim
            )
            # Fijación de la caja de paso (fila propia, porque el acumulador
            # general de tornillos ya se calculó y escribió antes de llegar
            # acá). En empalme independiente la caja va fijada al poste, así
            # que el criterio es tipo_poste (madera/metálico) — un poste
            # nunca es fibrocemento, no aplica tarugo acá.
            _tornillos_paso_alim = 4 * cajas_paso_alim
            if "madera" in str(tipo_poste).strip().lower():
                add_row(
                    desc='Tornillo para madera galvanizado 1/4" x 1 1/2" (caja de paso alimentador)',
                    marcas_txt=marcas.get("Tirafondo hexagonal madera", ""),
                    norma="-",
                    circuito="Alimentador - Caja de paso (tramo > 20m)",
                    unidad="u",
                    k=1,
                    longitud_m=f"{_tornillos_paso_alim} unid",
                    cantidad=_tornillos_paso_alim
                )
            elif "metal" in str(tipo_poste).strip().lower():
                add_row(
                    desc='Tornillo autoperforante punta de broca 10 x 3/4" (caja de paso alimentador)',
                    marcas_txt=marcas.get("Tornillo autoperforante broca", ""),
                    norma="-",
                    circuito="Alimentador - Caja de paso (tramo > 20m)",
                    unidad="u",
                    k=1,
                    longitud_m=f"{_tornillos_paso_alim} unid",
                    cantidad=_tornillos_paso_alim
                )
            else:
                add_row(
                    desc="Tornillo (definir según tipo de poste) — caja de paso alimentador",
                    marcas_txt=marcas.get("Tornillos", ""),
                    norma="-",
                    circuito="Alimentador - Caja de paso (tramo > 20m)",
                    unidad="u",
                    k=1,
                    longitud_m=f"{_tornillos_paso_alim} unid",
                    cantidad=_tornillos_paso_alim
                )
            # Sumar a acumuladores para otros usos (espuma SIP, etc.) aunque no
            # llegue a tiempo para la fila general de tornillos
            _cajas_total += cajas_paso_alim

        # caja de derivación metalica con tapa
        # Acometida aérea + alimentador aéreo
        cantidad_caja_derivacion = 1  # valor por defecto, por si ningún caso de abajo aplica
        if ("aer" in str(tipo_acometida).strip().lower() and "aer" in str(tipo_alimentador).strip().lower()):
            cantidad_caja_derivacion = 1
        # Acometida aérea + alimentador subterráneo
        elif ("aer" in str(tipo_acometida).strip().lower() and "sub" in str(tipo_alimentador).strip().lower()):
            cantidad_caja_derivacion = 2
        # Acometida subterránea + alimentador aéreo
        elif ("sub" in str(tipo_acometida).strip().lower() and "aer" in str(tipo_alimentador).strip().lower()):
            cantidad_caja_derivacion = 2
        # Acometida subterránea + alimentador subterráneo
        elif ("sub" in str(tipo_acometida).strip().lower() and "sub" in str(tipo_alimentador).strip().lower()):
            cantidad_caja_derivacion = 3
        add_row(
            desc="Caja de derivación metálica pregalvanizada 100x65x65mm con tapa",
            marcas_txt=marcas.get("Caja derivacion metalica", ""),
            norma="SEC",
            circuito="Empalme y/o alimentador",
            unidad="u",
            k=1,
            longitud_m=f"{cantidad_caja_derivacion} unid",
            cantidad=cantidad_caja_derivacion
        )
        add_row(
            desc="Tornillo autoperforante punta broca 8 x 1/2\" cabeza lenteja + golilla (conexión tierra caja metálica)",
            marcas_txt=marcas.get("Tornillos", ""),
            norma="RIC 4 (5.13)",
            circuito="Empalme y/o alimentador (caja de empalme + caja(s) de derivación)",
            unidad="u",
            k=1,
            longitud_m=f"{1 + cantidad_caja_derivacion} unid",
            cantidad=1 + cantidad_caja_derivacion
        )
        #Terminal de compresión tipo ojo: 1 por caja de empalme + 1 por cada caja de derivación
        cantidad_terminal_ojo = 1 + cantidad_caja_derivacion
        add_row(
            desc=f"Terminal de compresión tipo ojo {sec_pt}mm",
            marcas_txt=marcas.get("Terminal compresion tipo ojo", ""),
            norma="SEC",
            circuito="Tramos metálicos (caja metálica empalme, caja de derivación, acometida, puesta a tierra 1 y 2)",
            unidad="u",
            k=1,
            longitud_m=f"{cantidad_terminal_ojo} unid",
            cantidad=cantidad_terminal_ojo
        )
        # Material según forrado EXTERIOR de la casa
        hay_madera = "madera" in material_forrado_exterior
        hay_pvc = "pvc" in material_forrado_exterior  # ej. "siding pvc" — usa el mismo tornillo de madera, sin tarugo
        hay_siding_metalico = ("metál" in material_forrado_exterior) or ("metal" in material_forrado_exterior)  # ej. "siding metálico" o "siding metalico" (sin tilde)
        hay_fibro = "fibro" in material_forrado_exterior
        hay_forrado_valido = any(x in material_forrado_exterior for x in ("madera", "fibro", "siding"))
        # Fila de fachada: SOLO su propio aporte (abrazaderas pvc del
        # alimentador + caja(s) de derivación de la fachada + caddy de PT2) —
        # las abrazaderas/cajas/caddy que van EN EL POSTE ya tienen su propia
        # fila arriba, con el material del poste, no de la fachada.
        cant_grupo_B_total = int(cant_tornillos_fachada)
        if (hay_madera or hay_pvc) and hay_forrado_valido:
            if cant_grupo_B_total > 0:
                add_row(
                    desc='Tornillo 8x1" cabeza lenteja punta fina',
                    marcas_txt=marcas.get("Tornillo punta fina madera", ""),
                    norma="-",
                    circuito="Fachada: abrazaderas pvc del alimentador, caja(s) de derivación, y abrazadera caddy del tramo de puesta a tierra N°2",
                    unidad="u",
                    k=1,
                    longitud_m=f"{cant_grupo_B_total} unid",
                    cantidad=cant_grupo_B_total
                )
        elif hay_siding_metalico and hay_forrado_valido:
            if cant_grupo_B_total > 0:
                add_row(
                    desc='Tornillo cabeza lenteja 8x1 1/4" punta broca',
                    marcas_txt=marcas.get("Tornillo autoperforante broca", ""),
                    norma="-",
                    circuito="Fachada: abrazaderas pvc del alimentador, caja(s) de derivación, y abrazadera caddy del tramo de puesta a tierra N°2",
                    unidad="u",
                    k=1,
                    longitud_m=f"{cant_grupo_B_total} unid",
                    cantidad=cant_grupo_B_total
                )
        elif hay_fibro and hay_forrado_valido:
            # fibrocemento: necesita tarugo + tornillo juntos
            if cant_grupo_B_total > 0:
                add_row(
                    desc='Tornillo 8x1 1/2" punta fina cabeza lenteja',
                    marcas_txt=marcas.get("Tornillo punta fina madera", ""),
                    norma="-",
                    circuito="Fachada: abrazaderas pvc del alimentador, caja(s) de derivación, y abrazadera caddy del tramo de puesta a tierra N°2",
                    unidad="u",
                    k=1,
                    longitud_m=f"{cant_grupo_B_total} unid",
                    cantidad=cant_grupo_B_total
                )
                add_row(
                    desc="Tarugo paloma N°8",
                    marcas_txt=marcas.get("Tarugo paloma", ""),
                    norma="-",
                    circuito="Fachada: abrazaderas pvc del alimentador, caja(s) de derivación, y abrazadera caddy del tramo de puesta a tierra N°2",
                    unidad="u",
                    k=1,
                    longitud_m=f"{cant_grupo_B_total} unid",
                    cantidad=cant_grupo_B_total
                )
        elif not (hay_madera or hay_pvc or hay_siding_metalico or hay_fibro) and hay_forrado_valido:
            if cant_grupo_B_total > 0:
                add_row(
                    desc='Tornillo (definir según material de tabique)',
                    marcas_txt=marcas.get("Tornillos", ""),
                    norma="-",
                    circuito="Fachada: abrazaderas pvc del alimentador, caja(s) de derivación, y abrazadera caddy del tramo de puesta a tierra N°2",
                    unidad="u",
                    k=1,
                    longitud_m=f"{cant_grupo_B_total} unid",
                    cantidad=cant_grupo_B_total
                )

        # Conductor puesta a tierra camarilla N°2 (TDA - puesta a tierra 2)
        # +2m: 1 chicote en la camarilla N°2 + 1 chicote en el TDA
        metros_pt2 = math.ceil(float(dist_tda_pt2)) + 2
        add_row(
            desc=f"Conductor THWN-2 {sec_pt}mm2 verde",
            marcas_txt=marcas.get("Conductor THWN-2 verde", ""),
            norma="SEC",
            circuito="Puesta a tierra N°2 (Tda - Camarilla N°2)",
            unidad="m",
            k=1,
            longitud_m=f"{metros_pt2} m",
            cantidad=metros_pt2
        )

        # Tarugos solo si el forrado es volcanita o fibrocemento (ducto del alimentador), solo en fachada
        # hay_fibro ya se definió más arriba
        cantidad_tarugos = 0
        if hay_fibro:
            cantidad_tarugos = int((abrazaderas_fachada * 2) + (cajas_fachada * 4))
            if "sub" in str(tipo_acometida).strip().lower() and diam_pvc_sub:
                cantidad_tarugos += 1
            add_row(
                desc="Tarugo paloma 6mm",
                marcas_txt=marcas.get("Tarugo paloma", ""),
                norma="-",
                circuito="Abrazaderas PVC y caja metálica en fachada",
                unidad="u",
                k=1,
                longitud_m=f"{cantidad_tarugos} unid",
                cantidad=cantidad_tarugos
            )
            add_row(
                desc='Tornillo punta plana 6x1 1/4" para tarugo',
                marcas_txt=marcas.get("Tornillo para tarugo paloma", ""),
                norma="-",
                circuito="Abrazaderas PVC y caja metálica en fachada",
                unidad="u",
                k=1,
                longitud_m=f"{cantidad_tarugos} unid",
                cantidad=cantidad_tarugos
            )

        # Pilar para empalme independiente
        altura_visible_poste = 0
        if ("aer" in str(tipo_acometida).strip().lower()and "aer" in str(tipo_alimentador).strip().lower()):
            altura_visible_poste = (float(dist_empalme_pt1) + float(altura_acometida_aerea))
        elif ("aer" in str(tipo_acometida).strip().lower() and "sub" in str(tipo_alimentador).strip().lower()):
            altura_visible_poste = (float(dist_empalme_pt1) + float(altura_acometida_aerea))
        elif ("sub" in str(tipo_acometida).strip().lower() and "aer" in str(tipo_alimentador).strip().lower()):
            altura_visible_poste = (float(dist_empalme_pt1) + float(longitud_poste_alimentador_aereo))
        elif ("sub" in str(tipo_acometida).strip().lower() and "sub" in str(tipo_alimentador).strip().lower()):
            altura_visible_poste = 5
        longitud_total_poste = math.ceil(altura_visible_poste * 1.2)
        if "madera" in str(tipo_poste).strip().lower():
            add_row(
                desc=f"Poste de madera {longitud_total_poste}m, pino impregnado",
                marcas_txt=marcas.get("Poste madera", ""),
                norma="-",
                circuito="Empalme independiente",
                unidad="u",
                k=1,
                longitud_m=f"{longitud_total_poste} m",
                cantidad=1
            )
        elif "metal" in str(tipo_poste).strip().lower():
            add_row(
                desc=f"Pilar metálico cuadrado {longitud_total_poste}m, 100x100x3mm",
                marcas_txt=marcas.get("Pilar metalico", ""),
                norma="-",
                circuito="Empalme independiente",
                unidad="u",
                k=1,
                longitud_m=f"{longitud_total_poste} m",
                cantidad=1
            )

        # Portafusible aéreo loza según interruptor termomagnético del empalme
        tm_empalme_A = parse_in_tm(interruptor_texto)
        fusible_A = None
        if tm_empalme_A == 25:
            fusible_A = 30
        elif tm_empalme_A in [32, 40]:
            fusible_A = 60
        if fusible_A:
            add_row(
                desc=f"Portafusible de loza con fusibles cartucho {fusible_A}A",
                marcas_txt=marcas.get("Portafusible de loza", ""),
                norma="SEC",
                circuito="Empalme",
                unidad="u",
                k=1,
                longitud_m="1 unid",
                cantidad=1
            )


    # =========================
    # REORDENAR SECCIONES
    # =========================
    ORDEN_SECCIONES = [
        "Empalme",
        "Protecciones",
        "Borneras de conexión",
        "Cableado interior tablero",
        "Terminales ferrul interiores",
        "Canalizaciones",
        "Conductores",
        "Accesorios",
        "Conectores cónicos",
        "Iluminarias",
        "Tornillería",
        "Tarugos",
        "Sellos / Aislación (Panel SIP)",
    ]  # orden final en que deben aparecer las secciones en el Excel

    # Asignar grupo a cada fila según la última sección vista
    grupos = []
    seccion_actual = ""
    for fila in filas:
        desc = str(fila.get("Ítem", ""))
        # fila de sección: Ítem está vacío y Descripción técnica es el nombre de la sección
        if fila["Ítem"] == "":
            seccion_actual = fila["Descripción técnica"]
        grupos.append(seccion_actual)

    # Separar filas en bloques por sección (cabecera + sus filas)
    bloques = {}  # relaciona cada nombre_seccion con su lista [fila_header, fila1, fila2, ...]
    seccion_actual = ""
    for fila, grupo in zip(filas, grupos):
        if fila["Ítem"] == "" and fila["Descripción técnica"] != "":
            seccion_actual = fila["Descripción técnica"]
            bloques[seccion_actual] = [fila]
        elif seccion_actual:
            bloques[seccion_actual].append(fila)

    # Reconstruir filas en el orden deseado
    filas_ordenadas = []
    for nombre in ORDEN_SECCIONES:
        if nombre in bloques and len(bloques[nombre]) > 1:  # solo si tiene filas reales (no solo el título)
            filas_ordenadas.extend(bloques[nombre])
    # Agregar secciones que no estén en ORDEN_SECCIONES (por si acaso)
    for nombre, bloque in bloques.items():
        if nombre not in ORDEN_SECCIONES and len(bloque) > 1:
            filas_ordenadas.extend(bloque)

    # Renumerar Ítems (después de reordenar, los números tienen que quedar correlativos de nuevo)
    n = 1
    for fila in filas_ordenadas:
        if fila["Ítem"] != "":
            fila["Ítem"] = n
            n += 1

    return pd.DataFrame(
        filas_ordenadas,
        columns=[
            "Ítem",
            "Descripción técnica",
            "Marcas sugeridas",
            "Sello SEC",
            "Norma / RIC",
            "Circuito",
            "Unidad",
            "K",
            "Longitud (m) / Unidad",
            "Cantidad"
        ]
    )

# -------- PRIMERO: EMPALME (AUTO) --------
# El empalme se calculará más adelante usando la corriente CON factor de demanda.
# Aquí solo definimos valores iniciales para no bloquear el ingreso de datos.
proteccion_empalme = None  # (se deja por compatibilidad; NO se usa para cálculos)
dif_calibre = None
max_sum_tm = None
max_pot_especial = float("inf")  # no limita durante el ingreso; se valida al final
texto_omni = None

# -------- PARTE 1: INGRESO DE AMBIENTES Y CARGAS --------
cantidad_ambientes = int(input("\n Ingrese la cantidad de ambientes (living, baño, piezas, etc.): "))

ambientes_detalle = []       # va a terminar como el DataFrame ambientes_df
componentes_por_ambiente = {}
enchufes_por_ambiente = {}
luminarias_por_ambiente = {}

# recorre uno por uno todos los ambientes, preguntando sus datos
for i in range(1, cantidad_ambientes + 1):
    print(f"\n Ambiente N°{i}")
    nombre = input(f"   - Nombre del ambiente: ")
    dimension = float(input(f"   - Dimensión del ambiente '{nombre}' en m²: "))
    perimetro = float(input(f"   - Perímetro del ambiente '{nombre}' en m: "))
    material_tabique = input(
        f" - Tipo de material de construcción del tabique del ambiente '{nombre}' "
        "(Madera, Metalcon, Panel SIP): "
    ).strip().lower()

    material_forrado_interior = input(
        f" - Forrado interior del ambiente '{nombre}' (Volcanita, Madera): "
    ).strip().lower()


    # Iluminación: regla según área — ambientes chicos necesitan menos luminarias mínimas
    if dimension <= 10:
        min_luminarias = 1
        texto_condicion = "<= 10 m²"
    else:
        min_luminarias = 2
        texto_condicion = "> 10 m²"

    while True:
        try:
            cantidad_luminarias = int(input(
                f"   - Cantidad de luminarias del ambiente '{nombre}' "
                f"(mínimo {min_luminarias} para {texto_condicion}): "
            ))
            if cantidad_luminarias < min_luminarias:
                print(f"     ! Debe ingresar al menos {min_luminarias} luminaria(s) para {texto_condicion}.")
            else:
                break  # cumple el mínimo, sigue
        except:
            print("     ! Ingrese un número entero válido para la cantidad de luminarias.")

    detalle_luminarias = []
    potencia_luminarias = []
    luminarias_detalle_amb = []

    # pregunta los datos de cada luminaria del ambiente, una por una
    for j in range(1, cantidad_luminarias + 1):
        tipo = input(
            f"     • Tipo de iluminación {j} "
            f"(Foco LED, aplique LED, Tubo fluorescente, Tubo led, Ampolleta incandescente, ampolleta LED): "
        ).strip()

        while True:
            montaje = input(
                f"       • ¿La iluminación {j} es embutida o sobrepuesta? (embutida/sobrepuesta): "
            ).strip().lower()
            if montaje in ["embutida", "sobrepuesta"]:
                break
            print("         ! Responda solo 'embutida' o 'sobrepuesta'.")

        while True:
            conoce = input(
                f"       ¿Usted conoce el valor de la potencia de la luminaria {j}? (si/no)\n"
                f"       (Si no lo sabe, se ingresará automáticamente 100W por norma): "
            ).strip().lower()
            if conoce in ["si", "no"]:
                break
            print("         ! Responda solo 'si' o 'no'.")

        if conoce == "si":
            p = float(input(f"       Potencia de la luminaria {j} [W]: "))
        else:
            p = 100.0
            print("       -> No se conoce la potencia, se asignan 100 W por defecto según criterio.")

        potencia_luminarias.append(p)

        desc_final = desc_luminaria_auto(tipo, montaje, p)  # arma la descripción "bonita" de la luminaria

        detalle_luminarias.append(desc_final)
        luminarias_detalle_amb.append({"id": j, "tipo": tipo, "montaje": montaje, "potencia": p, "desc": desc_final})

    potencia_total_ilum = sum(potencia_luminarias)
    detalle_iluminacion = "\n".join(detalle_luminarias) if detalle_luminarias else "Ninguna"
    luminarias_por_ambiente[nombre] = luminarias_detalle_amb[:]

    # Conmutado: se pregunta una vez por ambiente
    while True:
        conmutado = input(f"   - ¿El ambiente '{nombre}' tiene conmutado (2 puntos)? (si/no): ").strip().lower()
        if conmutado in ["si", "no"]:
            break
        print("     ! Responda solo 'si' o 'no'.")

    # acumuladores de longitudes de cableado (se usan después en build_materiales_df
    # para calcular los metros exactos de cada color de conductor)
    L_viajeros_924 = 0.0
    L_retorno_lampara = 0.0
    L_fase_caja_primer_int = 0.0
    L_troncal_primera_oct_924 = 0.0  # distancia de la caja troncal a la primera caja octogonal (1 vez por ambiente, no por luminaria)

    L_caja_int_fase_ida = 0.0      # suma L_ida por grupo de interruptor (se mantiene por compatibilidad)
    L_troncal_oct1 = 0.0           # suma distancia de la troncal a la primera oct por grupo (idem)
    L_oct1_oct2 = 0.0              # suma distancia entre octs intermedias por grupo (idem)
    grupos_ilum_detalle = []       # detalle por grupo: [{"tipo","n_lum","L_ida","L_tr_oct1","L_o1_o2"}, ...]
                                    # permite calcular extra_R/extra_NT grupo por grupo, en vez de
                                    # con sumas totales (que sobre-cuentan si hay >1 interruptor
                                    # por ambiente con distancias distintas entre sí)

    N_conmutadas_924 = 0

    if conmutado == "si":
        while True:
            try:
                N_conmutadas_924 = int(input(
                    f"     • En '{nombre}' hay {cantidad_luminarias} luminarias.\n"
                    f"       ¿Cuántas luminarias serán conmutadas con 9/24? (0 a {cantidad_luminarias}): "
                ))
                if 0 <= N_conmutadas_924 <= int(cantidad_luminarias):
                    break
                print(f"       ! Debe estar entre 0 y {cantidad_luminarias}.")
            except:
                print("       ! Ingrese un número entero válido.")

        if N_conmutadas_924 > 0:
            # El 9/24 enciende TODAS las luminarias conmutadas del grupo al
            # mismo tiempo (mismo par de interruptores) — estas 4 longitudes
            # se preguntan 1 sola vez por ambiente, NO por cada luminaria.
            print("     • Conmutado (9/24): ingrese las longitudes REALES del grupo conmutado (ENTER=0.0)")
            L_viajeros_924 = pedir_float_opcional(
                "       • Longitud TOTAL entre interruptores (viajeros 9/24) [m] (ENTER=0.0): ", 0.0, 0.0
            )
            L_retorno_lampara = pedir_float_opcional(
                "       • Longitud de RETORNO (interruptor→lámpara) [m] (ENTER=0.0): ", 0.0, 0.0
            )
            L_fase_caja_primer_int = pedir_float_opcional(
                "       • Longitud de FASE (caja→primer interruptor) [m] (ENTER=0.0): ", 0.0, 0.0
            )
            L_troncal_primera_oct_924 = pedir_float_opcional(
                "       • Longitud de la CAJA TRONCAL a la PRIMERA CAJA OCTOGONAL [m] (ENTER=0.0): ", 0.0, 0.0
            )

        n_no_conmut = int(cantidad_luminarias) - int(N_conmutadas_924)
        if n_no_conmut > 0:
            # Descomponer en grupos de interruptor
            c12_g, c15_g, c32_g = descomponer_interruptores(n_no_conmut)
            grupos = []
            for _ in range(c32_g): grupos.append(('9/32', 3))
            for _ in range(c15_g): grupos.append(('9/15', 2))
            for _ in range(c12_g): grupos.append(('9/12', 1))

            # pregunta las longitudes de cada grupo de interruptor no conmutado
            print(f"     • Luminarias NO conmutadas: {n_no_conmut} luminarias → {len(grupos)} grupo(s) de interruptor")
            for g_idx, (tipo_g, n_lum_g) in enumerate(grupos, 1):
                print(f"       - Grupo {g_idx} ({tipo_g}, {n_lum_g} luminaria{'s' if n_lum_g>1 else ''}):")
                _L_ida_g = pedir_float_opcional(
                    f"         • Longitud troncal→interruptor {tipo_g} [m] (ENTER=0.0): ", 0.0, 0.0
                )
                _L_tr1_g = 0.0
                if n_lum_g > 1:
                    # Solo aplica a 9/15 y 9/32 (2+ luminarias) — para 9/12
                    # (1 sola luminaria) esta distancia nunca se usa en el
                    # cálculo, así que no se pregunta.
                    _L_tr1_g = pedir_float_opcional(
                        f"         • Longitud troncal→primera oct [m] (ENTER=0.0): ", 0.0, 0.0
                    )
                _L_o12_g = 0.0
                if n_lum_g > 2:
                    # Solo aplica a 9/32 (3+ luminarias) — para 9/15 (2
                    # luminarias) esta distancia nunca se usa en el cálculo,
                    # así que no se pregunta.
                    for j in range(1, n_lum_g - 1):
                        _L_o12_g += pedir_float_opcional(
                            f"         • Longitud oct{j}→oct{j+1} [m] (ENTER=0.0): ", 0.0, 0.0
                        )
                # Mantener las sumas totales por compatibilidad con otros usos
                L_caja_int_fase_ida += _L_ida_g
                L_troncal_oct1 += _L_tr1_g
                L_oct1_oct2 += _L_o12_g
                # Y guardar el detalle por grupo, para el cálculo grupo-por-grupo
                grupos_ilum_detalle.append({
                    "tipo": tipo_g, "n_lum": n_lum_g,
                    "L_ida": _L_ida_g, "L_tr_oct1": _L_tr1_g, "L_o1_o2": _L_o12_g
                })

    else:
        # sin conmutado: todas las luminarias van en grupos de interruptor normales
        n_no_conmut = int(cantidad_luminarias)
        c12_g, c15_g, c32_g = descomponer_interruptores(n_no_conmut)
        grupos = []
        for _ in range(c32_g): grupos.append(('9/32', 3))
        for _ in range(c15_g): grupos.append(('9/15', 2))
        for _ in range(c12_g): grupos.append(('9/12', 1))

        print(f"     • Sin conmutado: {n_no_conmut} luminarias → {len(grupos)} grupo(s) de interruptor")
        for g_idx, (tipo_g, n_lum_g) in enumerate(grupos, 1):
            print(f"       - Grupo {g_idx} ({tipo_g}, {n_lum_g} luminaria{'s' if n_lum_g>1 else ''}):")
            _L_ida_g = pedir_float_opcional(
                f"         • Longitud troncal→interruptor {tipo_g} [m] (ENTER=0.0): ", 0.0, 0.0
            )
            _L_tr1_g = 0.0
            if n_lum_g > 1:
                # Solo aplica a 9/15 y 9/32 (2+ luminarias) — para 9/12
                # (1 sola luminaria) esta distancia nunca se usa en el
                # cálculo, así que no se pregunta.
                _L_tr1_g = pedir_float_opcional(
                    f"         • Longitud troncal→primera oct [m] (ENTER=0.0): ", 0.0, 0.0
                )
            _L_o12_g = 0.0
            if n_lum_g > 2:
                # Solo aplica a 9/32 (3+ luminarias) — para 9/15 (2
                # luminarias) esta distancia nunca se usa en el cálculo,
                # así que no se pregunta.
                for j in range(1, n_lum_g - 1):
                    _L_o12_g += pedir_float_opcional(
                        f"         • Longitud oct{j}→oct{j+1} [m] (ENTER=0.0): ", 0.0, 0.0
                    )
            # Mantener las sumas totales por compatibilidad con otros usos
            L_caja_int_fase_ida += _L_ida_g
            L_troncal_oct1 += _L_tr1_g
            L_oct1_oct2 += _L_o12_g
            # Y guardar el detalle por grupo, para el cálculo grupo-por-grupo
            grupos_ilum_detalle.append({
                "tipo": tipo_g, "n_lum": n_lum_g,
                "L_ida": _L_ida_g, "L_tr_oct1": _L_tr1_g, "L_o1_o2": _L_o12_g
            })

    # Componentes especiales
    componentes_nombres = []
    componentes_potencias = []
    componentes_detalle_amb = []

    tiene_componentes = input(
        f"   - ¿El ambiente '{nombre}' tiene componentes especiales conectados a enchufes? (si/no): "
    ).strip().lower()

    if tiene_componentes == 'si':
        idx_comp = 1
        # pregunta uno por uno los equipos especiales, hasta que el usuario escriba "fin"
        while True:
            nombre_comp = input("     • Ingrese nombre del componente (o 'fin' para terminar): ")
            if nombre_comp.lower() == 'fin':
                break

            # Si es aire acondicionado, se pide el BTU y se convierte con un EER típico de 3.2
            # (dato estimado para potencia del ambiente; la placa real se ingresa en el circuito)
            _nombre_lower = nombre_comp.strip().lower()
            _es_clima_comp = any(k in _nombre_lower for k in ("aire", "clima", "split", "ac ", "a/c"))
            if _es_clima_comp:
                while True:
                    try:
                        _btu = float(input(f"       Capacidad de '{nombre_comp}' en BTU/h (ej: 9000, 12000, 18000, 24000): "))
                        if _btu <= 0:
                            print("         ! Debe ser mayor que 0.")
                            continue
                        break
                    except:
                        print("         ! Ingrese un número válido.")
                # Conversión con EER típico 3.2 (estimación; dato real se ingresa en el circuito)
                potencia_comp = round(_btu / (3.2 * 3.412), 1)
                print(f"       → Potencia estimada: {potencia_comp} W (dato referencial, se precisa en el circuito)")
            else:
                potencia_comp = float(input(f"       Potencia de '{nombre_comp}' en W: "))
            # Nota: no limitamos aquí por empalme, porque el empalme se calcula al final con factor de demanda.
            # La validación final se muestra como aviso (no bloqueante).
            componentes_nombres.append(nombre_comp)
            componentes_potencias.append(potencia_comp)
            componentes_detalle_amb.append({
                "id": idx_comp,
                "nombre": nombre_comp,
                "potencia": potencia_comp
            })
            idx_comp += 1
    else:
        print("     • No se ingresaron componentes especiales para este ambiente.")

    componentes_por_ambiente[nombre] = componentes_detalle_amb[:]
    potencia_total_comp = sum(componentes_potencias)

    # ---------- ENCHUFES COMUNES ----------
    nombre_lower = nombre.strip().lower()
    es_cocina = nombre_lower.startswith("cocina")
    es_lavadero = nombre_lower.startswith("lavadero")

    # ambientes que requieren 1 enchufe DOBLE/TRIPLE cada 8m de perímetro
    es_dormitorio = ("dormitorio" in nombre_lower) or nombre_lower.startswith("pieza") or nombre_lower.startswith("habitacion")
    es_living = nombre_lower.startswith("living")
    es_comedor = nombre_lower.startswith("comedor")
    es_sala_estar = ("sala de estar" in nombre_lower) or ("estar" in nombre_lower)
    es_bano = (nombre_lower.startswith("baño") or nombre_lower.startswith("bano"))
    es_pasillo = nombre_lower.startswith("pasillo")

    # Regla 1 enchufe cada 8m SOLO para: dormitorios/living/comedor/estar
    # NO aplicar en pasillo ni baño
    aplica_regla_dt = (es_dormitorio or es_living or es_comedor or es_sala_estar) and (not es_pasillo) and (not es_bano)
    req_dt = int(math.ceil(perimetro / 8.0)) if (aplica_regla_dt and perimetro > 0) else 0
    # mínimo por perímetro (solo si aplica regla)
    if aplica_regla_dt and perimetro > 0:
        min_ench_ric = max(1, math.ceil(perimetro / 8.0))
    else:
        min_ench_ric = 0

    # CASO ESPECIAL: PASILLO
    if es_pasillo:
        while True:
            resp_pasillo = input(
                f"   - ¿El pasillo '{nombre}' tiene enchufes? (si/no): "
            ).strip().lower()

            if resp_pasillo in ["si", "no"]:
                break
            print("     ! Responda solo 'si' o 'no'.")

        if resp_pasillo == "no":
            cantidad_enchufes = 0
        else:
           # si dice que SI, recién preguntamos cuántos
            while True:
                try:
                    cantidad_enchufes = int(
                        input(f"   - Cantidad de enchufes del pasillo '{nombre}': ")
                    )
                    if cantidad_enchufes < 0:
                        print("     ! No puede ser negativa.")
                        continue
                    break
                except:
                    print("     ! Ingrese un número entero válido.")
    else:
        while True:
            try:
                cantidad_enchufes = int(input(f"   - Cantidad de enchufes comunes del ambiente '{nombre}': "))
                if cantidad_enchufes < 0:
                    print("     ! La cantidad de enchufes no puede ser negativa.")
                    continue

                if es_cocina and cantidad_enchufes < 3:
                    print("     ! Según RIC N°10, la cocina debe tener al menos 3 enchufes.")
                    continue

                if es_lavadero and cantidad_enchufes < 1:
                    print("     ! El ambiente 'Lavadero' debe tener al menos 1 enchufe.")
                    continue

                if aplica_regla_dt and req_dt > 0 and cantidad_enchufes < req_dt:
                    print(f"     ! Según RIC, en '{nombre}' necesitas al menos {req_dt} enchufe(s) "
                        f"(doble o triple) (1 por cada 8m de perímetro).")
                    continue

                if min_ench_ric > 0 and cantidad_enchufes < min_ench_ric:
                    print(f"     ! Según RIC, en '{nombre}' se requiere al menos "
                        f"{min_ench_ric} enchufe(s) (1 cada 8 m de perímetro o fracción).")
                    continue
                break
            except:
                print("     ! Ingrese un número entero válido para la cantidad de enchufes.")

    potencia_enchufes = []
    enchufes_detalle_amb = []
    dobles_triples_cocina = 0
    lavadero_doble_triple = False
    # contador de enchufes dobles/triples para dormitorios/living/comedor/estar
    dobles_triples_dt = 0

    for j in range(1, cantidad_enchufes + 1):
        while True:
            try:
                mod = int(input(f"     • Módulos del enchufe común {j} (1=simple, 2=doble, 3=triple): "))
            except:
                print("         ! Valor inválido. Ingrese 1, 2 o 3.")
                continue
            if mod not in [1, 2, 3]:
                print("         ! Ingrese solo 1, 2 o 3.")
                continue

            if es_lavadero:
                restantes = cantidad_enchufes - j  # cuántos enchufes quedan por ingresar después de este
                if (not lavadero_doble_triple) and (restantes == 0) and (mod == 1):
                    print("         ! En Lavadero debe existir al menos 1 enchufe DOBLE (2) o TRIPLE (3) de 16A.")
                    print("           → Este último NO puede ser simple. Ingrese 2 o 3.")
                    continue

            if es_cocina:
                faltantes = 3 - dobles_triples_cocina
                restantes_incluyendo_este = cantidad_enchufes - j + 1
                max_dobles_posibles = restantes_incluyendo_este - 1

                if mod == 1 and faltantes > max_dobles_posibles:
                    print("         ! En cocina, para cumplir RIC N°10, aquí debes ingresar DOBLE (2) o TRIPLE (3).")
                    continue

            if aplica_regla_dt and req_dt > 0:
                faltan_dt = req_dt - dobles_triples_dt
                # cuántos enchufes quedan contando este (incluye el actual)
                restantes_incluyendo_este = cantidad_enchufes - j + 1
                # máximo de dobles/triples que aún podrías lograr si desde ahora TODOS fueran 2 o 3
                max_dt_posible = restantes_incluyendo_este
                # si el usuario pone simple y con eso ya no alcanza a cumplir, se bloquea
                if mod == 1 and faltan_dt > (max_dt_posible - 1):
                    print(f"         ! En '{nombre}' debes cumplir {req_dt} enchufe(s) DOBLE/TRIPLE "
                        f"(1 por cada 8m de perímetro).")
                    print("           → Aquí NO puedes ingresar simple. Ingrese 2 o 3.")
                    continue
            break

        if es_cocina and mod >= 2:
            dobles_triples_cocina += 1

        if es_lavadero and mod >= 2:
            lavadero_doble_triple = True

        if aplica_regla_dt and mod >= 2:
            dobles_triples_dt += 1

        while True:
            # UNA sola pregunta por enchufe
            while True:
                conoce_total = input(
                    f"       ¿Conoce la potencia TOTAL del enchufe {j}? (si/no)\n"
                    f"       (Si no lo sabe, se asignarán 250W por norma al enchufe completo): "
                ).strip().lower()
                if conoce_total in ["si", "no"]:
                    break
                print("         ! Responda solo 'si' o 'no'.")

            potencias_modulo = []

            if conoce_total == "no":
                # 250W POR ENCHUFE (no por módulo)
                potencia_total_enchufe = 250.0
                print("         -> No se conoce la potencia, se asignan 250 W por defecto (por enchufe).")
            else:
                # solo si conoce: preguntar por cada módulo
                for m in range(1, mod + 1):
                    p_mod = float(input(f"         Potencia del módulo {m} [W]: "))
                    potencias_modulo.append(p_mod)
                potencia_total_enchufe = sum(potencias_modulo)

            if (es_cocina or es_lavadero) and potencia_total_enchufe > 3200:
                print("       ! En cocina y lavadero la potencia total de cada enchufe "
                      "no puede ser mayor a 3200 W para asegurar circuitos de 16A.")
                print("         → Vuelva a ingresar los módulos y potencias de este enchufe.")
                continue
            break

        potencia_enchufes.append(potencia_total_enchufe)
        enchufes_detalle_amb.append({
            "id": j,
            "modulos": mod,
            "potencias_modulo": potencias_modulo,
            "potencia_total": potencia_total_enchufe
        })

    if es_cocina and dobles_triples_cocina < 3:
        print("\n      Según RIC N°10, en la cocina debe haber al menos 3 enchufes DOBLES o TRIPLES.")
        print("     → Debes reingresar los enchufes de cocina.")

    if es_lavadero and not lavadero_doble_triple:
        print("\n     ! Según norma, en Lavadero debe haber al menos 1 enchufe DOBLE o TRIPLE 16A.")
        print("       → Debes reingresar los enchufes del Lavadero.")

    if aplica_regla_dt and req_dt > 0 and dobles_triples_dt < req_dt:
        print(f"\n     ! Según RIC, en '{nombre}' debe haber al menos {req_dt} enchufe(s) DOBLE o TRIPLE "
            f"(1 por cada 8m de perímetro).")
        print("       → Debes reingresar los enchufes de este ambiente.")

    potencia_total_ench = sum(potencia_enchufes)
    detalle_enchufes = "\n".join([f"Enchufe {idx+1} ({p}W)" for idx, p in enumerate(potencia_enchufes)]) if potencia_enchufes else "Ninguno"
    enchufes_por_ambiente[nombre] = enchufes_detalle_amb[:]

    nombres_componentes = "\n".join([f"{n} ({p}W)" for n, p in zip(componentes_nombres, componentes_potencias)]) if componentes_nombres else "Ninguno"
    potencia_total_ambiente = potencia_total_ilum + potencia_total_ench + potencia_total_comp

    # --- CONTEOS REALES PARA MATERIALES ---
    n_luminarias = cantidad_luminarias
    n_enchufes = cantidad_enchufes
    n_modulos_enchufe = sum(e.get("modulos", 1) for e in enchufes_detalle_amb) if enchufes_detalle_amb else 0
    n_componentes_especiales = len(componentes_detalle_amb) if componentes_detalle_amb else 0

    # cuenta los enchufes por tipo (simple/doble/triple) y amperaje (10A o 10/16A
    # según si es cocina/lavadero, que llevan un amperaje distinto)
    ench_simple_10A = 0
    ench_doble_10A = 0
    ench_triple_10A = 0
    ench_simple_1016 = 0
    ench_doble_1016 = 0
    ench_triple_1016 = 0

    es_coc_lav_amb = es_cocina or es_lavadero
    for e in enchufes_detalle_amb:
        mod = int(e.get("modulos", 1))
        if es_coc_lav_amb:
            if mod == 1:
                ench_simple_1016 += 1
            elif mod == 2:
                ench_doble_1016 += 1
            elif mod == 3:
                ench_triple_1016 += 1
        else:
            if mod == 1:
                ench_simple_10A += 1
            elif mod == 2:
                ench_doble_10A += 1
            elif mod == 3:
                ench_triple_10A += 1

    # guarda todos los datos calculados de este ambiente, para armar después el DataFrame ambientes_df
    ambientes_detalle.append({
        "Ambiente": nombre,
        "Área (m²)": dimension,
        "Perímetro (m)": perimetro,
        "Material tabique": material_tabique,
        "Material forrado interior": material_forrado_interior,
        "Detalle iluminación": detalle_iluminacion,
        "Potencia iluminación (W)": potencia_total_ilum,

        "N_conmutadas_924 (u)": int(N_conmutadas_924),

        "L_viajeros_924 (m)": L_viajeros_924,
        "L_retorno_lampara (m)": L_retorno_lampara,
        "L_fase_caja_primer_int (m)": L_fase_caja_primer_int,
        "L_troncal_primera_oct_924 (m)": L_troncal_primera_oct_924,

        "L_caja_int_fase_ida (m)": L_caja_int_fase_ida,
        "L_troncal_oct1 (m)": L_troncal_oct1,
        "L_oct1_oct2 (m)": L_oct1_oct2,
        "Grupos_interruptor_ilum": grupos_ilum_detalle,

        "Detalle enchufes comunes": detalle_enchufes,
        "Potencia enchufes (W)": potencia_total_ench,
        "Componentes especiales": nombres_componentes,
        "Potencia comp. especiales (W)": potencia_total_comp,
        "Total por ambiente (W)": potencia_total_ambiente,
        "Cantidad luminarias (u)": n_luminarias,
        "Cantidad enchufes (u)": n_enchufes,
        "Cantidad módulos enchufe (u)": n_modulos_enchufe,
        "Cantidad comp. especiales (u)": n_componentes_especiales,
        "Enchufes simples 10A (u)": ench_simple_10A,
        "Enchufes dobles 10A (u)": ench_doble_10A,
        "Enchufes triples 10A (u)": ench_triple_10A,
        "Enchufes simples 10/16A (u)": ench_simple_1016,
        "Enchufes dobles 10/16A (u)": ench_doble_1016,
        "Enchufes triples 10/16A (u)": ench_triple_1016,
    })

# Elementos restantes para repartir entre circuitos
# (a medida que se van armando los circuitos, se van sacando de acá los que ya se usaron)
enchufes_restantes = {k.lower(): v[:] for k, v in enchufes_por_ambiente.items()}
luminarias_restantes = {k.lower(): v[:] for k, v in luminarias_por_ambiente.items()}
componentes_restantes = {k.lower(): v[:] for k, v in componentes_por_ambiente.items()}

# -------- PARTE 2: DATOS ADICIONALES DEL SISTEMA --------
print("\n Ahora se solicitarán datos generales del sistema")

area_total_vivienda = sum(a["Área (m²)"] for a in ambientes_detalle)
min_circuitos = 2 if area_total_vivienda < 30 else 3  # casas más grandes necesitan más circuitos mínimos

print(f"\n Superficie total aproximada de la vivienda: {area_total_vivienda:.2f} m²")
print(f" Según criterio, se requieren al menos {min_circuitos} circuitos.")

zona = input("- Ingrese las características de la zona (húmeda, seca): ")
while True:
    tipo_canalizacion = input("- Ingrese el tipo de canalización (embutida, sobrepuesta): ").strip().lower()
    if "embut" in tipo_canalizacion or "sobre" in tipo_canalizacion:
        break
    print("   ! Debe ingresar 'embutida' o 'sobrepuesta'.")
material_forrado_exterior = input("- Material del forrado exterior de la casa (Fibrocemento, Madera, Siding PVC, Siding metálico): ").strip().lower()

while True:
    try:
        cantidad_circuitos = int(input("- Ingrese la cantidad de circuitos (máx. 10): "))
        if cantidad_circuitos < min_circuitos:
            print(f"   ! La vivienda requiere al menos {min_circuitos} circuitos.")
            continue
        if cantidad_circuitos > 10:
            print("   ! El máximo permitido es 10 circuitos.")
            continue
        break
    except:
        print("   ! Ingrese un número entero válido.")

amb_idx = {a["Ambiente"].lower(): a for a in ambientes_detalle}  # para buscar rápido un ambiente por nombre
nombres_amb = [a["Ambiente"] for a in ambientes_detalle]

def normaliza_seleccion_ambientes(texto, disponibles):
    # toma lo que escribió el usuario (ambientes separados por coma) y
    # deja solo los nombres que existen de verdad, sin repetir
    disp_lower = {d.lower(): d for d in disponibles}  # relaciona el nombre en minúscula con el nombre real (con mayúsculas)
    salida = []
    for t in [s.strip() for s in texto.split(",") if s.strip()]:  # separa por coma y limpia espacios
        key = t.lower()
        if key in disp_lower and disp_lower[key] not in salida:  # existe y no está repetido
            salida.append(disp_lower[key])
    return salida

circuitos = []  # lista global donde se van guardando todos los circuitos creados

def add_circuito(nombre_circ, longitud, potencia_estimada, es_ilumin, es_enchufe, es_especial, items=None, detalle_asignacion="", tiene_tramo_20m=True):
    # agrega un circuito nuevo a la lista global "circuitos", que después
    # se convierte en circuitos_df y se usa en build_materiales_df()
    # tiene_tramo_20m: si el circuito NO tiene ningún tramo continuo >=20m
    # (aunque la longitud TOTAL del circuito sea mayor a 20m), no debe
    # contarse ninguna caja de paso (RIC 7.16.1.13) — por defecto True para
    # no romper los llamados que todavía no preguntan esto explícitamente.
    circuitos.append({
        "Circuito": nombre_circ,
        "Longitud (m)": float(longitud),
        "Potencia estimada (W)": round(float(potencia_estimada), 1),
        "Detalle asignación": detalle_asignacion,
        "es_ilumin": es_ilumin,
        "es_enchufe": es_enchufe,
        "tipo_dif": "especial" if es_especial else "general",  # define qué tipo de diferencial le corresponde
        "_items": items[:] if items else [],  # copia de los items (luminarias/enchufes) del circuito
        "_cajas_adic": _cajas_adic if es_enchufe else 0,  # cajas de derivación extra, solo aplica a enchufes
        "_tiene_tramo_20m": bool(tiene_tramo_20m),
    })

# cocina/lavadero por ambiente (cocina 1, cocina 2, etc.)
enchufes_coc_lav_por_amb = {}   # {"cocina 1":[...], "cocina 2":[...], "lavadero":[...]}
long_real_coc_lav_por_amb = {}  # {"cocina 1": 25.0, "cocina 2": 18.0, ...}
coc_lav_creado_manual = set()


CALIBRES_TM_CLIMA = [6, 10, 16, 20, 25, 32, 40, 50, 63]

def siguiente_calibre_clima(corriente):
    # busca el calibre de TM comercial más chico que aguante esta corriente
    for cal in CALIBRES_TM_CLIMA:
        if cal >= corriente:
            return cal  # este calibre ya aguanta la corriente
    return CALIBRES_TM_CLIMA[-1]  # ninguno alcanzó, se usa el más grande

def calcular_circuito_climatizacion(datos):
    """
    Calcula conductor, termomagnético y diferencial según RIC N°07.
    Lógica TM:
      - Conductor: I_max × 1.25 (RIC 7.3.4)
      - Curva: C para inverter (arranque suave, 5-10× In) / D para on-off
        (arranque fuerte, 10-20× In) — RIC N°10
      - TM: calibre sobre I_diseño, luego verificado contra LRA (RIC 5.6.2.2)
        Si el LRA real viene de la placa del equipo, se usa directo
        Si no hay LRA, se estima según el tipo de compresor on/off
        Si es inverter, no hay LRA, la curva C es suficiente
      - Si el LRA supera el umbral magnético del TM elegido (según su curva), se sube al siguiente calibre
    """
    V             = float(datos.get("tension", 220))
    I_max         = float(datos.get("corriente_maxima", 0))       # corriente máxima de placa (si la tiene)
    I_nom         = float(datos.get("corriente_nominal", 0))      # corriente nominal de placa
    P_nom_w       = float(datos.get("potencia_nominal_w", 0))     # potencia nominal, por si no hay corriente
    fp            = float(datos.get("factor_potencia", 0.95)) or 0.95
    tecnologia    = str(datos.get("tecnologia", "inverter")).lower()  # inverter u on/off
    lra           = datos.get("lra")   # corriente de arranque real (si el fabricante la entrega)
    mocp          = datos.get("mocp")  # protección máxima que indica el fabricante (si la entrega)
    tipo_compresor = str(datos.get("tipo_compresor", "rotativo")).lower()
    temp_sobre_30 = datos.get("temp_sobre_30", False)  # si el equipo queda expuesto a más de 30°C
    temp_ambiente = datos.get("temp_ambiente", 25.0)

    # --- Paso 1: corriente base desde placa (RIC 5.3.1) ---
    I_base = I_max if I_max > 0 else (P_nom_w / (V * fp) if P_nom_w > 0 else I_nom)

    # --- Paso 2: corriente de diseño conductor (RIC 7.3.4) ---
    I_diseno = I_base * 1.25

    # --- Paso 3: factor corrección temperatura (RIC N°4 art. 6.2.6 / Tabla N°4.7) ---
    # Fórmula RIC: Ic = Iz × ft  (ft MULTIPLICA sobre Iz, NO divide I_diseno)
    # La verificación se hace más abajo, en la PARTE 3.2: Ic = Iz × ft >= I_diseno
    if temp_sobre_30:
        ft = factor_temperatura_ft(temp_ambiente, "B1")
    else:
        ft = 1.0
    I_diseno_corr = I_diseno  # NO se divide por ft — ft va sobre Iz en verificación

    # --- Paso 4: sección mínima conductor ---
    seccion = 2.5  # mínimo RIC 7.3.4; más abajo, en la PARTE 3.2, se puede subir por caída de tensión

    # --- Paso 5: termomagnético ---
    # 5a. Si el fabricante indica un MOCP, ese valor no se puede superar (RIC 5.6.2.3)
    if mocp and float(mocp) > 0:
        In_tm = siguiente_calibre_clima(float(mocp))
        curva_tm = "C" if tecnologia == "inverter" else "D"
        nota_tm = f"Protección máxima indicada por fabricante: {float(mocp)} A → calibre comercial {In_tm} A (RIC 5.6.2.3)"
        aviso_lra = "Protección máxima según fabricante"
        # Diferencial y retorno directo
        In_dif = 25 if In_tm <= 25 else (40 if In_tm <= 40 else 63)
        I_cuadro = I_max if I_max > 0 else (I_nom if I_nom > 0 else (P_nom_w / (V * fp) if P_nom_w > 0 else I_base))
        return {
            "I_base_A": round(I_base, 2), "I_diseno_A": round(I_diseno_corr, 2),
            "seccion_mm2": seccion, "In_tm_A": In_tm, "curva_tm": curva_tm,
            "nota_tm": nota_tm, "aviso_lra": aviso_lra, "In_dif_A": In_dif,
            "I_cuadro_A": round(I_cuadro, 2), "ft": ft, "temp_sobre_30": temp_sobre_30,
        }

    # 5b. Calcular TM inicial sobre corriente de diseño
    In_tm = siguiente_calibre_clima(I_diseno_corr)
    # Inverter: arranque suave, no tiene LRA real, curva C directo.
    # On/Off: se calcula el LRA y se prueba primero con curva C (5-10× In,
    # dispara más rápido); si el LRA no cabe en ese umbral, se pasa a curva D
    # (10-20× In) antes de subir de calibre — RIC N°10.
    curva_tm = "C" if tecnologia == "inverter" else None  # on/off se define más abajo

    # --- Paso 6: verificación LRA (RIC 5.6.2.2) ---
    # Curva C: disparo magnético instantáneo entre 5× y 10× In_tm
    # Curva D: disparo magnético instantáneo entre 10× y 20× In_tm
    aviso_lra = ""

    if tecnologia == "inverter":
        # Inverter: arranque progresivo, no existe LRA real
        aviso_lra = "Inverter — arranque progresivo"

    else:
        # On/off: existe corriente de arranque real
        if lra and float(lra) > 0:
            # LRA real desde placa
            lra_val = float(lra)
            fuente_lra = f"LRA real de placa: {lra_val} A"
        else:
            # Estimar LRA según tipo de compresor (RIC 5.6.2.2 — criterio conservador)
            if "piston" in tipo_compresor or "pistón" in tipo_compresor:
                factor_lra = 7.0  # compresor de pistón: LRA típico 6-8× I_nom
                fuente_lra = f"LRA estimado (compresor pistón, factor 7×): {round(I_nom * factor_lra, 1)} A"
            else:
                factor_lra = 5.5  # compresor rotativo: LRA típico 4-6× I_nom
                fuente_lra = f"LRA estimado (compresor rotativo, factor 5.5×): {round(I_nom * factor_lra, 1)} A"
            lra_val = round(I_nom * factor_lra, 1)

        # Se prueba primero curva C (dispara más rápido, más protección) y
        # solo si el LRA no cabe en su umbral se pasa a curva D — subiendo de
        # calibre dentro de cada curva antes de descartarla.
        curva_tm = None
        for curva_candidata in ("C", "D"):
            In_tm_prueba = In_tm
            intentos = 0
            while intentos < len(CALIBRES_TM_CLIMA):
                umbral_max_c = In_tm_prueba * (20 if curva_candidata == "D" else 10)
                if lra_val <= umbral_max_c:
                    curva_tm = curva_candidata
                    In_tm = In_tm_prueba
                    break
                idx_actual = CALIBRES_TM_CLIMA.index(In_tm_prueba) if In_tm_prueba in CALIBRES_TM_CLIMA else 0
                if idx_actual < len(CALIBRES_TM_CLIMA) - 1:
                    In_tm_prueba = CALIBRES_TM_CLIMA[idx_actual + 1]
                    intentos += 1
                else:
                    break
            if curva_tm:
                break

        if curva_tm:
            subio = In_tm != siguiente_calibre_clima(I_diseno_corr)
            aviso_lra = f"{fuente_lra}" + (f" | TM subido a {In_tm} A por corriente de arranque" if subio else "")
        else:
            # Ni curva C ni D, en ningún calibre disponible, aguantan el LRA
            curva_tm = "D"
            aviso_lra = (f"AVISO: {fuente_lra} — supera el umbral incluso en el TM más grande "
                         f"disponible ({In_tm} A, curva {curva_tm}) — revisar manualmente")

    # --- Paso 7: diferencial exclusivo (RIC 7.4.5) ---
    In_dif = 25 if In_tm <= 25 else (40 if In_tm <= 40 else 63)

    # --- Paso 8: corriente cuadro de cargas ---
    # Se usa I_max (corriente absorbida máxima de placa) cuando está disponible
    # porque es el valor real de operación que define la demanda (RIC 7.3.4)
    I_cuadro = I_max if I_max > 0 else (I_nom if I_nom > 0 else (P_nom_w / (V * fp) if P_nom_w > 0 else I_base))

    nota_tm = (
        f"Curva {curva_tm} — arranque suave inverter (RIC 5.6.2.1)" if tecnologia == "inverter"
        else f"Curva {curva_tm} — verificado contra LRA {'real' if lra and float(lra)>0 else 'estimado'} (RIC 5.6.2.2)"
    )

    return {
        "I_base_A":      round(I_base, 2),
        "I_diseno_A":    round(I_diseno_corr, 2),
        "seccion_mm2":   seccion,
        "In_tm_A":       In_tm,
        "curva_tm":      curva_tm,
        "nota_tm":       nota_tm,
        "aviso_lra":     aviso_lra,
        "In_dif_A":      In_dif,
        "I_cuadro_A":    round(I_cuadro, 2),
        "ft":            ft,
        "temp_sobre_30": temp_sobre_30,
    }

def ingresar_equipo_climatizacion_inline(nombre_sugerido, ambientes_str, longitud):
    """
    Flujo de preguntas de climatización integrado en la parte que recorre los circuitos.
    nombre_sugerido, ambientes_str y longitud ya vienen de ahí.
    """
    nombre_circ = nombre_sugerido

    # pregunta si el equipo es inverter u on/off, porque cambia toda la lógica de cálculo
    while True:
        tec = input(
            "   - Tipo de tecnología del equipo:\n"
            "     (1) Inverter\n"
            "     (2) On/Off\n"
            "   Opción: "
        ).strip()
        if tec == "1":
            tecnologia = "inverter"
            break
        elif tec == "2":
            tecnologia = "on/off"
            break
        print("     ! Ingrese 1 o 2.")

    print("\n   Ingrese los datos de la PLACA o FICHA TÉCNICA del equipo")
    print("   (use el modo FRÍO que es el peor caso eléctrico):")

    I_nom = pedir_float_positivo(
        "   - Corriente nominal de placa — modo frío [A]: "
    )
    I_max = pedir_float_positivo(
        "   - Corriente máxima de placa — modo frío [A] "
        "(es la más alta que entrega la ficha): "
    )
    P_nom_w = pedir_float_positivo(
        "   - Potencia absorbida nominal — modo frío [W]: "
    )

    print("\n   Factor de potencia:")
    print("   (Si no tiene el dato en la placa, ingrese 0.95 como valor típico)")
    fp = pedir_float_positivo("   - Factor de potencia (ej: 0.95): ")

    # LRA y tipo de compresor solo para on/off
    lra = None
    tipo_compresor = "rotativo"  # default
    if tecnologia == "on/off":
        # los inverter no necesitan esto (arranque progresivo, no hay LRA real)
        print("\n   Tipo de compresor (para estimar corriente de arranque):")
        while True:
            tc = input(
                "   - Tipo de compresor:\n"
                "     (1) Rotativo — splits residenciales modernos (LRA típico 4-6× I_nom)\n"
                "     (2) Pistón — equipos más antiguos o industriales (LRA típico 6-8× I_nom)\n"
                "   Opción: "
            ).strip()
            if tc == "1":
                tipo_compresor = "rotativo"
                break
            elif tc == "2":
                tipo_compresor = "piston"
                break
            print("     ! Ingrese 1 o 2.")

        print("\n   Corriente de arranque del compresor (LRA):")
        print("   Busque en la placa de la unidad EXTERIOR el campo LRA o")
        print("   'Locked Rotor Amps'. Si no aparece, responda no y el")
        print("   programa lo estimará automáticamente.")
        while True:
            tiene_lra = input("   - ¿La placa exterior indica LRA? (si/no): ").strip().lower()
            if tiene_lra in ("si", "no"):
                break
            print("     ! Responda solo si o no.")
        if tiene_lra == "si":
            while True:
                lra = pedir_float_positivo("     • Ingrese el valor LRA de la placa [A]: ")
                # Validación física: LRA siempre debe ser mayor que I_nom
                # (LRA es la corriente con rotor trabado, siempre mayor que la de operación)
                if lra <= I_nom:
                    print(f"     ! El LRA ({lra} A) no puede ser menor o igual a la")
                    print(f"       corriente nominal ({I_nom} A). Esto indica que el dato")
                    print(f"       ingresado no es el LRA sino otro parámetro (RLA, FLA, etc).")
                    print(f"       Revise la placa y vuelva a ingresar, o responda no.")
                    while True:
                        reintentar = input("     • ¿Desea reingresar el LRA? (si/no): ").strip().lower()
                        if reintentar in ("si", "no"):
                            break
                    if reintentar == "no":
                        lra = None
                        print("     • Se usará estimación automática según tipo de compresor.")
                        break
                    # si reintentar == "si", no se hace nada más acá: vuelve a subir a la
                    # pregunta "Ingrese el valor LRA de la placa" y la repite
                else:
                    break  # LRA válido

    # MOCP para ambos tipos
    print("\n   Protección máxima recomendada por el fabricante:")
    print("   (MOCP = protección de sobrecorriente máxima — a veces indicada en placa)")
    print("   (Si no aparece en la placa ni en la ficha, responda no)")
    tiene_mocp = input("   - ¿La placa o ficha indica un valor MÁXIMO de protección? (si/no): ").strip().lower()
    mocp = None
    if tiene_mocp == "si":
        mocp = pedir_float_positivo(
            "     • Ingrese ese valor máximo de protección [A]: "
        )

    # Temperatura (RIC 7.5.2 — factor corrección si supera 30°C)
    temp_sobre_30 = False
    temp_ambiente = 25.0  # valor por defecto
    print("\n   Temperatura ambiente del lugar de instalación:")
    t_amb = input(
        "   - ¿La temperatura ambiente del recinto supera los 30°C? (si/no)\n"
        "     (RIC 7.5.2 exige corrección si supera 30°C): "
    ).strip().lower()
    if t_amb == "si":
        temp_sobre_30 = True
        while True:
            try:
                temp_ambiente = float(input(
                    "   - Ingrese la temperatura ambiente máxima del recinto (°C): "
                ).strip().replace(",", "."))
                if temp_ambiente > 30:
                    break
                print("     ! Debe ser mayor que 30°C.")
            except:
                print("     ! Ingrese un número válido.")

    datos_placa = {
        "tension":           220.0,
        "corriente_nominal":  I_nom,
        "corriente_maxima":   I_max,
        "potencia_nominal_w": P_nom_w,
        "factor_potencia":    fp,
        "tecnologia":         tecnologia,
        "tipo_compresor":     tipo_compresor,
        "lra":                lra,
        "mocp":               mocp,
        "temp_sobre_30":      temp_sobre_30,
        "temp_ambiente":      temp_ambiente,
    }  # junta todo lo que se preguntó, para pasárselo al cálculo

    resultado = calcular_circuito_climatizacion(datos_placa)  # acá se calcula TM, diferencial y sección

    # junta todos los datos del equipo en un solo paquete, listo para
    # usarse en build_materiales_df
    return {
        "nombre_circ":   nombre_circ,
        "ambientes_str": ambientes_str,
        "longitud":      longitud,
        "P_nom_w":       P_nom_w,
        "fp":            fp,
        "I_cuadro":      resultado["I_cuadro_A"],
        "I_diseno":      resultado["I_diseno_A"],
        "In_tm":         resultado["In_tm_A"],
        "curva_tm":      resultado["curva_tm"],
        "In_dif":        resultado["In_dif_A"],
        "seccion_mm2":   resultado["seccion_mm2"],
        "nota_tm":       resultado["nota_tm"],
        "aviso_lra":     resultado["aviso_lra"],
        "tecnologia":    tecnologia,
        "temp_sobre_30": temp_sobre_30,
    }


circuitos_climatizacion = []  # se va llenando más abajo, mientras se recorren los circuitos

# =========================================================
# AGUA CALIENTE — lista de equipos (análoga a circuitos_climatizacion)
# =========================================================
circuitos_agua_caliente = []  # se va llenando más abajo, mientras se recorren los circuitos


def calcular_circuito_agua_caliente(datos):
    """
    Calcula TM, diferencial y corriente de diseño para equipos de agua
    caliente (duchas eléctricas, termoelectros, calefones eléctricos).

    Reglas normativas aplicadas:
      - Conductor: Iz ≥ I_nom  (carga resistiva pura — sin factor 1.25)
        El factor 1.25 del RIC N°7 art. 7.3.4 aplica a motores y cargas
        de arranque difícil, no a cargas resistivas.
      - Sección mínima: 2.5 mm²       (RIC N°07 art. 7.3.4)
      - fp = 1.0 (carga resistiva pura — sin componente reactiva)
      - Factor temperatura si T > 30°C (RIC N°07 art. 7.5.2 / RIC N°04 6.2.5)
      - TM: calibre comercial sobre I_diseño_corr, Curva C
      - Diferencial:
            ≤ 10 mA si equipo en Volumen 1 (interior ducha) — RIC N°11 art. 6.4.3
            ≤ 30 mA en cualquier otro caso                  — RIC N°07 art. 7.4.5
      - Calibre diferencial ≥ calibre TM
    """
    V             = float(datos.get("tension", 220))
    P_nom_w       = float(datos.get("potencia_nominal_w", 0))
    temp_sobre_30 = datos.get("temp_sobre_30", False)
    temp_ambiente = float(datos.get("temp_ambiente", 25.0))
    vol1_bano     = datos.get("vol1_bano", False)   # si es True, va diferencial de 10 mA

    # --- Corriente nominal (fp=1.0, carga resistiva) ---
    I_nom = P_nom_w / V if V > 0 else 0.0

    # --- Corriente de diseño: si la carga es resistiva pura, I_diseño = I_nom
    # El factor 1.25 del RIC N°7 7.3.4 aplica a motores y cargas de arranque
    # difícil, NO a cargas resistivas (ducha, termoelectro, calefón).
    I_diseno = I_nom

    # --- Factor corrección temperatura (RIC N°4 art. 6.2.6 / Tabla N°4.7) ---
    # Fórmula RIC: Ic = Iz × ft  (ft MULTIPLICA sobre Iz, NO divide I_diseno)
    if temp_sobre_30:
        ft = factor_temperatura_ft(temp_ambiente, "B1")
    else:
        ft = 1.0
    I_diseno_corr = I_diseno  # NO se divide por ft

    # --- TM: seleccionar calibre tal que Iz_TM × ft >= I_diseno ---
    # Equivalente a: In_TM >= I_diseno / ft  (referencia para buscar calibre)
    _ref_tm = (I_diseno / ft) if ft > 0 else I_diseno
    In_tm = siguiente_calibre_clima(_ref_tm)   # primer calibre tal que Iz_TM × ft >= I_diseno

    # --- Diferencial: calibre ≥ TM, sensibilidad según volumen ---
    sensibilidad_dif = "10mA" if vol1_bano else "30mA"
    CALIBRES_DIF = [16, 25, 40, 63]
    In_dif = next((c for c in CALIBRES_DIF if c >= In_tm), CALIBRES_DIF[-1])  # primer calibre comercial que alcanza al TM

    return {
        "I_nom_A":        round(I_nom, 2),
        "I_diseno_A":     round(I_diseno_corr, 2),
        "In_tm":          In_tm,
        "curva_tm":       "C",
        "In_dif":         In_dif,
        "sensibilidad_dif": sensibilidad_dif,
        "ft":             ft,
        "temp_sobre_30":  temp_sobre_30,
    }


def ingresar_equipo_agua_caliente_inline(nombre_sugerido, ambientes_str, longitud):
    """
    Flujo de preguntas para circuitos de agua caliente.
    Retorna todos los datos necesarios juntos en un solo paquete, o None si hay error.
    """
    nombre_circ = nombre_sugerido
    amb_lower   = ambientes_str.lower()
    es_bano     = any(k in amb_lower for k in ("baño", "bano", "bathroom"))  # detecta si el ambiente es un baño

    # 1. Tipo de equipo
    print(
        "\n   Tipo de equipo:"
        "\n     (1) Ducha eléctrica"
        "\n     (2) Termoelectro"
        "\n     (3) Calefón eléctrico"
        "\n     (4) Otro calentador de agua"
    )
    while True:
        tipo = input("   Opción: ").strip()
        if tipo in ("1", "2", "3", "4"):
            break
        print("     ! Ingrese 1, 2, 3 o 4.")
    tipo_map   = {"1": "Ducha eléctrica", "2": "Termoelectro",
                  "3": "Calefón eléctrico", "4": "Calentador de agua"}
    tipo_equipo = tipo_map[tipo]

    # Validar: ducha eléctrica solo en baño
    if tipo == "1" and not es_bano:
        print(f"\n   ! La ducha eléctrica SOLO puede instalarse en un BAÑO.")
        print(f"     (RIC N°11 art. 6 + RIC N°07 art. 7.4.5)")
        print(f"     Ambiente indicado: '{ambientes_str}'. Corrija e intente de nuevo.")
        return None  # no cumple la norma, se cancela el ingreso

    # 2. Potencia nominal
    while True:
        try:
            P_nom_w = float(
                input("\n   - Potencia del equipo [W] (ej: 3300, 4400, 5500): ")
                .strip().replace(",", ".")
            )
            if P_nom_w > 0:
                break
            print("     ! Debe ser mayor que 0.")
        except:
            print("     ! Ingrese un número válido.")

    # 3. Volumen RIC N°11 (solo para equipos en baño que no sean ducha eléctrica)
    vol1_bano = False
    if tipo == "1":
        # Ducha eléctrica: siempre Volumen 1, siempre diferencial de 10 mA
        vol1_bano = True
    elif es_bano:
        # está en un baño pero no es ducha: hay que preguntar en qué volumen queda
        print(
            "\n   ¿El equipo está instalado dentro del Volumen 1?"
            "\n   (Vol.1 = dentro del perímetro de la ducha/bañera hasta 2,25m de altura)"
            "\n   → SÍ: diferencial 10 mA  |  NO: diferencial 30 mA"
        )
        while True:
            resp = input("   (si/no): ").strip().lower()
            if resp in ("si", "no"):
                vol1_bano = (resp == "si")
                break
            print("     ! Responda si o no.")

    # 4. Tablero externo de desconexión
    print("\n   ¿El equipo trae interruptor incorporado?")
    print("   → SÍ: sin tablero externo  |  NO: se genera tablero externo de desconexión")
    while True:
        resp_tab = input("   (si/no): ").strip().lower()
        if resp_tab in ("si", "no"):
            lleva_tablero_externo = (resp_tab == "no")
            break
        print("     ! Responda si o no.")

    # 5. Temperatura ambiente
    print("\n   ¿La temperatura del recinto supera los 30°C?")
    while True:
        t_amb = input("   (si/no): ").strip().lower()
        if t_amb in ("si", "no"):
            break
        print("     ! Responda si o no.")
    temp_sobre_30 = False
    temp_ambiente = 25.0
    if t_amb == "si":
        temp_sobre_30 = True
        while True:
            try:
                temp_ambiente = float(
                    input("   - Temperatura máxima [°C]: ").strip().replace(",", ".")
                )
                if temp_ambiente > 30:
                    break
                print("     ! Debe ser mayor que 30°C.")
            except:
                print("     ! Ingrese un número válido.")

    # Cálculo eléctrico
    resultado = calcular_circuito_agua_caliente({
        "tension":            220.0,
        "potencia_nominal_w": P_nom_w,
        "temp_sobre_30":      temp_sobre_30,
        "temp_ambiente":      temp_ambiente,
        "vol1_bano":          vol1_bano,
    })

    # junta todos los datos del equipo en un solo paquete, listo para
    # usarse en build_materiales_df
    return {
        "nombre_circ":          nombre_circ,
        "tipo_equipo":          tipo_equipo,
        "ambientes_str":        ambientes_str,
        "longitud":             longitud,
        "P_nom_w":              P_nom_w,
        "fp":                   1.0,
        "I_cuadro":             resultado["I_nom_A"],
        "I_diseno":             resultado["I_diseno_A"],
        "In_tm":                resultado["In_tm"],
        "curva_tm":             "C",
        "In_dif":               resultado["In_dif"],
        "sensibilidad_dif":     resultado["sensibilidad_dif"],
        "ft":                   resultado["ft"],
        "temp_sobre_30":        temp_sobre_30,
        "vol1_bano":            vol1_bano,
        "lleva_tablero_externo": lleva_tablero_externo,
    }
# -------- PARTE 2.1: CONSTRUCCIÓN DE CIRCUITOS --------
# recorre uno por uno los circuitos que pidió el usuario, preguntando su
# nombre, ambientes, longitud, y detectando de qué tipo es (iluminación,
# enchufes, climatización, agua caliente, o especial genérico)
for i in range(1, cantidad_circuitos + 1):

    base = input(
        f"\n   - Ingrese el nombre del circuito N°{i} "
        "(Ejemplos: iluminacion, enchufes generales, cocina-encimera, baño-lavadora, etc.): "
    ).strip()

    base = limpiar_nombre_circuito(base)
    base = sugerir_nombre_circuito(base)  # corrige errores de tipeo típicos (ej: deja "iluminacon" como "iluminacion")
    base_lower = base.lower()

    print("     • Ambientes disponibles:", ", ".join(nombres_amb))

    while True:
        sel = input("     • Escriba los ambientes que componen el circuito (ej: Living, Cocina): ").strip()
        sel_list = normaliza_seleccion_ambientes(sel, nombres_amb)
        if sel_list:
            break
        print("       ! Debe ingresar al menos un ambiente válido (respetando nombres).")

    ambientes_str = ", ".join(sel_list)

    # Detectar tipo de circuito ANTES de preguntar longitud para ajustar el texto
    _base_lower_prev = base.lower()
    _es_agua_prev = any(k in _base_lower_prev for k in
                        ("ducha", "termo", "termoelectro", "calefon", "calefón",
                         "calentador", "agua caliente"))

    if len(sel_list) == 1:
        # un solo ambiente: la longitud es "real", conocida con precisión
        if _es_agua_prev:
            longitud = float(input(
                f"     • Longitud del circuito '{base} ({ambientes_str})' en metros\n"
                f"       (desde el tablero principal hasta el equipo): "
            ))
        else:
            longitud = float(input(
                f"     • Ingrese la longitud REAL de la canalización en (m) del circuito '{base} ({ambientes_str})'\n"
                f"       (En tramos horizontales: recorridos a 0,30m del cielo y 0,20m del piso - RIC N°4 7.16.1.16): "
            ))
        longitud_unica = True
    else:
        # varios ambientes: la longitud es una estimación (recorre varios lugares)
        if _es_agua_prev:
            longitud = float(input(
                f"     • Longitud del circuito '{base} ({ambientes_str})' en metros\n"
                f"       (desde el tablero principal hasta el equipo): "
            ))
        else:
            longitud = float(input(
                f"     • Ingrese la longitud REAL de la canalización en (m) del circuito '{base} ({ambientes_str})'\n"
                f"       (En tramos horizontales: recorridos a 0,30m del cielo y 0,20m del piso - RIC N°4 7.16.1.16): "
            ))
        longitud_unica = False

    # Cajas de paso (RIC 7.16.1.13): solo se consideran si el circuito tiene
    # un tramo CONTINUO >=20m — no basta con que la longitud total del
    # circuito supere 20m (puede ser la suma de varios tramos cortos).
    # Se pregunta solo cuando la longitud total ya da para sospechar que
    # podría haber un tramo así (si es menor a 20m, es matemáticamente
    # imposible que exista un tramo continuo >=20m dentro de ese circuito).
    tiene_tramo_20m = False
    if longitud >= 20:
        _resp_tramo20 = input(
            f"     • ¿Existe algún tramo continuo dentro del circuito '{base} ({ambientes_str})' "
            f"mayor o igual a 20m? (si/no): "
        ).strip().lower()
        tiene_tramo_20m = _resp_tramo20 in ("si", "sí", "s", "y", "yes")

    # detecta el tipo de circuito buscando palabras clave en el nombre
    es_ilumin = "ilumin" in base_lower
    es_enchufe = "enchufe" in base_lower
    es_clima_circ = any(k in base_lower for k in ("clima", "aire", "split", "ac ", "a/c"))
    es_agua_circ  = any(k in base_lower for k in
                        ("ducha", "termo", "termoelectro", "calefon", "calefón",
                         "calentador", "agua caliente"))

    # ---- CIRCUITO DE CLIMATIZACIÓN detectado por nombre ----
    if es_clima_circ:
        datos_clima = ingresar_equipo_climatizacion_inline(
            nombre_sugerido=base,
            ambientes_str=ambientes_str,
            longitud=longitud
        )
        circuitos_climatizacion.append(datos_clima)
        items_clima = [{"amb": datos_clima["ambientes_str"], "potencia": float(datos_clima["P_nom_w"]), "nombre": datos_clima["nombre_circ"]}]
        # Formato detalle igual al resto: "ambiente: descripción"
        _amb_lower = datos_clima["ambientes_str"].lower()
        _lra_txt = f" | {datos_clima['aviso_lra']}" if datos_clima["aviso_lra"] else ""
        _detalle_clima = (
            f"{_amb_lower}: climatización {datos_clima['tecnologia'].upper()} "
            f"({datos_clima['I_cuadro']}A nominal{_lra_txt})"
        )
        add_circuito(
            nombre_circ       = datos_clima["nombre_circ"] + " (" + datos_clima["ambientes_str"] + ")",
            longitud          = datos_clima["longitud"],
            potencia_estimada = datos_clima["P_nom_w"],
            es_ilumin         = False,
            es_enchufe        = False,
            es_especial       = True,
            items             = items_clima,
            detalle_asignacion = _detalle_clima
        )
        circuitos[-1]["Disyuntor termomagnético"] = f"1x{datos_clima['In_tm']}A / 6kA / Curva {datos_clima['curva_tm']}"
        circuitos[-1]["_In_TM"]            = datos_clima["In_tm"]
        circuitos[-1]["_In_dif_clima"]     = datos_clima["In_dif"]
        circuitos[-1]["_es_climatizacion"] = True
        circuitos[-1]["_I_diseno_clima"]   = datos_clima["I_diseno"]
        circuitos[-1]["Corriente estimada (A)"] = round(float(datos_clima["I_cuadro"]), 2)
        continue  # este circuito ya quedó armado, pasa al siguiente

    # ---- CIRCUITO DE AGUA CALIENTE detectado por nombre ----
    elif es_agua_circ:
        # ── Validación de ambientes ──────────────────────────────────────────
        # Validación previa: rechaza ambientes claramente inválidos para
        # cualquier equipo de agua caliente (living, dormitorio, etc.)
        # La validación específica de que la ducha solo va en baño se hace dentro de
        # ingresar_equipo_agua_caliente_inline una vez que el usuario elige el tipo.
        _AMBIENTES_INVALIDOS_AGUA = (
            "living", "dormitorio", "dorm", "bedroom",
            "comedor", "dining", "estar", "sala",
            "hall", "estudio", "oficina", "terraza",
            "lounge", "studio", "office",
        )

        _ambs_invalidos_encontrados = []
        for _amb_check in sel_list:
            _amb_low = _amb_check.lower().strip()
            if any(k in _amb_low for k in _AMBIENTES_INVALIDOS_AGUA):
                _ambs_invalidos_encontrados.append(_amb_check)

        if _ambs_invalidos_encontrados:
            print(f"\n   ! ADVERTENCIA: Los siguientes ambientes NO son válidos")
            print(f"     para instalar equipos de agua caliente:")
            for _a in _ambs_invalidos_encontrados:
                print(f"       - {_a}")
            print(f"   ! Los equipos de agua caliente solo pueden instalarse en:")
            print(f"     baño, lavadero, cocina, bodega, pasillo de servicio o exterior.")
            print(f"   ! La ducha eléctrica SOLO puede instalarse en baño.")
            print(f"   ! Corrija el nombre del circuito o los ambientes e intente de nuevo.")
            continue  # vuelve a pedir los datos de este circuito desde el principio
        # ────────────────────────────────────────────────────────────────────
        datos_agua = ingresar_equipo_agua_caliente_inline(
            nombre_sugerido=base,
            ambientes_str=ambientes_str,
            longitud=longitud
        )
        if datos_agua is None:
            # Ambiente inválido para el tipo de equipo: se reinicia este circuito
            continue
        circuitos_agua_caliente.append(datos_agua)
        items_agua = [{"amb":     datos_agua["ambientes_str"],
                       "potencia": float(datos_agua["P_nom_w"]),
                       "nombre":   datos_agua["tipo_equipo"]}]
        _detalle_agua = (
            f"{datos_agua['ambientes_str'].lower()}: {datos_agua['tipo_equipo']} "
            f"({datos_agua['P_nom_w']:.0f} W)"
        )
        add_circuito(
            nombre_circ        = datos_agua["nombre_circ"] + " (" + datos_agua["ambientes_str"] + ")",
            longitud           = datos_agua["longitud"],
            potencia_estimada  = datos_agua["P_nom_w"],
            es_ilumin          = False,
            es_enchufe         = False,
            es_especial        = True,
            items              = items_agua,
            detalle_asignacion = _detalle_agua
        )
        circuitos[-1]["Disyuntor termomagnético"]  = f"1x{datos_agua['In_tm']}A / 6kA / Curva C"
        circuitos[-1]["_In_TM"]                      = datos_agua["In_tm"]
        circuitos[-1]["_In_dif_agua"]                = datos_agua["In_dif"]
        circuitos[-1]["_sensibilidad_dif_agua"]       = datos_agua["sensibilidad_dif"]
        circuitos[-1]["_es_agua_caliente"]            = True
        circuitos[-1]["_I_diseno_agua"]               = datos_agua["I_diseno"]
        circuitos[-1]["_lleva_tablero_externo_agua"]  = datos_agua["lleva_tablero_externo"]
        circuitos[-1]["_vol1_bano_agua"]              = datos_agua["vol1_bano"]
        circuitos[-1]["_tipo_equipo_agua"]            = datos_agua["tipo_equipo"]
        circuitos[-1]["Corriente estimada (A)"]       = round(float(datos_agua["I_cuadro"]), 2)
        continue  # este circuito ya quedó armado, pasa al siguiente

    # ni iluminación, ni enchufes, ni climatización, ni agua caliente:
    # se pregunta directamente si es un circuito especial (horno, lavadora, etc.)
    while True:
        resp = input("     • ¿Este circuito es ESPECIAL? (si/no): ").strip().lower()
        if resp in ["si", "no"]:
            break
        print("       ! Responda solo 'si' o 'no'.")
    es_especial = (resp == "si")

    if es_especial:
        # ---------- CIRCUITO ESPECIAL ----------
        # se le van mostrando al usuario los componentes especiales que ya
        # ingresó por ambiente, para que elija cuáles van en este circuito
        componentes_elegidos = []
        for amb in sel_list:
            key = amb.lower()
            disponibles = componentes_restantes.get(key, [])
            if not disponibles:
                print(f"       ! El ambiente '{amb}' no tiene componentes especiales disponibles.")
                continue

            lista_str = ", ".join([f"{e['id']} - {e['nombre']} ({e['potencia']}W)" for e in disponibles])
            print(f"       Componentes especiales disponibles en '{amb}': {lista_str}")

            while True:
                entrada = input(
                    "       • Ingrese los números de componentes que van en este circuito (ej: 1,3) o 0 si ninguno: "
                ).strip()
                if entrada in ["", "0"]:
                    ids_sel = []
                    break
                try:
                    ids_sel = sorted(set(int(x.strip()) for x in entrada.split(",") if x.strip()))
                except:
                    print("         ! Formato inválido. Use números separados por coma (ej: 1,2,3) o 0.")
                    continue

                ids_disponibles = {e["id"] for e in disponibles}
                if all(idx in ids_disponibles for idx in ids_sel):
                    break
                print("         ! Uno o más números no corresponden a componentes disponibles.")

            if ids_sel:
                # saca los componentes elegidos de la lista de "disponibles", para
                # que no se puedan volver a asignar a otro circuito por error
                asignados = [e for e in disponibles if e["id"] in ids_sel]
                componentes_restantes[key] = [e for e in disponibles if e["id"] not in ids_sel]
                for e in asignados:
                    componentes_elegidos.append({
                        "amb": amb,
                        "id": e["id"],
                        "nombre": e["nombre"],
                        "potencia": e["potencia"]
                    })

        if not componentes_elegidos:
            print("       ! No se eligieron componentes especiales. No se crea este circuito.")
            continue

        # calcula un TM aproximado según la potencia total de los componentes elegidos
        tension_ref = 220.0
        calibres_tm_ref = [6, 10, 16, 20, 25, 32, 40, 50, 63]
        potencia_total_comp = sum(e["potencia"] for e in componentes_elegidos)
        i_est_approx = potencia_total_comp / tension_ref
        i_necesaria = i_est_approx * 1.10  # 10% de margen
        tm_aprox = next((cal for cal in calibres_tm_ref if cal >= i_necesaria), calibres_tm_ref[-1])

        # OJO:
        # Aquí todavía NO se ha calculado el empalme final.
        # Por eso NO se debe comparar tm_aprox con max_sum_tm en este punto.
        # Primero se crean todos los circuitos, luego se calcula la demanda,
        # y recién después se obtiene el empalme calculado y sus protecciones.
        nombre_circ = f"{base} ({ambientes_str})"
        items = [{"amb": e["amb"], "potencia": float(e["potencia"]), "nombre": e["nombre"]} for e in componentes_elegidos]
        detalle = resumen_items_por_ambiente(items, modo="especial")

        add_circuito(
            nombre_circ,
            longitud,
            potencia_total_comp,
            es_ilumin=False,
            es_enchufe=False,
            es_especial=True,
            items=items,
            detalle_asignacion=detalle,
            tiene_tramo_20m=tiene_tramo_20m
        )

    else:
        # ---------- CIRCUITO GENERAL (iluminación o enchufes) ----------
        potencia_estimada = 0.0
        ambientes_str_final = ambientes_str
        items = []

        if es_ilumin:
            # va mostrando las luminarias disponibles de cada ambiente y
            # dejando que el usuario elija cuáles van en este circuito
            for amb in sel_list:
                key = amb.lower()
                disponibles = luminarias_restantes.get(key, [])
                if not disponibles:
                    print(f"       ! El ambiente '{amb}' no tiene luminarias disponibles.")
                    continue

                lista_str = ", ".join([f"{e['id']} ({e['potencia']}W)" for e in disponibles])
                print(f"       Luminarias disponibles en '{amb}': {lista_str}")

                while True:
                    entrada = input(
                        "       • Ingrese los números de luminaria que van en este circuito (ej: 1,3) o 0 si ninguna: "
                    ).strip()
                    if entrada in ["", "0"]:
                        ids_sel = []
                        break
                    try:
                        ids_sel = sorted(set(int(x.strip()) for x in entrada.split(",") if x.strip()))
                    except:
                        print("         ! Formato inválido.")
                        continue

                    ids_disponibles = {e["id"] for e in disponibles}
                    if all(idx in ids_disponibles for idx in ids_sel):
                        break
                    print("         ! Uno o más números no corresponden a luminarias disponibles.")

                if ids_sel:
                    asignados = [e for e in disponibles if e["id"] in ids_sel]
                    potencia_estimada += sum(e["potencia"] for e in asignados)
                    for e in asignados:
                        items.append({
                            "amb": amb,
                            "potencia": float(e["potencia"]),
                            "tipo_lum": str(e.get("tipo", "")).strip().lower(),
                            "desc_lum": str(e.get("desc", "")).strip().lower()
                        })
                    luminarias_restantes[key] = [e for e in disponibles if e["id"] not in ids_sel]  # ya no están disponibles para otro circuito

        elif es_enchufe:
            ambientes_con_enchufe_general = []
            agrego_coc_lav = False
            items_coc_lav = []

            # mismo patrón que luminarias: muestra los enchufes disponibles y
            # deja que el usuario elija cuáles van en este circuito
            for amb in sel_list:
                key = amb.lower()
                disponibles = enchufes_restantes.get(key, [])
                if not disponibles:
                    print(f"       ! El ambiente '{amb}' no tiene enchufes disponibles.")
                    continue

                lista_str = ", ".join([f"{e['id']} ({e['potencia_total']}W)" for e in disponibles])
                print(f"       Enchufes disponibles en '{amb}': {lista_str}")

                while True:
                    entrada = input(
                        "       • Ingrese los números de enchufe que van en este circuito (ej: 1,3) o 0 si ninguno: "
                    ).strip()
                    if entrada in ["", "0"]:
                        ids_sel = []
                        break
                    try:
                        ids_sel = sorted(set(int(x.strip()) for x in entrada.split(",") if x.strip()))
                    except:
                        print("         ! Formato inválido.")
                        continue

                    ids_disponibles = {e["id"] for e in disponibles}
                    if all(idx in ids_disponibles for idx in ids_sel):
                        break
                    print("         ! Uno o más números no corresponden a enchufes disponibles.")

                if ids_sel:
                    asignados = [e for e in disponibles if e["id"] in ids_sel]
                    aporto_general = False

                    for e in asignados:
                        kamb = key.strip().lower()
                        if kamb.startswith("cocina") or kamb.startswith("lavadero"):
                            # acumular por ambiente (cocina 1, cocina 2, lavadero, etc.)
                            if amb not in enchufes_coc_lav_por_amb:
                                enchufes_coc_lav_por_amb[amb] = []
                            enchufes_coc_lav_por_amb[amb].append({
                                "amb": amb,
                                "potencia": float(e["potencia_total"]),
                                "longitud": float(longitud),
                                "id_ench": int(e.get("id", 0)),
                                "modulos": int(e.get("modulos", 1))
                            })
                            agrego_coc_lav = True

                            # también se guarda acá para poder crear el circuito manual si hace falta
                            items_coc_lav.append({
                                "amb": amb,
                                "potencia": float(e["potencia_total"]),
                                "n_ench": 1,
                                "id_ench": int(e.get("id", 0)),
                                "modulos": int(e.get("modulos", 1))
                            })
                        else:
                            # enchufe normal: entra directo a este circuito
                            p = float(e["potencia_total"])
                            potencia_estimada += p
                            items.append({
                                "amb": amb,
                                "potencia": p,
                                "n_ench": 1,
                                "id_ench": int(e.get("id", 0)),
                                "modulos": int(e.get("modulos", 1))
                            })
                            aporto_general = True

                    if aporto_general:
                        ambientes_con_enchufe_general.append(amb)

                    enchufes_restantes[key] = [e for e in disponibles if e["id"] not in ids_sel]

            if ambientes_con_enchufe_general:
                ambientes_str_final = ", ".join(sorted(set(ambientes_con_enchufe_general)))
                _n_amb = len(set(ambientes_con_enchufe_general))
                _n_ench = len([it for it in items if it.get("amb", "") in ambientes_con_enchufe_general])
                if _n_amb <= 1 or _n_ench <= _n_amb:
                    _cajas_adic = 0
                else:
                    _cajas_adic = max(0, _n_amb - 1)  # 1 caja extra por cada ambiente adicional que se conecta
            else:
                _cajas_adic = 0

            if agrego_coc_lav:
                # pedir longitud real por cada cocina/lavadero que entró
                for amb_cl in [a for a in sel_list if a.lower().startswith("cocina") or a.lower().startswith("lavadero")]:
                    if amb_cl not in long_real_coc_lav_por_amb:
                        if len(sel_list) == 1:
                            long_real_coc_lav_por_amb[amb_cl] = longitud
                        else:
                            long_real_coc_lav_por_amb[amb_cl] = pedir_longitud_sub(f"Enchufes {amb_cl}", longitud)
        else:
            # circuito especial "genérico" que no calzó en ninguna categoría anterior:
            # se le asigna toda la potencia del ambiente completo
            for amb in sel_list:
                p = float(amb_idx[amb.lower()]["Total por ambiente (W)"])
                potencia_estimada += p
                items.append({"amb": amb, "potencia": p})

        # permite crear el circuito manualmente si solo tiene enchufes de cocina/lavadero
        if es_enchufe and potencia_estimada == 0 and (len(items) == 0) and (len(items_coc_lav) > 0):
            potencia_estimada = sum(it["potencia"] for it in items_coc_lav)
            items = items_coc_lav[:]
            # marcar cocina/lavadero como creado manualmente
            ambs_cl = set(it["amb"] for it in items_coc_lav)
            for a in ambs_cl:
                coc_lav_creado_manual.add(a.strip().lower())

        # Subdivisión inmediata si supera potencia máxima permitida
        if potencia_estimada > 0:
            tension_ref = 220.0

            # calibres de TM permitidos según el tipo de circuito
            if es_enchufe:
                allowed = [10, 16]
            elif es_ilumin:
                allowed = [6, 10, 16]
            else:
                allowed = [16]

            I_max = allowed[-1]  # el calibre más grande permitido
            P_max = I_max * tension_ref * 0.9  # potencia máxima que soporta ese calibre (con 10% de margen)
            if P_max <= 0:
                P_max = potencia_estimada

            if potencia_estimada > P_max and len(items) > 1:
                # se pasó del máximo: hay que partirlo en varios subcircuitos
                bins = binpack_items(items, P_max)
                for idx_bin, b in enumerate(bins, start=1):
                    pot_bin = float(b["potencia_total"])
                    ambs_bin = sorted(set(it["amb"] for it in b["items"]))
                    ambs_str_bin = ", ".join(ambs_bin)
                    nombre_sub = f"{base} ({ambs_str_bin}) - subcircuito {idx_bin}"
                    long_real = pedir_longitud_sub(nombre_sub, longitud)  # cada subcircuito pregunta su propia longitud real
                    detalle = ""
                    if es_enchufe:
                        detalle = resumen_items_por_ambiente(b["items"], modo="enchufe")
                    elif es_ilumin:
                        detalle = resumen_items_por_ambiente(b["items"], modo="iluminacion")
                    elif es_especial:
                        detalle = resumen_items_por_ambiente(items, modo="especial")
                    add_circuito(nombre_sub, long_real, pot_bin,
                                 es_ilumin=es_ilumin, es_enchufe=es_enchufe, es_especial=False,
                                 items=b["items"], detalle_asignacion=detalle,
                                 tiene_tramo_20m=tiene_tramo_20m)
            else:
                # cabe en un solo circuito, no hace falta dividir
                nombre_circ = f"{base} ({ambientes_str_final})"
                detalle = ""
                if items:
                    if es_enchufe:
                        detalle = resumen_items_por_ambiente(items, modo="enchufe")
                    elif es_ilumin:
                        detalle = resumen_items_por_ambiente(items, modo="iluminacion")
                    elif es_especial:
                        detalle = resumen_items_por_ambiente(items, modo="especial")

                long_final = longitud
                if es_enchufe and len(items) > 0:
                    ambs_items = {str(it.get("amb", "")).strip().lower() for it in items if str(it.get("amb", "")).strip()}
                    solo_coc_lav = all(
                        a.startswith("cocina") or a.startswith("lavadero")
                        for a in ambs_items
                    ) if ambs_items else False
                    # Si el circuito final es solo cocina/lavadero, usar la longitud ya ingresada antes
                    if solo_coc_lav and len(ambs_items) == 1:
                        amb_unico = next(iter(ambs_items))
                        long_final = float(long_real_coc_lav_por_amb.get(amb_unico, longitud))
                    # Si quedó un circuito general restante (baño/living/pasillo, etc.), preguntar su longitud real
                    elif agrego_coc_lav:
                        long_final = pedir_longitud_sub(nombre_circ, longitud)
                add_circuito(nombre_circ, long_final, potencia_estimada,
                            es_ilumin=es_ilumin, es_enchufe=es_enchufe, es_especial=False,
                            items=items, detalle_asignacion=detalle,
                            tiene_tramo_20m=tiene_tramo_20m)
        else:
            print("       ! No se asignaron cargas a este circuito (solo tenía enchufes de cocina/lavadero o nada). No se crea.")

# Circuito global para enchufes de cocina/lavadero (16A)

# Circuitos por ambiente para enchufes de cocina/lavadero (16A cada uno)
# esto corre DESPUÉS de terminar de recorrer todos los circuitos, porque
# junta los enchufes de cocina/lavadero que quedaron sueltos de todos los
# circuitos y arma un circuito dedicado 16A para cada uno (si no se armó ya
# a mano arriba)
if enchufes_coc_lav_por_amb:
    for amb_cl, lista_ench in enchufes_coc_lav_por_amb.items():

        # si ya se creó manual, NO crear automático
        if amb_cl.strip().lower() in coc_lav_creado_manual:
            continue

        if not lista_ench:
            continue

        potencia_total = sum(float(e["potencia"]) for e in lista_ench)
        if potencia_total <= 0:
            continue

        long_max = max(float(e.get("longitud", 0.0)) for e in lista_ench)
        long_final = float(long_real_coc_lav_por_amb.get(amb_cl, long_max))

        nombre_circ = f"Enchufes {amb_cl} (cocina/lavadero)"

        items_cl = [{
            "amb": e["amb"],
            "potencia": float(e["potencia"]),
            "n_ench": 1,
            "id_ench": int(e.get("id_ench", 0)),
            "modulos": int(e.get("modulos", 1))
        } for e in lista_ench]

        detalle_cl = resumen_items_por_ambiente(items_cl, modo="enchufe")

        add_circuito(
            nombre_circ, long_final, potencia_total,
            es_ilumin=False, es_enchufe=True, es_especial=False,
            items=items_cl, detalle_asignacion=detalle_cl
        )
        # Si detecta cocina/lavadero automáticamente, fuerza 16A
        if es_enchufe and any(
            it["amb"].lower().startswith(("cocina", "lavadero"))
            for it in items
        ):
            circuitos[-1]["es_coc_lav"] = True

        # Paso 4: si este circuito manual incluye cocina/lavadero, evitar duplicado automático
        if es_enchufe and (len(items_coc_lav) > 0):
            ambs_cl = set(it["amb"] for it in items_coc_lav)
            for amb_cl in ambs_cl:
                if amb_cl in enchufes_coc_lav_por_amb:
                    enchufes_coc_lav_por_amb[amb_cl] = []

        # Paso 5: forzar TM 16A
        if es_enchufe and (len(items_coc_lav) > 0):
            circuitos[-1]["es_coc_lav"] = True

        # marca para forzar TM 16A más adelante
        circuitos[-1]["es_coc_lav"] = True
        circuitos[-1]["enchufes_coc_lav"] = lista_ench[:]

# =========================================================
# PARTE 2.2: CIRCUITOS DE CLIMATIZACIÓN (RIC N°07 SEC. 7)
# =========================================================


# Guardar _items por nombre de circuito ANTES de PARTE 3
# (en PARTE 3 se hacen copias que pueden no incluir _items correctamente)
import copy as _copy
items_por_nombre = {str(c.get("Circuito","")): _copy.deepcopy(c.get("_items", [])) for c in circuitos}  # copia profunda, para no compartir referencias
cajas_adic_por_nombre = {str(c.get("Circuito","")): int(c.get("_cajas_adic", 0)) for c in circuitos}


# -------- PARTE 3: TABLAS + INTERRUPTOR TERMOMAGNÉTICO --------
tension = float(input("\n- Ingrese la tensión nominal [V]: "))
factor_potencia = float(input("- Ingrese el factor de potencia (ej. 0.92): "))
temperatura = float(input("- Ingrese la temperatura ambiente típica de la zona: "))
longitud_alimentador = float(input("Ingrese la longitud del alimentador (Empalme → Tablero) en metros: "))
longitud_transformador_empalme = float(input("Ingrese la longitud de la acometida (Transformador → Empalme) en metros: "))
while True:
    tipo_acometida = input("Ingrese el tipo de acometida (Aérea / subterránea): ").strip().lower()
    if "aer" in tipo_acometida or "sub" in tipo_acometida:
        tipo_acometida = "aerea" if "aer" in tipo_acometida else "subterranea"
        break
    print("  Valor no reconocido. Ingrese 'aerea' o 'subterranea'.")
# Temperatura del suelo: se pregunta más adelante, después de ingresar tipo_alimentador,
# ya que puede aplicar a acometida subterránea, alimentador subterráneo, o ambos.
# Ver bloque "TEMPERATURA DEL SUELO" más abajo (RIC 4 punto 6.2.5, nota tabla 4.7).
temperatura_suelo = temperatura  # valor por defecto hasta que se calcule abajo
# =========================
# DATOS DE UBICACIÓN DEL EMPALME
# =========================
while True:
    empalme_dentro_15m = input("¿El empalme se ubica dentro de 15 m del acceso a la propiedad? (si/no): ").strip().lower()
    if "si" in empalme_dentro_15m or "no" in empalme_dentro_15m:
        empalme_dentro_15m = "si" if "si" in empalme_dentro_15m else "no"
        break
    print("  Valor no reconocido. Ingrese 'si' o 'no'.")
while True:
    tipo_instalacion_empalme = input("¿El empalme se instalará en fachada o en estructura independiente? (fachada/independiente): ").strip().lower()
    if "fachada" in tipo_instalacion_empalme or "independiente" in tipo_instalacion_empalme:
        tipo_instalacion_empalme = "fachada" if "fachada" in tipo_instalacion_empalme else "independiente"
        # Validación RIC N°01: la distancia al acceso obliga a un tipo de instalación específico
        if empalme_dentro_15m == "si" and tipo_instalacion_empalme == "independiente":
            print("  Según RIC N°01 punto 7.2, un empalme dentro del radio de 15 m debe instalarse")
            print("    en la fachada de la vivienda. Por favor seleccione 'fachada'.")
            continue
        if empalme_dentro_15m == "no" and tipo_instalacion_empalme == "fachada":
            print("  Según RIC N°01 punto 7.3, un empalme fuera del radio de 15 m debe instalarse")
            print("    en estructura independiente cerca del cierre de la propiedad. Por favor seleccione 'independiente'.")
            continue
        break
    print("  Valor no reconocido. Ingrese 'fachada' o 'independiente'.")
tipo_poste = ""
altura_acometida_aerea = 0
longitud_subterraneo_medidor2 = 0
if tipo_instalacion_empalme == "independiente":
    while True:
        tipo_poste = input("¿El poste será de madera o metálico? (madera/metalico): ").strip().lower()
        if "madera" in tipo_poste or "metal" in tipo_poste:
            tipo_poste = "madera" if "madera" in tipo_poste else "metalico"
            break
        print("  Valor no reconocido. Ingrese 'madera' o 'metalico'.")
    # Solo para acometida aérea
    if "aer" in tipo_acometida:
        altura_acometida_aerea = float(input(
                "Ingrese la altura desde el empalme hasta el punto de llegada de la acometida aérea en metros: "))
    # Solo para acometida subterránea
    elif "sub" in tipo_acometida:
        longitud_subterraneo_medidor2 = float(input(
                "Ingrese la longitud desde la salida subterránea de la acometida hasta el medidor en metros: "))
requiere_mastil = ""
longitud_mastil = 0
if tipo_instalacion_empalme == "fachada" and ("aer" in tipo_acometida or "sub" in tipo_acometida):
    while True:
        tipo_txt = "aérea" if "aer" in tipo_acometida else "subterránea"
        requiere_mastil = input(f"¿La acometida {tipo_txt} requiere mástil? (si/no): ").strip().lower()
        if "sub" in tipo_acometida and requiere_mastil == "si":
            print(" No se puede utilizar mástil en un empalme en fachada con acometida subterránea.")
            continue
        break
    if requiere_mastil == "si":
        longitud_mastil = float(input("Ingrese el largo del mástil (Caja empalme - extremo del mástil en metros): "))

# tipo de alimentador: cada combinación con el tipo de empalme tiene reglas propias
while True:
    while True:
        _resp_alim = input("Ingrese el tipo de alimentador (Aéreo / Subterráneo / En ducto): ").strip().lower()
        if "aer" in _resp_alim:
            tipo_alimentador = "aereo"
            break
        elif "sub" in _resp_alim:
            tipo_alimentador = "subterraneo"
            break
        elif "duct" in _resp_alim or "duc" in _resp_alim:
            tipo_alimentador = "en ducto"
            break
        else:
            print("  Valor no reconocido. Ingrese 'aereo', 'subterraneo' o 'en ducto'.")
    # Si es fachada, solo se permite en ducto
    if (tipo_instalacion_empalme == "fachada" and ("aer" in tipo_alimentador or "sub" in tipo_alimentador)):
        print(" No se puede poner un alimentador aéreo o subterráneo en fachada sin mástil. Debe seleccionar: En ducto.")
        continue
    # Si es independiente, no se permite en ducto
    elif (tipo_instalacion_empalme == "independiente" and "duct" in tipo_alimentador):
        print("No se puede poner un alimentador en ducto en empalme independiente. Debe seleccionar: aéreo o subterráneo.")
        continue
    break
# Poste independiente con acometida AÉREA + alimentador SUBTERRÁNEO también
# necesita longitud_subterraneo_medidor2: la usan el tubo galvanizado del
# poste (L = altura_acometida_aerea + este tramo), el conduit PVC del
# alimentador (se resta este tramo) y las cámaras tipo C del alimentador.
if (tipo_instalacion_empalme == "independiente" and "aer" in tipo_acometida and "sub" in tipo_alimentador):
    longitud_subterraneo_medidor2 = float(input(
        "Ingrese la longitud desde la salida subterránea del alimentador hasta el medidor en metros: "))
#Alimentador aereo, empalme independiente
longitud_llegada_aerea_tda = 0
longitud_poste_alimentador_aereo = 0
if (tipo_instalacion_empalme == "independiente" and "aer" in tipo_alimentador):
    longitud_llegada_aerea_tda = float(
    input("Ingrese la longitud desde la llegada del alimentador aéreo en la casa hasta el TDA en metros: "))
    longitud_poste_alimentador_aereo = float(
    input("Ingrese la longitud de subida del tramo del alimentador aéreo en el poste en metros: "))
# Alimentador subteraneo, empalme independiente, para calculo de abrazaderas pvc
longitud_abrazaderas_alimentador = 0
if (tipo_instalacion_empalme == "independiente" and "sub" in tipo_alimentador):
    longitud_abrazaderas_alimentador = float(
        input("Ingrese la longitud desde la salida subterránea del alimentador hasta el TDA en metros: "))
# Longitud salida subterranea en empalme en fachada
longitud_subterraneo_medidor = 0
if (tipo_instalacion_empalme == "fachada" and requiere_mastil == "no" and "sub" in tipo_acometida
    and "duct" in tipo_alimentador):
    longitud_subterraneo_medidor = float(
    input("Ingrese la longitud desde la salida subterránea de la acometida hasta el medidor en metros: ")
    )
# Distancia vertical acometida aérea en fachada sin mástil
dist_vertical_acometida = 0
if (tipo_instalacion_empalme == "fachada" and requiere_mastil == "no" and "aer" in tipo_acometida):
    dist_vertical_acometida = float(
        input("Ingrese la distancia vertical (caja del empalme - punto de llegada de la acometida en metros: "))
dist_empalme_pt1 = float(input("Ingrese la distancia del empalme a la puesta a tierra 1 (camarilla N°1) en metros: ") or 0)
dist_tda_pt2 = float(input("Ingrese la distancia del TDA a la puesta a tierra 2 (camarilla N°2) en metros: ") or 0)

# =========================
# TEMPERATURA DEL SUELO (RIC 4 punto 6.2.5, nota tabla 4.7)
# Se pregunta si acometida O alimentador son subterráneos
# =========================
_acom_sub = "sub" in str(tipo_acometida).lower()
_alim_sub = "sub" in str(tipo_alimentador).lower()

if _acom_sub or _alim_sub:
    _tramos_sub = []
    if _acom_sub:
        _tramos_sub.append("acometida subterránea")
    if _alim_sub:
        _tramos_sub.append("alimentador subterráneo")
    _tramos_str = " y ".join(_tramos_sub)
    print(f"\n  Tiene tramos subterráneos: {_tramos_str}.")
    print("  La temperatura del suelo afecta el factor de corrección de ampacidad (ft) del método D1.")
    while True:
        try:
            temperatura_suelo = float(input(
                f"  Ingrese la temperatura del suelo para tramos subterráneos [°C] "
                f"(RIC 4 punto 6.2.5, nota tabla 4.7) [ENTER = {temperatura}°C]: "
            ).strip() or temperatura)
            break
        except ValueError:
            print("  ! Ingrese un número válido.")
else:
    temperatura_suelo = temperatura  # si no hay tramos subterráneos, queda igual a la T° ambiente

# =========================
# PUESTA A TIERRA - CÁLCULO RIC 6
# =========================

# Largo de barra copperweld
print("\n  Largo de barra copperweld:")
print("    1) 3 metros (recomendado)")
print("    2) 1,5 metros (solo si hay restricción de profundidad)")
while True:
    op_barra = input("  Seleccione opción (ENTER = 1): ").strip()
    if op_barra in ["", "1"]:
        largo_barra_pt = 3.0
        desc_barra_pt  = "Barra copperweld 5/8 3mts + conector de bronce"
        break
    elif op_barra == "2":
        largo_barra_pt = 1.5
        desc_barra_pt  = "Barra copperweld 5/8 1,5mts + conector de bronce"
        break
    else:
        print("  ! Ingrese 1 o 2.")

def _pedir_resistividad(etiqueta):
    """Pide resistividad del terreno para una PT dada."""
    print(f"\n  Resistividad del terreno (ρ) - {etiqueta}:")
    print("    1) Ingresar valor medido en Ohm·m")
    print("    2) Estimar por tipo de terreno (Tabla 6.3 RIC 6)")
    while True:
        op_rho = input("  Seleccione opción: ").strip()
        if op_rho == "1":
            # el usuario ya midió la resistividad en terreno
            while True:
                try:
                    rho = float(input("  Ingrese resistividad medida (Ohm·m): ").strip())
                    if rho > 0:
                        return rho, f"{rho} Ohm·m - Medido"
                    print("  ! Debe ser mayor que 0.")
                except:
                    print("  ! Ingrese un número válido.")
        elif op_rho == "2":
            # no tiene medición: usa un valor típico según el tipo de terreno
            print("    1) Terreno fértil / húmedo         →  50 Ohm·m")
            print("    2) Terraplén poco fértil            → 500 Ohm·m")
            print("    3) Pedregoso / arena seca           → 3000 Ohm·m")
            while True:
                op_ter = input("  Seleccione tipo de terreno: ").strip()
                if op_ter == "1":
                    return 50.0, "50 Ohm·m - Fértil/húmedo"
                elif op_ter == "2":
                    return 500.0, "500 Ohm·m - Poco fértil"
                elif op_ter == "3":
                    return 3000.0, "3000 Ohm·m - Pedregoso/seco"
                else:
                    print("  ! Ingrese 1, 2 o 3.")
        else:
            print("  ! Ingrese 1 o 2.")

# Resistividad PT1 (empalme - camarilla N°1)
rho_pt1, desc_rho_pt1 = _pedir_resistividad("PT N°1 (empalme - camarilla N°1)")
# Resistividad PT2 (tablero - camarilla N°2)
rho_pt2, desc_rho_pt2 = _pedir_resistividad("PT N°2 (tablero - camarilla N°2)")

# Para compatibilidad con tabla de resumen
rho_terreno = rho_pt1
desc_rho    = desc_rho_pt1

def _calcular_pt(rho, largo_barra):
    """Calcula N° barras, R final, separación y longitud conductor desnudo."""
    R_1b  = rho / largo_barra  # resistencia si se usara solo 1 barra
    n     = max(1, math.ceil(R_1b / 20.0))  # cuántas barras se necesitan para bajar de 20 ohm
    R_fin = R_1b / n  # resistencia final con esa cantidad de barras (en paralelo)
    sep   = 2 * largo_barra  # separación mínima recomendada entre barras
    long_desnudo = (n - 1) * sep  # metros de conductor desnudo para unir las barras
    return n, R_1b, R_fin, sep, long_desnudo

# Cálculo PT1
n_barras_pt1, R_1barra_pt1, R_final_pt1, sep_min_pt1, long_cond_desnudo_pt1 = _calcular_pt(rho_pt1, largo_barra_pt)
# Cálculo PT2
n_barras_pt2, R_1barra_pt2, R_final_pt2, sep_min_pt2, long_cond_desnudo_pt2 = _calcular_pt(rho_pt2, largo_barra_pt)

# Para compatibilidad con código existente (usa el mayor de los dos)
n_barras_pt       = max(n_barras_pt1, n_barras_pt2)
R_1barra          = R_1barra_pt1
R_final_pt        = R_final_pt1
sep_min_pt        = sep_min_pt1
long_cond_desnudo_pt = long_cond_desnudo_pt1

# Advertencias si se necesitan demasiadas barras (puede requerir estudio especial)
for n, label in [(n_barras_pt1, "PT1"), (n_barras_pt2, "PT2")]:
    if largo_barra_pt == 1.5 and n > 6:
        print(f"\n  ADVERTENCIA {label}: Se requieren {n} barras de 1,5m.")
        print("    Se recomienda usar barras de 3m o realizar un estudio especial de puesta a tierra.")
    elif largo_barra_pt == 3.0 and n > 6:
        print(f"\n  ADVERTENCIA {label}: Se requieren {n} barras de 3m.")
        print("    Se recomienda realizar un estudio especial de puesta a tierra (RIC 6, punto 5.1).")

potencia_total = sum(a["Total por ambiente (W)"] for a in ambientes_detalle)
# Agregar potencia de equipos de climatización (no están en ambientes_detalle)
if circuitos_climatizacion:
    potencia_total += sum(float(eq["P_nom_w"]) for eq in circuitos_climatizacion)
if tension <= 0:
    tension = 1.0  # evita división por cero más adelante

calibres_tm = [6, 10, 16, 20, 25, 32, 40, 50, 63]

# Tabla de corriente admisible a 70°C para los circuitos interiores (Tabla
# N°4.4 RIC N°4): por cada sección de conductor, "A1" es la ampacidad si va
# embutido/en conducto y "B1" si va en ducto superficial.
tabla_70 = {
    1.5:  {"A1": 14, "B1": 16},
    2.08: {"A1": 16, "B1": 19},
    2.5:  {"A1": 18, "B1": 21},
    3.31: {"A1": 21, "B1": 25},
    4.0:  {"A1": 24, "B1": 28},
    5.26: {"A1": 28, "B1": 34},
    6.0:  {"A1": 31, "B1": 36},
    8.37: {"A1": 38, "B1": 45},
    10.0: {"A1": 42, "B1": 50},
    13.3: {"A1": 50, "B1": 60},
    16.0: {"A1": 56, "B1": 68},
    21.1: {"A1": 66, "B1": 80},
    25.0: {"A1": 73, "B1": 89},
    26.7: {"A1": 76, "B1": 93},
    33.6: {"A1": 87, "B1": 108},
    35.0: {"A1": 89, "B1": 110},
    42.4: {"A1": 99, "B1": 125},
    50.0: {"A1": 108, "B1": 134},
    53.5: {"A1": 116, "B1": 144},
    67.4: {"A1": 133, "B1": 167},
}

nuevo_circuitos = []
tm_corrientes = []

# recorre todos los circuitos y le calcula/asigna el termomagnético a cada uno
for c in circuitos:
    pot = float(c.get("Potencia estimada (W)", 0.0))
    es_ilumin = c.get("es_ilumin", False)
    es_enchufe = c.get("es_enchufe", False)
    tipo_dif = c.get("tipo_dif", "general")
    curva = "B" if es_ilumin else "C"  # iluminación usa curva B, el resto curva C

    nombre = str(c.get("Circuito", "")).lower()
    items_c = c.get("_items", [])

    # CLIMATIZACIÓN: TM ya calculado por RIC N°07 — NO recalcular
    if c.get("_es_climatizacion", False):
        in_tm_clima = int(c.get("_In_TM", 16))
        i_cuadro = float(c.get("Corriente estimada (A)", 0.0))
        if i_cuadro <= 0:
            i_cuadro = float(c.get("Potencia estimada (W)", 0.0)) / tension if tension else 0.0
        nuevo = c.copy()
        nuevo["Corriente estimada (A)"] = round(i_cuadro, 2)
        nuevo["_In_TM"] = in_tm_clima
        nuevo["_I_diseno_clima"] = c.get("_I_diseno_clima", in_tm_clima)
        nuevo_circuitos.append(nuevo)
        tm_corrientes.append(in_tm_clima)
        continue

    # AGUA CALIENTE: TM ya calculado por RIC N°07 7.3.4 — NO recalcular
    if c.get("_es_agua_caliente", False):
        in_tm_agua = int(c.get("_In_TM", 20))
        i_cuadro = float(c.get("Corriente estimada (A)", 0.0))
        if i_cuadro <= 0:
            i_cuadro = float(c.get("Potencia estimada (W)", 0.0)) / tension if tension else 0.0
        nuevo = c.copy()
        nuevo["Corriente estimada (A)"] = round(i_cuadro, 2)
        nuevo["_In_TM"]        = in_tm_agua
        nuevo["_I_diseno_agua"] = c.get("_I_diseno_agua", in_tm_agua)
        nuevo_circuitos.append(nuevo)
        tm_corrientes.append(in_tm_agua)
        continue

    # detectar si el circuito incluye baño (por nombre o por ambientes en items)
    tiene_bano = ("baño" in nombre) or ("bano" in nombre) or any(
        isinstance(it, dict) and str(it.get("amb", "")).strip().lower().startswith(("baño", "bano"))
        for it in (items_c if isinstance(items_c, list) else [])
    )
    # forzar 16A en enchufes de cocina/lavadero/baño
    if tipo_dif == "general" and es_enchufe and (("cocina" in nombre) or ("lavadero" in nombre) or tiene_bano):
        i_est = pot / tension
        nuevo = c.copy()
        nuevo["Corriente estimada (A)"] = round(i_est, 2)
        nuevo["Disyuntor termomagnético"] = "1x16A / 6kA / Curva C"
        nuevo["_In_TM"] = 16
        nuevo_circuitos.append(nuevo)
        tm_corrientes.append(16)
        continue

    if tipo_dif == "general" and (es_ilumin or es_enchufe):
        i_est = pot / tension
        # RIC N°10 art. 5.1.4.1: la capacidad del circuito de alumbrado estará
        # determinada por la potencia requerida más un 10% de capacidad adicional.
        # El valor nominal del TM será el valor nominal de corriente de la protección
        # inmediatamente superior disponible en el mercado.
        i_necesaria = i_est * 1.10
        allowed = [10, 16] if es_enchufe else [6, 10, 16]

        in_tm = allowed[-1]
        for cal in allowed:
            if cal >= i_necesaria:  # primer calibre >= i_est × 1.10
                in_tm = cal
                break

        nuevo = c.copy()
        nuevo["Corriente estimada (A)"] = round(i_est, 2)
        nuevo["Disyuntor termomagnético"] = f"1x{in_tm}A / 6kA / Curva {curva}"
        nuevo["_In_TM"] = in_tm
        nuevo_circuitos.append(nuevo)
        tm_corrientes.append(in_tm)
    else:
        # circuitos especiales genéricos: mismo criterio, pero con toda la lista de calibres
        i_est = pot / tension
        i_necesaria = i_est * 1.10

        in_tm = calibres_tm[-1]
        for cal in calibres_tm:
            if cal >= i_necesaria:
                in_tm = cal
                break

        if c.get("tipo_dif") == "especial" and in_tm < 16:
            in_tm = 16  # los especiales nunca van con TM menor a 16A

        nuevo = c.copy()
        nuevo["Corriente estimada (A)"] = round(i_est, 2)
        nuevo["Disyuntor termomagnético"] = f"1x{in_tm}A / 6kA / Curva {curva}"
        nuevo["_In_TM"] = in_tm
        nuevo_circuitos.append(nuevo)
        tm_corrientes.append(in_tm)

circuitos = nuevo_circuitos

# -------- PARTE 3.1: DIFERENCIALES SEGÚN EMPALME --------
def optimizar_agrupacion(indices, max_sum):
    """
    Agrupa indices minimizando la cantidad de diferenciales,
    restricciones:
      - máx 3 circuitos por diferencial
      - suma In_TM por diferencial <= max_sum
    """
    items = sorted(indices, key=lambda i: tm_corrientes[i], reverse=True)  # de mayor a menor corriente
    best_bins = None  # acá se va guardando la mejor agrupación encontrada

    def score(bins):
        # bins = la agrupación completa: una lista de grupos.
        # cada grupo (b) es una lista con los circuitos que comparten
        # el mismo diferencial. Ej: bins = [[0,1], [2,3,4]] son 2 grupos:
        # el primero con 2 circuitos, el segundo con 3.
        #
        # entre 2 opciones de "bins", gana la que use menos diferenciales
        # (o sea, menos grupos da un len(bins) más chico).
        # si empatan, gana la que deja los grupos más llenos
        # (el "-" es solo para que "más lleno" cuente como "mejor puntaje")
        return (len(bins), -sum(len(b) for b in bins))

    def can_place(bin_list, idx):
        # bin_list = un grupo (lista de circuitos que ya están juntos en un diferencial)
        # idx = el circuito que se quiere agregar a ese grupo
        # tm_corrientes = lista con la corriente del TM de cada circuito
        #
        # revisa si el circuito "idx" cabe en este grupo, sin pasarse
        # de 3 circuitos ni de la corriente máxima del diferencial
        if len(bin_list) >= 3:
            return False
        return (sum(tm_corrientes[j] for j in bin_list) + tm_corrientes[idx]) <= max_sum

    def backtrack(pos, bins):
        # va probando dónde meter cada circuito (en los grupos ya armados o
        # en uno nuevo) y se queda con la mejor combinación encontrada
        nonlocal best_bins

        if best_bins is not None and score(bins) >= score(best_bins):
            return  # esta combinación ya no puede ser mejor que la que tenemos, corta acá (poda)

        if pos == len(items):
            best_bins = [b[:] for b in bins]  # llegó al final: guarda esta combinación como la mejor
            return

        idx = items[pos]  # el circuito que toca ubicar ahora

        # probar primero en bins existentes (más llenos primero)
        order = sorted(
            range(len(bins)),
            key=lambda k: (len(bins[k]), sum(tm_corrientes[j] for j in bins[k])),
            reverse=True
        )  # ordena los grupos ya armados, del más lleno al más vacío

        for k in order:
            if can_place(bins[k], idx):
                bins[k].append(idx)          # prueba metiendo el circuito en este grupo
                backtrack(pos + 1, bins)      # sigue probando con el siguiente circuito
                bins[k].pop()                 # deshace, para probar otra combinación

        # abrir bin nuevo
        bins.append([idx])       # prueba abriendo un grupo nuevo solo para este circuito
        backtrack(pos + 1, bins)  # sigue probando con el siguiente circuito
        bins.pop()                # deshace, para no arrastrar el grupo nuevo a otras pruebas

    backtrack(0, [])  # arranca la búsqueda de la mejor combinación

    if best_bins is None:
        return [[i] for i in items]  # no encontró nada (no debería pasar), cada circuito en su propio grupo

    return [sorted(b) for b in best_bins]  # devuelve la mejor agrupación encontrada


def parametros_desde_empalme(interruptor_empalme):
    """
    Deriva los parámetros del tablero a partir del interruptor termomagnético del empalme (A).
    - diferencial recomendado
    - suma máxima de ITM por diferencial (para tu agrupación)
    - potencia máxima sugerida para un componente especial (solo referencia/validación)
    - texto del omnipolar general
    """
    try:
        Ie = int(interruptor_empalme)  # corriente del interruptor del empalme
    except:
        Ie = 25  # si viene mal, usa un valor por defecto razonable

    # según el calibre del empalme, define el calibre del diferencial general,
    # el tope de suma de TM por diferencial y una potencia especial de referencia
    if Ie <= 25:
        dif_calibre = 25
        max_sum_tm = 25
        max_pot_especial = 5000
    elif Ie <= 32:
        dif_calibre = 40
        max_sum_tm = 32
        max_pot_especial = 6400
    elif Ie <= 40:
        dif_calibre = 40
        max_sum_tm = 40
        max_pot_especial = 8000
    else:
        # Si en el futuro agregas empalmes mayores, ajusta aquí.
        dif_calibre = 63
        max_sum_tm = Ie
        max_pot_especial = Ie * 220  # aproximación a 220V

    texto_omni = f"2x{Ie}A / 10kA / Curva C"  # texto del interruptor general omnipolar
    return dif_calibre, max_sum_tm, max_pot_especial, texto_omni


# =========================================================
# EMPALME Y PROTECCIONES (AUTO DESDE FACTOR DE DEMANDA)
# =========================================================
# Nota: aquí se calcula el empalme con la corriente con factor de demanda.
# Luego se derivan diferencial/omnipolar y se asignan a los circuitos ANTES
# de crear dataframes/materiales.

# --- Factor de demanda (RIC 3 Tabla 3.1) ---
# Criterio:
#   corriente_sin = suma de corrientes individuales del cuadro de cargas
#                   (cada circuito ya tiene su corriente calculada con su fp real)
#   I_primeros    = (primeros_kw / pot_kw) x corriente_sin, multiplicado por 1,0
#   I_resto       = (resto_kw   / pot_kw) x corriente_sin, multiplicado por 0,35
#   corriente_con = I_primeros x 1,0 + I_resto x 0,35
tension_nominal = float(tension)
fp = float(factor_potencia) if factor_potencia not in [None, ""] else 1.0
if fp == 0:
    fp = 1.0

# Potencias separadas (para tabla Excel)
# potencia_sin_clima_w debe excluir climatización y agua caliente — esos
# componentes quedan guardados dentro de "Total por ambiente (W)" con la
# potencia ESTIMADA que se ingresó al crear el ambiente (no la real de la
# placa, que se pide después al crear el circuito), así que se restan acá
# para que pot_kw_alum_ench_global (base de primeros/resto) no las arrastre.
import re
_palabras_clima_agua = ("aire", "clima", "split", "ac ", "a/c", "ducha",
                         "termo", "calefon", "calefón", "calentador", "agua caliente")
potencia_clima_agua_en_ambientes = 0.0
for _amb in ambientes_detalle:
    for _linea in str(_amb.get("Componentes especiales", "") or "").split("\n"):
        _linea_l = _linea.strip().lower()
        if any(palabra in _linea_l for palabra in _palabras_clima_agua):
            _match_w = re.search(r'\(([\d.,]+)\s*W\)', _linea, re.IGNORECASE)
            if _match_w:
                try:
                    potencia_clima_agua_en_ambientes += float(_match_w.group(1).replace(",", "."))
                except ValueError:
                    pass
potencia_sin_clima_w = sum(a["Total por ambiente (W)"] for a in ambientes_detalle) - potencia_clima_agua_en_ambientes
potencia_clima_w = 0.0
if circuitos_climatizacion:
    for eq_cl in circuitos_climatizacion:
        potencia_clima_w += float(eq_cl.get("P_nom_w", 0))
# Potencia agua caliente (no está en ambientes_detalle)
potencia_agua_w = 0.0
if circuitos_agua_caliente:
    for eq_ac in circuitos_agua_caliente:
        potencia_agua_w += float(eq_ac.get("P_nom_w", 0))
potencia_total_instalacion_w = potencia_sin_clima_w + potencia_clima_w + potencia_agua_w
pot_kw = potencia_total_instalacion_w / 1000.0

# kW primeros/resto — solo alumbrado+enchufes (para tabla Excel y corriente_con)
pot_kw_alum_ench_global = potencia_sin_clima_w / 1000.0
primeros_kw = min(pot_kw_alum_ench_global, 3.0)  # los primeros 3kW van al 100%
resto_kw    = max(pot_kw_alum_ench_global - 3.0, 0.0)  # el resto va al 35%
factor_primero = 1.0
factor_resto   = 0.35
con_factor_primero = primeros_kw * factor_primero
con_factor_resto   = resto_kw   * factor_resto
total_kw = con_factor_primero + con_factor_resto

# --- Corriente alumbrado+enchufes SIN fd (factor de demanda) ---
_keywords_clima_agua = ("climatiz","aire","split","ac ","a/c",
                        "ducha","termo","calefon","calefón","calentador","agua caliente")
corriente_sin_alum_ench = sum(
    float(c.get("Corriente estimada (A)", 0.0)) for c in circuitos
    if not any(k in str(c.get("Circuito","")).lower() for k in _keywords_clima_agua)
)
corriente_sin_clima = sum(
    float(c.get("Corriente estimada (A)", 0.0)) for c in circuitos
    if any(k in str(c.get("Circuito","")).lower() for k in ("climatiz","aire","split","ac ","a/c"))
)
corriente_sin_agua = sum(
    float(c.get("Corriente estimada (A)", 0.0)) for c in circuitos
    if any(k in str(c.get("Circuito","")).lower() for k in ("ducha","termo","calefon","calefón","calentador","agua caliente"))
)
corriente_sin = corriente_sin_alum_ench + corriente_sin_clima + corriente_sin_agua

# --- Corriente CON factor de demanda (RIC 3 art. 6.1, 6.2, 6.3) ---
# Alumbrado+enchufes: aplica Tabla N°3.1 (fd primeros 3kW=1,0; resto=0,35)
# Climatización y agua caliente: corriente plena (fd=1,0) según RIC 7
pot_kw_alum_ench = pot_kw_alum_ench_global
if pot_kw_alum_ench <= 3.0:
    I_alum_primeros = corriente_sin_alum_ench
    I_alum_resto    = 0.0
else:
    I_alum_primeros = (min(pot_kw_alum_ench, 3.0) * 1000.0) / tension_nominal if tension_nominal else 0.0
    I_alum_resto    = (max(pot_kw_alum_ench - 3.0, 0.0) * 1000.0) / tension_nominal if tension_nominal else 0.0

corriente_con = (I_alum_primeros * factor_primero
               + I_alum_resto    * factor_resto
               + corriente_sin_clima   # fd=1,0
               + corriente_sin_agua)   # fd=1,0

# --- Selección interruptor empalme (normalizado a calibre comercial) ---
# Se aplica un 10% extra de holgura sobre la corriente con factor de
# demanda antes de elegir el calibre comercial.
corriente_empalme = corriente_con * 1.1

if corriente_empalme <= 25:
    interruptor_empalme = 25
elif corriente_empalme <= 32:
    interruptor_empalme = 32
elif corriente_empalme <= 40:
    interruptor_empalme = 40
elif corriente_empalme <= 50:
    interruptor_empalme = 50
elif corriente_empalme <= 63:
    interruptor_empalme = 63
else:
    interruptor_empalme = 63  # tope máximo que soporta el programa

# ── Aviso límite del programa ─────────────────────────────────────────────
if interruptor_empalme > 40:
    print(
        f"\n{'='*65}"
        f"\n  AVISO — Empalme calculado: {interruptor_empalme} A"
        f"\n  Este programa está diseñado para instalaciones con empalme"
        f"\n  hasta 40 A (tipo A-9 o S-9)."
        f"\n  La instalación ingresada supera ese límite ({corriente_empalme:.1f} A de demanda)."
        f"\n  Se continuará el cálculo con {interruptor_empalme} A, pero se recomienda"
        f"\n  revisar el proyecto manualmente."
        f"\n{'='*65}\n"
    )

interruptor_texto = f"1x{interruptor_empalme}A / 6kA / Curva D"
dif_calibre, max_sum_tm, max_pot_especial, texto_omni = parametros_desde_empalme(interruptor_empalme)

# Bloque desactivado (queda como texto, no se ejecuta): validación
# informativa que compara componentes especiales contra el empalme
# calculado, solo para referencia, no detiene el programa.
"""
try:
    if isinstance(componentes_potencias, list) and len(componentes_potencias) > 0:
        pmax_esp = max([float(p) for p in componentes_potencias])
        if pmax_esp > float(max_pot_especial):
            print(
                f"\n[AVISO] Hay un componente especial de {pmax_esp:.0f} W que supera "
                f"la referencia ({max_pot_especial} W) para un empalme calculado de {interruptor_empalme}A."
            )
            print("       Revisa si corresponde circuito dedicado y/o aumentar empalme según criterio/SEC-RIC.")
except Exception:
    pass
"""

# =========================================================
# DIFERENCIALES + OMNIPOLAR (DESDE EMPALME CALCULADO)
# =========================================================
modo_dif = input(
    "\n- ¿Desea 1 diferencial por circuito (1) o agrupar hasta 3 circuitos por diferencial (3)?: "
).strip()

group_info = {}
gid = 0

if modo_dif == "1":
    # cada circuito con su propio diferencial exclusivo
    for idx in range(len(circuitos)):
        es_clima = circuitos[idx].get("_es_climatizacion", False)
        es_agua  = circuitos[idx].get("_es_agua_caliente", False)
        if es_clima:
            dif_val = max(circuitos[idx].get("_In_dif_clima", dif_calibre), dif_calibre)
            group_info[gid] = {"indices": [idx], "dif": dif_val, "sensibilidad_dif": "30mA"}
        elif es_agua:
            dif_val = circuitos[idx].get("_In_dif_agua", dif_calibre)
            sens    = circuitos[idx].get("_sensibilidad_dif_agua", "30mA")
            group_info[gid] = {"indices": [idx], "dif": dif_val, "sensibilidad_dif": sens}
        else:
            group_info[gid] = {"indices": [idx], "dif": dif_calibre, "sensibilidad_dif": "30mA"}
        gid += 1
else:
    # Separar circuitos de climatización y agua caliente (siempre diferencial
    # exclusivo) del resto — los "especiales" (horno, lavadora, encimera,
    # etc.) ahora SÍ se agrupan igual que iluminación/enchufes.
    idx_normales = [i for i in range(len(circuitos))
                    if not circuitos[i].get("_es_climatizacion", False)
                    and not circuitos[i].get("_es_agua_caliente", False)]
    idx_clima    = [i for i in range(len(circuitos)) if circuitos[i].get("_es_climatizacion", False)]
    idx_agua     = [i for i in range(len(circuitos)) if circuitos[i].get("_es_agua_caliente", False)]

    # Agrupar circuitos normales + especiales (usa el algoritmo de
    # backtracking para minimizar diferenciales)
    if idx_normales:
        grupos_norm = optimizar_agrupacion(idx_normales, max_sum_tm)
        for g in grupos_norm:
            group_info[gid] = {"indices": g, "dif": dif_calibre, "sensibilidad_dif": "30mA"}
            gid += 1

    # Cada circuito de climatización tiene diferencial EXCLUSIVO (RIC 7.1.2 + 7.4.5)
    for idx in idx_clima:
        dif_val = max(circuitos[idx].get("_In_dif_clima", dif_calibre), dif_calibre)
        group_info[gid] = {"indices": [idx], "dif": dif_val, "sensibilidad_dif": "30mA"}
        gid += 1

    # Cada circuito de agua caliente tiene diferencial EXCLUSIVO (RIC N°07 7.4.5 + RIC N°11 6.4.3)
    for idx in idx_agua:
        dif_val = circuitos[idx].get("_In_dif_agua", dif_calibre)
        sens    = circuitos[idx].get("_sensibilidad_dif_agua", "30mA")
        group_info[gid] = {"indices": [idx], "dif": dif_val, "sensibilidad_dif": sens}
        gid += 1

# Reordenar circuitos para que queden agrupados por diferencial
# (así en el Excel salen juntos los circuitos que comparten diferencial)
idx_to_gid = {}
for g, meta in group_info.items():
    for idx in meta["indices"]:
        idx_to_gid[idx] = g

new_order = sorted(range(len(circuitos)), key=lambda i: (idx_to_gid.get(i, 999), i))
circuitos = [circuitos[i] for i in new_order]
tm_corrientes = [tm_corrientes[i] for i in new_order]

# Recalcular group_info con nuevos índices (porque el orden de los circuitos cambió)
old_to_new = {old_idx: new_pos for new_pos, old_idx in enumerate(new_order)}
new_group_info = {}
new_gid = 0
for g, meta in group_info.items():
    nuevos = sorted(old_to_new[i] for i in meta["indices"])
    new_group_info[new_gid] = {
        "indices":          nuevos,
        "dif":              meta["dif"],
        "sensibilidad_dif": meta.get("sensibilidad_dif", "30mA"),
    }
    new_gid += 1
group_info = new_group_info

# ── REGLA: mínimo 2 diferenciales en la instalación ──────────────────────────
# Si quedó solo 1 grupo (1 diferencial), se divide en 2 grupos
if len(group_info) == 1:
    gid_unico = list(group_info.keys())[0]
    meta_unica = group_info[gid_unico]
    indices = meta_unica["indices"]
    dif_val = meta_unica["dif"]
    if len(indices) >= 2:
        # Dividir a la mitad
        mitad = len(indices) // 2
        sens_val = meta_unica.get("sensibilidad_dif", "30mA")
        group_info = {
            0: {"indices": indices[:mitad], "dif": dif_val, "sensibilidad_dif": sens_val},
            1: {"indices": indices[mitad:], "dif": dif_val, "sensibilidad_dif": sens_val},
        }
    # Si solo hay 1 circuito total, no se puede dividir, así que queda 1 diferencial
    # (caso borde: instalación con 1 solo circuito)

# Asignar texto de diferencial y omnipolar a cada circuito (para Excel y materiales)
for g, meta in group_info.items():
    val  = meta["dif"]
    sens = meta.get("sensibilidad_dif", "30mA")
    dif_text = f"2X{val} {sens} / Tipo A"
    for idx in meta["indices"]:
        circuitos[idx]["Interruptor diferencial"] = dif_text

for c in circuitos:
    c["Interruptor general omnipolar"] = texto_omni

ambientes_df = pd.DataFrame(ambientes_detalle)

# -------- PARTE 3.2: CÁLCULO DE SECCIÓN Y CAÍDA DE TENSIÓN --------
zona_lower = zona.strip().lower()
canal_lower = tipo_canalizacion.strip().lower()

if "humed" in zona_lower:
    tabla_corriente = tabla_70  # zona húmeda: THWN-2 usa tabla de 70°C (la de 90°C es para THHN)
    tipo_cable_default = "THWN-2"
else:
    tabla_corriente = tabla_70  # zona seca usa la tabla de aislación 70°C
    tipo_cable_default = "H07Z1-K"

metodo_inst = "A1" if "embut" in canal_lower else "B1"

secciones_ordenadas = sorted(tabla_corriente.keys())
rho_cobre = 0.0179  # [ohm·mm²/m]

# recorre cada circuito y le calcula la sección de conductor que cumple
# tanto la capacidad de corriente como la caída de tensión máxima
for c in circuitos:
    I_circ = float(c.get("Corriente estimada (A)", 0.0))
    L_circuito = float(c.get("Longitud (m)", 0.0))
    In_tm = float(c.get("_In_TM", 0.0))
    es_ilumin = c.get("es_ilumin", False)
    es_enchufe = c.get("es_enchufe", False)
    es_clima = c.get("_es_climatizacion", False)

    # La sección/caída de tensión se calcula con la longitud REAL de la
    # canalización (L_circuito) — los chicotes son cable extra dentro de
    # las cajas, sin distancia eléctrica real, así que no deben inflar el
    # cálculo de caída de tensión. (El largo con chicotes, para saber
    # cuántos metros comprar, se recalcula aparte más adelante.)

    # Para climatización usar corriente de diseño (I_max × 1.25 ya calculada)
    # que quedó guardada en _I_diseno_clima
    if es_clima:
        I_circ_diseno = float(c.get("_I_diseno_clima", In_tm))
        if I_circ <= 0:
            c["Conductor"] = ""
            c["Caída de tensión (%)"] = np.nan
            c["Canalización"] = ""
            continue
    # Para agua caliente usar corriente de diseño (I_nom × 1.25 / ft ya calculada)
    elif c.get("_es_agua_caliente", False):
        I_circ_diseno = float(c.get("_I_diseno_agua", In_tm))
        if I_circ <= 0:
            c["Conductor"] = ""
            c["Caída de tensión (%)"] = np.nan
            c["Canalización"] = ""
            continue
    else:
        I_circ_diseno = I_circ  # circuitos normales: usan directo su corriente estimada

    if I_circ <= 0 or L_circuito <= 0 or In_tm <= 0:
        c["Conductor"] = ""
        c["Caída de tensión (%)"] = np.nan
        c["Canalización"] = ""
        continue

    sec_min = 1.5 if es_ilumin else 2.5

    # ===== PASO 1: SECCIÓN MÍNIMA POR CAÍDA DE TENSIÓN =====
    dv_max_volts = tension * 0.03
    Smin = (2 * L_circuito * I_circ * rho_cobre) / dv_max_volts if dv_max_volts > 0 else sec_min

    # respetar mínimo normativo
    Sbase = max(sec_min, Smin)

    # ===== PASO 2: BUSCAR SECCIÓN COMERCIAL =====
    seccion = None
    for s in secciones_ordenadas:
        if s >= Sbase:
            seccion = s
            break

    if seccion is None:
        seccion = secciones_ordenadas[-1]

    # ===== PASO 3: VERIFICAR AMPACIDAD Y CAÍDA =====
    while True:
        Iz = tabla_corriente[seccion][metodo_inst]

        # Corrección temperatura RIC N°4 art. 6.2.6 / Tabla N°4.7: Ic = Iz × ft
        ft_circ = factor_temperatura_ft(temperatura, metodo_inst)
        Ic = Iz * ft_circ  # ampacidad real corregida (fórmula RIC correcta)

        delta_v = 2 * I_circ * rho_cobre * L_circuito / seccion
        pct = (delta_v / tension) * 100.0 if tension > 0 else 0.0

        cumple_ampacidad = (Ic > I_circ_diseno) and (Ic >= In_tm)
        cumple_caida = pct <= 3.0

        if cumple_ampacidad and cumple_caida:
            break

        idx_sec = secciones_ordenadas.index(seccion)
        if idx_sec == len(secciones_ordenadas) - 1:
            break

        seccion = secciones_ordenadas[idx_sec + 1]

    c["Conductor"] = f"{tipo_cable_default} {seccion} mm²"
    c["Caída de tensión (%)"] = pct
    # --- N° conductores para canalización ---
    if es_ilumin:
        n_cond = n_conductores_iluminacion_para_circuito(c.get("_items", []), ambientes_df)
    elif c.get("es_enchufe", False):
        # Enchufes: máximo 6 conductores en el tramo entre cajas
        # (3 que llegan + 3 que salen hacia siguiente caja)
        # Si solo hay 1 enchufe, son 3 conductores
        n_ench_circ = sum(
            int(it.get("n_ench", 1) or 1)
            for it in c.get("_items", [])
            if isinstance(it, dict) and ("id_ench" in it or "modulos" in it or "n_ench" in it)
        )
        n_cond = 6 if n_ench_circ > 1 else 3
    else:
        n_cond = 3  # climatización, agua caliente, especiales: 3 conductores (F + N + T)

    # ── Agua caliente en Vol.1: canalización SIEMPRE conduit PVC ────────────
    # RIC N°11 art. 6.4.3 tabla Vol.1: exige mínimo IPX4, y la canaleta PVC no cumple ese requisito
    # Art. 6.5.3: cable bajo tubo aislante garantizando IPX5
    # Aplica a: ducha eléctrica (siempre Vol.1) y cualquier equipo en Vol.1
    _es_agua_vol1 = (
        c.get("_es_agua_caliente", False) and
        c.get("_vol1_bano_agua", False)
    )
    if _es_agua_vol1 and "sobre" in tipo_canalizacion.strip().lower():
        _tipo_can_agua = "Embutida"   # forzar conduit PVC
    else:
        _tipo_can_agua = tipo_canalizacion
    c["Canalización"] = canalizacion_recomendada_por_conductores(_tipo_can_agua, seccion, n_cond)

# -------- PARTE 3.3: DATAFRAMES --------
# arma la tabla de resumen general que va en la hoja "Informe" del Excel
parametros_generales = pd.DataFrame([
    ["Zona", zona],
    ["Tipo de canalización", tipo_canalizacion],
    ["Protección de empalme (A) (calculada)", interruptor_empalme],
    ["Tensión nominal (V)", tension],
    ["Factor de potencia", factor_potencia],
    ["Temperatura ambiente", temperatura],
    ["Potencia total estimada (W)", potencia_total],
    ["**SEP** PUESTA A TIERRA  (RIC 6, Tabla 6.4)", ""],
    ["Largo barra copperweld (m)", largo_barra_pt],
    ["**SEP** PT N°1 (empalme - camarilla N°1)", ""],
    ["Resistividad terreno PT1 (Ohm·m)", f"{desc_rho_pt1}"],
    ["R por barra PT1 (Ohm)", round(R_1barra_pt1, 2)],
    ["N° barras PT1", n_barras_pt1],
    ["R final PT1 (Ohm)", f"{round(R_final_pt1, 2)} ≤ 20 Ohm  cumple" if R_final_pt1 <= 20 else f"{round(R_final_pt1, 2)} > 20 Ohm  NO cumple"],
    *([ ["Separación mín. entre barras PT1 (m)", f"{sep_min_pt1:.1f}  (RIC 6, punto 8.3.2)"] ] if n_barras_pt1 > 1 else []),
    ["**SEP** PT N°2 (tablero - camarilla N°2)", ""],
    ["Resistividad terreno PT2 (Ohm·m)", f"{desc_rho_pt2}"],
    ["R por barra PT2 (Ohm)", round(R_1barra_pt2, 2)],
    ["N° barras PT2", n_barras_pt2],
    ["R final PT2 (Ohm)", f"{round(R_final_pt2, 2)} ≤ 20 Ohm  cumple" if R_final_pt2 <= 20 else f"{round(R_final_pt2, 2)} > 20 Ohm  NO cumple"],
    *([ ["Separación mín. entre barras PT2 (m)", f"{sep_min_pt2:.1f}  (RIC 6, punto 8.3.2)"] ] if n_barras_pt2 > 1 else []),
], columns=["Parámetro", "Valor"])

circuitos_df = pd.DataFrame(circuitos)
ambientes_df = pd.DataFrame(ambientes_detalle)

for columna in ["Longitud (m)", "Potencia estimada (W)", "Corriente estimada (A)"]:
    if columna in circuitos_df.columns:
        circuitos_df[columna] = pd.to_numeric(circuitos_df[columna], errors="coerce")

# Guardar tipo_dif antes del pop para usarlo en build_materiales_df
tipo_dif_por_circ = {str(c.get("Circuito","")).strip(): c.get("tipo_dif","general") for c in circuitos}

# Limpiar auxiliares antes de exportar (NO borrar "Detalle asignación")
# estos campos con "_" al inicio eran solo para uso interno del cálculo,
# no deben quedar visibles en la hoja Excel de circuitos
for c in circuitos:
    for k in ["es_ilumin", "es_enchufe", "tipo_dif", "_In_TM", "es_coc_lav", "enchufes_coc_lav", "_items",
              "_In_dif_clima", "_es_climatizacion", "_I_diseno_clima",
              "_In_dif_agua", "_sensibilidad_dif_agua", "_es_agua_caliente",
              "_I_diseno_agua", "_lleva_tablero_externo_agua", "_vol1_bano_agua", "_tipo_equipo_agua"]:
        c.pop(k, None)

# circuitos_df_con_items: después del pop para tener Canalización
# _items se provee por items_por_nombre (deepcopy hecho en línea 5989)
circuitos_df_con_items = pd.DataFrame(circuitos)

# Re-hacer circuitos_df sin _items (para hoja Informe)
circuitos_df = pd.DataFrame(circuitos)
for columna in ["Longitud (m)", "Potencia estimada (W)", "Corriente estimada (A)"]:
    if columna in circuitos_df.columns:
        circuitos_df[columna] = pd.to_numeric(circuitos_df[columna], errors="coerce")

# Reorden columnas (si existen)
columnas = list(circuitos_df.columns)
if "Interruptor diferencial" in columnas and "Interruptor general omnipolar" in columnas:
    for extra_col in ["Conductor", "Caída de tensión (%)", "Canalización"]:
        if extra_col in columnas:
            columnas.remove(extra_col)

    columnas.remove("Interruptor general omnipolar")
    idx_ins = columnas.index("Interruptor diferencial") + 1
    columnas.insert(idx_ins, "Interruptor general omnipolar")

    insert_pos = idx_ins + 1
    for extra_col in ["Conductor", "Caída de tensión (%)", "Canalización"]:
        if extra_col in circuitos_df.columns:
            columnas.insert(insert_pos, extra_col)
            insert_pos += 1

    circuitos_df = circuitos_df[columnas]

# Eliminar cualquier columna interna residual (que empiece con _) antes de exportar
cols_visibles = [c for c in circuitos_df.columns if not str(c).startswith("_")]
circuitos_df = circuitos_df[cols_visibles]

nombre_archivo = f"Informe_Instalacion_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

# -------- PARTE 4: FORMATO + AUTOAJUSTE DE ANCHO --------
with pd.ExcelWriter(nombre_archivo, engine="openpyxl") as writer:
    # a partir de acá se escribe todo al Excel: primero los DataFrames (ambientes,
    # parámetros generales, circuitos, materiales), y después se les da formato
    # con openpyxl (colores, bordes, hipervínculos, anchos de columna, etc.)
    start_amb = 0
    ambientes_df.to_excel(writer, sheet_name="Informe", index=False, startrow=start_amb)

    start_param = start_amb + len(ambientes_df) + 3
    parametros_generales.to_excel(writer, sheet_name="Informe", index=False, startrow=start_param)
    ws = writer.sheets["Informe"]
    # =========================================================
    # SECCIÓN FACTOR DE DEMANDA (al lado de Parámetro / Valor)
    # =========================================================
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    thin = Side(border_style="thin", color="AAAAAA")
    fill_header = PatternFill(start_color="C6E0B4",
                              end_color="C6E0B4",
                              fill_type="solid")
    # =========================================================
    # CUADRO CÁLCULO PROTECCIÓN GENERAL
    # =========================================================
    # --- Potencia alumbrado + enchufes (sin clima ni agua caliente) ---
    # primeros_kw, resto_kw, con_factor_primero, con_factor_resto
    # ya están calculados correctamente en el bloque global (solo alumbrado+enchufes)

    # Corrientes alumbrado + enchufes
    corriente_sin_alum = sum(
        float(c.get("Corriente estimada (A)", 0.0)) for c in circuitos
        if not any(k in str(c.get("Circuito","")).lower()
                   for k in ("climatiz","aire","split","ac ","a/c",
                              "ducha","termo","calefon","calefón","calentador","agua caliente"))
    )
    # Si toda la carga cabe en los primeros 3kW, no hay "resto":
    # toda la corriente real medida va a los primeros 3kW.
    # Si hay carga más allá de los 3kW, se convierte cada tramo
    # de potencia a corriente usando la tensión nominal.
    if pot_kw_alum_ench_global <= 3.0 or resto_kw <= 0.0:
        I_prim_sin = corriente_sin_alum
        I_rest_sin = 0.0
    else:
        I_prim_sin = (primeros_kw * 1000.0) / tension_nominal if tension_nominal else 0.0
        I_rest_sin = (resto_kw   * 1000.0) / tension_nominal if tension_nominal else 0.0
    I_prim_con = I_prim_sin * factor_primero
    I_rest_con = I_rest_sin * factor_resto

    # --- Filas dinámicas clima (fd=1,0) ---
    filas_clima = []
    for eq in (circuitos_climatizacion or []):
        kw_eq = float(eq.get("P_nom_w", 0)) / 1000.0
        I_eq  = float(eq.get("I_cuadro", eq.get("I_diseno", kw_eq * 1000.0 / tension_nominal if tension_nominal else 0.0)))
        nombre_eq = str(eq.get("nombre_circ", "Climatización")).split("(")[0].strip()
        filas_clima.append((nombre_eq, kw_eq, I_eq, 1.0, kw_eq, I_eq))

    # --- Filas dinámicas agua caliente (fd=1,0) ---
    filas_agua = []
    for eq in (circuitos_agua_caliente or []):
        kw_eq = float(eq.get("P_nom_w", 0)) / 1000.0
        I_eq  = kw_eq * 1000.0 / tension_nominal if tension_nominal else 0.0
        nombre_eq = str(eq.get("tipo_equipo", "Agua caliente")).capitalize()
        filas_agua.append((nombre_eq, kw_eq, I_eq, 1.0, kw_eq, I_eq))

    # --- Totales ---
    # cada fila de filas_clima/filas_agua trae 6 datos juntos, en este orden:
    # (nombre, kw_sin_fd, I_sin_fd, factor_demanda, kw_con_fd, I_con_fd)
    # o sea f[1]/f[2] = sin factor de demanda, f[4]/f[5] = con factor de demanda

    # kW totales SIN factor de demanda (alumbrado+enchufes + clima + agua, tal cual consumen)
    # f[1] = kW de cada equipo (clima/agua)
    total_kw_sin = pot_kw_alum_ench_global + sum(f[1] for f in filas_clima) + sum(f[1] for f in filas_agua)
    # corriente total SIN factor de demanda
    # f[2] = corriente de cada equipo (clima/agua)
    total_I_sin  = corriente_sin_alum + sum(f[2] for f in filas_clima) + sum(f[2] for f in filas_agua)
    # kW totales CON factor de demanda ya aplicado (esto es lo que realmente exige el empalme)
    # f[4] = kW con fd de cada equipo (clima/agua)
    total_kw_fd  = con_factor_primero + con_factor_resto + sum(f[4] for f in filas_clima) + sum(f[4] for f in filas_agua)
    # corriente total CON factor de demanda ya aplicado
    # f[5] = corriente con fd de cada equipo (clima/agua)
    total_I_con  = I_prim_con + I_rest_con + sum(f[5] for f in filas_clima) + sum(f[5] for f in filas_agua)

    fill_titulo  = PatternFill(start_color="9DC3E6", end_color="9DC3E6", fill_type="solid")
    fill_gris    = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    fill_blanco  = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    borde_gris   = Border(top=Side(style="thin", color="BFBFBF"),
                          left=Side(style="thin", color="BFBFBF"),
                          right=Side(style="thin", color="BFBFBF"),
                          bottom=Side(style="thin", color="BFBFBF"))

    row_base   = start_param + 1  # fila donde arranca el cuadro (justo debajo de los parámetros generales)
    col_inicio = 4  # columna D

    # ---- TÍTULO ----
    ws.merge_cells(start_row=row_base, start_column=col_inicio,
                   end_row=row_base,   end_column=col_inicio + 5)
    tc = ws.cell(row=row_base, column=col_inicio)
    tc.value     = "CUADRO CÁLCULO PROTECCIÓN GENERAL"
    tc.font      = Font(bold=True, color="000000")
    tc.alignment = Alignment(horizontal="center", vertical="center")
    tc.fill      = fill_titulo
    for c in range(col_inicio, col_inicio + 6):
        ws.cell(row=row_base, column=c).border = borde_gris
        ws.cell(row=row_base, column=c).fill   = fill_titulo
    ws.row_dimensions[row_base].height = 18

    # ---- ENCABEZADOS ----
    col_headers = ["Tablero", "KW", "In [A]", "f/d", "KW", "In [A]"]
    for j, h in enumerate(col_headers):
        cell = ws.cell(row=row_base+1, column=col_inicio+j)
        cell.value     = h
        cell.font      = Font(bold=True, color="000000")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill      = fill_gris
        cell.border    = borde_gris

    # ---- FILAS DE DATOS ----
    filas_datos = [
        ("T.D.A PRIMEROS 3KW", primeros_kw, I_prim_sin, factor_primero, con_factor_primero, I_prim_con),
        (f"T.D.A RESTO {resto_kw:.3f} KW", resto_kw, I_rest_sin, factor_resto, con_factor_resto, I_rest_con),
    ] + filas_clima + filas_agua

    for i, (nombre, kw_sin, in_sin, fd, kw_con, in_con) in enumerate(filas_datos):
        r = row_base + 2 + i
        kw_con_r  = round(kw_con, 3)
        in_con_r  = round(in_con, 2)
        kw_sin_r  = round(kw_sin, 3)
        in_sin_r  = round(in_sin, 2)
        valores = [nombre, kw_sin_r, in_sin_r, fd, kw_con_r, in_con_r]
        for j, v in enumerate(valores):
            cell = ws.cell(row=r, column=col_inicio+j)
            cell.value     = v
            cell.font      = Font(bold=(j==0), color="000000")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.fill      = fill_blanco
            cell.border    = borde_gris
        ws.row_dimensions[r].height = 28

    # ---- FILA TOTAL ----
    row_total = row_base + 2 + len(filas_datos)
    total_kw_sin_r = round(total_kw_sin, 3)
    total_kw_fd_r  = round(total_kw_fd, 3)
    total_I_sin_r  = round(total_I_sin, 2)
    total_I_con_r  = round(total_I_con, 2)
    total_vals = ["TOTAL", total_kw_sin_r, total_I_sin_r, "", total_kw_fd_r, total_I_con_r]
    for j, v in enumerate(total_vals):
        cell = ws.cell(row=row_total, column=col_inicio+j)
        cell.value     = v
        cell.font      = Font(bold=True, color="000000")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill      = fill_blanco
        cell.border    = borde_gris
    ws.row_dimensions[row_total].height = 18


    # =========================================================
    # CÁLCULO ACOMETIDA
    # =========================================================
    # Calculo acometida
    tipo_acometida_calc = "aereo" if tipo_acometida == "aerea" else "subterraneo"
    res_acom = seleccionar_acometida(
        L_m=float(longitud_transformador_empalme),
        I_empalme_A=float(interruptor_empalme),
        fp=float(fp),
        V_nom=float(tension_nominal),
        tipo_acometida=tipo_acometida_calc,
        temp_override=temperatura_suelo if "sub" in str(tipo_acometida_calc).lower() else None
    )
    # Texto final
    acometida_txt = f"Concéntrico Cu 2x {res_acom['S']} mm²"

    # =========================================================
    # CANALIZACIÓN ACOMETIDA
    # =========================================================
    tac = str(tipo_acometida).strip().lower()
    if "sub" in tac:
        n_cond_acom = 2  # Concéntrico = 2 conductores (F+N)
        d_ducto_acom = ducto_nominal_tablas(res_acom["S"], n_cond_acom, "subterraneo")
        canalizacion_acom_txt = f"Ø PVC Conduit {d_ducto_acom} mm" if d_ducto_acom else ""
    else:
        canalizacion_acom_txt = "-"

    #==========================================================
    #CÁLCULO DE CAÍDA DE TENSIÓN ACOMETIDA
    #==========================================================
    caida_acometida_txt = f"{res_acom['dV_pct']:.2f}%"

    # =========================================================
    # EMPALME NORMALIZADO
    # =========================================================
    # Determinar número del empalme
    if interruptor_empalme <= 30:
        emp_num = 6
    elif interruptor_empalme <= 40:
        emp_num = 9
    else:
        emp_num = 16
    # Letra según acometida
    if tipo_acometida == "aerea":
        empalme_txt = f"A-{emp_num}"
    else:
        empalme_txt = f"S-{emp_num}"

    # =========================================================
    # POTENCIA NOMINAL SEGÚN EMPALME
    # =========================================================
    tabla_potencia_empalme = {6: 1, 10: 2, 16: 3, 20: 4, 25: 5, 30: 6, 32: 6.5, 35: 7, 40: 8, 50: 10, 63: 13}
    potencia_nominal_empalme = tabla_potencia_empalme.get(interruptor_empalme, "")

    # =========================================================
    # CÁLCULO ALIMENTADOR
    # =========================================================
    # Caídas de tensión de cada circuito (para verificar ΔV total <= 5%)
    _dv_circs = []
    for _c in circuitos:
        _dv = _c.get("Caída de tensión (%)", None)
        try:
            _dv_f = float(_dv)
            if _dv_f == _dv_f:  # filtrar NaN
                _dv_circs.append(_dv_f)
        except (TypeError, ValueError):
            pass

    res_alim = seleccionar_alimentador(
        L_m=float(longitud_alimentador),
        I_demanda_A=float(corriente_con),
        I_empalme_A=float(interruptor_empalme),
        fp=float(fp),
        V_nom=float(tension_nominal),
        tipo_alimentador=tipo_alimentador,
        dv_circuitos=_dv_circs,
        temp_override=temperatura_suelo if "sub" in str(tipo_alimentador).lower() else None
    )
    alim_txt = f"RV-K Cu 3x{res_alim['S']} mm²"
    canal_txt = f"{tipo_alimentador} (Método {res_alim['metodo']})"
    caida_txt = f"{res_alim['dV_pct']:.2f}%"
    # Canalización real (Ø ducto) según tablas
    n_cond_alim = 3  # F + N + PE
    d_ducto = ducto_nominal_tablas(res_alim["S"], n_cond_alim, tipo_alimentador)
    if d_ducto is None:
        canalizacion_txt = "-"              # si es aéreo, no lleva canalización
    else:
        canalizacion_txt = f"Ø PVC Conduit {d_ducto} mm"

    # >>> CREAR MATERIALES ANTES DE BORRAR _items (PASANDO tipo_canalizacion)
    # acá se llama a la función más grande del programa: recorre todos los
    # circuitos ya calculados y arma la lista completa de materiales a comprar
    materiales_df = build_materiales_df(circuitos_df_con_items, texto_omni, ambientes_df, group_info, tipo_canalizacion,
                                    tipo_alimentador, tipo_acometida, tipo_instalacion_empalme, requiere_mastil,
                                    acometida_txt, interruptor_texto, longitud_transformador_empalme, canalizacion_txt,
                                    dist_empalme_pt1, dist_tda_pt2, longitud_subterraneo_medidor, longitud_mastil,
                                    altura_acometida_aerea, longitud_subterraneo_medidor2, longitud_llegada_aerea_tda,
                                    longitud_alimentador, longitud_abrazaderas_alimentador, longitud_poste_alimentador_aereo,
                                    dist_vertical_acometida, alim_txt,
                                    circuitos_climatizacion=circuitos_climatizacion,
                                    items_por_nombre=items_por_nombre,
                                    cajas_adic_por_nombre=cajas_adic_por_nombre,
                                    tipo_dif_por_circ=tipo_dif_por_circ,
                                    n_barras_pt1=n_barras_pt1, n_barras_pt2=n_barras_pt2,
                                    long_cond_desnudo_pt1=long_cond_desnudo_pt1,
                                    long_cond_desnudo_pt2=long_cond_desnudo_pt2)

    # =========================================================
    # TABLA DE FÓRMULAS (al lado derecho del bloque factor demanda)
    # =========================================================
    col_form = col_inicio + 7   # deja espacio después de la tabla de 6 columnas (col_inicio a col_inicio+5) + 1 libre

    fill_form_hdr  = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    fill_form_title= PatternFill(start_color="9DC3E6", end_color="9DC3E6", fill_type="solid")

    formulas_data = [
        # (título de sección, fórmula, descripción corta)
        ("CORRIENTE DE CIRCUITO",
         "I = P / (V · fp)",
         "I: corriente [A]  |  P: potencia [W]  |  V: tensión nominal [V]  |  fp: factor de potencia (1.0 para cargas resistivas)"),
        ("TERMOMAGNÉTICO CIRCUITO",
         "In_TM ≥ I_circ",
         "In_TM: calibre termomagnético [A]  |  I_circ: corriente de circuito [A]  |  Calibres: 6, 10, 16 A (ilum.) / 10, 16 A (ench.)  |  RIC 5, Tabla 5.1"),
        ("TERMOMAGNÉTICO ESPECIAL",
         "In_TM ≥ 1.10 × I_circ  (mín. 16 A)",
         "In_TM: calibre termomagnético [A]  |  I_circ: corriente de circuito [A]  |  10% de reserva  |  Mínimo 16 A para agua caliente y climatización"),
        ("SECCIÓN MÍNIMA POR ΔV",
         "Smin = (2 · L · I · ρ) / ΔVmax",
         "Smin: sección mínima [mm²]  |  L: longitud tramo [m]  |  I: corriente [A]  |  ρ = 0.0179 Ω·mm²/m (Cu)  |  ΔVmax = V × 3% [V]  |  Mínimo: 1.5 mm² (ilum.) / 2.5 mm² (ench.)"),
        ("CAÍDA DE TENSIÓN",
         "ΔV% = (2 · ρ · L · I) / (S · V) × 100",
         "ΔV%: caída de tensión [%]  |  ρ = 0.0179 Ω·mm²/m (Cu)  |  L: longitud tramo [m]  |  I: corriente [A]  |  S: sección conductor [mm²]  |  V: tensión nominal [V]  |  Límite: ≤ 3%"),
        ("AMPACIDAD CONDUCTOR",
         "Iz ≥ I_circ  y  Iz ≥ In_TM",
         "Iz: ampacidad corregida [A]  |  Iz = Iz_tabla × ft  |  Iz_tabla: ampacidad tabulada [A]  |  ft: factor corrección temperatura [-]  |  I_circ: corriente circuito [A]  |  In_TM: calibre TM [A]  |  RIC 5, Tabla 5.4"),
        ("FACTOR DE DEMANDA",
         "Pd = P₁ × 1.0 + P₂ × 0.35",
         "Pd: potencia de demanda [W]  |  P₁ = min(Ptotal, 3000) [W]: primeros 3 kW al 100%  |  P₂ = max(Ptotal − 3000, 0) [W]: resto al 35%  |  RIC 3, Art. 6.1, 6.2, 6.3"),
        ("CORRIENTE EMPALME",
         "I_emp = Pd / (V × fp)",
         "I_emp: corriente de empalme [A]  |  Pd: potencia de demanda [W]  |  V: tensión nominal [V]  |  fp: factor de potencia [-]  |  In normalizado: 25 / 32 / 40 / 50 / 63 A"),
        ("ALIMENTADOR (Smin)",
         "Smin = (2 · L · I_dem · fp · ρ) / ΔVmax",
         "Smin: sección mínima [mm²]  |  L: longitud alimentador [m]  |  I_dem: corriente de demanda [A]  |  fp: factor de potencia [-]  |  ρ = 0.0179 Ω·mm²/m  |  ΔVmax = V × 3% [V]  |  Cumple: Iz≥I_dem, Iz≥I_emp y ΔV≤3%  |  Mínimo 4 mm²"),
        ("ACOMETIDA (Smin)",
         "Smin = (2 · L · I_emp · fp · ρ) / ΔVmax",
         "Smin: sección mínima [mm²]  |  L: longitud acometida [m]  |  I_emp: corriente de empalme [A]  |  fp: factor de potencia [-]  |  ρ = 0.0179 Ω·mm²/m  |  ΔVmax = V × 3% [V]  |  Método E (aéreo) o D1 (subterráneo)  |  Mínimo 4 mm²"),
        ("PUESTA A TIERRA (RIC 6)", "", ""),
        ("RESIST. UNA PICA VERTICAL",
         "R₁ = ρ / L",
         "R₁: resistencia una barra [Ohm]  |  ρ: resistividad del terreno [Ohm·m]  |  L: largo de la barra [m]  |  RIC 6, Tabla 6.4"),
        ("N° BARRAS NECESARIAS",
         "N = ⌈R₁ / 20⌉  (mín. 1)",
         "N: número de barras [-]  |  R₁: resistencia una barra [Ohm]  |  20: resistencia máxima permitida [Ohm]  |  Barras en paralelo  |  RIC 6, punto 6.1"),
        ("RESIST. FINAL (N barras)",
         "R_final = R₁ / N",
         "R_final: resistencia sistema [Ohm]  |  R₁: resistencia una barra [Ohm]  |  N: número de barras [-]  |  Separación mínima entre barras = 2 × L  |  RIC 6, punto 8.3.2"),
        ("CONDUCTOR DESNUDO (unión barras)",
         "L_desnudo = (N − 1) × (2 · L_barra)",
         "L_desnudo: longitud conductor desnudo [m]  |  N: número de barras [-]  |  L_barra: largo de cada barra [m]  |  Material: Cu desnudo 16 mm²  |  RIC 6"),
    ]

    # Título principal de la tabla
    ws.merge_cells(start_row=row_base, start_column=col_form,
                   end_row=row_base, end_column=col_form + 2)
    tc = ws.cell(row=row_base, column=col_form)
    tc.value = "FÓRMULAS UTILIZADAS EN LOS CÁLCULOS"
    tc.font = Font(bold=True, color="000000")
    tc.alignment = Alignment(horizontal="center", vertical="center")
    tc.fill = fill_form_title
    tc.border = Border(top=thin, left=thin, right=thin, bottom=thin)
    for _c in range(col_form + 1, col_form + 3):
        _cell = ws.cell(row=row_base, column=_c)
        _cell.fill = fill_form_title
        _cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    # Encabezados de columna
    for _j, _hdr in enumerate(["Parámetro / Etapa", "Fórmula", "Notas"]):
        _cell = ws.cell(row=row_base + 1, column=col_form + _j)
        _cell.value = _hdr
        _cell.font = Font(bold=True)
        _cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        _cell.fill = fill_form_hdr
        _cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    # Filas de datos (una fila más ancha y en negrita para el separador "PUESTA A TIERRA")
    fill_form_sep = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    for _i, (etapa, formula, notas) in enumerate(formulas_data):
        _row = row_base + 2 + _i
        es_separador = etapa.startswith("PUESTA A TIERRA")
        for _j, _val in enumerate([etapa, formula, notas]):
            _cell = ws.cell(row=_row, column=col_form + _j)
            _cell.value = _val
            _cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            if es_separador:
                _cell.font = Font(bold=True, color="000000")
                _cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                _cell.fill = fill_form_sep
            else:
                _cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                if _j == 1:   # columna fórmula: Calibri, sin negrita, centrada
                    _cell.font = Font(bold=False, name="Calibri")
                    _cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if es_separador:
            ws.merge_cells(start_row=_row, start_column=col_form,
                           end_row=_row, end_column=col_form + 2)

    from openpyxl.utils import get_column_letter as _gcl

    # ahora escribe la tabla de circuitos, debajo de todo lo anterior
    start_circ = start_param + len(parametros_generales) + 8
    circuitos_df.to_excel(writer, sheet_name="Informe", index=False, startrow=start_circ)

    wb = writer.book
    ws = writer.sheets["Informe"]

    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    def aplicar_estilo(fila_inicio, nrows, ncols, filtro=False):
        # le da formato a una tabla de la hoja "Informe": encabezado en negrita,
        # bordes, ancho de columna automático y filtro si se pide
        hdr = fila_inicio + 1  # fila del encabezado (justo debajo de donde empieza la tabla)
        fila_final = hdr + nrows  # última fila de datos de la tabla
        thin = Side(border_style="thin", color="AAAAAA")  # estilo de borde fino gris

        # formatea la fila de encabezado: negrita, centrado, fondo celeste, con borde
        for ccol in range(1, ncols + 1):
            cell = ws.cell(row=hdr, column=ccol)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.fill = PatternFill(start_color="E6F2FF", end_color="E6F2FF", fill_type="solid")
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

        # recorre columna por columna: le pone borde a los datos y calcula el ancho
        for ccol in range(1, ncols + 1):
            col_letter = get_column_letter(ccol)
            max_len = 0  # va guardando el texto más largo de la columna

            for r in range(hdr + 1, fila_final + 1):
                cell = ws.cell(row=r, column=ccol)
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                cell.alignment = Alignment(vertical="top", wrap_text=True)

                val = cell.value
                if val is not None:
                    max_len = max(max_len, max(len(str(x)) for x in str(val).split("\n")))  # por si el texto tiene saltos de línea

            val_hdr = ws.cell(row=hdr, column=ccol).value
            if val_hdr is not None:
                max_len = max(max_len, len(str(val_hdr)))  # el título también cuenta para el ancho

            ws.column_dimensions[col_letter].width = min(max(12, max_len + 4), 80)  # ancho mínimo 12, máximo 80

        if filtro:
            ws.auto_filter.ref = f"A{hdr}:{get_column_letter(ncols)}{fila_final}"  # agrega el autofiltro de Excel

    aplicar_estilo(start_amb, len(ambientes_df), ambientes_df.shape[1], filtro=True)
    aplicar_estilo(start_param, len(parametros_generales), parametros_generales.shape[1])
    fila_ini = start_param + 2
    fila_fin = start_param + 1 + len(parametros_generales)
    # recorre la tabla de parámetros generales buscando filas marcadas con
    # "**SEP**" (así se armaron en parametros_generales, más arriba): esas
    # filas no son un dato más, son un título de subsección, así que se
    # fusionan en 1 sola celda centrada en vez de quedar como fila normal
    for r in range(fila_ini, fila_fin + 1):
        val_param = ws.cell(row=r, column=1).value
        if val_param and str(val_param).startswith("**SEP**"):
            texto_sep = str(val_param).replace("**SEP** ", "")
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
            cell_sep = ws.cell(row=r, column=1)
            cell_sep.value = texto_sep
            cell_sep.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
            cell_sep.font = Font(bold=True)
            thin = Side(border_style="thin", color="AAAAAA")
            cell_sep.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            ws.row_dimensions[r].height = 15
        else:
            ws.cell(row=r, column=1).alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
            ws.cell(row=r, column=2).alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
            ws.row_dimensions[r].height = 15

    aplicar_estilo(start_circ, len(circuitos_df), circuitos_df.shape[1])
    # Anchos fijos DESPUÉS de todos los aplicar_estilo para que no sean pisados
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 55
    # Anchos tabla de fórmulas (forzar AQUÍ para que no sean pisados por aplicar_estilo)
    ws.column_dimensions[_gcl(col_form)].width     = 32
    ws.column_dimensions[_gcl(col_form + 1)].width = 35
    ws.column_dimensions[_gcl(col_form + 2)].width = 65
    # Columnas E-I
    for _col_letra in ["E", "F", "G", "H", "I"]:
        ws.column_dimensions[_col_letra].width = 23
    # Altura filas tabla de fórmulas (individual)
    _form_heights = [33, 33, 33, 53, 53, 53, 50, 53, 63, 63, 15, 33, 48, 50, 33]
    for _i, _h in enumerate(_form_heights):
        ws.row_dimensions[row_base + 2 + _i].height = _h

    # Formato porcentaje numérico para "Caída de tensión (%)"
    if "Caída de tensión (%)" in circuitos_df.columns:
        col_caida = list(circuitos_df.columns).index("Caída de tensión (%)") + 1
        fila_ini = start_circ + 2
        fila_fin = start_circ + 1 + len(circuitos_df)
        for r in range(fila_ini, fila_fin + 1):
            ws.cell(row=r, column=col_caida).number_format = '0.00"%"'

  # =========================================================
    # MERGE de celdas en "Informe" (SIN tocar Materiales)
    # - Omnipolar: una sola celda para todos los circuitos
    # - Diferencial: merge por grupos (mismo texto consecutivo)
    # =========================================================

    def _find_col_idx_by_name(df, nombre_columna):
        """Devuelve índice 1-based de la columna 'nombre_columna' en el DataFrame, o None si no existe."""
        try:
            return list(df.columns).index(nombre_columna) + 1  # +1 porque Excel empieza en columna 1, no 0
        except:
            return None  # esa columna no existe en el DataFrame

    def merge_vertical_runs(ws, col_idx, first_data_row, last_data_row, merge_all=False):
        """
        Combina verticalmente celdas en una columna:
        - Si merge_all=True: merge desde first_data_row a last_data_row.
        - Si merge_all=False: merge por 'runs' consecutivos con el mismo valor (no vacío).
        """
        if col_idx is None:
            return  # no hay columna, no hace nada

        col_letter = get_column_letter(col_idx)

        if merge_all:
            # modo simple: fusiona todo el rango de una sola vez
            if last_data_row > first_data_row:
                ws.merge_cells(f"{col_letter}{first_data_row}:{col_letter}{last_data_row}")
                c = ws.cell(row=first_data_row, column=col_idx)
                c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            return

        # modo "por tramos": va agrupando filas seguidas que tengan el mismo valor
        fila_tramo_inicio = first_data_row  # primera fila del tramo actual
        prev_val = ws.cell(row=first_data_row, column=col_idx).value  # valor del tramo actual

        def _is_empty(v):
            # revisa si la celda está vacía (None o solo espacios)
            return v is None or str(v).strip() == ""

        for r in range(first_data_row + 1, last_data_row + 1):
            cur_val = ws.cell(row=r, column=col_idx).value

            if str(cur_val).strip() != str(prev_val).strip():
                # cambió el valor: cierra el tramo anterior y lo fusiona (si tenía más de 1 fila)
                if (r - 1) > fila_tramo_inicio and (not _is_empty(prev_val)):
                    ws.merge_cells(f"{col_letter}{fila_tramo_inicio}:{col_letter}{r-1}")
                    c = ws.cell(row=fila_tramo_inicio, column=col_idx)
                    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                fila_tramo_inicio = r  # empieza un tramo nuevo
                prev_val = cur_val

        # cierra el último tramo, después de terminar el for
        if last_data_row > fila_tramo_inicio and (not _is_empty(prev_val)):
            ws.merge_cells(f"{col_letter}{fila_tramo_inicio}:{col_letter}{last_data_row}")
            c = ws.cell(row=fila_tramo_inicio, column=col_idx)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # --- calcular filas de la tabla de circuitos en la hoja "Informe"
    hdr_circ_row = start_circ + 1                    # header
    first_data_row = hdr_circ_row + 1                # primera fila datos
    last_data_row = hdr_circ_row + len(circuitos_df) # última fila datos

    # --- ubicar columnas por nombre
    col_dif = _find_col_idx_by_name(circuitos_df, "Interruptor diferencial")
    col_omni = _find_col_idx_by_name(circuitos_df, "Interruptor general omnipolar")

    # merge omnipolar (una celda para todos los circuitos)
    merge_vertical_runs(ws, col_omni, first_data_row, last_data_row, merge_all=True)

    # merge diferenciales por grupo (group_info)
    def merge_diferenciales_por_grupo(ws, col_idx, first_data_row, group_info):
        # junta en una sola celda las filas de circuitos que comparten el
        # mismo diferencial, para que no se repita el texto
        if col_idx is None or not group_info:
            return  # no hay columna o no hay grupos, no hay nada que hacer

        col_letter = get_column_letter(col_idx)  # letra de columna Excel (ej: "F")

        for gid, meta in group_info.items():
            idxs = sorted(meta.get("indices", []))  # filas (posiciones) de los circuitos de este grupo
            if not idxs:
                continue  # grupo vacío, se salta

            start_r = first_data_row + idxs[0]  # primera fila del grupo en la hoja
            end_r   = first_data_row + idxs[-1]  # última fila del grupo en la hoja

            dif_txt = f"2X{meta['dif']} {meta.get('sensibilidad_dif', '30mA')} / Tipo A"  # texto del diferencial

            if end_r > start_r:
                ws.merge_cells(f"{col_letter}{start_r}:{col_letter}{end_r}")  # fusiona las celdas del grupo

            c = ws.cell(row=start_r, column=col_idx)
            c.value = dif_txt  # el texto queda solo en la primera celda fusionada
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # aplicar merge por grupo
    merge_diferenciales_por_grupo(ws, col_dif, first_data_row, group_info)


    # =========================================================
    # ENCABEZADO SECCIÓN EMPALME (sin merge)
    # =========================================================
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    thin = Side(border_style="thin", color="AAAAAA")
    fill_header = PatternFill(start_color="E6F2FF",
                              end_color="E6F2FF",
                              fill_type="solid")
    # última fila real de la tabla de circuitos
    hdr_circ_row = start_circ + 1
    last_data_row = hdr_circ_row + len(circuitos_df)
    # dejar 1 fila en blanco y escribir encabezado
    row_emp = last_data_row + 2
    headers_emp = [
        "Empalme",
        "Tarifa",
        "Pot. Nominal (kW)",
        "Acometida",
        "Longitud (m)",
        "Disyuntor termomagnético",
        "Caída de tensión",
        "Canalización"
    ]
    for columna, txt in enumerate(headers_emp, start=1):
        cell = ws.cell(row=row_emp, column=columna)
        cell.value = txt
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center",
                                   vertical="center",
                                   wrap_text=True)
        cell.fill = fill_header
        cell.border = Border(top=thin,
                             left=thin,
                             right=thin,
                             bottom=thin)
        row_emp_data = row_emp + 1

    # escribe los datos del empalme, celda por celda, debajo de cada encabezado
    ws.cell(row=row_emp_data, column=1).value = empalme_txt
    cell_empalme = ws.cell(row=row_emp_data, column=1)
    cell_empalme.alignment = Alignment(horizontal="center", vertical="center")
    cell_empalme.border = Border(top=thin, left=thin, right=thin, bottom=thin)
    ws.cell(row=row_emp_data, column=2).value = "BT-1"
    ws.cell(row=row_emp_data, column=2).alignment = Alignment(horizontal="center")
    cell_pot = ws.cell(row=row_emp_data, column=3)
    cell_pot.value = potencia_nominal_empalme
    cell_pot.alignment = Alignment(horizontal="center", vertical="center")
    cell_pot.border = Border(top=thin, left=thin, right=thin, bottom=thin)
    cell_acom = ws.cell(row=row_emp_data, column=4)
    cell_acom.value = acometida_txt
    cell_acom.alignment = Alignment(horizontal="center", vertical="center")
    cell_acom.border = Border(top=thin, left=thin, right=thin, bottom=thin)
    ws.cell(row=row_emp_data, column=6).value = interruptor_texto
    ws.cell(row=row_emp_data, column=6).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=row_emp_data, column=6).border = Border(top=thin, left=thin, right=thin, bottom=thin)
    cel_long = ws.cell(row=row_emp_data, column=5)
    cel_long.value = f"{longitud_transformador_empalme:.2f}"
    cel_long.alignment = Alignment(horizontal="center", vertical="center")
    cel_long.border = Border(top=thin, left=thin, right=thin, bottom=thin)
    cell_tarifa = ws.cell(row=row_emp_data, column=2)
    cell_tarifa.border = Border(top=thin, left=thin, right=thin, bottom=thin)
    cell_dv_acom = ws.cell(row=row_emp_data, column=7)
    cell_dv_acom.value = caida_acometida_txt
    cell_dv_acom.alignment = Alignment(horizontal="center", vertical="center")
    cell_dv_acom.border = Border(top=thin, left=thin, right=thin, bottom=thin)
    cell_canal_acom = ws.cell(row=row_emp_data, column=8)
    cell_canal_acom.value = canalizacion_acom_txt
    cell_canal_acom.alignment = Alignment(horizontal="center", vertical="center")
    cell_canal_acom.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    # =========================================================
    # SECCIÓN ALIMENTADOR (con 1 fila en blanco antes)
    # =========================================================
    row_alim = row_emp_data + 2
    headers_alim = ["Alimentador", "Tipo de alimentador", "Canalización", "Longitud (m)", "Caída de tensión"]
    # Encabezado con mismo formato
    for columna, txt in enumerate(headers_alim, start=1):
        cell = ws.cell(row=row_alim, column=columna)
        cell.value = txt
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = fill_header
        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
    # Fila de datos debajo (con bordes)
    row_alim_data = row_alim + 1
    for columna in range(1, len(headers_alim) + 1):
        cell = ws.cell(row=row_alim_data, column=columna)
        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.cell(row=row_alim_data, column=1).value = alim_txt
    ws.cell(row=row_alim_data, column=2).value = f"{tipo_alimentador} (Método {res_alim['metodo']})"
    ws.cell(row=row_alim_data, column=3).value = canalizacion_txt
    ws.cell(row=row_alim_data, column=4).value = f"{longitud_alimentador:.2f}"
    ws.cell(row=row_alim_data, column=5).value = caida_txt
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 35

    # =========================================================
    # TABLA VERIFICACION CAIDA DE TENSION TOTAL (alim + circ <= 5%)
    # =========================================================
    _dv_alim_pct = res_alim.get("dV_pct", 0.0)
    _fill_title_dv  = PatternFill(start_color="E6F2FF", end_color="E6F2FF", fill_type="solid")
    _fill_hdr_dv    = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    _font_calibri_b = Font(name="Calibri", size=10, bold=True)
    _font_calibri   = Font(name="Calibri", size=10)
    _hdrs_dv = ["Circuito", "ΔV circuito (%)", "ΔV alimentador (%)", "ΔV total (%)", "¿Cumple ≤ 5%?"]
    _ncols_dv = len(_hdrs_dv)

    row_dv_title = row_alim_data + 2
    ws.merge_cells(start_row=row_dv_title, start_column=1,
                   end_row=row_dv_title, end_column=_ncols_dv)
    _tc = ws.cell(row=row_dv_title, column=1)
    _tc.value     = "VERIFICACIÓN CAÍDA DE TENSIÓN TOTAL"
    _tc.font      = Font(name="Calibri", size=10, bold=True)
    _tc.alignment = Alignment(horizontal="center", vertical="center")
    _tc.fill      = _fill_title_dv
    _tc.border    = Border(top=thin, left=thin, right=thin, bottom=thin)
    for _c in range(2, _ncols_dv + 1):
        _cell = ws.cell(row=row_dv_title, column=_c)
        _cell.fill   = _fill_title_dv
        _cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    row_dv_hdr = row_dv_title + 1
    for _j, _h in enumerate(_hdrs_dv, start=1):
        _cell = ws.cell(row=row_dv_hdr, column=_j)
        _cell.value     = _h
        _cell.font      = _font_calibri_b
        _cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        _cell.fill      = _fill_hdr_dv
        _cell.border    = Border(top=thin, left=thin, right=thin, bottom=thin)

    row_dv_data = row_dv_hdr + 1
    if "Circuito" in circuitos_df.columns and "Caída de tensión (%)" in circuitos_df.columns:
        # recorre cada circuito y verifica que la caída de tensión total
        # (circuito + alimentador) no supere el 5% máximo permitido
        for _idx, _row in circuitos_df.iterrows():
            _nombre  = _row.get("Circuito", "")
            _dv_circ = _row.get("Caída de tensión (%)", None)
            try:
                _dv_circ_f = float(_dv_circ)
                if _dv_circ_f != _dv_circ_f:  # es NaN
                    continue
            except (TypeError, ValueError):
                continue
            _dv_total_f = round(_dv_circ_f + _dv_alim_pct, 2)
            _cumple_txt = "cumple" if _dv_total_f <= 5.0 else "NO cumple"
            _vals = [_nombre, f"{_dv_circ_f:.2f}%", f"{_dv_alim_pct:.2f}%", f"{_dv_total_f:.2f}%", _cumple_txt]
            for _j, _v in enumerate(_vals, start=1):
                _cell = ws.cell(row=row_dv_data, column=_j)
                _cell.value     = _v
                _cell.font      = _font_calibri
                _cell.alignment = Alignment(
                    horizontal="left" if _j == 1 else "center",
                    vertical="center", wrap_text=True)
                _cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            row_dv_data += 1
    # =========================
    # HOJA 2: MATERIALES (PRO)
    # =========================
    # Calcular Sello SEC basado en descripción Y norma original
    def _requiere_sello_sec(desc, norma):
        # decide qué poner en la columna "Sello SEC": si ya venía marcado como
        # SEC lo deja igual, y si es un consumible (cinta, sellador, etc.)
        # no le pide sello
        d = str(desc or "").lower()
        n = str(norma or "").strip()
        if n == "SEC":
            return "SEC"
        if not d or d in ["nan", "none", ""]:
            return ""
        # Los consumibles de instalación no llevan Sello SEC
        if any(x in d for x in ["tubo de estaño", "pasta para soldar", "cinta aislante", "cinta autofundente", "sellador", "espuma", "teflón", "teflon"]):
            return "-"
        # Conductores y cables
        if any(x in d for x in ["conductor", "cable", "thwn-2", "rv-k", "nyyj", "nyy",
                                  "concéntrico", "concentrico", "h07z1", "desnudo cu"]):
            return "SEC"
        # Protecciones
        if any(x in d for x in ["interruptor", "diferencial", "termomagnético",
                                  "termomagnetico", "disyuntor", "omnipolar",
                                  "automático", "automatico"]):
            return "SEC"
        # Enchufes
        if "enchufe" in d:
            return "SEC"
        # Tableros
        if "tablero" in d:
            return "SEC"
        # Cajas de derivación y empalme
        if any(x in d for x in ["caja de derivación", "caja de derivacion",
                                  "caja de empalme"]):
            return "SEC"
        # Tapas de cajas
        if "tapa ciega" in d:
            return "SEC"
        # Conduit y accesorios
        if any(x in d for x in ["conduit", "canaleta pvc", "tubo conduit",
                                  "tubo galvanizado", "salida de caja",
                                  "unión copla", "union copla",
                                  "curva", "terminal pvc conduit"]):
            return "SEC"
        # Abrazaderas
        if "abrazadera" in d:
            return "SEC"
        # Prensaestopas y boquillas
        if any(x in d for x in ["prensaestopa", "boquilla"]):
            return "SEC"
        # Medidor
        if any(x in d for x in ["medidor", "medida monofásica", "medida trifásica"]):
            return "SEC"
        # Portafusibles y fusibles
        if any(x in d for x in ["portafusible", "fusible"]):
            return "SEC"
        # Barras repartidoras y unipolares
        if any(x in d for x in ["barra unipolar", "barra repartidora"]):
            return "SEC"
        # Borneras
        if "bornera" in d:
            return "SEC"
        # Terminales
        if any(x in d for x in ["terminal ferrul", "terminal de compresión",
                                  "terminal de compresion"]):
            return "SEC"
        # Supresor de transiente
        if "supresor" in d:
            return "SEC"
        # Protector sobrevoltaje
        if "protector sobrevoltaje" in d or "protector de sobrevoltaje" in d:
            return "SEC"
        # Portalámparas y luminarias
        if any(x in d for x in ["portalámpara", "portalampara", "luminaria",
                                  "ampolleta", "led", "foco"]):
            return "SEC"
        # Riel DIN
        if "riel din" in d:
            return "SEC"
        # Mordaza
        if "mordaza" in d:
            return "SEC"
        # Hub / conector acometida
        if any(x in d for x in ["hub", "conector hub"]):
            return "SEC"
        # Alimentador
        if "alimentador" in d:
            return "SEC"
        # Punto conexión fija / directa
        if "punto para conexión" in d or "punto para conexion" in d:
            return "SEC"
        # Todo lo demás no requiere Sello SEC
        return "-"

    if "Sello SEC" in materiales_df.columns:
        # aplica la función a cada fila para decidir si necesita Sello SEC
        materiales_df["Sello SEC"] = materiales_df.apply(
            lambda rr: _requiere_sello_sec(rr.get("Descripción técnica", ""), rr.get("Norma / RIC", "")),
            axis=1
        )

    # Limpia la columna "Norma / RIC":
    # Ejemplo: "RIC 4.7.2" se deja como "RIC 4" para que sea legible y luego tenga hipervínculo.
    if "Norma / RIC" in materiales_df.columns:
        materiales_df["Norma / RIC"] = materiales_df.apply(
            lambda rr: normalizar_ric_materiales(
                rr.get("Norma / RIC", ""),
                rr.get("Descripción técnica", "")
            ),
            axis=1
        )

    materiales_df.to_excel(writer, sheet_name="Materiales", index=False, startrow=2)
    ws2 = writer.sheets["Materiales"]

    # título de la hoja
    ws2.merge_cells("A1:J1")
    ws2["A1"] = "CUBICACIÓN DE MATERIALES PARA INSTALACIONES RESIDENCIALES EN CASAS PREFABRICADAS"
    ws2["A1"].font = Font(bold=True, size=14)
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 24

    header_row = 3
    thin = Side(border_style="thin", color="AAAAAA")
    fill_header = PatternFill(start_color="E6F2FF", end_color="E6F2FF", fill_type="solid")
    fill_section = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    ncols2 = materiales_df.shape[1]
    nrows2 = materiales_df.shape[0]
    last_row2 = header_row + nrows2

    gris_fill = PatternFill(
    fill_type="solid",
    start_color="F2F2F2",
    end_color="F2F2F2"
    )

    # busca la sección "Espuma expansiva" y le pinta de gris la fila siguiente,
    # para separar visualmente esa sección del resto (es la última del listado)
    for fila in range(4, last_row2 + 1):
        texto = ws2.cell(row=fila, column=2).value
        if texto and "Espuma expansiva" in str(texto):
            fila_gris = fila + 1
            for columna in range(1, ncols2 + 1):
                ws2.cell(row=fila_gris, column=columna).fill = gris_fill
            break

    # formato del encabezado de la tabla de materiales
    for ccol in range(1, ncols2 + 1):
        cell = ws2.cell(row=header_row, column=ccol)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = fill_header
        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    # recorre cada fila de datos: si es una fila "separadora" de sección
    # (sin Ítem pero con Descripción), la resalta con fondo gris y negrita
    for r in range(header_row + 1, last_row2 + 1):
        is_section = (ws2.cell(row=r, column=1).value in [None, ""]) and (ws2.cell(row=r, column=2).value not in [None, ""])
        max_lineas = 1  # cuántas líneas (separadas por \n) tiene el texto más largo de la fila
        for ccol in range(1, ncols2 + 1):
            cell = ws2.cell(row=r, column=ccol)
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            if ccol in [8, 9, 10]:  # K, Longitud/Unidad y Cantidad
                cell.alignment = Alignment(horizontal="right", vertical="top", wrap_text=True)
            else:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            if ccol == 8 and isinstance(cell.value, (int, float)):  # a la columna K se le ponen 2 decimales
                cell.number_format = "0.00"
            if is_section:
                cell.fill = fill_section
                cell.font = Font(bold=True)
            if isinstance(cell.value, str) and "\n" in cell.value:
                max_lineas = max(max_lineas, cell.value.count("\n") + 1)

        if is_section:
            # las filas de sección ocupan todo el ancho, fusionadas en una sola celda
            ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols2)
            ws2.cell(row=r, column=1).alignment = Alignment(horizontal="left", vertical="center")
            ws2.row_dimensions[r].height = 18
        else:
            # Altura dinámica: ~15 puntos por línea de texto, para que el
            # detalle multi-línea (ferrules, cónicos, etc.) se vea completo
            # sin tener que agrandar la fila a mano en Excel.
            ws2.row_dimensions[r].height = max(15, max_lineas * 15)

    # ancho de columna automático según el contenido más largo
    for ccol in range(1, ncols2 + 1):
        col_letter = get_column_letter(ccol)
        max_len = 0
        for r in range(1, last_row2 + 1):
            v = ws2.cell(row=r, column=ccol).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws2.column_dimensions[col_letter].width = min(max(10, max_len + 4), 55)

    # anchos fijos definidos a mano para las columnas principales (sobreescribe el automático)
    ws2.column_dimensions["A"].width = 6
    ws2.column_dimensions["B"].width = 48
    ws2.column_dimensions["C"].width = 30
    ws2.column_dimensions["D"].width = 12
    ws2.column_dimensions["E"].width = 20
    ws2.column_dimensions["F"].width = 45
    ws2.column_dimensions["G"].width = 8
    ws2.column_dimensions["H"].width = 5
    ws2.column_dimensions["I"].width = 22
    ws2.column_dimensions["J"].width = 10

    ws2.auto_filter.ref = f"A{header_row}:{get_column_letter(ncols2)}{last_row2}"
    ws2.freeze_panes = f"A{header_row+1}"  # deja fija la fila de encabezado cuando se mueve hacia abajo en la pantalla

    # =========================
    # HOJA 3: BASE NORMATIVA + HIPERVÍNCULOS DESDE MATERIALES
    # =========================
    aplicar_base_normativa_e_hipervinculos(
        writer,
        materiales_df,
        sheet_materiales="Materiales",
        sheet_base="Base Normativa"
    )

# al terminar de escribir el Excel, intenta abrirlo automáticamente
# (funciona distinto según el sistema operativo: Windows, Mac o Linux)
try:
    if sys.platform.startswith("win"):
        os.startfile(nombre_archivo)
    elif sys.platform == "darwin":
        os.system(f'open "{nombre_archivo}"')
    else:
        os.system(f'xdg-open "{nombre_archivo}"')
except Exception:
    pass  # si no se pudo abrir automáticamente, no es grave, el archivo ya quedó guardado

print(f"\n Informe generado correctamente: {nombre_archivo}")
